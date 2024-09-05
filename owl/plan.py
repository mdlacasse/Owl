'''

Owl/plan
---

A retirement planner using linear programming optimization.

See companion document for a complete explanation and description
of all variables and parameters.

Copyright -- Martin-D. Lacasse (2024)

Disclaimer: This program comes with no guarantee. Use at your own risk.
'''

###########################################################################
import numpy as np
from datetime import date

from owl import utils as u
from owl import tax2024 as tx
from owl import rates
from owl import timelists
from owl import abcapi as abc


def setVerbose(self, state=True):
    '''
    Control verbosity of calculations. True or False for now.
    Return previous state.
    '''
    return u.setVerbose(state)


def _gamma_n(tau, Nn):
    '''
    Utility function to generate inflation multiplier.
    Return time series of cumulative inflation multiplier
    at year n with respect to the current year.
    '''
    gamma = np.ones(Nn)

    for n in range(1, Nn):
        gamma[n] = gamma[n - 1] * (1 + tau[3, n - 1])

    return gamma


def _xi_n(profile, frac, n_d, N_n, a=15, b=12):
    '''
    Utility function to generate spending profile.
    Return time series of spending profile.
    Value is reduced by frac starting in year n_d,
    after the passing of shortest-lived spouse.
    Series is unadjusted for inflation.
    '''
    xi = np.ones(N_n)
    if profile == 'flat':
        if n_d < N_n:
            xi[n_d:] *= frac
    elif profile == 'smile':
        x = np.linspace(0, N_n - 1, N_n)
        a /= 100
        b /= 100
        # Use a cosine +/- 15% combined with a gentle +12% linear increase.
        xi = xi + a * np.cos((2 * np.pi / (N_n - 1)) * x) + (b / (N_n - 1)) * x
        # Normalize to be sum-neutral with respect to a flat profile.
        neutralSum = N_n
        # Reduce income needs after passing of one spouse.
        if n_d < N_n:
            neutralSum -= (1 - frac) * (N_n - n_d)  # Account for flat spousal reduction.
            xi[n_d:] *= frac
        xi *= neutralSum / xi.sum()
    else:
        u.xprint('Unknown profile', profile)

    return xi


def _qC(C, N1, N2=1, N3=1, N4=1):
    '''
    Index range accumulator.
    '''
    return C + N1 * N2 * N3 * N4


def _q1(C, l1, N1=None):
    '''
    Index mapping function. 1 argument.
    '''
    return C + l1


def _q2(C, l1, l2, N1, N2):
    '''
    Index mapping function. 2 arguments.
    '''
    return C + l1 * N2 + l2


def _q3(C, l1, l2, l3, N1, N2, N3):
    '''
    Index mapping function. 3 arguments.
    '''
    return C + l1 * N2 * N3 + l2 * N3 + l3


def _q4(C, l1, l2, l3, l4, N1, N2, N3, N4):
    '''
    Index mapping function. 4 arguments.
    '''
    return C + l1 * N2 * N3 * N4 + l2 * N3 * N4 + l3 * N4 + l4


############################################################################


class Plan:
    '''
    This is the main class of the Owl Project.
    '''

    def __init__(self, yobs, expectancy, name):
        '''
        Constructor requires two lists: the first one is
        the year of birth of each spouse, and the second
        the life expectancy. Last argument is a name for
        the plan.
        '''
        self._name = name

        # 7 tax brackets, 3 types of accounts, 4 classes of assets.
        self.N_t = 7
        self.N_j = 3
        self.N_k = 4
        # 3 binary variables.
        self.N_z = 3

        # Default interpolation parameters for allocation ratios.
        self.interpMethod = 'linear'
        self._interpolator = self._linInterp
        self.interpCenter = 15
        self.interpWidth = 5

        self.defaultPlots = 'nominal'
        self.solver = 'HiGHS'

        self.N_i = len(yobs)
        assert 0 < self.N_i and self.N_i <= 2, 'Cannot support %d individuals.' % self.N_i
        assert self.N_i == len(expectancy), 'Expectancy must have %d entries.' % self.N_i

        self.filingStatus = ['single', 'married'][self.N_i - 1]

        self.yobs = yobs
        self.expectancy = expectancy

        thisyear = date.today().year
        self.horizons = [yobs[i] + expectancy[i] - thisyear + 1 for i in range(self.N_i)]
        self.N_n = max(self.horizons)
        self.year_n = np.linspace(thisyear, thisyear + self.N_n - 1, self.N_n, dtype=int)
        # Handle passing of one spouse before the other.
        if self.N_i == 2:
            self.n_d = min(self.horizons)
            self.i_d = self.horizons.index(self.n_d)
            self.i_s = (self.i_d + 1) % 2
        else:
            self.n_d = self.N_n + 2  # Push beyond upper bound.
            self.i_d = 0
            self.i_s = -1

        # Default parameters:
        self.psi = 0.15  # Long-term income tax rate (decimal)
        self.chi = 0.6  # Survivor fraction
        self.mu = 0.02  # Dividend rate (decimal)
        self.nu = 0.30  # Heirs tax rate (decimal)
        self.eta = 0    # Spousal deposit/withdrawal ratio
        self.phi_j = [1, 1, 1]  # Fraction left to other spouse at death

        # Placeholder for before reading contributions file.
        self.inames = ['Person 1', 'Person 2']

        # Default to zero pension and social security.
        self.pi_in = np.zeros((self.N_i, self.N_n))
        self.zeta_in = np.zeros((self.N_i, self.N_n))

        # Parameters from timeLists.
        self.omega_in = np.zeros((self.N_i, self.N_n))
        self.Lambda_in = np.zeros((self.N_i, self.N_n))
        self.myRothX_in = np.zeros((self.N_i, self.N_n))
        self.kappa_ijn = np.zeros((self.N_i, self.N_j, self.N_n))

        u.vprint(
            'Preparing scenario of %d years for %d individual%s.' % (self.N_n - 1, self.N_i, ['', 's'][self.N_i - 1])
        )
        for i in range(self.N_i):
            u.vprint('%s: life horizon from %d -> %d.' % (self.inames[i], thisyear, thisyear + self.horizons[i] - 1))

        u.vprint('Name of individual(s) will be read with readContributions(file).')

        # Prepare income tax and RMD time series.
        self.rho_in = tx.rho_in(self.yobs, self.N_n)
        self.sigma_n, self.theta_tn, self.Delta_tn = tx.taxParams(self.yobs, self.i_d, self.n_d, self.N_n)

        self._adjustedParameters = False
        self._caseStatus = 'unsolved'
        self._buildOffsetMap()
        self.timeListsFileName = None

        return

    def setSolver(self, solver):
        '''
        Select solver.
        '''
        solvers = ['HiGHS', 'MOSEK']
        assert solver in solvers, 'Solver %s not supported.'%solver
        self.solver = solver

    def _checkSolverStatus(self, funcName):
        '''
        Check if problem was solved successfully.
        '''
        if self._caseStatus == 'solved':
            return False

        u.vprint('Preventing to run %s() while solution is %s.' % (funcName, self._caseStatus))

        return True

    def setName(self, name):
        '''
        Override name of the plan. Name is used
        to distinguish graph outputs.
        '''
        self._name = name

        return

    def setSpousalFraction(self, eta, bound='>'):
        '''
        Set spousal deposit and withdrawal fraction. Default 0.
        A fraction x is given then
            d0/(d0 + d1) >= x,
        or
            d0/(d0 + d1) <= x, 
        depending on the bound '>' or '<' provided.
        Here d0 is the deposit amount into the taxable account
        of first spouse while d1 is for the second spouse.
        If x is 1, and bound is '>', all surplus get deposited
        into the taxable account of the first spouse. If x is 0, and bound is '<',
        then all surplus get deposited into the taxable account of the second spouse.
        '''
        assert 0 <= eta and eta <= 1, 'Fraction must be between 0 and 1.'
        u.vprint('Setting spousal fraction to %.1f.'%eta)
        assert bound in ['>', '<'], 'Bound must be one of "<" or ">".'
        if bound == '>':
            self.eta = eta
        else:
            self.eta = -eta

        return

    def _checkValue(self, value):
        '''
        Short utility function to parse and check arguments for plotting.
        '''
        if value is None:
            return self.defaultPlots

        opts = ['nominal', 'today']
        if value in opts:
            return value

        u.xprint('Value type must be one of:', *opts)

        return

    def setDefaultPlots(self, value):
        '''
        Set plots between nominal values or today's $.
        '''

        self.defaultPlots = self._checkValue(value)

        return

    def setDividendRate(self, mu):
        '''
        Set dividend rate on equities. Rate is in percent. Default 2%.
        '''
        assert 0 <= mu and mu <= 100, 'Rate must be between 0 and 100.'
        mu /= 100
        u.vprint('Dividend return rate on equities set to', u.pc(mu, f=1))
        self.mu = mu
        self._caseStatus = 'modified'

        return

    def setLongTermCapitalTaxRate(self, psi):
        '''
        Set long-term income tax rate. Rate is in percent. Default 15%.
        '''
        assert 0 <= psi and psi <= 100, 'Rate must be between 0 and 100.'
        psi /= 100
        u.vprint('Long-term income tax set to', u.pc(psi, f=0))
        self.psi = psi
        self._caseStatus = 'modified'

        return

    def setBeneficiaryFraction(self, phi):
        '''
        Set fractions of savings accounts that is left to surviving spouse.
        Default is [1, 1, 1] for taxable, tax-deferred, adn tax-exempt accounts.
        '''
        assert len(phi) == self.N_j, 'Fractions must have %d entries.' % self.N_j
        for j in range(self.N_j):
            assert 0 <= phi[j] <= 1, 'Fractions must be between 0 and 1.'

        u.vprint('Beneficiary spousal beneficiary fractions set to', phi)
        self.phi_j = phi
        self._caseStatus = 'modified'

        return

    def setHeirsTaxRate(self, nu):
        '''
        Set the heirs tax rate on the tax-deferred portion of the estate.
        Rate is in percent. Default is 30%.
        '''
        assert 0 <= nu and nu <= 100, 'Rate must be between 0 and 100.'
        nu /= 100
        u.vprint('Heirs tax rate on tax-deferred portion of estate set to %s.'%u.pc(nu, f=0))
        self.nu = nu
        self._caseStatus = 'modified'

        return

    def setPension(self, amounts, ages, units='k'):
        '''
        Set value of pension for each individual and commencement age.
        Units are in $k, unless specified otherwise: 'k', 'M', or '1'.
        '''
        assert len(amounts) == self.N_i, 'Amounts must have %d entries.' % self.N_i
        assert len(ages) == self.N_i, 'Ages must have %d entries.' % self.N_i

        fac = u.getUnits(units)
        u.rescale(amounts, fac)

        u.vprint(
            'Setting pension of',
            [u.d(amounts[i]) for i in range(self.N_i)],
            'at age(s)',
            ages
        )

        thisyear = date.today().year
        # Use zero array freshly initialized.
        self.pi_in = np.zeros((self.N_i, self.N_n))
        for i in range(self.N_i):
            if amounts[i] != 0:
                ns = max(0, self.yobs[i] + ages[i] - thisyear)
                nd = self.horizons[i]
                self.pi_in[i, ns:nd] = amounts[i]

        self._caseStatus = 'modified'

        return

    def setSocialSecurity(self, amounts, ages, units='k'):
        '''
        Set value of social security for each individual and commencement age.
        Units are in $k, unless specified otherwise: 'k', 'M', or '1'.
        '''
        assert len(amounts) == self.N_i, 'Amounts must have %d entries.' % self.N_i
        assert len(ages) == self.N_i, 'Ages must have %d entries.' % self.N_i

        fac = u.getUnits(units)
        u.rescale(amounts, fac)

        u.vprint(
            'Setting social security benefits of',
            [u.d(amounts[i]) for i in range(self.N_i)],
            'at age(s)',
            ages,
        )

        self._adjustedParameters = False
        thisyear = date.today().year
        self.zeta_in = np.zeros((self.N_i, self.N_n))
        for i in range(self.N_i):
            ns = max(0, self.yobs[i] + ages[i] - thisyear)
            nd = self.horizons[i]
            self.zeta_in[i, ns:nd] = amounts[i]

        if self.N_i == 2:
            # Approximate calculation for spousal benefit (only valid at FRA).
            self.zeta_in[self.i_s, self.n_d :] = max(amounts[self.i_s], amounts[self.i_d] / 2)

        self._caseStatus = 'modified'

        return

    def setSpendingProfile(self, profile, percent=60):
        '''
        Generate time series for spending profile.
        Surviving spouse fraction can be specified
        as a second argument. Default value is 60%.
        '''
        self.chi = percent / 100

        u.vprint('Setting', profile, 'spending profile.')
        if self.N_i == 2:
            u.vprint('Using ', u.pc(self.chi, f=0), 'spending for surviving spouse.')

        self.xi_n = _xi_n(profile, self.chi, self.n_d, self.N_n)
        self.spendingProfile = profile
        self._caseStatus = 'modified'

        return

    def setRates(self, method, frm=None, to=None, values=None):
        '''
        Generate rates for return and inflation based on the method and
        years selected. Note that last bound is included.

        The following methods are available:
        default, fixed, realistic, conservative, average, stochastic,
        and historical.

        For 'fixed', rate values must be provided.
        For 'average', 'stochastic', and 'historical', a starting year
        must be provided.

        Valid year range is from 1928 to last year.
        '''
        if frm != None and to == None:
            to = frm + self.N_n - 1  # 'to' is inclusive: subtract 1

        dr = rates.rates()
        self.rateValues = dr.setMethod(method, frm, to, values)
        self.rateMethod = method
        self.rateFrm = frm
        self.rateTo = to
        self.tau_kn = dr.genSeries(self.N_n).transpose()
        u.vprint(
            'Generating rate series of',
            len(self.tau_kn[0]),
            'years using',
            method,
            'method.',
        )

        # Once rates are selected, (re)build cumulative inflation multipliers.
        self.gamma_n = _gamma_n(self.tau_kn, self.N_n + 1)
        self._adjustedParameters = False
        self._caseStatus = 'modified'

        return

    def _adjustParameters(self):
        '''
        Adjust parameters that follow inflation or allocations.
        '''
        if self._adjustedParameters == False:
            u.vprint('Adjusting parameters for inflation.')
            self.DeltaBar_tn = self.Delta_tn * self.gamma_n[:-1]
            self.zetaBar_in = self.zeta_in * self.gamma_n[:-1]
            self.sigmaBar_n = self.sigma_n * self.gamma_n[:-1]
            self.xiBar_n = self.xi_n * self.gamma_n[:-1]

            self._adjustedParameters = True

        return

    def setAccountBalances(self, *, taxable, taxDeferred, taxFree, units='k'):
        '''
        Three lists containing the balance of all assets in each category for
        each spouse.  For single individuals, these lists will contain only
        one entry. Units are in $k, unless specified otherwise: 'k', 'M', or '1'.
        '''
        assert len(taxable) == self.N_i, 'taxable must have %d entries.' % self.N_i
        assert len(taxDeferred) == self.N_i, 'taxDeferred must have %d entries.' % self.N_i
        assert len(taxFree) == self.N_i, 'taxFree must have %d entries.' % self.N_i

        fac = u.getUnits(units)
        u.rescale(taxable, fac)
        u.rescale(taxDeferred, fac)
        u.rescale(taxFree, fac)

        self.b_ji = np.zeros((self.N_j, self.N_i))
        self.b_ji[0][:] = taxable
        self.b_ji[1][:] = taxDeferred
        self.b_ji[2][:] = taxFree
        self.beta_ij = self.b_ji.transpose()
        self._caseStatus = 'modified'

        u.vprint('Taxable balances:', *[u.d(taxable[i]) for i in range(self.N_i)])
        u.vprint('Tax-deferred balances:', *[u.d(taxDeferred[i]) for i in range(self.N_i)])
        u.vprint('Tax-free balances:', *[u.d(taxFree[i]) for i in range(self.N_i)])
        u.vprint(
            'Total post-tax wealth of approximately',
            u.d(np.sum(taxable) + 0.7 * np.sum(taxDeferred) + np.sum(taxFree)),
        )

        return

    def setInterpolationMethod(self, method, center=15, width=5):
        '''
        Interpolate assets allocation ratios from initial value (today) to
        final value (at the end of horizon).

        Two interpolation methods are supported: linear and s-curve.
        Linear is a straight line between now and the end of the simulation.
        Hyperbolic tangent give a smooth "S" curve centered at point "center"
        with a width "width". Center point defaults to 15 years and width to
        5 years. This means that the transition from initial to final
        will start occuring in 10 years (15-5) and will end in 20 years (15+5).
        '''
        if method == 'linear':
            self._interpolator = self._linInterp
        elif method == 's-curve':
            self._interpolator = self._tanhInterp
            self.interpCenter = center
            self.interpWidth = width
        else:
            u.xprint('Method', method, 'not supported.')

        self.interpMethod = method
        self._caseStatus = 'modified'

        u.vprint('Asset allocation interpolation method set to', method)

        return

    def setAllocationRatios(self, allocType, taxable=None, taxDeferred=None, taxFree=None, generic=None):
        '''
        Single function for setting all types of asset allocations.
        Allocation types are 'account', 'individual', and 'spouses'.

        For 'account' the three different account types taxable, taxDeferred,
        and taxFree need to be set to a list. For spouses,
        taxable = [[[ko00, ko01, ko02, ko03], [kf00, kf01, kf02, kf02]],
                   [[ko10, ko11, ko12, ko13], [kf10, kf11, kf12, kf12]]]
        where ko is the initial allocation while kf is the final.
        The order of [initial, final] pairs is the same as for the birth
        years and longevity provided. Single only provide one pair for each
        type of savings account.

        For the 'individual' allocation type, only one generic list needs
        to be provided:
        generic = [[[ko00, ko01, ko02, ko03], [kf00, kf01, kf02, kf02]],
                  [[ko10, ko11, ko12, ko13], [kf10, kf11, kf12, kf12]]].
        while for 'spouses' only one pair needs to be given as follows:
        generic = [[ko00, ko01, ko02, ko03], [kf00, kf01, kf02, kf02]]
        as assets are coordinated between accounts and spouses.
        '''
        self.alpha_ijkn = np.zeros((self.N_i, self.N_j, self.N_k, self.N_n + 1))
        if allocType == 'account':
            # Make sure we have proper input.
            for item in [taxable, taxDeferred, taxFree]:
                assert len(item) == self.N_i, '%s must one entry per individual.' % (item)
                for i in range(self.N_i):
                    # Initial and final.
                    assert len(item[i]) == 2, '%s[%d] must have 2 lists (initial and final).' % (
                        item,
                        i,
                    )
                    for z in range(2):
                        assert len(item[i][z]) == self.N_k, '%s[%d][%d] must have %d entries.' % (
                            item,
                            i,
                            z,
                            self.N_k,
                        )
                        assert abs(sum(item[i][z]) - 100) < 0.01, 'Sum of percentages must add to 100.'

            for i in range(self.N_i):
                u.vprint(
                    self.inames[i],
                    ': Setting gliding allocation ratios (%) to',
                    allocType,
                )
                u.vprint('      taxable:', taxable[i][0], '->', taxable[i][1])
                u.vprint('  taxDeferred:', taxDeferred[i][0], '->', taxDeferred[i][1])
                u.vprint('      taxFree:', taxFree[i][0], '->', taxFree[i][1])

            # Order in alpha is j, i, 0/1, k.
            alpha = {}
            alpha[0] = np.array(taxable)
            alpha[1] = np.array(taxDeferred)
            alpha[2] = np.array(taxFree)
            for i in range(self.N_i):
                Nin = self.horizons[i] + 1
                for j in range(self.N_j):
                    for k in range(self.N_k):
                        start = alpha[j][i, 0, k] / 100
                        end = alpha[j][i, 1, k] / 100
                        dat = self._interpolator(start, end, Nin)
                        self.alpha_ijkn[i, j, k, :Nin] = dat[:]

        elif allocType == 'individual':
            assert len(generic) == self.N_i, 'generic must have one list per individual.'
            for i in range(self.N_i):
                # Initial and final.
                assert len(generic[i]) == 2, 'generic[%d] must have 2 lists (initial and final).' % i
                for z in range(2):
                    assert len(generic[i][z]) == self.N_k, 'generic[%d][%d] must have %d entries.' % (
                        i,
                        z,
                        self.N_k,
                    )
                    assert abs(sum(generic[i][z]) - 100) < 0.01, 'Sum of percentages must add to 100.'

            for i in range(self.N_i):
                u.vprint(
                    self.inames[i],
                    ': Setting gliding allocation ratios (%) to',
                    allocType,
                )
                u.vprint('\t', generic[i][0], '->', generic[i][1])

            for i in range(self.N_i):
                Nin = self.horizons[i] + 1
                for k in range(self.N_k):
                    start = generic[i][0][k] / 100
                    end = generic[i][1][k] / 100
                    dat = self._interpolator(start, end, Nin)
                    for j in range(self.N_j):
                        self.alpha_ijkn[i, j, k, :Nin] = dat[:]

        elif allocType == 'spouses':
            assert len(generic) == 2, 'generic must have 2 entries (initial and final).'
            for z in range(2):
                assert len(generic[z]) == self.N_k, 'generic[%d] must have %d entries.' % (
                    z,
                    self.N_k,
                )
                assert abs(sum(generic[z]) - 100) < 0.01, 'Sum of percentages must add to 100.'

            u.vprint('Setting gliding allocation ratios (%) to', allocType)
            u.vprint('\t', generic[0], '->', generic[1])

            # Use longest-lived spouse for both time scales.
            Nxn = max(self.horizons) + 1

            for k in range(self.N_k):
                start = generic[0][k] / 100
                end = generic[1][k] / 100
                dat = self._interpolator(start, end, Nxn)
                for i in range(self.N_i):
                    for j in range(self.N_j):
                        self.alpha_ijkn[i, j, k, :Nxn] = dat[:]

        self.ARCoord = allocType
        self._caseStatus = 'modified'

        u.vprint('Interpolating assets allocation ratios using', self.interpMethod, 'method.')

        return

    def _linInterp(self, a, b, numPoints):
        '''
        Utility function to interpolate allocations using
        a linear interpolation.
        '''
        # num goes one more year as endpoint=True.
        dat = np.linspace(a, b, numPoints)

        return dat

    def _tanhInterp(self, a, b, numPoints):
        '''
        Utility function to interpolate allocations using a hyperbolic
        tangent interpolation. "c" is the year where the inflection point
        is happening, and "w" is the width of the transition.
        '''
        c = self.interpCenter
        w = self.interpWidth
        t = np.linspace(0, numPoints, numPoints)
        # Solve 2x2 system to match end points exactly.
        th0 = np.tanh((t[0] - c) / w)
        thN = np.tanh((t[numPoints - 1] - c) / w)
        k11 = 0.5 - 0.5 * th0
        k21 = 0.5 - 0.5 * thN
        k12 = 0.5 + 0.5 * th0
        k22 = 0.5 + 0.5 * thN
        _b = (b - (k21 / k11) * a) / (k22 - (k21 / k11) * k12)
        _a = (a - k12 * _b) / k11
        dat = _a + 0.5 * (_b - _a) * (1 + np.tanh((t - c) / w))

        return dat

    def readContributions(self, filename):
        '''
        Provide the name of the file containing the financial events
        over the anticipated life span determined by the
        assumed longevity. File can be an excel, or odt file with one
        tab named after each spouse and must have the following
        column headers:

                'year',
                'anticipated wages',
                'ctrb taxable',
                'ctrb 401k',
                'ctrb Roth 401k',
                'ctrb IRA',
                'ctrb Roth IRA',
                'Roth X',
                'big ticket items'

        in any order. A template is provided as an example.
        Missing rows (years) are populated with zero values.
        '''
        self.inames, self.timeLists = timelists.read(filename, self.N_i, self.horizons)

        timelists.check(self.inames, self.timeLists, self.horizons)
        self.timeListsFileName = filename

        # Now fill in parameters.
        for i in range(self.N_i):
            h = self.horizons[i]
            self.omega_in[i, :h] = self.timeLists[i]['anticipated wages'][:h]
            self.Lambda_in[i, :h] = self.timeLists[i]['big ticket items'][:h]
            self.myRothX_in[i, :h] = self.timeLists[i]['Roth X'][:h]
            self.kappa_ijn[i, 0, :h] = self.timeLists[i]['ctrb taxable'][:h]
            self.kappa_ijn[i, 1, :h] = self.timeLists[i]['ctrb 401k'][:h]
            self.kappa_ijn[i, 1, :h] += self.timeLists[i]['ctrb IRA'][:h]
            self.kappa_ijn[i, 2, :h] = self.timeLists[i]['ctrb Roth 401k'][:h]
            self.kappa_ijn[i, 2, :h] += self.timeLists[i]['ctrb Roth IRA'][:h]

        self._caseStatus = 'modified'

        return

    def _buildOffsetMap(self):
        '''
        Utility function to map variables to block vector.
        Refer to companion document for explanations.
        '''
        # Stack variables in block vector.
        C = {}
        C['b'] = 0
        C['d'] = _qC(C['b'], self.N_i, self.N_j, self.N_n + 1)
        C['f'] = _qC(C['d'], self.N_i, self.N_n)
        C['g'] = _qC(C['f'], self.N_t, self.N_n)
        C['w'] = _qC(C['g'], self.N_n)
        C['x'] = _qC(C['w'], self.N_i, self.N_j, self.N_n)
        C['z'] = _qC(C['x'], self.N_i, self.N_n)
        C['Z'] = _qC(C['z'], self.N_i, self.N_n, self.N_z)
        self.nvars = _qC(C['Z'], 1)

        self.C = C
        u.vprint(
            'Problem has',
            len(C),
            'distinct time series forming',
            self.nvars,
            'decision variables.',
        )

        return

    def _buildConstraints(self, objective, options):
        '''
        Utility function that builds constraint matrix and vectors.
        Refer to companion document for notation and detailed explanations.
        '''
        # Bounds values.
        zero = 0
        inf = np.inf

        # Simplified notation.
        Ni = self.N_i
        Nj = self.N_j
        Nk = self.N_k
        Nn = self.N_n
        Nt = self.N_t
        Nz = self.N_z
        i_d = self.i_d
        i_s = self.i_s
        n_d = self.n_d

        Cb = self.C['b']
        Cd = self.C['d']
        Cf = self.C['f']
        Cg = self.C['g']
        Cw = self.C['w']
        Cx = self.C['x']
        Cz = self.C['z']
        CZ = self.C['Z']

        tau_ijn = np.zeros((Ni, Nj, Nn))
        for i in range(Ni):
            for j in range(Nj):
                for n in range(Nn):
                    tau_ijn[i, j, n] = np.sum(self.alpha_ijkn[i, j, :, n] * self.tau_kn[:, n], axis=0)

        # Weights are normalized on k. [alpha*(1 + tau) = 1 + alpha*tau)].
        Tau1_ijn = 1 + tau_ijn
        Tauh_ijn = 1 + tau_ijn / 2

        if 'units' in options:
            units = u.getUnits(options['units'])
        else:
            units = 1000

        bigM = 5e6
        if 'bigM' in options:
            bigM = units * options['bigM']

        ###################################################################
        # Inequality constraint matrix with upper and lower bound vectors.
        A = abc.ConstraintMatrix(self.nvars)
        B = abc.Bounds(self.nvars)

        # RMDs inequalities.
        for i in range(Ni):
            for n in range(self.horizons[i]):
                rowDic = {
                    _q3(Cw, i, 1, n, Ni, Nj, Nn): 1,
                    _q3(Cb, i, 1, n, Ni, Nj, Nn + 1): -self.rho_in[i, n],
                }
                A.addNewRow(rowDic, zero, inf)

        # Income tax bracket range inequalities.
        for n in range(Nn):
            for t in range(Nt):
                B.set0_Ub(_q2(Cf, t, n, Nt, Nn), self.DeltaBar_tn[t, n])

        # Roth conversions equalities/inequalities.
        if 'maxRothConversion' in options:
            if options['maxRothConversion'] == 'file':
                # u.vprint('Fixing Roth conversions to those from file %s.' % self.timeListsFileName)
                for i in range(Ni):
                    for n in range(self.horizons[i]):
                        rhs = self.myRothX_in[i][n]
                        B.setRange(_q2(Cx, i, n, Ni, Nn), rhs, rhs)
            else:
                rhsopt = options['maxRothConversion']
                assert isinstance(rhsopt, (int, float)) == True, 'Specified maxConversion is not a number.'
                rhsopt *= units
                if rhsopt < 0:
                    # u.vprint('Unlimited Roth conversions (<0)')
                    pass
                else:
                    # u.vprint('Limiting Roth conversions to:', u.d(rhsopt))
                    for i in range(Ni):
                        for n in range(self.horizons[i]):
                            #  Adjust cap with inflation?
                            B.set0_Ub(_q2(Cx, i, n, Ni, Nn), rhsopt)

        if Ni == 2:
            # No activity for i_d after year of passing.
            for n in range(n_d, Nn):
                B.set0_Ub(_q2(Cd, i_d, n, Ni, Nn), zero)
                for j in range(Nj):
                    B.set0_Ub(_q3(Cw, i_d, j, n, Ni, Nj, Nn), zero)
            # Process noRothConversions option.
            if 'noRothConversions' in options:
                rhsopt = options['noRothConversions']
                try:
                    i_x = self.inames.index(rhsopt)
                except ValueError:
                    u.xprint('Unknown individual for noRothConversions:', rhsopt)

                for n in range(Nn):
                    B.set0_Ub(_q2(Cx, i_x, n, Ni, Nn), zero)


        # Deposits in taxable account during last year are a tax loophole.
        for i in range(Ni):
            B.set0_Ub(_q2(Cd, i, Nn-1, Ni, Nn), zero)

        ###################################################################
        # Equalities.

        if objective == 'maxSpending':
            # if 'netSpending' in options:
            #   u.vprint('Ignoring netSpending option provided.')
            # Impose requested constraint on estate, if any.
            if 'estate' in options:
                estate = options['estate']
                assert isinstance(estate, (int, float)) == True, 'Desired estate provided not a number.'
                estate *= units * self.gamma_n[-1]
            else:
                # If not specified, default to $1.
                estate = 1

            row = A.newRow()
            for i in range(Ni):
                row.addElem(_q3(Cb, i, 0, Nn, Ni, Nj, Nn + 1), 1)
                row.addElem(_q3(Cb, i, 1, Nn, Ni, Nj, Nn + 1), 1 - self.nu)
                # Nudge could be added (e.g. 1.02) to artificially favor tax-exempt account
                # as heirs's benefits of 10y tax-free is not weighted in?
                row.addElem(_q3(Cb, i, 2, Nn, Ni, Nj, Nn + 1), 1)
            A.addRow(row, estate, estate)
            # u.vprint('Adding estate constraint of:', u.d(estate))
        elif objective == 'maxBequest':
            # if 'estate' in options:
            #   u.vprint('Ignoring estate option provided.')
            spending = options['netSpending']
            assert isinstance(spending, (int, float)) == True, 'Desired spending provided not a number.'
            spending *= units
            # u.vprint('Maximizing bequest with desired net spending of:', u.d(spending))
            A.addNewRow({_q1(Cg, 0): 1}, spending, spending)
        else:
            u.xprint('Unknown objective function:', objective)

        # Set initial balances through bounds or constraints.
        for i in range(Ni):
            for j in range(Nj):
                rhs = self.beta_ij[i, j]
                A.addNewRow({_q3(Cb, i, j, 0, Ni, Nj, Nn + 1): 1}, rhs, rhs)

        # Account balances carried from year to year.
        # Considering spousal asset transfer.
        # Using hybrid approach with 'if' statement and Kronecker deltas.
        for i in range(Ni):
            for j in range(Nj):
                for n in range(Nn):
                    fac1 = 1 - u.krond(n, n_d - 1) * u.krond(i, i_d)
                    rhs = fac1 * self.kappa_ijn[i, j, n] * Tauh_ijn[i, j, n]

                    row = A.newRow()
                    row.addElem(_q3(Cb, i, j, n + 1, Ni, Nj, Nn + 1), 1)
                    row.addElem(_q3(Cb, i, j, n, Ni, Nj, Nn + 1), -fac1 * Tau1_ijn[i, j, n])
                    row.addElem(_q2(Cx, i, n, Ni, Nn), -fac1 * (u.krond(j, 2) - u.krond(j, 1)) * Tauh_ijn[i, j, n])
                    row.addElem(_q3(Cw, i, j, n, Ni, Nj, Nn), fac1 * Tau1_ijn[i, j, n])
                    row.addElem(_q2(Cd, i, n, Ni, Nn), -fac1 * u.krond(j, 0) * Tau1_ijn[i, j, n])

                    if Ni == 2 and i == i_s and n == n_d - 1:
                        fac2 = self.phi_j[j]
                        rhs += fac2 * self.kappa_ijn[i_d, j, n] * Tauh_ijn[i_d, j, n]
                        row.addElem(_q3(Cb, i_d, j, n, Ni, Nj, Nn + 1), -fac2 * Tau1_ijn[i_d, j, n])
                        row.addElem(_q2(Cx, i_d, n, Ni, Nn), -fac2 * (u.krond(j, 2) - u.krond(j, 1)) * Tauh_ijn[i_d, j, n])
                        row.addElem(_q3(Cw, i_d, j, n, Ni, Nj, Nn), fac2 * Tau1_ijn[i_d, j, n])
                        # row.addElem(_q2(Cd, i_d, n, Ni, Nn), -fac2 * u.krond(j, 0))
                    A.addRow(row, rhs, rhs)

        tau_0prev = np.roll(self.tau_kn[0, :], 1)
        tau_0prev[tau_0prev < 0] = 0

        # Net cash flow.
        for n in range(Nn):
            rhs = -self.M_n[n]
            row = A.newRow({_q1(Cg, n, Nn): 1})
            for i in range(Ni):
                fac = self.psi * self.alpha_ijkn[i, 0, 0, n]
                rhs += (
                    self.omega_in[i, n]
                    + self.zetaBar_in[i, n]
                    + self.pi_in[i, n]
                    + self.Lambda_in[i, n]
                    - 0.5 * fac * self.mu * self.kappa_ijn[i, 0, n]
                )

                row.addElem(_q3(Cb, i, 0, n, Ni, Nj, Nn + 1), fac * self.mu)
                row.addElem(_q2(Cd, i, n, Ni, Nn), fac * self.mu + 1)
                # Minus capital gains on taxable withdrawals using last year's rate if >=0.
                # Plus taxable account withdrawals, and all other withdrawals.
                row.addElem(_q3(Cw, i, 0, n, Ni, Nj, Nn), fac * (tau_0prev[n] - self.mu) - 1)
                row.addElem(_q3(Cw, i, 1, n, Ni, Nj, Nn), -1)
                row.addElem(_q3(Cw, i, 2, n, Ni, Nj, Nn), -1)

            # Minus tax on ordinary income. Tn
            for t in range(Nt):
                row.addElem(_q2(Cf, t, n, Nt, Nn), self.theta_tn[t, n])
            A.addRow(row, rhs, rhs)

        # Impose income profile.
        for n in range(1, Nn):
            rowDic = {_q1(Cg, 0, Nn): -self.xiBar_n[n], _q1(Cg, n, Nn): self.xiBar_n[0]}
            A.addNewRow(rowDic, zero, zero)

        # Balance deposits and withdrawals while both spouses are alive.
        if Ni == 2:
            pm1 = np.sign(self.eta)
            for n in range(n_d - 1):
                rowDic = {_q2(Cd, 0, n, Ni, Nn): (pm1 - self.eta), _q2(Cd, 1, n, Ni, Nn): -self.eta}
                A.addNewRow(rowDic, zero, inf)
                '''
                rowDic = {_q3(Cw, 0, 1, n, Ni, Nj, Nn): (pm1 - self.eta), 
                          _q3(Cw, 1, 1, n, Ni, Nj, Nn): -self.eta
                         }
                rhs = self.rho_in[0, n]*(pm1 - self.eta) - self.rho_in[1, n]*self.eta
                A.addNewRow(rowDic, rhs, inf)
                '''

        # Impose account balance max on all withdrawals. This helps converging.
        for i in range(Ni):
            for j in range(Nj):
                for n in range(Nn):
                    rowDic = {_q3(Cw, i, j, n, Ni, Nj, Nn): -1, _q3(Cb, i, j, n, Ni, Nj, Nn + 1): 1}
                    A.addNewRow(rowDic, zero, inf)

        # Taxable ordinary income.
        for n in range(Nn):
            row = A.newRow()
            rhs = -self.sigmaBar_n[n]
            for i in range(Ni):
                rhs += self.omega_in[i, n] + 0.85 * self.zetaBar_in[i, n] + self.pi_in[i, n]
                row.addElem(_q3(Cw, i, 1, n, Ni, Nj, Nn), -1)
                row.addElem(_q2(Cx, i, n, Ni, Nn), -1)

                # Returns on securities in taxable account.
                fak = np.sum(self.tau_kn[1:Nk, n] * self.alpha_ijkn[i, 0, 1:Nk, n], axis=0)
                rhs += 0.5 * fak * self.kappa_ijn[i, 0, n]
                row.addElem(_q3(Cb, i, 0, n, Ni, Nj, Nn + 1), -fak)
                row.addElem(_q3(Cw, i, 0, n, Ni, Nj, Nn), fak)
                row.addElem(_q2(Cd, i, n, Ni, Nn), -fak)

            for t in range(Nt):
                row.addElem(_q2(Cf, t, n, Nt, Nn), 1)
            A.addRow(row, rhs, rhs)

        # Configure all binary variables.
        for i in range(Ni):
            for n in range(Nn):
                for z in range(Nz):
                    B.setBinary(_q3(Cz, i, n, z, Ni, Nn, Nz))

        B.setBinary(_q1(CZ, 0, 1))

        # Exclude simultaneous deposits and withdrawals from taxable or tax-free accounts.
        for i in range(Ni):
            for n in range(Nn):
                A.addNewRow(
                    {_q3(Cz, i, n, 0, Ni, Nn, Nz): bigM, _q2(Cd, i, n, Ni, Nn): -1},
                    zero,
                    bigM,
                )

                A.addNewRow(
                    {_q3(Cz, i, n, 0, Ni, Nn, Nz): bigM, _q3(Cw, i, 0, n, Ni, Nj, Nn): 1},
                    zero,
                    bigM,
                )

                A.addNewRow(
                    {_q3(Cz, i, n, 1, Ni, Nn, Nz): bigM, _q2(Cd, i, n, Ni, Nn): -1},
                    zero,
                    bigM,
                )

                A.addNewRow(
                    {_q3(Cz, i, n, 1, Ni, Nn, Nz): bigM, _q3(Cw, i, 2, n, Ni, Nj, Nn): 1},
                    zero,
                    bigM
                )

        # Exclude simultaneous Roth conversions and tax-exempt withdrawals.
        for i in range(Ni):
            for n in range(Nn):
                A.addNewRow(
                    {_q3(Cz, i, n, 2, Ni, Nn, Nz): bigM, _q2(Cx, i, n, Ni, Nn): -1},
                    zero,
                    bigM,
                )

                A.addNewRow(
                    {_q3(Cz, i, n, 2, Ni, Nn, Nz): bigM, _q3(Cw, i, 2, n, Ni, Nj, Nn): 1},
                    zero,
                    bigM,
                )

        # Exclude taxable withdrawals from passing spouse to deposits in surviving.
        # A.addNewRow({_q1(CZ, 0, 1): bigM, _q2(Cd, i_s, n_d - 1, Ni, Nn): -1}, zero, bigM)
        # A.addNewRow({_q1(CZ, 0, 1): bigM, _q3(Cw, i_d, 0, n_d - 1, Ni, Nj, Nn): 1}, zero, bigM)

        # Now build objective vector. Slight 1% favor to tax-free to avoid null space.
        c = abc.Objective(self.nvars)
        if objective == 'maxSpending':
            c.setElem(_q1(Cg, 0, Nn), -1)
        elif objective == 'maxBequest':
            for i in range(Ni):
                c.setElem(_q3(Cb, i, 0, Nn, Ni, Nj, Nn + 1), -1)
                c.setElem(_q3(Cb, i, 1, Nn, Ni, Nj, Nn + 1), -(1 - self.nu))
                # Add nudge to tax-exempt account
                c.setElem(_q3(Cb, i, 2, Nn, Ni, Nj, Nn + 1), -1)
        else:
            u.xprint('Internal error in objective function.')

        self.A = A
        self.B = B
        self.c = c

        return

    def solve(self, objective, options={}):
        '''
        This function builds the necessary constaints and
        runs the optimizer.

        - objective can be 'maxSpending' or 'maxBequest'.

        - options is a dictionary which can include:
            - maxRothConversion: Only allow conversion smaller than amount specified.
            - netSpending: Desired spending amount when optimizing with maxBequest.
            - estate: Value of bequest in today's $ when optimizing with maxSpending.
            - units: Units to use for amounts (1, k, or M).

        All units are in $k, unless specified otherwise.

        Refer to companion document for implementation details.
        '''
        # Assume unsuccessful until problem solved.
        self._caseStatus = 'unsuccessful'

        knownOptions = ['units', 'maxRothConversion', 'netSpending', 'estate', 'bigM', 'noRothConversions']
        for opt in options:
            if opt not in knownOptions:
                u.xprint('Option', opt, 'not one of', knownOptions)
        knownObjectives = ['maxBequest', 'maxSpending']
        if objective not in knownObjectives:
            u.xprint('Objective', objective, 'not one of', knownObjectives)
        if objective == 'maxBequest' and 'netSpending' not in options:
            u.xprint('Objective', objective, 'needs netSpending option.')

        self._adjustParameters()

        if self.solver == 'HiGHS':
            self._milpSolve(objective, options)
        elif self.solver == 'MOSEK':
            self._mosekSolve(objective, options)
        else:
            u.xprint('Unknown solver %s.'%self.solver)

        return

    def _milpSolve(self, objective, options):
        '''
        Solve problem using scipy HiGHS solver
        '''
        from scipy import optimize
        
        milpOptions = {'disp': True, 'mip_rel_gap': 1e-10}

        # 1$ tolerance over all values.
        it = 0
        diff = np.inf
        old_delta = 0
        old_x = np.zeros(self.nvars)
        self._estimateMedicare()
        while diff > 1:
            self._buildConstraints(objective, options)
            Alu, lbvec, ubvec = self.A.arrays()
            Lb, Ub = self.B.arrays()
            integrality = self.B.integralityArray()
            c = self.c.arrays()

            bounds = optimize.Bounds(Lb, Ub)
            constraint = optimize.LinearConstraint(Alu, lbvec, ubvec)
            solution = optimize.milp(
                c, integrality=integrality, constraints=constraint, bounds=bounds, options=milpOptions
            )
            it += 1

            if solution.success != True:
                break

            self._estimateMedicare(solution.x)
            delta = solution.x - old_x
            # u.vprint('Iteration:', it, 'delta:', np.sum(delta)/1000)

            diff = np.sum(np.abs(delta), axis=0)
            old_x = solution.x
            delta = np.sum(delta, axis=0)
            if abs(delta + old_delta) < 1e-3 or it > 32:
                print('WARNING: Detected oscilating solution.')
                print('    Try again with slightly different input parameters.')
                break
            old_delta = delta

        if solution.success == True:
            u.vprint('Self-consistent Medicare loop returned after %d iterations.' % it)
            u.vprint(solution.message)
            # u.vprint('Upper bound:', u.d(-solution.mip_dual_bound))
            self._aggregateResults(solution.x)
            self._caseStatus = 'solved'
        else:
            u.vprint('WARNING: Optimization failed:', solution.message, solution.success)
            self._caseStatus = 'unsuccessful'

        self.objective = objective
        self.solverOptions = options

        return

    def _mosekSolve(self, objective, options):
        '''
        Solve problem using MOSEK solver.
        '''
        import mosek

        bdic = {'fx': mosek.boundkey.fx,
                'fr': mosek.boundkey.fr,
                'lo': mosek.boundkey.lo,
                'ra': mosek.boundkey.ra,
                'up': mosek.boundkey.up,
               }

        # 1$ tolerance over all values.
        it = 0
        diff = np.inf
        old_delta = 0
        old_x = np.zeros(self.nvars)
        self._estimateMedicare()
        while diff > 1:
            self._buildConstraints(objective, options)
            Aind, Aval, clb, cub = self.A.lists()
            ckeys = self.A.keys()
            vlb,  vub = self.B.arrays()
            integrality = self.B.integralityList()
            vkeys = self.B.keys()
            cind, cval = self.c.lists()

            task = mosek.Task()
            # task.set_Stream(mosek.streamtype.msg, _streamPrinter)
            task.appendcons(self.A.ncons)
            task.appendvars(self.A.nvars)

            for ii in range(len(cind)):
                task.putcj(cind[ii], cval[ii])

            for ii in range(self.nvars):
                task.putvarbound(ii, bdic[vkeys[ii]], vlb[ii], vub[ii])

            for ii in range(len(integrality)):
                task.putvartype(integrality[ii], mosek.variabletype.type_int)

            for ii in range(self.A.ncons):
                task.putarow(ii, Aind[ii], Aval[ii])
                task.putconbound(ii, bdic[ckeys[ii]], clb[ii], cub[ii])

            task.putobjsense(mosek.objsense.minimize)
            task.optimize()

            solsta = task.getsolsta(mosek.soltype.itg) 
            prosta = task.getprosta(mosek.soltype.itg) 
            it += 1

            if solsta != mosek.solsta.integer_optimal:
                break

            xx = np.array(task.getxx(mosek.soltype.itg))

            self._estimateMedicare(xx)
            delta = xx - old_x
            # u.vprint('Iteration:', it, 'delta:', np.sum(delta)/1000)

            diff = np.sum(np.abs(delta), axis=0)
            old_x = xx
            delta = np.sum(delta, axis=0)
            if abs(delta + old_delta) < 1e-3 or it > 32:
                print('WARNING: Detected oscilating solution.')
                print('    Try again with slightly different input parameters.')
                break
            old_delta = delta

        task.set_Stream(mosek.streamtype.msg, _streamPrinter)
        # task.writedata(self._name+'.ptf')
        if solsta == mosek.solsta.integer_optimal:
            u.vprint('Self-consistent Medicare loop returned after %d iterations.' % it)
            task.solutionsummary(mosek.streamtype.msg)
            # u.vprint('Upper bound:', u.d(-solution.mip_dual_bound))
            self._aggregateResults(xx)
            self._caseStatus = 'solved'
        else:
            u.vprint('WARNING: Optimization failed:', 'Infeasible or unbounded.')
            task.solutionsummary(mosek.streamtype.msg)
            self._caseStatus = 'unsuccessful'

        self.objective = objective
        self.solverOptions = options

        return

    def _estimateMedicare(self, x=None):
        '''
        Compute rough MAGI and Medicare costs.
        '''
        if x is None:
            self.G_n = np.zeros(self.N_n)
        else:
            self.f_tn = np.array(x[self.C['f'] : self.C['g']])
            self.f_tn = self.f_tn.reshape((self.N_t, self.N_n))
            self.G_n = np.sum(self.f_tn, axis=0)

        self.M_n = tx.mediCosts(self.yobs, self.horizons, self.G_n + self.sigmaBar_n, self.gamma_n[:-1], self.N_n)

        return

    def _aggregateResults(self, x):
        '''
        Utility function to aggregate results from solver.
        Process all results from solution vector.
        '''
        # Define shortcuts.
        Ni = self.N_i
        Nj = self.N_j
        Nk = self.N_k
        Nn = self.N_n
        Nt = self.N_t
        Nz = self.N_z

        Cb = self.C['b']
        Cd = self.C['d']
        Cf = self.C['f']
        Cg = self.C['g']
        Cw = self.C['w']
        Cx = self.C['x']
        Cz = self.C['z']
        CZ = self.C['Z']

        x = u.roundCents(x)

        # Allocate, slice in, and reshape variables.
        self.b_ijn = np.array(x[Cb:Cd])
        self.b_ijn = self.b_ijn.reshape((Ni, Nj, Nn + 1))
        self.b_ijkn = np.zeros((Ni, Nj, Nk, Nn + 1))
        for k in range(Nk):
            self.b_ijkn[:, :, k, :] = self.b_ijn[:, :, :] * self.alpha_ijkn[:, :, k, :]

        self.d_in = np.array(x[Cd:Cf])
        self.d_in = self.d_in.reshape((Ni, Nn))

        self.f_tn = np.array(x[Cf:Cg])
        self.f_tn = self.f_tn.reshape((Nt, Nn))

        self.g_n = np.array(x[Cg:Cw])
        # self.g_n = self.g_n.reshape((Nn))

        self.w_ijn = np.array(x[Cw:Cx])
        self.w_ijn = self.w_ijn.reshape((Ni, Nj, Nn))

        self.x_in = np.array(x[Cx:Cz])
        self.x_in = self.x_in.reshape((Ni, Nn))

        # self.z_inz = np.array(x[Cz:CZ])
        # self.z_inz = self.z_inz.reshape((Ni, Nn, Nz))
        # print(self.z_inz)

        sourcetypes = [
            'wages',
            'ssec',
            'pension',
            'dist',
            'rmd',
            'RothX',
            'div',
            'wdrwl taxable',
            'wdrwl tax-free',
        ]

        self.rmd_in = self.rho_in * self.b_ijn[:, 1, :-1]
        self.dist_in = self.w_ijn[:, 1, :] - self.rmd_in
        self.dist_in[self.dist_in < 0] = 0
        self.G_n = np.sum(self.f_tn, axis=0)
        T_tn = self.f_tn * self.theta_tn
        self.T_n = np.sum(T_tn, axis=0)

        tau_0 = np.array(self.tau_kn[0, :])
        tau_0[tau_0 < 0] = 0
        # Last year's rates.
        tau_0prev = np.roll(tau_0, 1)
        self.Q_n = np.sum(
            (
                self.mu
                * (self.b_ijn[:, 0, :-1] - self.w_ijn[:, 0, :] + self.d_in[:, :] + 0.5 * self.kappa_ijn[:, 0, :])
                + +tau_0prev * self.w_ijn[:, 0, :]
            )
            * self.alpha_ijkn[:, 0, 0, :-1],
            axis=0,
        )
        self.U_n = self.psi * self.Q_n

        # Make derivative variables.
        # Putting it all together in a dictionary.
        sources = {}
        sources['wages'] = self.omega_in
        sources['ssec'] = self.zetaBar_in
        sources['pension'] = self.pi_in
        sources['wdrwl taxable'] = self.w_ijn[:, 0, :]
        sources['rmd'] = self.rmd_in
        sources['dist'] = self.dist_in
        sources['RothX'] = self.x_in
        sources['wdrwl tax-free'] = self.w_ijn[:, 2, :]
        sources['bti'] = self.Lambda_in

        savings = {}
        savings['taxable'] = self.b_ijn[:, 0, :]
        savings['tax-deferred'] = self.b_ijn[:, 1, :]
        savings['tax-free'] = self.b_ijn[:, 2, :]

        self.sources_in = sources
        self.savings_in = savings

        return

    def estate(self):
        '''
        Reports final account balances.
        '''
        if self._checkSolverStatus('estate'):
            return

        _estate = np.sum(self.b_ijn[:, :, :, self.N_n], axis=(0, 2))
        _estate[1] *= 1 - self.nu
        u.vprint('Estate value of %s at the end of year %s.' % (u.d(sum(_estate)), self.year_n[-1]))

        return

    def summary(self):
        '''
        Print summary of values.
        '''
        if self._checkSolverStatus('summary'):
            return

        now = self.year_n[0]
        print('SUMMARY ================================================================')
        print('Plan name:', self._name)
        for i in range(self.N_i):
            u.vprint('%12s\'s life horizon: %d -> %d' % (self.inames[i], now, now + self.horizons[i] - 1))
        print('Contributions file:', self.timeListsFileName)
        print('Return rates:', self.rateMethod)
        if self.rateMethod in ['historical', 'average', 'stochastic']:
            print('Rates used: from', self.rateFrm, 'to', self.rateTo)
        else:
            print('Rates used:', *[u.pc(self.rateValues[k], f=1) for k in range(self.N_k)])
        print('Optimized for:', self.objective)
        print('Solver options:', self.solverOptions)
        print('Solver used:', self.solver)
        print('Number of decision variables:', self.A.nvars)
        print('Number of constraints:', self.A.ncons)
        print('Spending profile:', self.spendingProfile)
        if self.N_i == 2:
            print('Surviving spouse percent income:', u.pc(self.chi, f=0))

        print('Net yearly spending basis in %d$: %s' % (now, u.d(self.g_n[0]/self.xi_n[0])))

        totIncome = np.sum(self.g_n, axis=0)
        totIncomeNow = np.sum(self.g_n / self.gamma_n[:-1], axis=0)
        print('Total net spending in %d$: %s (%s nominal)' % (now, u.d(totIncomeNow), u.d(totIncome)))

        totRoth = np.sum(self.x_in, axis=(0, 1))
        totRothNow = np.sum(np.sum(self.x_in, axis=0) / self.gamma_n[:-1], axis=0)
        print('Total Roth conversions in %d$: %s (%s nominal)' % (now, u.d(totRothNow), u.d(totRoth)))

        taxPaid = np.sum(self.T_n, axis=0)
        taxPaidNow = np.sum(self.T_n / self.gamma_n[:-1], axis=0)
        print('Total ordinary income tax paid in %d$: %s (%s nominal)' % (now, u.d(taxPaidNow), u.d(taxPaid)))

        taxPaid = np.sum(self.U_n, axis=0)
        taxPaidNow = np.sum(self.U_n / self.gamma_n[:-1], axis=0)
        print('Total dividend tax paid in %d$: %s (%s nominal)' % (now, u.d(taxPaidNow), u.d(taxPaid)))

        taxPaid = np.sum(self.M_n, axis=0)
        taxPaidNow = np.sum(self.M_n / self.gamma_n[:-1], axis=0)
        print('Total Medicare premiums paid in %d$: %s (%s nominal)' % (now, u.d(taxPaidNow), u.d(taxPaid)))

        estate = np.sum(self.b_ijn[:, :, self.N_n], axis=0)
        estate[1] *= 1 - self.nu
        print('Assumed heirs tax rate:', u.pc(self.nu, f=0))
        print('Final account post-tax nominal values:')
        print('    taxable: %s  tax-def: %s  tax-free: %s' % (u.d(estate[0]), u.d(estate[1]), u.d(estate[2])))

        totEstate = np.sum(estate)
        totEstateNow = totEstate / self.gamma_n[-1]
        print('Total estate value in %d$: %s (%s nominal)' % (now, u.d(totEstateNow), u.d(totEstate)))
        print('Final inflation factor:', u.pc(self.gamma_n[-1], f=1))

        print('------------------------------------------------------------------------')

        return

    def showRates(self, tag=''):
        '''
        Plot rate values used over the time horizon.

        A tag string can be set to add information to the title of the plot.
        '''
        import matplotlib.pyplot as plt
        import matplotlib.ticker as tk

        fig, ax = plt.subplots(figsize=(6, 4))
        plt.grid(visible='both')
        title = self._name + '\nReturn & Inflation Rates (' + str(self.rateMethod)
        if self.rateMethod in ['historical', 'stochastic', 'average']:
            title += ' ' + str(self.rateFrm) + '-' + str(self.rateTo)
        elif self.rateMethod == 'fixed':
            title += str(self.rateMethod)
        title += ')'

        if tag != '':
            title += ' - ' + tag

        rateName = [
            'S&P500 including dividends',
            'Baa Corporate bonds',
            '10-y Treasury notes',
            'Inflation',
        ]
        ltype = ['-', '-.', ':', '--']
        for k in range(self.N_k):
            data = 100 * self.tau_kn[k]
            label = rateName[k] + ' <' + '{:.2f}'.format(np.mean(data)) + '%>'
            ax.plot(self.year_n, data, label=label, ls=ltype[k % self.N_k])

        ax.xaxis.set_major_locator(tk.MaxNLocator(integer=True))
        ax.legend(loc='upper left', reverse=False, fontsize=8, framealpha=0.7)
        # ax.legend(loc='upper left')
        ax.set_title(title)
        ax.set_xlabel('year')
        ax.set_ylabel('%')

        # plt.show()
        # return fig, ax
        return

    def showProfile(self, tag=''):
        '''
        Plot income profile over time.

        A tag string can be set to add information to the title of the plot.
        '''
        title = self._name + '\nIncome Profile'
        if tag != '':
            title += ' - ' + tag

        # style = {'net': '-', 'target': ':'}
        style = {'profile': '-'}
        series = {'profile': self.xi_n}
        _lineIncomePlot(self.year_n, series, style, title, yformat='xi')

        return

    def showNetSpending(self, tag='', value=None):
        '''
        Plot net available spending and target over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        '''
        if self._checkSolverStatus('showNetSpending'):
            return

        value = self._checkValue(value)

        title = self._name + '\nNet Available Spending'
        if tag != '':
            title += ' - ' + tag

        style = {'net': '-', 'target': ':'}
        if value == 'nominal':
            series = {'net': self.g_n, 'target': (self.g_n[0] / self.xi_n[0]) * self.xiBar_n}
            yformat = 'k$ (nominal)'
        else:
            series = {'net': self.g_n / self.gamma_n[:-1], 'target': (self.g_n[0] / self.xi_n[0]) * self.xi_n}
            yformat = 'k$ (' + str(self.year_n[0]) + '\$)'

        _lineIncomePlot(self.year_n, series, style, title, yformat)

        return

    def showAssetDistribution(self, tag='', value=None):
        '''
        Plot the distribution of each savings account in thousands of dollars
        during the simulation time. This function will generate three
        graphs, one for taxable accounts, one the tax-deferred accounts,
        and one for tax-free accounts.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        '''
        if self._checkSolverStatus('showAssetDistribution'):
            return

        value = self._checkValue(value)

        if value == 'nominal':
            yformat = 'k$ (nominal)'
            infladjust = 1
        else:
            yformat = 'k$ (' + str(self.year_n[0]) + '\$)'
            infladjust = self.gamma_n

        years_n = np.array(self.year_n)
        years_n = np.append(years_n, [years_n[-1] + 1])
        y2stack = {}
        jDic = {'taxable': 0, 'tax-deferred': 1, 'tax-free': 2}
        kDic = {'stocks': 0, 'C bonds': 1, 'T notes': 2, 'common': 3}
        for jkey in jDic:
            stackNames = []
            for kkey in kDic:
                name = kkey + ' / ' + jkey
                stackNames.append(name)
                y2stack[name] = np.zeros((self.N_i, self.N_n + 1))
                for i in range(self.N_i):
                    y2stack[name][i][:] = self.b_ijkn[i][jDic[jkey]][kDic[kkey]][:] / infladjust

            title = self._name + '\nAssets Distribution - ' + jkey
            if tag != '':
                title += ' - ' + tag

            _stackPlot(years_n, self.inames, title, range(self.N_i), y2stack, stackNames, 'upper left', yformat)

        return

    def showAllocations(self, tag=''):
        '''
        Plot desired allocation of savings accounts in percentage
        over simulation time and interpolated by the selected method
        through the interpolateAR() method.

        A tag string can be set to add information to the title of the plot.
        '''
        count = self.N_i
        if self.ARCoord == 'spouses':
            acList = [self.ARCoord]
            count = 1
        elif self.ARCoord == 'individual':
            acList = [self.ARCoord]
        elif self.ARCoord == 'account':
            acList = ['taxable', 'tax-deferred', 'tax-free']
        else:
            u.xprint('Unknown coordination', self.ARCoord)

        assetDic = {'stocks': 0, 'C bonds': 1, 'T notes': 2, 'common': 3}
        for i in range(count):
            y2stack = {}
            for acType in acList:
                stackNames = []
                for key in assetDic:
                    aname = key + ' / ' + acType
                    stackNames.append(aname)
                    y2stack[aname] = np.zeros((count, self.N_n))
                    y2stack[aname][i][:] = self.alpha_ijkn[i, acList.index(acType), assetDic[key], : self.N_n]

                    title = self._name + '\nAssets Allocations (%) - ' + acType
                    if self.ARCoord == 'spouses':
                        title += ' spouses'
                    else:
                        title += ' ' + self.inames[i]

                if tag != '':
                    title += ' - ' + tag

                _stackPlot(
                    self.year_n,
                    self.inames,
                    title,
                    [i],
                    y2stack,
                    stackNames,
                    'upper left',
                    'percent',
                )

        return

    def showAccounts(self, tag='', value=None):
        '''
        Plot values of savings accounts over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        '''
        if self._checkSolverStatus('showAccounts'):
            return

        value = self._checkValue(value)

        title = self._name + '\nSavings Balance'
        if tag != '':
            title += ' - ' + tag

        stypes = self.savings_in.keys()
        # Add one year for estate.
        year_n = np.append(self.year_n, [self.year_n[-1] + 1])

        if value == 'nominal':
            yformat = 'k$ (nominal)'
            savings_in = self.savings_in
        else:
            yformat = 'k$ (' + str(self.year_n[0]) + '\$)'
            savings_in = {}
            for key in self.savings_in:
                savings_in[key] = self.savings_in[key] / self.gamma_n

        _stackPlot(year_n, self.inames, title, range(self.N_i), savings_in, stypes, 'upper left', yformat)

        return

    def showSources(self, tag='', value=None):
        '''
        Plot income over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        '''
        if self._checkSolverStatus('showSources'):
            return

        value = self._checkValue(value)

        title = self._name + '\nRaw Income Sources'
        stypes = self.sources_in.keys()

        if tag != '':
            title += ' - ' + tag

        if value == 'nominal':
            yformat = 'k$ (nominal)'
            sources_in = self.sources_in
        else:
            yformat = 'k$ (' + str(self.year_n[0]) + '\$)'
            sources_in = {}
            for key in self.sources_in:
                sources_in[key] = self.sources_in[key] / self.gamma_n[:-1]

        _stackPlot(self.year_n, self.inames, title, range(self.N_i), sources_in, stypes, 'upper left', yformat)

        return

    def _showFeff(self, tag=''):
        '''
        Plot income tax paid over time.

        A tag string can be set to add information to the title of the plot.
        '''
        if self._checkSolverStatus('showEff'):
            return

        title = self._name + '\nEff f '
        if tag != '':
            title += ' - ' + tag

        various = ['-', '--', '-.', ':']
        style = {}
        series = {}
        q = 0
        for t in range(self.N_t):
            key = 'f ' + str(t)
            series[key] = self.f_tn[t] / self.DeltaBar_tn[t]
            print(key, series[key])
            style[key] = various[q % len(various)]
            q += 1

        fig, ax = _lineIncomePlot(self.year_n, series, style, title, yformat='')

        return

    def showTaxes(self, tag='', value=None):
        '''
        Plot income tax paid over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        '''
        if self._checkSolverStatus('showTaxes'):
            return

        value = self._checkValue(value)

        style = {'income taxes': '-', 'Medicare': '-.'}

        if value == 'nominal':
            series = {'income taxes': self.T_n, 'Medicare': self.M_n}
            yformat = 'k$ (nominal)'
        else:
            series = {'income taxes': self.T_n / self.gamma_n[:-1], 'Medicare': self.M_n / self.gamma_n[:-1]}
            yformat = 'k$ (' + str(self.year_n[0]) + '\$)'

        title = self._name + '\nIncome Tax'
        if tag != '':
            title += ' - ' + tag

        fig, ax = _lineIncomePlot(self.year_n, series, style, title, yformat)

        return

    def showGrossIncome(self, tag='', value=None):
        '''
        Plot income tax and taxable income over time horizon.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        '''
        import matplotlib.pyplot as plt

        if self._checkSolverStatus('showGrossIncome'):
            return

        value = self._checkValue(value)

        style = {'taxable income': '-'}

        if value == 'nominal':
            series = {'taxable income': self.G_n}
            yformat = 'k$ (nominal)'
            infladjust = self.gamma_n[:-1]
        else:
            series = {'taxable income': self.G_n / self.gamma_n[:-1]}
            yformat = 'k$ (' + str(self.year_n[0]) + '\$)'
            infladjust = 1

        title = self._name + '\nTaxable Ordinary Income vs. Tax Brackets'
        if tag != '':
            title += ' - ' + tag

        fig, ax = _lineIncomePlot(self.year_n, series, style, title, yformat)

        data = tx.taxBrackets(self.N_i, self.n_d, self.N_n)
        for key in data:
            data_adj = data[key] * infladjust
            ax.plot(self.year_n, data_adj, label=key, ls=':')

        plt.grid(visible='both')
        ax.legend(loc='upper left', reverse=True, fontsize=8, framealpha=0.3)

        return

    def saveWorkbook(self, basename, overwrite=False):
        '''
        Save instance in an Excel spreadsheet.
        The first worksheet will contain income in the following
        fields in columns:
            - net spending
            - taxable ordinary income
            - taxable dividends
            - tax bills (federal only, including IRMAA)
        for all the years for the time span of the plan.

        The second worksheet contains the rates
        used for the plan as follows:
            - S&P 500
            - Corporate Baa bonds
            - Treasury notes (10y)
            - Inflation.

        The subsequent worksheets has the sources for each
        spouse as follows:
            - taxable account withdrawals
            - RMDs
            - distributions
            - Roth conversions
            - tax-free withdrawals
            - big-ticket items.

        The subsequent worksheets contains the balances
        and input/ouput to the savings accounts for each spouse:
            - taxable savings account
            - tax-deferred account
            - tax-free account.

        Last worksheet contains cash flow.

        '''
        if self._checkSolverStatus('saveWorkbook'):
            return

        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        # Income.
        ws = wb.active
        ws.title = 'Income'

        incomeDic = {
            'net spending': self.g_n,
            'taxable ord. income': self.G_n,
            'taxable dividends': self.Q_n,
            'Tax bills + Med.': self.T_n + self.U_n + self.M_n,
        }

        rawData = {}
        rawData['year'] = self.year_n
        for key in incomeDic:
            rawData[key] = incomeDic[key]

        # We need to work by row.
        df = pd.DataFrame(rawData)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        _formatSpreadsheet(ws, 'currency')

        # Cash flow.
        cashFlowDic = {
            'net spending': self.g_n,
            'all wages': np.sum(self.omega_in, axis=0),
            'all pensions': np.sum(self.pi_in, axis=0),
            'all soc sec': np.sum(self.zetaBar_in, axis=0),
            'all bti\'s': np.sum(self.Lambda_in, axis=0),
            'all wdrwls': np.sum(self.w_ijn, axis=(0, 1)),
            'all deposits': -np.sum(self.d_in, axis=0),
            'ord taxes': -self.T_n,
            'div taxes': -self.U_n,
            'Medicare': -self.M_n,
        }
        sname = 'Cash Flow'
        ws = wb.create_sheet(sname)
        rawData = {}
        rawData['year'] = self.year_n
        for key in cashFlowDic:
            rawData[key] = cashFlowDic[key]
        df = pd.DataFrame(rawData)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        _formatSpreadsheet(ws, 'currency')

        # Sources.
        srcDic = {
            'wages': 'wages',
            'social sec': 'ssec',
            'pension': 'pension',
            'txbl acc. wdrwl': 'wdrwl taxable',
            'RMDs': 'rmd',
            'distribution': 'dist',
            'Roth conversion': 'RothX',
            'tax-free wdrwl': 'wdrwl tax-free',
            'big-ticket items': 'bti',
        }

        for i in range(self.N_i):
            sname = self.inames[i] + '\'s Sources'
            ws = wb.create_sheet(sname)
            rawData = {}
            rawData['year'] = self.year_n
            for key in srcDic:
                rawData[key] = self.sources_in[srcDic[key]][i]

            df = pd.DataFrame(rawData)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

            _formatSpreadsheet(ws, 'currency')

        # Account balances except final year.
        accDic = {
            'taxable bal': self.b_ijn[:, 0, :-1],
            'taxable dep': self.d_in,
            'taxable wdrwl': self.w_ijn[:, 0, :],
            'tax-deferred bal': self.b_ijn[:, 1, :-1],
            'tax-deferred ctrb': self.kappa_ijn[:, 1, :],
            'tax-deferred wdrwl': self.w_ijn[:, 1, :],
            'tax-deferred (rmd)': self.rmd_in[:, :],
            'Roth conversion': self.x_in,
            'tax-free bal': self.b_ijn[:, 2, :-1],
            'tax-free ctrb': self.kappa_ijn[:, 2, :],
            'tax-free wdrwl': self.w_ijn[:, 2, :],
        }
        for i in range(self.N_i):
            sname = self.inames[i] + '\'s Accounts'
            ws = wb.create_sheet(sname)
            rawData = {}
            rawData['year'] = self.year_n
            for key in accDic:
                rawData[key] = accDic[key][i]
            df = pd.DataFrame(rawData)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            lastRow = [self.year_n[-1]+1,
                       self.b_ijn[i][0][-1], 0, 0,
                       self.b_ijn[i][1][-1], 0, 0, 0, 0,
                       self.b_ijn[i][2][-1], 0, 0]
            ws.append(lastRow)

            _formatSpreadsheet(ws, 'currency')

        # Allocations.
        jDic = {'taxable': 0, 'tax-deferred': 1, 'tax-free': 2}
        kDic = {'stocks': 0, 'C bonds': 1, 'T notes': 2, 'common': 3}

        # Add one year for estate.
        year_n = np.append(self.year_n, [self.year_n[-1] + 1])
        for i in range(self.N_i):
            sname = self.inames[i] + '\'s Allocations'
            ws = wb.create_sheet(sname)
            rawData = {}
            rawData['year'] = year_n
            for jkey in jDic:
                for kkey in kDic:
                    rawData[jkey + '/' + kkey] = self.alpha_ijkn[i, jDic[jkey], kDic[kkey], :]
            df = pd.DataFrame(rawData)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

            _formatSpreadsheet(ws, 'percent1')

        # Rates on last sheet.
        ws = wb.create_sheet('Rates')
        rawData = {}
        rawData['year'] = self.year_n
        ratesDic = {'S&P 500': 0, 'Corporate Baa': 1, 'T Bonds': 2, 'inflation': 3}

        for key in ratesDic:
            rawData[key] = self.tau_kn[ratesDic[key]]

        df = pd.DataFrame(rawData)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        _formatSpreadsheet(ws, 'percent2')

        _saveWorkbook(wb, basename, overwrite)

        return

    def saveWorkbookCSV(self, basename):
        '''
        Function similar to saveWorkbook(), but saving information in csv format
        instead of an Excel worksheet.
        See saveWorkbook() sister function for more information.
        '''
        import pandas as pd

        planData = {}
        planData['year'] = self.year_n
        planData['net spending'] = self.g_n
        planData['taxable ord. income'] = self.G_n
        planData['taxable dividends'] = self.Q_n
        planData['tax bill'] = self.T_n

        for i in range(self.N_i):
            planData[self.inames[i] + ' txbl bal'] = self.b_ijn[i, 0, :-1]
            planData[self.inames[i] + ' txbl dep'] = self.d_in[i, :]
            planData[self.inames[i] + ' txbl wrdwl'] = self.w_ijn[i, 0, :]
            planData[self.inames[i] + ' tx-def bal'] = self.b_ijn[i, 1, :-1]
            planData[self.inames[i] + ' tx-def ctrb'] = self.kappa_ijn[i, 1, :]
            planData[self.inames[i] + ' tx-def wdrl'] = self.w_ijn[i, 1, :]
            planData[self.inames[i] + ' (RMD)'] = self.rmd_in[i, :]
            planData[self.inames[i] + ' Roth conversion'] = self.x_in[i, :]
            planData[self.inames[i] + ' tx-free bal'] = self.b_ijn[i, 2, :-1]
            planData[self.inames[i] + ' tx-free ctrb'] = self.kappa_ijn[i, 2, :]
            planData[self.inames[i] + ' tax-free wdrwl'] = self.w_ijn[i, 2, :]
            planData[self.inames[i] + ' big-ticket items'] = self.Lambda_in[i, :]

        ratesDic = {'S&P 500': 0, 'Corporate Baa': 1, 'T Bonds': 2, 'inflation': 3}
        for key in ratesDic:
            planData[key] = self.tau_kn[ratesDic[key]]

        df = pd.DataFrame(planData)

        while True:
            try:
                fname = 'worksheet' + '_' + basename + '.csv'
                df.to_csv(fname)
                break
            except PermissionError:
                print('Failed to save "%s": %s.' % (fname, 'Permission denied'))
                key = input('Close file and try again? [Yn] ')
                if key == 'n':
                    break
            except Exception:
                u.xprint('Unanticipated exception', Exception)

        return


def _lineIncomePlot(x, series, style, title, yformat='k$'):
    '''
    Core line plotter function.
    '''
    import matplotlib.pyplot as plt
    import matplotlib.ticker as tk

    fig, ax = plt.subplots(figsize=(6, 4))
    plt.grid(visible='both')

    for sname in series:
        ax.plot(x, series[sname], label=sname, ls=style[sname])

    ax.legend(loc='upper left', reverse=True, fontsize=8, framealpha=0.3)
    ax.set_title(title)
    ax.set_xlabel('year')
    ax.set_ylabel(yformat)
    ax.xaxis.set_major_locator(tk.MaxNLocator(integer=True))
    if 'k$' in yformat:
        ax.get_yaxis().set_major_formatter(tk.FuncFormatter(lambda x, p: format(int(x / 1000), ',')))
        # Give range to y values in unindexed flat profiles.
        ymin, ymax = ax.get_ylim()
        if ymax - ymin < 5000:
            ax.set_ylim((ymin*0.95 , ymax*1.05))

    return fig, ax


def _stackPlot(x, inames, title, irange, series, snames, location, yformat='k$'):
    '''
    Core function for stacked plots.
    '''
    import matplotlib.pyplot as plt
    import matplotlib.ticker as tk

    nonzeroSeries = {}
    for sname in snames:
        for i in irange:
            tmp = series[sname][i]
            if sum(tmp) > 1.0:
                nonzeroSeries[sname + ' ' + inames[i]] = tmp

    if len(nonzeroSeries) == 0:
        print('Nothing to plot for', title)
        return

    fig, ax = plt.subplots(figsize=(6, 4))
    plt.grid(visible='both')

    ax.stackplot(x, nonzeroSeries.values(), labels=nonzeroSeries.keys(), alpha=0.6)
    ax.legend(loc=location, reverse=True, fontsize=8, ncol=2, framealpha=0.5)
    ax.set_title(title)
    ax.set_xlabel('year')
    ax.xaxis.set_major_locator(tk.MaxNLocator(integer=True))
    if 'k$' in yformat:
        ax.set_ylabel(yformat)
        ax.get_yaxis().set_major_formatter(tk.FuncFormatter(lambda x, p: format(int(x / 1000), ',')))
    elif yformat == 'percent':
        ax.set_ylabel('%')
        ax.get_yaxis().set_major_formatter(tk.FuncFormatter(lambda x, p: format(int(100 * x), ',')))
    else:
        u.xprint('Unknown yformat:', yformat)

        return fig, ax


def showRateDistributions(frm=rates.FROM, to=rates.TO):
    '''
    Plot histograms of the rates distributions.
    '''
    import matplotlib.pyplot as plt

    title = 'Rates from ' + str(frm) + ' to ' + str(to)
    # Bring year values to indices.
    frm -= rates.FROM
    to -= rates.FROM

    nbins = int((to - frm) / 4)
    fig, ax = plt.subplots(1, 4, sharey=True, sharex=True, tight_layout=True)

    dat0 = np.array(rates.SP500[frm:to])
    dat1 = np.array(rates.BondsBaa[frm:to])
    dat2 = np.array(rates.TNotes[frm:to])
    dat3 = np.array(rates.Inflation[frm:to])

    fig.suptitle(title)
    ax[0].set_title('S&P500')
    label = '<>: ' + u.pc(np.mean(dat0), 2, 1)
    ax[0].hist(dat0, bins=nbins, label=label)
    ax[0].legend(loc='upper left', fontsize=8, framealpha=0.7)

    ax[1].set_title('BondsBaa')
    label = '<>: ' + u.pc(np.mean(dat1), 2, 1)
    ax[1].hist(dat1, bins=nbins, label=label)
    ax[1].legend(loc='upper left', fontsize=8, framealpha=0.7)

    ax[2].set_title('TNotes')
    label = '<>: ' + u.pc(np.mean(dat2), 2, 1)
    ax[2].hist(dat1, bins=nbins, label=label)
    ax[2].legend(loc='upper left', fontsize=8, framealpha=0.7)

    ax[3].set_title('Inflation')
    label = '<>: ' + u.pc(np.mean(dat3), 2, 1)
    ax[3].hist(dat3, bins=nbins, label=label)
    ax[3].legend(loc='upper left', fontsize=8, framealpha=0.7)

    plt.show()

    # return fig, ax
    return


def _saveWorkbook(wb, basename, overwrite=False):
    '''
    Utility function to save XL workbook.
    '''
    from os.path import isfile
    from pathlib import Path

    if Path(basename).suffixes == []:
        fname = 'workbook' + '_' + basename + '.xlsx'
    else:
        fname = basename

    if overwrite is False and isfile(fname):
        print('File "%s" already exists.' % fname)
        key = input('Overwrite? [Ny] ')
        if key != 'y':
            print('Skipping save and returning.')
            return

    while True:
        try:
            u.vprint('Saving plan as "%s"' % fname)
            wb.save(fname)
            break
        except PermissionError:
            print('Failed to save "%s": %s.' % (fname, 'Permission denied'))
            key = input('Close file and try again? [Yn] ')
            if key == 'n':
                break
        except Exception:
            u.xprint('Unanticipated exception', Exception)

    return


def _formatSpreadsheet(ws, ftype):
    '''
    Utility function to beautify spreadsheet.
    '''
    if ftype == 'currency':
        fstring = u'$#,##0_);[Red]($#,##0)'
    elif ftype == 'percent2':
        fstring = u'#.00%'
    elif ftype == 'percent1':
        fstring = u'#.0%'
    elif ftype == 'percent0':
        fstring = u'#0%'
    else:
        u.xprint('Unknown format:', ftype)

    for cell in ws[1] + ws['A']:
        cell.style = 'Pandas'
    for col in ws.columns:
        column = col[0].column_letter
        # col[0].style = 'Title'
        width = max(len(str(col[0].value)) + 4, 10)
        ws.column_dimensions[column].width = width
        if column != 'A':
            for cell in col:
                cell.number_format = fstring

    return


def _streamPrinter(text):
    '''
    Define a stream printer to grab output from MOSEK.
    '''
    import sys
    sys.stdout.write(text)
    sys.stdout.flush()

