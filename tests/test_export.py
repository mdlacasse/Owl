"""
Tests for export.py: Excel/CSV export utilities.

Coverage targets:
  - _FORMAT_STRINGS dict
  - _format_spreadsheet() — all branches including "summary" and error path
  - _format_col_sheet() — default_fmt, lowercase flag, column widths, header style
  - _format_debts_sheet() / _format_fixed_assets_sheet() — column format mapping
  - _format_federal_income_tax_sheet() — default_fmt fallback (SS % taxed / currency mix)
  - plan_to_excel() / saveWorkbook() — sheet names, with_config options
  - build_summary_dic() — partial-bequest, final-bequest, and debt branches
  - plan_to_csv() — file written with expected columns

Copyright (C) 2025-2026 The Owlplanner Authors

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import os

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import owlplanner as owl
from owlplanner.export import (
    _FORMAT_STRINGS,
    _format_col_sheet,
    _format_debts_sheet,
    _format_federal_income_tax_sheet,
    _format_fixed_assets_sheet,
    _format_spreadsheet,
    build_summary_dic,
    plan_to_csv,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_ws(headers, data_rows=None):
    """Return a minimal openpyxl worksheet with a header row and optional data."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in (data_rows or []):
        ws.append(row)
    return ws


def cell_fmt(ws, col_letter, row=2):
    """Return the number_format of the first data cell in a given column."""
    for row_cells in ws.iter_rows(min_row=row, max_row=row):
        for cell in row_cells:
            if cell.column_letter == col_letter:
                return cell.number_format
    return None


# ---------------------------------------------------------------------------
# Plan fixtures — module-scoped to avoid re-solving in every test
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def joe_plan():
    """Solved single-person plan (joe)."""
    exdir = "./examples/"
    p = owl.readConfig(os.path.join(exdir, "Case_joe"))
    p.readHFP(os.path.join(exdir, "HFP_joe.xlsx"))
    p.resolve()
    assert p.caseStatus == "solved"
    return p


@pytest.fixture(scope="module")
def alex_jamie_plan():
    """Solved two-person plan with partial bequest, debts, fixed assets, and SS income."""
    exdir = "./examples/"
    p = owl.readConfig(os.path.join(exdir, "Case_alex+jamie"))
    p.readHFP(os.path.join(exdir, "HFP_alex+jamie.xlsx"))
    p.resolve()
    assert p.caseStatus == "solved"
    return p


# ---------------------------------------------------------------------------
# _FORMAT_STRINGS
# ---------------------------------------------------------------------------

def test_format_strings_has_expected_keys():
    assert set(_FORMAT_STRINGS.keys()) == {"currency", "percent2", "percent1", "percent0", "pct_value"}


def test_format_strings_currency_value():
    assert _FORMAT_STRINGS["currency"] == "$#,##0_);[Red]($#,##0)"


def test_format_strings_all_values_are_strings():
    assert all(isinstance(v, str) for v in _FORMAT_STRINGS.values())


# ---------------------------------------------------------------------------
# _format_spreadsheet
# ---------------------------------------------------------------------------

def test_format_spreadsheet_unknown_type_raises():
    ws = make_ws(["year", "value"], [[2026, 100]])
    with pytest.raises(RuntimeError, match="Unknown format"):
        _format_spreadsheet(ws, "bogus")


def test_format_spreadsheet_summary_sets_wide_columns():
    label = "SUMMARY =============================="
    ws = make_ws([label], [["some value"]])
    _format_spreadsheet(ws, "summary")
    assert ws.column_dimensions["A"].width >= 40


def test_format_spreadsheet_summary_returns_none():
    ws = make_ws(["SUMMARY"], [["x"]])
    result = _format_spreadsheet(ws, "summary")
    assert result is None


def test_format_spreadsheet_currency_col_a_gets_year_format():
    ws = make_ws(["year", "amount"], [[2026, 50000], [2027, 60000]])
    _format_spreadsheet(ws, "currency")
    assert cell_fmt(ws, "A") == "0"


def test_format_spreadsheet_currency_col_b_gets_currency_format():
    ws = make_ws(["year", "amount"], [[2026, 50000]])
    _format_spreadsheet(ws, "currency")
    assert cell_fmt(ws, "B") == _FORMAT_STRINGS["currency"]


@pytest.mark.parametrize("ftype", ["percent2", "percent1", "percent0", "pct_value"])
def test_format_spreadsheet_percent_types_applied_to_col_b(ftype):
    ws = make_ws(["year", "rate"], [[2026, 0.05]])
    _format_spreadsheet(ws, ftype)
    assert cell_fmt(ws, "B") == _FORMAT_STRINGS[ftype]


def test_format_spreadsheet_styles_header_and_year_column():
    ws = make_ws(["year", "amount"], [[2026, 100]])
    _format_spreadsheet(ws, "currency")
    # Both ws[1] (header row) and ws["A"] (year column) should have "Pandas" style
    for cell in ws[1]:
        assert cell.style == "Pandas"
    for cell in ws["A"]:
        assert cell.style == "Pandas"


# ---------------------------------------------------------------------------
# _format_col_sheet
# ---------------------------------------------------------------------------

def test_format_col_sheet_maps_known_columns():
    ws = make_ws(["year", "rate", "amount"], [[2026, 3.5, 50000]])
    _format_col_sheet(ws, col_formats={"year": "0", "rate": "#,##0.00", "amount": "$#,##0"})
    assert cell_fmt(ws, "A") == "0"
    assert cell_fmt(ws, "B") == "#,##0.00"
    assert cell_fmt(ws, "C") == "$#,##0"


def test_format_col_sheet_skips_unlisted_when_no_default():
    ws = make_ws(["year", "notes"], [[2026, "text"]])
    _format_col_sheet(ws, col_formats={"year": "0"})
    # "notes" not in col_formats and no default → cell keeps openpyxl default "General"
    assert cell_fmt(ws, "B") == "General"


def test_format_col_sheet_applies_default_fmt_to_unlisted():
    default = "$#,##0_);[Red]($#,##0)"
    ws = make_ws(["year", "income", "other"], [[2026, 10000, 5000]])
    _format_col_sheet(ws, col_formats={"year": "0"}, default_fmt=default)
    assert cell_fmt(ws, "A") == "0"
    assert cell_fmt(ws, "B") == default
    assert cell_fmt(ws, "C") == default


def test_format_col_sheet_lowercase_false_preserves_case():
    ws = make_ws(["year", "SS % taxed", "Income"], [[2026, 0.85, 30000]])
    _format_col_sheet(
        ws,
        col_formats={"year": "0", "SS % taxed": "#.0%"},
        default_fmt="$#,##0_);[Red]($#,##0)",
        lowercase=False,
    )
    assert cell_fmt(ws, "A") == "0"
    assert cell_fmt(ws, "B") == "#.0%"
    assert cell_fmt(ws, "C") == "$#,##0_);[Red]($#,##0)"


def test_format_col_sheet_lowercase_true_matches_mixed_case_headers():
    # With lowercase=True (default), "Year" and "Amount" are matched as "year"/"amount"
    ws = make_ws(["Year", "Amount"], [[2026, 50000]])
    _format_col_sheet(ws, col_formats={"year": "0", "amount": "$#,##0"})
    assert cell_fmt(ws, "A") == "0"
    assert cell_fmt(ws, "B") == "$#,##0"


def test_format_col_sheet_sets_column_widths():
    ws = make_ws(["year", "a_very_long_column_name"], [[2026, 100]])
    _format_col_sheet(ws, col_formats={"year": "0"})
    # Width must be at least 10 (the minimum) and should reflect header length for column B
    assert ws.column_dimensions["A"].width >= 10
    assert ws.column_dimensions["B"].width >= len("a_very_long_column_name") + 4


def test_format_col_sheet_styles_header_row():
    ws = make_ws(["year", "amount"], [[2026, 50000]])
    _format_col_sheet(ws, col_formats={"year": "0", "amount": "$#,##0"})
    for cell in ws[1]:
        assert cell.style == "Pandas"


# ---------------------------------------------------------------------------
# _format_debts_sheet
# ---------------------------------------------------------------------------

def test_format_debts_sheet_year_and_term_are_integers():
    ws = make_ws(["year", "term", "rate", "amount", "description"], [[2026, 30, 3.5, 250000, "mortgage"]])
    _format_debts_sheet(ws)
    assert cell_fmt(ws, "A") == "0"  # year
    assert cell_fmt(ws, "B") == "0"  # term


def test_format_debts_sheet_rate_is_decimal():
    ws = make_ws(["year", "term", "rate", "amount"], [[2026, 30, 3.5, 250000]])
    _format_debts_sheet(ws)
    assert cell_fmt(ws, "C") == "#,##0.00"


def test_format_debts_sheet_amount_is_currency():
    ws = make_ws(["year", "term", "rate", "amount"], [[2026, 30, 3.5, 250000]])
    _format_debts_sheet(ws)
    assert cell_fmt(ws, "D") == "$#,##0_);[Red]($#,##0)"


def test_format_debts_sheet_unlisted_column_unchanged():
    ws = make_ws(["year", "term", "rate", "amount", "description"], [[2026, 30, 3.5, 250000, "mortgage"]])
    _format_debts_sheet(ws)
    assert cell_fmt(ws, "E") == "General"  # "description" not in col_formats


# ---------------------------------------------------------------------------
# _format_fixed_assets_sheet
# ---------------------------------------------------------------------------

def test_format_fixed_assets_sheet_yod_is_integer():
    ws = make_ws(["yod", "rate", "commission", "basis", "value"], [[2030, 5.0, 1.0, 100000, 150000]])
    _format_fixed_assets_sheet(ws)
    assert cell_fmt(ws, "A") == "0"


def test_format_fixed_assets_sheet_rate_and_commission_are_decimal():
    ws = make_ws(["yod", "rate", "commission", "basis", "value"], [[2030, 5.0, 1.0, 100000, 150000]])
    _format_fixed_assets_sheet(ws)
    assert cell_fmt(ws, "B") == "#,##0.00"
    assert cell_fmt(ws, "C") == "#,##0.00"


def test_format_fixed_assets_sheet_basis_and_value_are_currency():
    ws = make_ws(["yod", "rate", "commission", "basis", "value"], [[2030, 5.0, 1.0, 100000, 150000]])
    _format_fixed_assets_sheet(ws)
    assert cell_fmt(ws, "D") == "$#,##0_);[Red]($#,##0)"
    assert cell_fmt(ws, "E") == "$#,##0_);[Red]($#,##0)"


def test_format_fixed_assets_sheet_unlisted_column_unchanged():
    ws = make_ws(["yod", "rate", "commission", "basis", "value", "notes"], [[2030, 5.0, 1.0, 100000, 150000, "rental"]])
    _format_fixed_assets_sheet(ws)
    assert cell_fmt(ws, "F") == "General"


# ---------------------------------------------------------------------------
# _format_federal_income_tax_sheet
# ---------------------------------------------------------------------------

def test_format_federal_income_tax_year_is_integer():
    ws = make_ws(["year", "SS % taxed", "total tax"], [[2026, 0.85, 12000]])
    _format_federal_income_tax_sheet(ws)
    assert cell_fmt(ws, "A") == "0"


def test_format_federal_income_tax_ss_percent_col_is_percent_format():
    ws = make_ws(["year", "SS % taxed", "total tax"], [[2026, 0.85, 12000]])
    _format_federal_income_tax_sheet(ws)
    assert cell_fmt(ws, "B") == "#.0%"


def test_format_federal_income_tax_other_cols_get_currency_default():
    # Columns not in col_formats get the default_fmt (currency)
    ws = make_ws(["year", "SS % taxed", "total tax", "LTCG tax"], [[2026, 0.85, 12000, 3000]])
    _format_federal_income_tax_sheet(ws)
    assert cell_fmt(ws, "C") == "$#,##0_);[Red]($#,##0)"
    assert cell_fmt(ws, "D") == "$#,##0_);[Red]($#,##0)"


def test_format_federal_income_tax_case_sensitive_ss_col():
    # lowercase=False: "ss % taxed" (wrong case) should NOT match "SS % taxed"
    ws = make_ws(["year", "ss % taxed", "total tax"], [[2026, 0.85, 12000]])
    _format_federal_income_tax_sheet(ws)
    # "ss % taxed" (lowercase) doesn't match "SS % taxed" key → gets default_fmt
    assert cell_fmt(ws, "B") == "$#,##0_);[Red]($#,##0)"


# ---------------------------------------------------------------------------
# plan_to_excel / saveWorkbook
# ---------------------------------------------------------------------------

def test_save_workbook_returns_workbook_object(joe_plan):
    wb = joe_plan.saveWorkbook(saveToFile=False)
    assert wb is not None
    assert isinstance(wb, Workbook)


def test_save_workbook_has_core_sheets(joe_plan):
    wb = joe_plan.saveWorkbook(saveToFile=False)
    titles = {ws.title for ws in wb.worksheets}
    for expected in ("Income", "Cash Flow", "Federal Income Tax", "Rates", "Summary"):
        assert expected in titles, f"Missing sheet: {expected}"


def test_save_workbook_has_per_individual_sheets(joe_plan):
    wb = joe_plan.saveWorkbook(saveToFile=False)
    titles = {ws.title for ws in wb.worksheets}
    for name in joe_plan.inames:
        assert any(name in t and "Sources" in t for t in titles), f"Missing Sources sheet for {name}"
        assert any(name in t and "Accounts" in t for t in titles), f"Missing Accounts sheet for {name}"


def test_save_workbook_with_config_first(joe_plan):
    wb = joe_plan.saveWorkbook(saveToFile=False, with_config="first")
    assert wb.worksheets[0].title == "Config (.toml)"


def test_save_workbook_with_config_last(joe_plan):
    wb = joe_plan.saveWorkbook(saveToFile=False, with_config="last")
    assert wb.worksheets[-1].title == "Config (.toml)"


def test_save_workbook_with_config_invalid_raises(joe_plan):
    with pytest.raises(ValueError, match="Invalid with_config"):
        joe_plan.saveWorkbook(saveToFile=False, with_config="invalid")


def test_save_workbook_federal_income_tax_has_ss_percent_col(alex_jamie_plan):
    wb = alex_jamie_plan.saveWorkbook(saveToFile=False)
    tax_ws = next(ws for ws in wb.worksheets if ws.title == "Federal Income Tax")
    headers = [cell.value for cell in tax_ws[1]]
    assert "SS % taxed" in headers


def test_save_workbook_federal_income_tax_ss_col_format(alex_jamie_plan):
    wb = alex_jamie_plan.saveWorkbook(saveToFile=False)
    tax_ws = next(ws for ws in wb.worksheets if ws.title == "Federal Income Tax")
    headers = [cell.value for cell in tax_ws[1]]
    ss_col_letter = get_column_letter(headers.index("SS % taxed") + 1)
    ss_fmt = cell_fmt(tax_ws, ss_col_letter)
    assert ss_fmt == "#.0%"


def test_save_workbook_federal_income_tax_currency_col_format(alex_jamie_plan):
    wb = alex_jamie_plan.saveWorkbook(saveToFile=False)
    tax_ws = next(ws for ws in wb.worksheets if ws.title == "Federal Income Tax")
    headers = [cell.value for cell in tax_ws[1]]
    # "total" column should get default currency format
    total_col_letter = get_column_letter(headers.index("total") + 1)
    assert cell_fmt(tax_ws, total_col_letter) == "$#,##0_);[Red]($#,##0)"


# ---------------------------------------------------------------------------
# saveContributions — exercises _format_debts_sheet / _format_fixed_assets_sheet
# via plan.py on a plan that has actual debt and fixed-asset data
# ---------------------------------------------------------------------------

def test_save_contributions_has_debts_sheet(alex_jamie_plan):
    wb = alex_jamie_plan.saveContributions()
    assert wb is not None
    titles = {ws.title for ws in wb.worksheets}
    assert "Debts" in titles


def test_save_contributions_has_fixed_assets_sheet(alex_jamie_plan):
    wb = alex_jamie_plan.saveContributions()
    titles = {ws.title for ws in wb.worksheets}
    assert "Fixed Assets" in titles


# ---------------------------------------------------------------------------
# build_summary_dic — bequest and debt branches
# ---------------------------------------------------------------------------

def test_build_summary_dic_includes_partial_bequest_keys(alex_jamie_plan):
    """Two-person plan where one dies early: partial bequest block is populated."""
    p = alex_jamie_plan
    assert p.N_i == 2 and p.n_d < p.N_n
    dic = build_summary_dic(p)  # N defaults to N_n → triggers bequest block
    assert "Year of partial bequest" in dic
    survivor = p.inames[p.i_s]
    assert any(survivor in k for k in dic if "spousal transfer" in k.lower())


def test_build_summary_dic_includes_final_bequest_keys(alex_jamie_plan):
    """Full-horizon summary includes final bequest section and fixed-assets line."""
    dic = build_summary_dic(alex_jamie_plan)
    assert "Year of final bequest" in dic
    assert " Total after-tax value of final bequest" in dic
    assert "» Fixed assets liquidated at end of plan" in dic


def test_build_summary_dic_includes_debt_keys(alex_jamie_plan):
    """Plan with non-zero debt payments includes debt totals in summary."""
    p = alex_jamie_plan
    assert p.debt_payments_n.sum() > 0
    dic = build_summary_dic(p)
    assert " Total debt payments" in dic
    assert "[Total debt payments]" in dic


def test_build_summary_dic_partial_n_omits_bequest_sections(alex_jamie_plan):
    """Summary computed over fewer than N_n years omits both bequest sections."""
    p = alex_jamie_plan
    dic = build_summary_dic(p, N=p.n_d)
    assert "Year of partial bequest" not in dic
    assert "Year of final bequest" not in dic


# ---------------------------------------------------------------------------
# plan_to_csv
# ---------------------------------------------------------------------------

def test_plan_to_csv_creates_file(joe_plan, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    plan_to_csv(joe_plan, joe_plan._name, joe_plan.mylog)
    fname = tmp_path / f"worksheet_{joe_plan._name}.csv"
    assert fname.exists()


def test_plan_to_csv_has_expected_columns(joe_plan, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    plan_to_csv(joe_plan, joe_plan._name, joe_plan.mylog)
    df = pd.read_csv(tmp_path / f"worksheet_{joe_plan._name}.csv", index_col=0)
    assert "year" in df.columns
    assert "net spending" in df.columns
    assert "taxable ord. income" in df.columns


def test_plan_to_csv_row_count_matches_plan_years(joe_plan, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    plan_to_csv(joe_plan, joe_plan._name, joe_plan.mylog)
    df = pd.read_csv(tmp_path / f"worksheet_{joe_plan._name}.csv", index_col=0)
    assert len(df) == joe_plan.N_n


def test_plan_to_csv_has_all_net_inv_column(joe_plan, tmp_path, monkeypatch):
    """'all net inv' column is present in CSV output."""
    monkeypatch.chdir(tmp_path)
    plan_to_csv(joe_plan, joe_plan._name, joe_plan.mylog)
    df = pd.read_csv(tmp_path / f"worksheet_{joe_plan._name}.csv", index_col=0)
    assert "all net inv" in df.columns


# ---------------------------------------------------------------------------
# Cash Flow balance
# ---------------------------------------------------------------------------

def test_cash_flow_sheet_has_all_net_inv_column(joe_plan):
    """'all net inv' column is present in the Cash Flow worksheet."""
    wb = joe_plan.saveWorkbook(saveToFile=False)
    cf_ws = next(ws for ws in wb.worksheets if ws.title == "Cash Flow")
    headers = [cell.value for cell in cf_ws[1]]
    assert "all net inv" in headers


def test_cash_flow_sheet_balances_to_zero(joe_plan):
    """Each row of the Cash Flow sheet sums to zero (net spending == sum of all other columns).

    Convention: inflows are positive, outflows (taxes, deposits, debt, spending) are
    included with their natural sign in the dict, so the algebraic row sum is zero.
    Equivalently: 'net spending' equals the sum of all remaining columns.
    """
    import numpy as np

    p = joe_plan
    # Reconstruct the cash-flow dict the same way export.py does, then check balance.
    inflows = (
        np.sum(p.omega_in, axis=0)        # all wages
        + np.sum(p.other_inc_in, axis=0)  # all other inc
        + np.sum(p.netinv_in, axis=0)     # all net inv
        + np.sum(p.piBar_in, axis=0)      # all pensions
        + np.sum(p.zetaBar_in, axis=0)    # all soc sec
        + np.sum(p.Lambda_in, axis=0)     # all BTI's
        + p.fixed_assets_ordinary_income_n
        + p.fixed_assets_capital_gains_n
        + p.fixed_assets_tax_free_n
        - p.debt_payments_n               # debt pmts (negative flow)
        + np.sum(p.w_ijn, axis=(0, 1))    # all wdrwls
        - np.sum(p.d_in, axis=0)          # all deposits (negative flow)
        - p.T_n - p.J_n                   # ord taxes (negative flow)
        - p.U_n                           # div taxes (negative flow)
        - p.m_n - p.M_n                   # Medicare (negative flow)
    )
    np.testing.assert_allclose(inflows, p.g_n, atol=1.0,
                               err_msg="Cash Flow sheet does not balance: sum of columns != net spending")
