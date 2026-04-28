"""
Core retirement planning module using linear programming optimization.

This module implements the main Plan class and optimization logic for retirement
financial planning. See companion PDF document for an explanation of the underlying
mathematical model and a description of all variables and parameters.

Copyright (C) 2025-2026 The Owlplanner Authors

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

###########################################################################
import numpy as np
import pandas as pd
from datetime import date, datetime
from functools import wraps
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import textwrap

from . import utils as u
from . import tax2026 as tx
from . import abcapi as abc
from . import rates
from . import config
from . import timelists
from . import export
from . import pension
from . import socialsecurity as socsec
from . import spending
from . import debts as debts
from . import fixedassets as fxasst
from . import mylogging as log
from .config.plan_bridge import clone                                         # noqa: F401
from .plotting.factory import PlotFactory
from .rate_models.constants import HISTORICAL_RANGE_METHODS
from .stresstests import run_historical_range, run_mc, run_stochastic_spending
from .varmap import VarMap


def _mosek_available():
    import importlib.util
    import os
    return (importlib.util.find_spec("mosek") is not None
            and os.environ.get("MOSEKLM_LICENSE_FILE") is not None)


# Default values
BIGM_AMO = 5e7     # 100 times large withdrawals or conversions
GAP = 1e-4
_PSI_DAMP = 0.3    # SC-loop damping weight for new Psi_n estimate (blend 30% new / 70% old)
MILP_GAP = 30 * GAP
MAX_ITERATIONS = 29
STAGNATION_WINDOW = 8   # SC iterations without best-objective improvement before early exit
ABS_TOL = 100
REL_TOL = 5e-5
TIME_LIMIT = 900
EPSILON = 1e-8


############################################################################


def _checkCaseStatus(func):
    """
    Decorator to check if problem was solved successfully and
    prevent method from running if not.
    """

    @wraps(func)
    def wrapper(self, *args, **kwargs):
        if self.caseStatus != "solved":
            self.mylog.print(f"Preventing to run method {func.__name__}() while case is {self.caseStatus}.")
            return None
        return func(self, *args, **kwargs)

    return wrapper


def _checkConfiguration(func):
    """
    Decorator to check if problem was configured successfully and
    prevent method from running if not.
    """

    @wraps(func)
    def wrapper(self, *args, **kwargs):
        if self.xi_n is None:
            msg = f"You must define a spending profile before calling {func.__name__}()."
            self.mylog.print(msg)
            raise RuntimeError(msg)
        if self.alpha_ijkn is None:
            msg = f"You must define an allocation profile before calling {func.__name__}()."
            self.mylog.print(msg)
            raise RuntimeError(msg)
        return func(self, *args, **kwargs)

    return wrapper


def _timer(func):
    """
    Decorator to report CPU and Wall time.
    """

    @wraps(func)
    def wrapper(self, *args, **kwargs):
        pt0 = time.process_time()
        rt0 = time.time()
        result = func(self, *args, **kwargs)
        pt = time.process_time() - pt0
        rt = time.time() - rt0
        self.mylog.vprint(f"CPU time used: {int(pt / 60)}m{pt % 60:.1f}s, Wall time: {int(rt / 60)}m{rt % 60:.1f}s.",
                          tag="INFO")
        return result

    return wrapper


class Plan:
    """
    This is the main class of the Owl Project.
    """
    # Class-level counter for unique Plan IDs
    _id_counter = 0

    @classmethod
    def get_next_id(cls):
        cls._id_counter += 1
        return cls._id_counter

    @classmethod
    def get_current_id(cls):
        return cls._id_counter

    def __init__(self, inames, dobs, expectancy, name, *, verbose=False, logstreams=None):
        """
        Constructor requires three lists: the first
        one contains the name(s) of the individual(s),
        the second one is the year of birth of each individual,
        and the third the life expectancy. Last argument is a name for
        the case.
        """
        if name == "":
            raise ValueError("Plan must have a name")

        # Generate unique ID for this Plan instance using the class method
        self._id = Plan.get_next_id()

        self._name = name
        self.setLogstreams(verbose, logstreams)

        # 7 tax brackets, 6 IRMAA (Medicare) brackets, 3 LTCG brackets,
        # 4 types of accounts (j=3 is HSA), 4 classes of assets.
        self.N_t = 7
        self.N_irmaa = 6
        self.N_p = 3
        self.N_j = 4
        self.N_k = 4
        # 4 binary variables for exclusions.
        self.N_zx = 4

        # Default interpolation parameters for allocation ratios.
        self.interpMethod = "linear"
        self._interpolator = self._linInterp
        self.interpCenter = 15
        self.interpWidth = 5

        self._description = ''
        self._config_extra = None
        self.defaultPlots = "nominal"
        self.worksheetShowAges = False
        self.worksheetHideZeroColumns = False
        self.worksheetRealDollars = False
        self.defaultSolver = "default"
        self._plotterName = None
        # Pick a default plotting backend here.
        # self.setPlotBackend("matplotlib")
        self.setPlotBackend("plotly")

        self.N_i = len(dobs)
        if not (0 <= self.N_i <= 2):
            raise ValueError(f"Cannot support {self.N_i} individuals.")
        if self.N_i != len(expectancy):
            raise ValueError(f"Expectancy must have {self.N_i} entries.")
        if self.N_i != len(inames):
            raise ValueError(f"Names for individuals must have {self.N_i} entries.")
        if inames[0] == "" or (self.N_i == 2 and inames[1] == ""):
            raise ValueError("Name for each individual must be provided.")

        self.filingStatus = ("single", "married")[self.N_i - 1]

        # Default year OBBBA speculated to be expired and replaced by pre-TCJA rates.
        self.yOBBBA = 2032
        self.inames = inames
        self.yobs, self.mobs, self.tobs = u.parseDobs(dobs)
        self.dobs = dobs
        self.expectancy = np.array(expectancy, dtype=np.int32)
        self.sexes = (["M", "F"] if self.N_i == 2 else ["F"])  # default; overridden by setSexes()
        self.mortality_table = "SSA2025"  # default; overridden by setMortalityTable()

        # Reference time is starting date in the current year and all passings are assumed at the end.
        thisyear = date.today().year
        self.horizons = self.yobs + self.expectancy - thisyear + 1
        self.N_n = np.max(self.horizons)
        if self.N_n <= 1:
            raise ValueError(f"Plan needs more than {self.N_n} years.")

        self.year_n = np.linspace(thisyear, thisyear + self.N_n - 1, self.N_n, dtype=np.int32)
        # Year index when each individual turns 59½ (IRS threshold). Born Jul–Dec: +1 year.
        self.n595 = 59 - thisyear + self.yobs + (self.mobs > 6).astype(np.int32)
        self.n595[self.n595 < 0] = 0
        # Handle passing of one spouse before the other.
        if self.N_i == 2 and np.min(self.horizons) != np.max(self.horizons):
            self.n_d = np.min(self.horizons)
            self.i_d = np.argmax(self.horizons == self.n_d)
            self.i_s = (self.i_d + 1) % 2
        else:
            self.n_d = self.N_n  # Push at upper bound and check for n_d < Nn.
            self.i_d = 0
            self.i_s = -1

        # Default parameters:
        # Fraction of SS benefits subject to federal income tax (initial: 0.85).
        # Refined each SC-loop iteration by _update_Psi_n() using the IRS provisional income formula.
        self.Psi_n = np.ones(self.N_n) * 0.85
        self.chi = 0.60   # Survivor fraction
        self.mu = 0.0172  # Dividend rate (decimal)
        self.nu = 0.300   # Heirs tax rate (decimal)
        self.bequest = 0.0            # After-tax bequest in today's dollars (set after full solve)
        self.effectiveTaxRate = 0.20  # Effective tax rate for for spending-to-savings ratio
        self.eta = (self.N_i - 1) / 2  # Spousal deposit ratio (0 or .5)
        self.phi_j = np.array([1, 1, 1, 1])  # Fractions left to other spouse at death (j=3: HSA)
        self.n_hsa_i = np.full(self.N_i, self.N_n, dtype=int)  # Year HSA contributions stop (default: never)
        self.slcsp_annual = 0.0             # Today-dollar annual ACA benchmark Silver plan premium ($)
        self.ACA_n = np.zeros(self.N_n)    # Net ACA cost (after subsidy) per year (plan $)
        self._aca_lp = False               # True when withACA="optimize" is active
        self.maca_n = np.zeros(self.N_n)   # ACA LP cost variable extraction result
        self.n_aca = 0                     # Number of ACA-eligible plan years (LP mode)
        self.smileDip = 15  # Percent to reduce smile profile
        self.smileIncrease = 12  # Percent to increse profile over time span

        # Default to zero pension and social security.
        self.pi_in = np.zeros((self.N_i, self.N_n))
        self.zeta_in = np.zeros((self.N_i, self.N_n))
        self.pensionAmounts = np.zeros(self.N_i, dtype=np.int32)
        self.pensionAges = 65 * np.ones(self.N_i, dtype=np.int32)
        self.pensionIsIndexed = [False] * self.N_i
        self.pensionSurvivorFraction = np.zeros(self.N_i)
        self.ssecAmounts = np.zeros(self.N_i, dtype=np.int32)
        self.ssecAges = 67 * np.ones(self.N_i, dtype=np.int32)
        self.ssecTrimPct = 0
        self.ssecTrimYear = None

        # Parameters from timeLists initialized to zero.
        self.omega_in = np.zeros((self.N_i, self.N_n))
        self.other_inc_in = np.zeros((self.N_i, self.N_n))
        self.netinv_in = np.zeros((self.N_i, self.N_n))
        self.Lambda_in = np.zeros((self.N_i, self.N_n))
        # Go back 5 years for maturation rules on IRA and Roth.
        self.myRothX_in = np.zeros((self.N_i, self.N_n + 5))
        self.kappa_ijn = np.zeros((self.N_i, self.N_j, self.N_n + 5))

        # Debt payments array (length N_n)
        self.debt_payments_n = np.zeros(self.N_n)

        # Fixed assets arrays (length N_n)
        self.fixed_assets_tax_free_n = np.zeros(self.N_n)
        self.fixed_assets_ordinary_income_n = np.zeros(self.N_n)
        self.fixed_assets_capital_gains_n = np.zeros(self.N_n)
        # Fixed assets bequest value (assets with yod past plan end)
        self.fixed_assets_bequest_value = 0.0

        # Remaining debt balance at end of plan
        self.remaining_debt_balance = 0.0

        # Previous 2 years of MAGI needed for Medicare.
        self.prevMAGI = np.zeros((2))
        self.MAGI_n = np.zeros(self.N_n)
        self.solverOptions = {}

        # Init current balances to none.
        self.beta_ij = None
        self.startDate = None

        # Default slack on profile.
        self.lambdha = 0

        # Scenario starts at the beginning of this year and ends at the end of the last year.
        s = ("", "s")[self.N_i - 1]
        self.mylog.vprint(f"Preparing scenario '{self._id}' of {self.N_n} years for {self.N_i} individual{s}.")
        for i in range(self.N_i):
            endyear = thisyear + self.horizons[i] - 1
            self.mylog.vprint(f"{self.inames[i]:>14}: life horizon from {thisyear} -> {endyear}.")

        # Prepare RMD time series.
        self.rho_in = tx.rho_in(self.yobs, self.expectancy, self.N_n)

        # Initialize guardrails to ensure proper configuration.
        self._adjustedParameters = False
        self.timeListsFileName = "None"
        self.timeLists = {}
        self.houseLists = {}
        self.rawHFP = {}  # raw dict of DataFrames from the HFP xlsx (horizon-independent)
        self.zeroWagesAndContributions()
        self.caseStatus = "unsolved"
        # "monotonic", "oscillatory", "max iteration", or "undefined" - how solution was obtained
        self.convergenceType = "undefined"
        self.rateMethod = None
        self.reproducibleRates = False
        self.rateSeed = None
        self.rateReverse = False
        self.rateRoll = 0

        # for plugins and core models
        self.rateModel = None
        self.rateMethodFile = None

        self.ARCoord = None
        self.objective = "unknown"

        # Placeholders values used to check if properly configured.
        self.xi_n = None
        self.alpha_ijkn = None

        return None

    def setLogger(self, logger):
        self.mylog = logger

    def setLogstreams(self, verbose, logstreams):
        self.mylog = log.Logger(verbose, logstreams)
        self.mylog.vprint(f"Setting logger with logstreams {logstreams}.")

    def logger(self):
        return self.mylog

    def setVerbose(self, state=True):
        """
        Control verbosity of calculations. True or False for now.
        Return previous state of verbosity.
        -``state``: Boolean selecting verbosity level.
        """
        return self.mylog.setVerbose(state)

    def _setStartingDate(self, mydate):
        """
        Set the date when the case starts in the current year.
        This is mostly for reproducibility purposes and back projecting known balances to Jan 1st.
        String format of mydate is 'MM/DD', 'MM-DD', 'YYYY-MM-DD', or 'YYYY/MM/DD'. Year is ignored.
        """
        import calendar

        thisyear = date.today().year

        if isinstance(mydate, date):
            mydate = mydate.strftime("%Y-%m-%d")

        if mydate is None or mydate == "today":
            refdate = date.today()
            self.startDate = refdate.strftime("%Y-%m-%d")
        else:
            mydatelist = mydate.replace("/", "-").split("-")
            if len(mydatelist) == 2 or len(mydatelist) == 3:
                self.startDate = mydate
                # Ignore the year provided.
                refdate = date(thisyear, int(mydatelist[-2]), int(mydatelist[-1]))
            else:
                raise ValueError('Date must be "MM-DD" or "YYYY-MM-DD".')

        lp = calendar.isleap(thisyear)
        # Take midnight as the reference.
        self.yearFracLeft = 1 - (refdate.timetuple().tm_yday - 1) / (365 + lp)

        self.mylog.vprint(f"Setting 1st-year starting date to {self.startDate}.")

        return None

    def _checkValueType(self, value):
        """
        Short utility function to parse and check arguments for plotting.
        """
        if value is None:
            return self.defaultPlots

        opts = ("nominal", "today")
        if value not in opts:
            raise ValueError(f"Value type must be one of: {opts}")

        return value

    def rename(self, newname):
        """
        Override name of the case. Case name is used
        to distinguish graph outputs and as base name for
        saving configurations and workbooks.
        """
        self.mylog.vprint(f"Renaming case '{self._name}' -> '{newname}'.")
        self._name = newname

    def setDescription(self, description):
        """
        Set a text description of the case.
        """
        self._description = description

    def setSexes(self, sexes):
        """
        Set the biological sex for each individual, used for SSA mortality table lookups.

        Parameters
        ----------
        sexes : list of str or None
            'M' (male) or 'F' (female) for each individual.  If None, defaults
            are preserved ('M' for all).  Must have N_i entries.
        """
        if sexes is None:
            return
        if len(sexes) != self.N_i:
            raise ValueError(f"sexes must have {self.N_i} entries, got {len(sexes)}.")
        for s in sexes:
            if s not in ("M", "F"):
                raise ValueError(f"Each sex must be 'M' or 'F', got {s!r}.")
        self.sexes = list(sexes)

    def setMortalityTable(self, table):
        """
        Select the actuarial mortality table used for longevity risk sampling.

        Parameters
        ----------
        table : str
            One of "SSA2025", "RP2014", "IAM2012", "VBT2015-NS", "VBT2015-SM".
        """
        from .data.mortality_tables import MORTALITY_TABLE_KEYS
        if table not in MORTALITY_TABLE_KEYS:
            raise ValueError(f"Unknown mortality table {table!r}. Valid: {MORTALITY_TABLE_KEYS}.")
        self.mortality_table = table

    def setSpousalDepositFraction(self, eta):
        """
        Set spousal deposit and withdrawal fraction. Default 0.5.
        Fraction eta is use to split surplus deposits between spouses as
        d_0n = (1 - eta)*s_n,
        and
        d_1n = eta*s_n,
        where s_n is the surplus amount. Here d_0n is the taxable account
        deposit for the first spouse while d_1n is for the second spouse.
        """
        if not (0 <= eta <= 1):
            raise ValueError("Fraction must be between 0 and 1.")
        if self.N_i != 2:
            self.mylog.print("Deposit fraction can only be 0 for single individuals.")
            eta = 0
        else:
            self.mylog.vprint(f"Setting spousal surplus deposit fraction to {eta:.1f}.")
            self.mylog.vprint(f"\t{self.inames[0]}: {1-eta:.1f}, {self.inames[1]}: {eta:.1f}")
            self.eta = eta

    def setDefaultPlots(self, value):
        """
        Set plots between nominal values or today's $.
        """

        self.defaultPlots = self._checkValueType(value)
        self.mylog.vprint(f"Setting plots default value to '{value}'.")

    def setWorksheetShowAges(self, value):
        """Enable or disable age columns in Streamlit worksheet tables."""
        self.worksheetShowAges = bool(value)
        self.mylog.vprint(f"Setting worksheet show ages to {self.worksheetShowAges}.")

    def setWorksheetHideZeroColumns(self, value):
        """Enable or hide all-zero numeric columns in Streamlit worksheet tables."""
        self.worksheetHideZeroColumns = bool(value)
        self.mylog.vprint(f"Setting worksheet hide zero columns to {self.worksheetHideZeroColumns}.")

    def setWorksheetRealDollars(self, value):
        """Enable or disable real-dollar (inflation-adjusted) worksheet display and save."""
        self.worksheetRealDollars = bool(value)
        self.mylog.vprint(f"Setting worksheet real dollars to {self.worksheetRealDollars}.")

    def setPlotBackend(self, backend: str):
        """
        Set plotting backend.
        """

        if backend not in ("matplotlib", "plotly"):
            raise ValueError(f"Backend '{backend}' not a valid option.")

        if backend != self._plotterName:
            self._plotter = PlotFactory.createBackend(backend)
            self._plotterName = backend
            self.mylog.vprint(f"Setting plotting backend to '{backend}'.")

    def setDividendRate(self, mu):
        """
        Set dividend tax rate. Rate is in percent. Default 1.8%.
        """
        if not (0 <= mu <= 5):
            raise ValueError("Rate must be between 0 and 5.")
        mu /= 100
        self.mylog.vprint(f"Dividend tax rate set to {u.pc(mu, f=0)}.")
        self.mu = mu
        self.caseStatus = "modified"

    def setExpirationYearOBBBA(self, yOBBBA):
        """
        Set year at which OBBBA is speculated to expire and rates go back to something like pre-TCJA.
        """
        self.mylog.vprint(f"Setting OBBBA expiration year to {yOBBBA}.")
        self.yOBBBA = yOBBBA
        self.caseStatus = "modified"
        self._adjustedParameters = False

    def setBeneficiaryFractions(self, phi):
        """
        Set fractions of savings accounts that is left to surviving spouse.
        Default is [1, 1, 1, 1] for taxable, tax-deferred, tax-free, and HSA accounts.
        A 3-element list (legacy) is auto-extended with 1.0 for the HSA account.
        """
        if len(phi) == self.N_j - 1:
            phi = list(phi) + [1.0]
        if len(phi) != self.N_j:
            raise ValueError(f"Fractions must have {self.N_j} entries.")
        for j in range(self.N_j):
            if not (0 <= phi[j] <= 1):
                raise ValueError("Fractions must be between 0 and 1.")
        self.phi_j = np.array(phi, dtype=np.float32)
        self.mylog.vprint("Spousal beneficiary fractions set to",
                          ["{:.2f}".format(self.phi_j[j]) for j in range(self.N_j)])
        self.caseStatus = "modified"

        if np.any(self.phi_j != 1):
            self.mylog.print("Consider changing spousal deposit fraction for better convergence.")
            self.mylog.print(f"\tRecommended: setSpousalDepositFraction({self.i_d}.)")

    def setHeirsTaxRate(self, nu):
        """
        Set the heirs tax rate on the tax-deferred portion of the estate.
        Rate is in percent. Default is 30%.
        """
        if not (0 <= nu <= 100):
            raise ValueError("Rate must be between 0 and 100.")
        nu /= 100
        self.mylog.vprint(f"Heirs tax rate on tax-deferred portion of estate set to {u.pc(nu, f=0)}.")
        self.nu = nu
        self.caseStatus = "modified"

    def setEffectiveTaxRate(self, rate):
        """
        Set the effective tax rate used to discount tax-deferred assets when computing
        the spending-to-savings ratio. Rate is in percent. Default is 20%.
        """
        if not (0 <= rate <= 100):
            raise ValueError("Rate must be between 0 and 100.")
        self.effectiveTaxRate = rate / 100
        self.mylog.vprint(
            f"Effective tax rate for spending-to-savings ratio set to {u.pc(self.effectiveTaxRate, f=0)}."
        )

    def _after_tax_savings(self, etr=None):
        """Return total after-tax initial savings, discounting tax-deferred by effective tax rate."""
        if etr is None:
            etr = self.effectiveTaxRate
        taxable = np.sum(self.beta_ij[:, 0])
        tax_deferred = np.sum(self.beta_ij[:, 1])
        roth = np.sum(self.beta_ij[:, 2])
        hsa = np.sum(self.beta_ij[:, 3])
        return taxable + tax_deferred * (1 - etr) + roth + hsa

    def setPension(self, amounts, ages, indexed=None, survivor_fraction=None):
        """
        Set value of pension for each individual and commencement age.
        Units are in $.

        Parameters
        ----------
        amounts : array-like
            Monthly pension amounts per individual ($)
        ages : array-like
            Commencement ages per individual
        indexed : list of bool, optional
            Whether each pension is inflation-indexed
        survivor_fraction : list of float, optional
            Fraction of pension continuing to surviving spouse (0-1).
            Default: [0] * N_i (single-life, no survivor).
        """
        if len(amounts) != self.N_i:
            raise ValueError(f"Amounts must have {self.N_i} entries.")
        if len(ages) != self.N_i:
            raise ValueError(f"Ages must have {self.N_i} entries.")
        if indexed is None:
            indexed = [False] * self.N_i
        if survivor_fraction is None:
            survivor_fraction = [0.0] * self.N_i
        if len(survivor_fraction) != self.N_i:
            raise ValueError(f"survivor_fraction must have {self.N_i} entries.")

        self.mylog.vprint("Setting monthly pension of", [u.d(amounts[i]) for i in range(self.N_i)],
                          "at age(s)", [int(ages[i]) for i in range(self.N_i)])

        thisyear = date.today().year
        self.pi_in = pension.compute_pension_benefits(
            amounts, ages, self.yobs, self.mobs, self.horizons,
            self.N_i, self.N_n, thisyear=thisyear
        )

        self.pensionAmounts = np.array(amounts, dtype=np.int32)
        self.pensionAges = np.array(ages)
        self.pensionIsIndexed = indexed
        self.pensionSurvivorFraction = np.array(survivor_fraction, dtype=np.float64)
        self.caseStatus = "modified"
        self._adjustedParameters = False

    def setSocialSecurity(self, pias, ages, trim_pct=0, trim_year=None, tax_fraction=None):
        """
        Set value of social security for each individual and claiming age.

        Note: Social Security benefits are paid in arrears (one month after eligibility).
        The zeta_in array represents when checks actually arrive, not when eligibility starts.

        Args:
            tax_fraction: Optional fixed SS taxability fraction in [0, 1] (default None).
                If provided, this overrides the self-consistent-loop computation of Psi_n
                and forces a constant value throughout the planning horizon.
                Useful for testing and for households whose provisional income (PI) is
                well within a single IRS bracket:
                  - 0.0  if PI < $32k (MFJ) / $25k (single) — benefits fully non-taxable
                  - 0.5  if PI is in the $32k–$44k (MFJ) / $25k–$34k (single) range
                  - 0.85 if PI > $44k (MFJ) / $34k (single) — up to 85% taxable (default)
                When None (default), Psi_n is computed dynamically each SC-loop iteration.
        """
        if len(pias) != self.N_i:
            raise ValueError(f"Principal Insurance Amount must have {self.N_i} entries.")
        if len(ages) != self.N_i:
            raise ValueError(f"Ages must have {self.N_i} entries.")

        if trim_pct != 0:
            if not (0 <= trim_pct <= 100):
                raise ValueError(f"trim_pct {trim_pct} outside range [0, 100].")
            if trim_year is None:
                raise ValueError("trim_year required when trim_pct > 0.")
            if not isinstance(trim_year, int):
                raise ValueError("trim_year must be an integer.")

        if tax_fraction is not None:
            if not (0 <= tax_fraction <= 1):
                raise ValueError(f"tax_fraction {tax_fraction} outside range [0, 1].")

        pias = np.array(pias, dtype=np.int32)
        ages = np.array(ages)
        ages_orig = ages.copy()

        fras = socsec.getFRAs(self.yobs, self.mobs, self.tobs)
        self.mylog.vprint("SS monthly PIAs set to", [u.d(pias[i]) for i in range(self.N_i)])
        self.mylog.vprint("SS FRAs(s)", [fras[i] for i in range(self.N_i)])

        thisyear = date.today().year
        self.zeta_in, ages = socsec.compute_social_security_benefits(
            pias, ages, self.yobs, self.mobs, self.tobs, self.horizons,
            self.N_i, self.N_n, trim_pct=trim_pct, trim_year=trim_year, thisyear=thisyear
        )

        for i in range(self.N_i):
            if ages[i] != ages_orig[i]:
                eligible = 62 if (self.tobs[i] <= 2) else 62 + 1/12
                self.mylog.print(f"Resetting SS claiming age of {self.inames[i]} to {eligible}.")

        self.mylog.vprint("SS benefits claimed at age(s)", [ages[i] for i in range(self.N_i)])

        if trim_pct > 0:
            self.mylog.print(f"Reducing Social Security by {trim_pct}% starting in year {trim_year}.")

        self.ssecAmounts = pias
        self.ssecAges = ages
        self.ssecTrimPct = trim_pct
        self.ssecTrimYear = trim_year
        self.ssecTaxFraction = tax_fraction
        self.caseStatus = "modified"
        self._adjustedParameters = False

    def setSpendingProfile(self, profile, percent=60, dip=15, increase=12, delay=0):
        """
        Generate time series for spending profile. Surviving spouse fraction can be specified
        as a second argument. Default value is 60%.
        Dip and increase are percent changes in the smile profile.
        """
        if not (0 <= percent <= 100):
            raise ValueError(f"Survivor value {percent} outside range.")
        if not (0 <= dip <= 100):
            raise ValueError(f"Dip value {dip} outside range.")
        if not (-100 <= increase <= 100):
            raise ValueError(f"Increase value {increase} outside range.")
        if not (0 <= delay <= self.N_n - 2):
            raise ValueError(f"Delay value {delay} outside year range.")

        self.chi = percent / 100

        self.mylog.vprint("Setting", profile, "spending profile.")
        if self.N_i == 2:
            self.mylog.vprint("Securing", u.pc(self.chi, f=0), "of spending amount for surviving spouse.")

        self.xi_n = spending.gen_spending_profile(
            profile, self.chi, self.n_d, self.N_n, dip=dip, increase=increase, delay=delay
        )

        self.spendingProfile = profile
        self.smileDip = dip
        self.smileIncrease = increase
        self.smileDelay = delay
        self.caseStatus = "modified"

    def setReproducible(self, reproducible, seed=None):
        """
        Set whether rates should be reproducible for stochastic methods.
        This should be called before setting rates. It only sets configuration
        and does not regenerate existing rates.

        Args:
            reproducible: Boolean indicating if rates should be reproducible.
            seed: Optional seed value. If None and reproducible is True,
                  generates a new seed from current time. If None and
                  reproducible is False, generates a seed but won't reuse it.
        """
        self.reproducibleRates = bool(reproducible)
        if reproducible:
            if seed is None:
                if self.rateSeed is not None:
                    # Reuse existing seed if available
                    seed = self.rateSeed
                else:
                    # Generate new seed from current time
                    seed = int(time.time() * 1_000_000)  # Use microseconds
            else:
                seed = int(seed)
            self.rateSeed = seed
        else:
            # For non-reproducible rates, clear the seed
            # setRates() will generate a new seed each time it's called
            self.rateSeed = None

    def setRates(
        self,
        method,
        frm=None,
        to=None,
        values=None,
        stdev=None,
        corr=None,
        df=None,
        method_file=None,
        override_reproducible=False,
        reverse=False,
        roll=0,
        **kwargs,
    ):
        """
        Generate rates using pluggable rate model architecture.

        Fully metadata-driven:
            - No method-specific filtering
            - Validation handled inside each RateModel
            - Supports built-in and plugin models

        Unit convention:
            values: rates in percent (e.g. 7.0 = 7%), matching the format
                returned by getRatesDistributions() and consistent with
                setDividendRate/setHeirsTaxRate.
            stdev: standard deviations in percent (e.g. 17.0 = 17%).
        """

        # --------------------------------------------------
        # Determine seed handling
        # --------------------------------------------------

        if self.reproducibleRates and not override_reproducible:
            seed = self.rateSeed
        elif override_reproducible:
            seed = int(time.time() * 1_000_000)
        else:
            seed = None

        # --------------------------------------------------
        # Legacy compatibility: historical shorthand
        # --------------------------------------------------

        if method in HISTORICAL_RANGE_METHODS:
            if frm is not None and to is None:
                to = frm + self.N_n - 1

        # --------------------------------------------------
        # Build model configuration dictionary
        # --------------------------------------------------

        model_config = {"method": method}

        # Only include parameters that are not None
        base_args = {
            "frm": frm,
            "to": to,
            "values": values,
            "stdev": stdev,
            "corr": corr,
            "df": df,
        }

        for k, v in base_args.items():
            if v is not None:
                model_config[k] = v

        # Include any additional keyword arguments
        model_config.update(kwargs)

        if method == "dataframe":
            model_config["n_years"] = self.N_n

        # --------------------------------------------------
        # Load rate model class
        # --------------------------------------------------

        from owlplanner.rate_models.loader import load_rate_model

        ModelClass = load_rate_model(method, method_file)

        model = ModelClass(
            config=model_config,
            seed=seed,
            logger=self.mylog,
        )

        # --------------------------------------------------
        # Generate series
        # --------------------------------------------------

        series = model.generate(self.N_n)

        if series.shape != (self.N_n, 4):
            raise RuntimeError(
                f"Rate model returned shape {series.shape}, expected ({self.N_n}, 4)"
            )

        # --------------------------------------------------
        # Store model + metadata
        # --------------------------------------------------

        self.rateModel = model
        self.rateMethod = method
        self.rateMethodFile = method_file
        self.rateReverse = bool(reverse)
        self.rateRoll = int(roll)

        # Store frm/to if present in model config
        self.rateFrm = model.config.get("frm")
        self.rateTo = model.config.get("to")

        # Backward compatibility fields (for built-in stochastic/user)
        self.rateValues = model.params.get("values")
        self.rateStdev = model.params.get("stdev")
        self.rateCorr = model.params.get("corr")

        if self.rateValues is not None:
            self.rateValues = np.array(self.rateValues)

        if self.rateStdev is not None:
            self.rateStdev = np.array(self.rateStdev)

        if self.rateCorr is not None:
            self.rateCorr = np.array(self.rateCorr)

        # --------------------------------------------------
        # Apply reverse / roll
        # --------------------------------------------------

        # model.generate returns (N, 4)
        # tau_kn must be (4, N)
        series_kn = series.transpose()

        if getattr(model, "constant", False):
            if reverse or roll != 0:
                self.mylog.print(
                    "Warning: reverse and roll are ignored for constant (fixed) rate methods."
                )
        else:
            series_kn = rates.apply_rate_sequence_transform(
                series_kn,
                reverse,
                roll,
            )

        self.tau_kn = series_kn

        # --------------------------------------------------
        # Inflation multiplier
        # --------------------------------------------------

        self.gamma_n = rates.gen_gamma_n(self.tau_kn)

        self._adjustedParameters = False
        self.caseStatus = "modified"

        self.mylog.vprint(f"Generated {self.N_n} years of rates using model '{method}'.")

    def regenRates(self, override_reproducible=False):
        """
        Regenerate stochastic rate series using stored model.
        """

        if not hasattr(self, "rateModel") or self.rateModel is None:
            return

        # Do not regenerate deterministic models
        if getattr(self.rateModel, "deterministic", False):
            return

        # Respect reproducibility setting
        if self.reproducibleRates and not override_reproducible:
            return

        # Generate new series
        series = self.rateModel.generate(self.N_n)

        if series.shape != (self.N_n, 4):
            raise RuntimeError(
                f"Rate model returned shape {series.shape}, expected ({self.N_n}, 4)"
            )

        series_kn = series.transpose()

        if not getattr(self.rateModel, "constant", False):
            series_kn = rates.apply_rate_sequence_transform(
                series_kn,
                self.rateReverse,
                self.rateRoll,
            )

        self.tau_kn = series_kn
        self.gamma_n = rates.gen_gamma_n(self.tau_kn)

        self.mylog.vprint("Regenerated stochastic rate series.")

    def setAccountBalances(self, *, taxable, taxDeferred, taxFree, hsa=None, startDate=None, units="k"):
        """
        Three lists (plus optional HSA) containing the balance of all assets in each category
        for each spouse.  For single individuals, these lists will contain only one entry.
        Units are in $k, unless specified otherwise: 'k', 'M', or '1'.
        """
        plurals = ["", "y", "ies"][self.N_i]
        if len(taxable) != self.N_i:
            raise ValueError(f"taxable must have {self.N_i} entr{plurals}.")
        if len(taxDeferred) != self.N_i:
            raise ValueError(f"taxDeferred must have {self.N_i} entr{plurals}.")
        if len(taxFree) != self.N_i:
            raise ValueError(f"taxFree must have {self.N_i} entr{plurals}.")

        fac = u.getUnits(units)
        taxable = u.rescale(taxable, fac)
        taxDeferred = u.rescale(taxDeferred, fac)
        taxFree = u.rescale(taxFree, fac)

        self.bet_ji = np.zeros((self.N_j, self.N_i))
        self.bet_ji[0][:] = taxable
        self.bet_ji[1][:] = taxDeferred
        self.bet_ji[2][:] = taxFree
        if hsa is not None:
            if len(hsa) != self.N_i:
                raise ValueError(f"hsa must have {self.N_i} entr{plurals}.")
            self.bet_ji[3][:] = [v * fac for v in hsa]
        self.beta_ij = self.bet_ji.transpose()

        # If none was given, default is to begin plan on today's date.
        self._setStartingDate(startDate)

        self.caseStatus = "modified"

        self.mylog.vprint("Taxable balances:", *[u.d(taxable[i]) for i in range(self.N_i)])
        self.mylog.vprint("Tax-deferred balances:", *[u.d(taxDeferred[i]) for i in range(self.N_i)])
        self.mylog.vprint("Tax-free balances:", *[u.d(taxFree[i]) for i in range(self.N_i)])
        if hsa is not None:
            hsa_arr = self.bet_ji[3]
            self.mylog.vprint("HSA balances:", *[u.d(hsa_arr[i]) for i in range(self.N_i)])
        self.mylog.vprint("Sum of all savings accounts:", u.d(np.sum(taxable) + np.sum(taxDeferred) + np.sum(taxFree)))
        self.mylog.vprint(
            "Post-tax total wealth of approximately",
            u.d(np.sum(taxable) + 0.7 * np.sum(taxDeferred) + np.sum(taxFree)),
        )

    def setHSA(self, balances, medicare_ages=None, units="k"):
        """
        Set HSA (Health Savings Account) initial balances and Medicare enrollment ages.
        HSA contributions stop when Medicare enrollment begins (~age 65).

        Parameters
        ----------
        balances : list
            Initial HSA balance per individual (in $k by default).
        medicare_ages : list, optional
            Age at which HSA contributions stop for each individual (default: 65).
        units : str, optional
            Units for balances: 'k' (default), 'M', or '1'.
        """
        self.setAccountBalances(
            taxable=list(self.bet_ji[0]),
            taxDeferred=list(self.bet_ji[1]),
            taxFree=list(self.bet_ji[2]),
            hsa=balances,
            units=units,
        )
        thisyear = date.today().year
        ages = medicare_ages if medicare_ages is not None else [65] * self.N_i
        for i in range(self.N_i):
            n_hsa = self.yobs[i] + ages[i] - thisyear
            self.n_hsa_i[i] = min(max(0, n_hsa), self.N_n)
        self.mylog.vprint("HSA contribution stop years:", [int(self.n_hsa_i[i]) for i in range(self.N_i)])

    def setACA(self, slcsp, units="k"):
        """
        Configure ACA marketplace health insurance premium for pre-Medicare years.

        Sets the annual benchmark Silver plan (SLCSP) premium for this household.
        The Premium Tax Credit reduces this by the amount that household income exceeds
        the required self-contribution (a piecewise-linear % of MAGI keyed to FPL).
        ACA costs are only assessed in years where at least one individual is under 65
        and within their planning horizon.

        Parameters
        ----------
        slcsp : float or list of float
            Annual benchmark Silver plan premium in today's dollars (default units: $k).
            If a scalar, applied uniformly across all plan years (inflation-adjusted).
            If a list of length N_n, used as-is (each entry inflated for that year).
        units : str
            Unit multiplier: 'k' ($k, default), 'M' ($M), '1' (dollars).
        """
        fac = u.getUnits(units)
        if np.isscalar(slcsp):
            self.slcsp_annual = float(slcsp) * fac
            self.mylog.vprint(f"ACA benchmark premium set to ${self.slcsp_annual / 1000:.1f}k/year (today's $).")
        else:
            raise ValueError("setACA: slcsp must be a scalar (today's $). For per-year amounts use a list "
                             "with a future per-year API.")
        self.caseStatus = "modified"

    def setInterpolationMethod(self, method, center=15, width=5):
        """
        Interpolate asset allocation ratios from initial value (today) to
        final value (at the end of horizon).

        Two interpolation methods are supported: linear and s-curve.
        Linear is a straight line between now and the end of the simulation.
        Hyperbolic tangent give a smooth "S" curve centered at point "center"
        with a width "width". Center point defaults to 15 years and width to
        5 years. This means that the transition from initial to final
        will start occuring in 10 years (15-5) and will end in 20 years (15+5).
        """
        if method == "linear":
            self._interpolator = self._linInterp
        elif method == "s-curve":
            self._interpolator = self._tanhInterp
            self.interpCenter = center
            self.interpWidth = width
        else:
            raise ValueError(f"Method '{method}' not supported.")

        self.interpMethod = method
        self.caseStatus = "modified"

        self.mylog.vprint(f"Asset allocation interpolation method set to '{method}'.")

    def setAllocationRatios(self, allocType, taxable=None, taxDeferred=None,  # noqa: C901
                            taxFree=None, hsa=None, generic=None):
        """
        Single function for setting all types of asset allocations.
        Allocation types are 'account', 'individual', and 'spouses'.

        For 'account' the three different account types taxable, taxDeferred,
        qand taxFree need to be set to a list. For spouses,
        taxable = [[[ko00, ko01, ko02, ko03], [kf00, kf01, kf02, kf02]],
        [[ko10, ko11, ko12, ko13], [kf10, kf11, kf12, kf12]]]
        where ko is the initial allocation while kf is the final.
        The order of [initial, final] pairs is the same as for the birth
        years and longevity provided. Single only provide one pair for each
        type of savings account.

        For the 'individual' allocation type, only one generic list needs
        to be provided:
        generic = [[[ko00, ko01, ko02, ko03], [kf00, kf01, kf02, kf02]],
        [[ko10, ko11, ko12, ko13], [kf10, kf11, kf12, kf12]]].
        while for 'spouses' only one pair needs to be given as follows:
        generic = [[ko00, ko01, ko02, ko03], [kf00, kf01, kf02, kf02]]
        as assets are coordinated between accounts and spouses.
        """
        # Validate allocType parameter
        validTypes = ["account", "individual", "spouses"]
        if allocType not in validTypes:
            raise ValueError(f"allocType must be one of {validTypes}, got '{allocType}'.")

        self.boundsAR = {}
        self.alpha_ijkn = np.zeros((self.N_i, self.N_j, self.N_k, self.N_n + 1))
        if allocType == "account":
            # Make sure we have proper input.
            for item in [taxable, taxDeferred, taxFree]:
                if len(item) != self.N_i:
                    raise ValueError(f"{item} must have one entry per individual.")
                for i in range(self.N_i):
                    # Initial and final.
                    if len(item[i]) != 2:
                        raise ValueError(f"{item}[{i}] must have 2 lists (initial and final).")
                    for z in range(2):
                        if len(item[i][z]) != self.N_k:
                            raise ValueError(f"{item}[{i}][{z}] must have {self.N_k} entries.")
                        if abs(sum(item[i][z]) - 100) > 0.01:
                            raise ValueError("Sum of percentages must add to 100.")

            for i in range(self.N_i):
                self.mylog.vprint(f"{self.inames[i]}: Setting gliding allocation ratios (%) to '{allocType}'.")
                self.mylog.vprint(f"      taxable: {taxable[i][0]} -> {taxable[i][1]}")
                self.mylog.vprint(f"  taxDeferred: {taxDeferred[i][0]} -> {taxDeferred[i][1]}")
                self.mylog.vprint(f"      taxFree: {taxFree[i][0]} -> {taxFree[i][1]}")

            # Order in alpha is j, i, 0/1, k.
            alpha = {}
            alpha[0] = np.array(taxable)
            alpha[1] = np.array(taxDeferred)
            alpha[2] = np.array(taxFree)
            alpha[3] = np.array(hsa) if hsa is not None else np.array(taxFree)  # HSA inherits tax-free by default
            for i in range(self.N_i):
                Nin = self.horizons[i] + 1
                for j in range(self.N_j):
                    for k in range(self.N_k):
                        start = alpha[j][i, 0, k] / 100
                        end = alpha[j][i, 1, k] / 100
                        dat = self._interpolator(start, end, Nin)
                        self.alpha_ijkn[i, j, k, :Nin] = dat[:]

            self.boundsAR["taxable"] = taxable
            self.boundsAR["tax-deferred"] = taxDeferred
            self.boundsAR["tax-free"] = taxFree
            self.boundsAR["hsa"] = hsa if hsa is not None else taxFree

        elif allocType == "individual":
            if len(generic) != self.N_i:
                raise ValueError("generic must have one list per individual.")
            for i in range(self.N_i):
                # Initial and final.
                if len(generic[i]) != 2:
                    raise ValueError(f"generic[{i}] must have 2 lists (initial and final).")
                for z in range(2):
                    if len(generic[i][z]) != self.N_k:
                        raise ValueError(f"generic[{i}][{z}] must have {self.N_k} entries.")
                    if abs(sum(generic[i][z]) - 100) > 0.01:
                        raise ValueError("Sum of percentages must add to 100.")

            for i in range(self.N_i):
                self.mylog.vprint(f"{self.inames[i]}: Setting gliding allocation ratios (%) to '{allocType}'.")
                self.mylog.vprint(f"\t{generic[i][0]} -> {generic[i][1]}")

            for i in range(self.N_i):
                Nin = self.horizons[i] + 1
                for k in range(self.N_k):
                    start = generic[i][0][k] / 100
                    end = generic[i][1][k] / 100
                    dat = self._interpolator(start, end, Nin)
                    self.alpha_ijkn[i, :, k, :Nin] = dat[:]

            self.boundsAR["generic"] = generic

        elif allocType == "spouses":
            if len(generic) != 2:
                raise ValueError("generic must have 2 entries (initial and final).")
            for z in range(2):
                if len(generic[z]) != self.N_k:
                    raise ValueError(f"generic[{z}] must have {self.N_k} entries.")
                if abs(sum(generic[z]) - 100) > 0.01:
                    raise ValueError("Sum of percentages must add to 100.")

            for i in range(self.N_i):
                Nin = self.horizons[i] + 1
                for k in range(self.N_k):
                    start = generic[0][k] / 100
                    end = generic[1][k] / 100
                    dat = self._interpolator(start, end, Nin)
                    self.alpha_ijkn[i, :, k, :Nin] = dat[:]

            self.boundsAR["generic"] = generic

        self.ARCoord = allocType
        self.caseStatus = "modified"

        self.mylog.vprint(f"Interpolating asset allocation ratios using '{self.interpMethod}' method.")

    def readHFP(self, filename, filename_for_logging=None):
        """
        Load the Household Financial Profile (HFP) from file.

        The HFP file contains wages, contributions, Roth conversions,
        big-ticket items (per individual), and optionally Debts and Fixed Assets.
        File can be an excel, or odt file with one tab named after each
        spouse and must have the following column headers (all required;
        use 0 where a concept does not apply):

                'year',
                'anticipated wages',
                'other inc',
                'net inv',
                'taxable ctrb',
                '401k ctrb',
                'Roth 401k ctrb',
                'IRA ctrb',
                'Roth IRA ctrb',
                'HSA ctrb',
                'Roth conv',
                'big-ticket items'

        in any order. Legacy header 'other inc.' is read as 'other inc'.
        Optional workbook sheets 'Debts' and 'Fixed Assets' follow timelists formats.
        A template is provided as an example.
        Missing rows (years) are populated with zero values.

        Parameters
        ----------
        filename : file-like object, str, or dict
            Input file or dictionary of DataFrames
        filename_for_logging : str, optional
            Explicit filename for logging purposes. If provided, this will be used
            in log messages instead of trying to extract it from filename.
        """
        try:
            returned_filename, self.timeLists, self.houseLists, self.rawHFP = timelists.read(
                filename, self.inames, self.horizons, self.mylog, filename=filename_for_logging
            )
        except Exception as e:
            raise Exception(f"Unsuccessful read of Household Financial Profile: {e}") from e

        # Use filename_for_logging if provided, otherwise use returned filename
        self.timeListsFileName = filename_for_logging if filename_for_logging is not None else returned_filename
        self.setContributions()

        return True

    def setContributions(self, timeLists=None):
        """
        If no argument is given, use the values that have been stored in self.timeLists.
        """
        if timeLists is not None:
            self.timeLists = timeLists

        # Now fill in parameters which are in $.
        for i, iname in enumerate(self.inames):
            h = self.horizons[i]
            self.omega_in[i, :h] = self.timeLists[iname]["anticipated wages"].iloc[5:5+h]
            self.other_inc_in[i, :h] = self.timeLists[iname]["other inc"].iloc[5:5+h]
            self.netinv_in[i, :h] = self.timeLists[iname]["net inv"].iloc[5:5+h]
            self.Lambda_in[i, :h] = self.timeLists[iname]["big-ticket items"].iloc[5:5+h]

            # Values for last 5 years of Roth conversion and contributions stored at the end
            # of array and accessed with negative index.
            self.kappa_ijn[i, 0, :h] = self.timeLists[iname]["taxable ctrb"][5:h+5]
            self.kappa_ijn[i, 1, :h] = self.timeLists[iname]["401k ctrb"][5:h+5]
            self.kappa_ijn[i, 1, :h] += self.timeLists[iname]["IRA ctrb"][5:h+5]
            self.kappa_ijn[i, 2, :h] = self.timeLists[iname]["Roth 401k ctrb"][5:h+5]
            self.kappa_ijn[i, 2, :h] += self.timeLists[iname]["Roth IRA ctrb"][5:h+5]
            self.kappa_ijn[i, 3, :h] = self.timeLists[iname]["HSA ctrb"][5:h+5]
            # Zero HSA contributions after Medicare enrollment year.
            # If n_hsa_i was never set (still at default N_n), initialize from yobs and age 65
            # so programmatic plans that bypass config still get correct Medicare cutoff.
            if self.n_hsa_i[i] >= self.N_n:
                thisyear = date.today().year
                n_hsa = self.yobs[i] + 65 - thisyear
                self.n_hsa_i[i] = min(max(0, n_hsa), self.N_n)
            n_stop = self.n_hsa_i[i]
            if n_stop < h:
                self.kappa_ijn[i, 3, n_stop:h] = 0.0
            self.myRothX_in[i, :h] = self.timeLists[iname]["Roth conv"][5:h+5]

            # Last 5 years are at the end of the N_n array.
            self.kappa_ijn[i, 0, -5:] = self.timeLists[iname]["taxable ctrb"][:5]
            self.kappa_ijn[i, 1, -5:] = self.timeLists[iname]["401k ctrb"][:5]
            self.kappa_ijn[i, 1, -5:] += self.timeLists[iname]["IRA ctrb"][:5]
            self.kappa_ijn[i, 2, -5:] = self.timeLists[iname]["Roth 401k ctrb"][:5]
            self.kappa_ijn[i, 2, -5:] += self.timeLists[iname]["Roth IRA ctrb"][:5]
            self.kappa_ijn[i, 3, -5:] = self.timeLists[iname]["HSA ctrb"][:5]
            self.myRothX_in[i, -5:] = self.timeLists[iname]["Roth conv"][:5]

        self.caseStatus = "modified"

        return self.timeLists

    def processDebtsAndFixedAssets(self):
        """
        Process debts and fixed assets from houseLists and populate arrays.
        Should be called after setContributions() and before solve().
        """
        thisyear = date.today().year

        # Process debts
        if "Debts" in self.houseLists and not u.is_dataframe_empty(self.houseLists["Debts"]):
            self.debt_payments_n = debts.get_debt_payments_array(
                self.houseLists["Debts"], self.N_n, thisyear
            )
            self.remaining_debt_balance = debts.get_remaining_debt_balance(
                self.houseLists["Debts"], self.N_n, thisyear
            )
        else:
            self.debt_payments_n = np.zeros(self.N_n)
            self.remaining_debt_balance = 0.0

        # Process fixed assets
        if "Fixed Assets" in self.houseLists and not u.is_dataframe_empty(self.houseLists["Fixed Assets"]):
            filing_status = "married" if self.N_i == 2 else "single"
            (self.fixed_assets_tax_free_n,
             self.fixed_assets_ordinary_income_n,
             self.fixed_assets_capital_gains_n) = fxasst.get_fixed_assets_arrays(
                self.houseLists["Fixed Assets"], self.N_n, thisyear, filing_status
            )
            # Calculate bequest value for assets with yod past plan end
            self.fixed_assets_bequest_value = fxasst.get_fixed_assets_bequest_value(
                self.houseLists["Fixed Assets"], self.N_n, thisyear
            )
        else:
            self.fixed_assets_tax_free_n = np.zeros(self.N_n)
            self.fixed_assets_ordinary_income_n = np.zeros(self.N_n)
            self.fixed_assets_capital_gains_n = np.zeros(self.N_n)
            self.fixed_assets_bequest_value = 0.0

    def getFixedAssetsBequestValueInTodaysDollars(self):
        """
        Return the fixed assets bequest value in today's dollars.
        This requires rates to be set to calculate gamma_n (inflation factor).

        Returns:
        --------
        float
            Fixed assets bequest value in today's dollars. HFP monetary values
            (basis, value) are stored in dollars; the UI divides by 1000 for k$ display.
            Returns 0.0 if rates not set, gamma_n not calculated, or no fixed assets.
        """
        if self.fixed_assets_bequest_value == 0.0:
            return 0.0

        # Check if we can calculate gamma_n
        if self.rateMethod is None or not hasattr(self, 'tau_kn'):
            # Rates not set yet - return 0
            return 0.0

        # Calculate gamma_n if not already calculated
        if not hasattr(self, 'gamma_n') or self.gamma_n is None:
            self.gamma_n = rates.gen_gamma_n(self.tau_kn)

        # Convert: today's dollars = nominal dollars / inflation_factor
        return self.fixed_assets_bequest_value / self.gamma_n[-1]

    def saveContributions(self):
        """
        Return workbook on wages and contributions, including Debts and Fixed Assets.
        """
        if self.timeLists is None:
            return None

        self.mylog.vprint("Preparing wages and contributions workbook.")

        def fillsheet(sheet, i):
            sheet.title = self.inames[i]
            df = self.timeLists[self.inames[i]]
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)
            export._format_spreadsheet(sheet, "currency")

        wb = Workbook()
        ws = wb.active
        fillsheet(ws, 0)

        if self.N_i == 2:
            ws = wb.create_sheet(self.inames[1])
            fillsheet(ws, 1)

        # Add Debts sheet if available
        if "Debts" in self.houseLists and not u.is_dataframe_empty(self.houseLists["Debts"]):
            ws = wb.create_sheet("Debts")
            df = self.houseLists["Debts"]
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            export._format_debts_sheet(ws)
        else:
            # Create empty Debts sheet with proper columns
            ws = wb.create_sheet("Debts")
            df = pd.DataFrame(columns=timelists._debtItems)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            export._format_debts_sheet(ws)

        # Add Fixed Assets sheet if available
        if "Fixed Assets" in self.houseLists and not u.is_dataframe_empty(self.houseLists["Fixed Assets"]):
            ws = wb.create_sheet("Fixed Assets")
            df = self.houseLists["Fixed Assets"]
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            export._format_fixed_assets_sheet(ws)
        else:
            # Create empty Fixed Assets sheet with proper columns
            ws = wb.create_sheet("Fixed Assets")
            df = pd.DataFrame(columns=timelists._fixedAssetItems)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            export._format_fixed_assets_sheet(ws)

        return wb

    def zeroWagesAndContributions(self):
        """
        Zero wages, contributions, Roth conversions, and big-ticket items.
        Resets timeLists; does not modify Debts or Fixed Assets.
        """
        self.mylog.vprint("Resetting wages and contributions to zero.")

        # Reset parameters with zeros.
        self.omega_in[:, :] = 0.0
        self.other_inc_in[:, :] = 0.0
        self.netinv_in[:, :] = 0.0
        self.Lambda_in[:, :] = 0.0
        self.myRothX_in[:, :] = 0.0
        self.kappa_ijn[:, :, :] = 0.0

        cols = [
            "year",
            "anticipated wages",
            "other inc",
            "net inv",
            "taxable ctrb",
            "401k ctrb",
            "Roth 401k ctrb",
            "IRA ctrb",
            "Roth IRA ctrb",
            "HSA ctrb",
            "Roth conv",
            "big-ticket items",
        ]
        for i, iname in enumerate(self.inames):
            h = self.horizons[i]
            df = pd.DataFrame(0, index=np.arange(0, h+5), columns=cols)
            df["year"] = np.arange(self.year_n[0] - 5, self.year_n[h-1]+1)
            self.timeLists[iname] = df

        self.caseStatus = "modified"

        return self.timeLists

    def _linInterp(self, a, b, numPoints):
        """
        Utility function to interpolate allocations using
        a linear interpolation.
        """
        # num goes one more year as endpoint=True.
        return np.linspace(a, b, numPoints)

    def _tanhInterp(self, a, b, numPoints):
        """
        Utility function to interpolate allocations using a hyperbolic
        tangent interpolation. "c" is the year where the inflection point
        is happening, and "w" is the width of the transition.
        """
        c = self.interpCenter
        w = self.interpWidth + 0.0001  # Avoid division by zero.
        t = np.linspace(0, numPoints, numPoints)
        # Solve 2x2 system to match end points exactly.
        th0 = np.tanh((t[0] - c) / w)
        thN = np.tanh((t[numPoints - 1] - c) / w)
        k11 = 0.5 - 0.5 * th0
        k21 = 0.5 - 0.5 * thN
        k12 = 0.5 + 0.5 * th0
        k22 = 0.5 + 0.5 * thN
        _b = (b - (k21 / k11) * a) / (k22 - (k21 / k11) * k12)
        _a = (a - k12 * _b) / k11
        dat = _a + 0.5 * (_b - _a) * (1 + np.tanh((t - c) / w))

        return dat

    def _adjustParameters(self, gamma_n, MAGI_n):
        """
        Adjust parameters that follow inflation or depend on MAGI.
        Separate variables depending on MAGI (exemptions now depends on MAGI).
        """
        if self.rateMethod is None:
            raise RuntimeError("A rate method needs to be first selected using setRates(...).")

        self.sigmaBar_n, self.theta_tn, self.Delta_tn = tx.taxParams(self.yobs, self.i_d, self.n_d,
                                                                     self.N_n, gamma_n,
                                                                     MAGI_n, self.yOBBBA)

        if not self._adjustedParameters:
            self.mylog.vprint("Adjusting parameters for inflation.")
            self.DeltaBar_tn = self.Delta_tn * gamma_n[:-1]
            self.zetaBar_in = self.zeta_in * gamma_n[:-1]
            self.xiBar_n = self.xi_n * gamma_n[:-1]
            self.piBar_in = pension.compute_piBar_in(
                self.pi_in, gamma_n[:-1], self.pensionIsIndexed,
                self.pensionSurvivorFraction, self.n_d, self.i_d, self.i_s,
                self.horizons, self.N_i, self.N_n,
            )

            # Part D: include by default; base premium optional (monthly -> annual).
            self._include_medicare_part_d = self.solverOptions.get("includeMedicarePartD", True)
            part_d_base_monthly = self.solverOptions.get("medicarePartDBasePremium")
            self._medicare_part_d_base_annual_per_person = (
                float(part_d_base_monthly) * 12 if part_d_base_monthly is not None else 0.0
            )

            self.nm, self.Lbar_nq, self.Cbar_nq = tx.mediVals(
                self.yobs, self.horizons, gamma_n, self.N_n, self.N_irmaa,
                include_part_d=self._include_medicare_part_d,
                part_d_base_annual_per_person=self._medicare_part_d_base_annual_per_person,
            )

            if self.slcsp_annual > 0:
                self.n_aca, self.Lbar_aca_nr, self.cap_pct_aca_r, self.slcsp_aca_n = \
                    tx.acaVals(self.yobs, self.horizons, gamma_n, self.slcsp_annual, self.N_n)
            else:
                self.n_aca = 0

            self._adjustedParameters = True

        # return None

    def _buildOffsetMap(self, options):
        """
        Utility function to map variables to a block vector.
        Refer to companion document for explanations.
        All binary variables must be lumped at the end of the vector.
        """
        medi = options.get("withMedicare", "loop") == "optimize"
        ss_lp = options.get("withSSTaxability", "loop") == "optimize"
        aca_lp = options.get("withACA", "loop") == "optimize"
        ltcg_lp = options.get("withLTCG", "loop") == "optimize"
        niit_lp = options.get("withNIIT", "loop") == "optimize"
        # withSSAges: "fixed"/"none" → no opt; "optimize" → all;
        # individual name or list of names → optimize only those individuals.
        _ssa_opt = options.get("withSSAges", "fixed")
        if isinstance(_ssa_opt, (list, tuple)):
            _ssa_optimize_set = set()
            for name in _ssa_opt:
                try:
                    _ssa_optimize_set.add(self.inames.index(name))
                except ValueError as e:
                    raise ValueError(f"Unknown individual '{name}' for withSSAges:") from e
            self._ssa_optimize_set = _ssa_optimize_set
        elif _ssa_opt == "optimize":
            self._ssa_optimize_set = set(range(self.N_i))
        elif _ssa_opt not in ("fixed", "none"):
            # Single name string (e.g. "Jack").
            try:
                self._ssa_optimize_set = {self.inames.index(_ssa_opt)}
            except ValueError as e:
                raise ValueError(f"Unknown individual '{_ssa_opt}' for withSSAges:") from e
        else:
            self._ssa_optimize_set = set()
        ssa_lp = bool(self._ssa_optimize_set)
        self._aca_lp = aca_lp and self.slcsp_annual > 0 and self.n_aca > 0
        self._ltcg_lp = ltcg_lp
        self._niit_lp = niit_lp
        self._bigMltcg = options.get("bigMltcg", None)   # None → use T20_n per year
        self._bigMniit = options.get("bigMniit", None)   # None → use 3*T20_n per year
        Nmed = self.N_n - self.nm

        # SS claiming-age optimization: precompute benefit table and initialize SC offset.
        pias = getattr(self, "ssecAmounts", None)
        ssa_lp = ssa_lp and pias is not None and bool(np.any(pias > 0))
        self._ssa_lp = ssa_lp
        self._ssa_N_K = 97   # monthly grid from 62.0 to 70.0 (inclusive)
        if ssa_lp:
            fras = socsec.getFRAs(self.yobs, self.mobs, self.tobs)
            trim_pct = getattr(self, "ssecTrimPct", 0) or 0
            trim_year = getattr(self, "ssecTrimYear", None)
            self._ssa_B_own, self._ssa_ages_k = socsec.build_own_benefit_table(
                pias, fras, self.yobs, self.mobs, self.tobs, self.horizons,
                self.N_i, self.N_n, self.gamma_n[:-1],
                trim_pct=trim_pct, trim_year=trim_year,
                N_K=self._ssa_N_K, thisyear=None,
            )
            # Initialize spousal/survivor offset from initial claiming ages.
            # B_own at initial claiming age k_init; offset = total zetaBar - own benefit.
            self._ssa_spousal_offset = np.zeros((self.N_i, self.N_n))
            for i in range(self.N_i):
                k_init = int(round((float(self.ssecAges[i]) - 62.0) * 12))
                k_init = max(0, min(self._ssa_N_K - 1, k_init))
                self._ssa_spousal_offset[i, :] = self.zetaBar_in[i, :] - self._ssa_B_own[i, k_init, :]

        # Stack all variables in a single block vector with all binary variables at the end.
        vm = VarMap()
        vm.add("b", self.N_i, self.N_j, self.N_n + 1)
        vm.add("d", self.N_i, self.N_n)
        vm.add("e", self.N_n)
        vm.add("f", self.N_t, self.N_n)
        vm.add("g", self.N_n)
        vm.add_if(medi, "h", Nmed, self.N_irmaa)    # IRMAA bracket portions (Medicare optimize)
        vm.add_if(self._aca_lp, "haca", self.n_aca, tx.N_ACA_R)  # ACA MAGI bracket portions (optimize)
        vm.add_if(self._aca_lp, "maca", self.N_n)  # ACA LP cost variable (optimize mode only)
        vm.add("m", self.N_n)
        vm.add("q", self.N_p, self.N_n)             # q_{pn}: LTCG bracket allocations (p=0,1,2)
        vm.add("s", self.N_n)
        vm.add("w", self.N_i, self.N_j, self.N_n)
        vm.add("x", self.N_i, self.N_n)
        # SS taxability LP variables (continuous) must precede the binary block.
        vm.add_if(ss_lp, "plo", self.N_n)           # p^lo_n = max(0, Π_n − 𝒫^lo)
        vm.add_if(ss_lp, "phi", self.N_n)           # p^hi_n = max(0, Π_n − 𝒫^hi)
        vm.add_if(ss_lp, "pmin", self.N_n)          # p^{σ,min}_n = min(𝒫^hi−𝒫^lo, p^lo_n)
        vm.add_if(ss_lp, "tss", self.N_n)           # t^σ_n  = min(0.85·ζ̄_n, 0.5·p^{σ,min}_n + 0.85·p^hi_n)
        vm.add_if(ltcg_lp, "gn", self.N_n)          # G_n: ordinary taxable income (LTCG MILP)
        vm.add_if(niit_lp, "magi", self.N_n)        # MAGI_n LP variable (NIIT MILP)
        vm.add_if(niit_lp, "Jn", self.N_n)          # J_n: NIIT tax LP variable (NIIT MILP)
        vm.add_if(ssa_lp, "ssb", self.N_i, self.N_n)   # SS own-benefit LP var (SS age optimize)
        vm.mark_binary_start()
        vm.add("zx", self.N_n, self.N_zx)           # Roth exclusion binaries
        vm.add_if(medi, "zm", Nmed, self.N_irmaa)   # IRMAA bracket selection binaries
        vm.add_if(ss_lp, "zs", self.N_n, 2)         # z^σ family (2 per year) for SS min() ops
        vm.add_if(self._aca_lp, "za", self.n_aca, tx.N_ACA_R)   # ACA bracket selection binaries
        vm.add_if(ltcg_lp, "zl", 2, self.N_n)       # 2×N_n regime binaries (LTCG MILP)
        vm.add_if(niit_lp, "zj", self.N_n)          # N_n NIIT threshold binaries
        vm.add_if(ssa_lp, "zssa", self.N_i, self._ssa_N_K)  # claiming-month selectors (SS age)
        self.vm = vm

        self.nvars = vm.nvars
        self.nbins = vm.nbins
        self.nconts = vm.nconts
        self.nbals = vm.nbals

        nseries = len(vm._blocks)
        self.mylog.vprint(
            f"Problem has {nseries} distinct series, {self.nvars} decision variables (including {self.nbins} binary).")

    def _buildConstraints(self, objective, options):
        """
        Utility function that builds constraint matrix and vectors.
        Refactored for clarity and maintainability.
        """
        # Ensure parameters are adjusted for inflation and MAGI.
        self._adjustParameters(self.gamma_n, self.MAGI_n)

        self.A = abc.ConstraintMatrix(self.nvars)
        self.B = abc.Bounds(self.nvars, self.nbins)

        self._add_rmd_inequalities()
        self._add_tax_bracket_bounds()
        self._add_standard_exemption_bounds()
        self._add_defunct_constraints()
        self._add_roth_conversion_constraints(options)
        self._add_safety_net(options)
        self._add_roth_maturation_constraints()
        self._add_withdrawal_limits()
        self._add_objective_constraints(objective, options)
        self._add_initial_balances()
        self._add_surplus_deposit_linking(options)
        self._add_account_balance_carryover()
        self._add_net_cash_flow(options)
        self._add_income_profile(objective)
        self._add_taxable_income(options)
        self._configure_ss_taxability_lp(options)
        self._configure_ss_age_variables()
        self._configure_ltcg_constraints()
        self._configure_Medicare_binary_variables(options)
        self._add_Medicare_costs(options)
        self._configure_ACA_binary_variables(options)
        self._add_ACA_costs(options)
        self._add_magi_lp(options)
        self._configure_NIIT_binary_variables(options)
        self._configure_exclusion_binary_variables(options)
        self._build_objective_vector(objective, options)

    def _add_rmd_inequalities(self):
        """
        Enforce Required Minimum Distributions (RMDs) on tax-deferred accounts (j=1) only.

        RMD rules:
        - Traditional IRA, SEP-IRA, and 401(k) balances (aggregated in j=1) are subject to RMDs
          starting at the age specified by rho_in() (SECURE 1.0/2.0 birth-year cohorts).
        - Roth IRA accounts (j=2) are exempt from RMDs during the original owner's lifetime
          (IRC §408A(c)(5)).
        - Roth 401(k) accounts were subject to RMDs prior to 2024, but SECURE 2.0 Act §325
          eliminated Roth 401(k) RMDs effective for tax years beginning after December 31, 2023.
          Since plans modeled here start in 2024 or later, treating all Roth (j=2) as exempt
          from RMDs is correct for all currently supported scenarios.
        - Inherited IRA / beneficiary RMDs are not modeled.
        """
        for i in range(self.N_i):
            if self.beta_ij[i, 1] > 0:
                for n in range(self.horizons[i]):
                    rowDic = {
                        self.vm["w"].idx(i, 1, n): 1,
                        self.vm["b"].idx(i, 1, n): -self.rho_in[i, n],
                    }
                    self.A.addNewRow(rowDic, 0, np.inf)

    def _add_tax_bracket_bounds(self):
        for t in range(self.N_t):
            for n in range(self.N_n):
                self.B.setRange(self.vm["f"].idx(t, n), 0, self.DeltaBar_tn[t, n])

    def _add_standard_exemption_bounds(self):
        for n in range(self.N_n):
            self.B.setRange(self.vm["e"].idx(n), 0, self.sigmaBar_n[n])

    def _add_defunct_constraints(self):
        if self.N_i == 2:
            for n in range(self.n_d, self.N_n):
                self.B.setRange(self.vm["d"].idx(self.i_d, n), 0, 0)
                self.B.setRange(self.vm["x"].idx(self.i_d, n), 0, 0)
                for j in range(self.N_j):
                    self.B.setRange(self.vm["w"].idx(self.i_d, j, n), 0, 0)

    def _add_roth_maturation_constraints(self):
        """
        Enforce the Roth 5-year seasoning rule for conversions and contribution gains.

        IRS rules (simplified here):
        - Roth contribution *principal* can be withdrawn tax- and penalty-free at any time.
        - Roth *earnings* on contributions require both (a) account age ≥ 5 years AND
          (b) age ≥ 59½ (or other exception) to be penalty-free.
        - Each Roth *conversion* carries its own 5-year clock; conversion principal converted
          before age 59½ is subject to the 10% penalty if withdrawn within 5 years of conversion.

        Simplification: This implementation applies a single unified 5-year lookback that retains
        all recent conversions (at compounded value) and the gains-only portion of recent
        contributions as a minimum balance floor. Contribution *principal* is not separately tracked
        and freed — the constraint treats it as part of the 5-year retainer, making it intentionally
        conservative (preventing some valid early withdrawals of contribution principal). This never
        allows a withdrawal that would violate IRS rules; it only restricts some withdrawals that
        would technically be permitted. Exact per-dollar basis tracking is out of scope for an LP.
        """
        # Assume 10% per year for contributions and conversions for past 5 years.
        # Future years will use the assumed returns.
        oldTau1 = 1.10
        for i in range(self.N_i):
            h = self.horizons[i]
            for n in range(h):
                rhs = 0
                # To add compounded gains to cumulative amounts. Always keep cgains >= 1.
                cgains = 1
                row = self.A.newRow()
                row.addElem(self.vm["b"].idx(i, 2, n), 1)
                row.addElem(self.vm["w"].idx(i, 2, n), -1)
                for dn in range(1, 6):
                    nn = n - dn
                    if nn >= 0:   # Past of future is now or in the future: use variables or parameters.
                        Tau1 = 1 + np.sum(self.alpha_ijkn[i, 2, :, nn] * self.tau_kn[:, nn], axis=0)
                        # Ignore market downs.
                        cgains *= max(1, Tau1)
                        row.addElem(self.vm["x"].idx(i, nn), -cgains)
                        # If a contribution, it has only penalty on gains, not on deposited amount.
                        rhs += (cgains - 1) * self.kappa_ijn[i, 2, nn]
                    else:  # Past of future is in the past:
                        cgains *= oldTau1
                        # Past years are stored at the end of contributions and conversions arrays.
                        # Use negative index to access tail of array.
                        # Past years are stored at the end of arrays, accessed via negative indexing
                        rhs += (cgains - 1) * self.kappa_ijn[i, 2, nn] + cgains * self.myRothX_in[i, nn]

                self.A.addRow(row, rhs, np.inf)

    def _add_roth_conversion_constraints(self, options):
        """
        Enforce Roth conversion limits and add converted amounts to taxable income.

        Tax treatment: Roth conversions are fully taxable as ordinary income (IRC §408A(d)(3)).

        Pro-rata rule (not modeled): If the tax-deferred account (j=1) contains a mix of
        pre-tax and after-tax (nondeductible) contributions, IRS Form 8606 requires the taxable
        fraction of each conversion to be computed pro-rata across all IRA balances. This tool
        assumes 100% of the j=1 balance is pre-tax (no cost basis from nondeductible contributions),
        so all conversions are treated as fully taxable. Users with significant IRA basis should
        be aware of this limitation and consult a tax professional.
        """
        # Values in file supercedes everything.
        if "maxRothConversion" in options and options["maxRothConversion"] == "file":
            for i in range(self.N_i):
                for n in range(self.horizons[i]):
                    rhs = self.myRothX_in[i][n]
                    self.B.setRange(self.vm["x"].idx(i, n), rhs, rhs)
        else:
            # Don't exclude anyone by default.
            i_xcluded = -1
            if "noRothConversions" in options and options["noRothConversions"] not in ("none", "None"):
                rhsopt = options["noRothConversions"]
                try:
                    i_xcluded = self.inames.index(rhsopt)
                except ValueError as e:
                    raise ValueError(f"Unknown individual '{rhsopt}' for noRothConversions:") from e
                for n in range(self.horizons[i_xcluded]):
                    self.B.setRange(self.vm["x"].idx(i_xcluded, n), 0, 0)

            if "maxRothConversion" in options:
                rhsopt = u.get_monetary_option(options, "maxRothConversion", 0)

                if rhsopt >= 0:
                    for i in range(self.N_i):
                        if i == i_xcluded:
                            continue
                        for n in range(self.horizons[i]):
                            # Apply the cap per individual.
                            self.B.setRange(self.vm["x"].idx(i, n), 0, rhsopt)

            if "startRothConversions" in options:
                rhsopt = int(u.get_numeric_option(options, "startRothConversions", 0))
                thisyear = date.today().year
                yearn = max(rhsopt - thisyear, 0)
                for i in range(self.N_i):
                    if i == i_xcluded:
                        continue
                    nstart = min(yearn, self.horizons[i])
                    for n in range(0, nstart):
                        self.B.setRange(self.vm["x"].idx(i, n), 0, 0)

            if "swapRothConverters" in options and i_xcluded == -1:
                rhsopt = int(u.get_numeric_option(options, "swapRothConverters", 0))
                if self.N_i == 2 and rhsopt != 0:
                    thisyear = date.today().year
                    absrhsopt = abs(rhsopt)
                    yearn = max(absrhsopt - thisyear, 0)
                    i_x = 0 if rhsopt > 0 else 1
                    i_y = (i_x + 1) % 2

                    transy = min(yearn, self.horizons[i_y])
                    for n in range(0, transy):
                        self.B.setRange(self.vm["x"].idx(i_y, n), 0, 0)

                    transx = min(yearn, self.horizons[i_x])
                    for n in range(transx, self.horizons[i_x]):
                        self.B.setRange(self.vm["x"].idx(i_x, n), 0, 0)

            # Disallow Roth conversions in last two years alive.
            for i in range(self.N_i):
                if i == i_xcluded:
                    continue
                for n in range(max(0, self.horizons[i] - 2), self.horizons[i]):
                    self.B.setRange(self.vm["x"].idx(i, n), 0, 0)

    def _add_safety_net(self, options):
        """
        Enforce minimum taxable account balances (safety net) for each individual.
        Amounts are in today's $ and indexed for inflation. Constraints apply
        from year 2 onward through each individual's life horizon (not year 1).
        """
        if "minTaxableBalance" not in options:
            return
        min_bal = u.get_monetary_list_option(
            options, "minTaxableBalance", self.N_i, min_value=0
        )
        for i in range(self.N_i):
            min_dollar = min_bal[i]
            if min_dollar <= 0:
                continue
            # From year 2 onward; last year = min(horizons[i], N_n) for survivor,
            # horizons[i]-1 for deceased (last year alive)
            for n in range(1, self.horizons[i]):
                rhs = min_dollar * self.gamma_n[n]
                self.B.setRange(self.vm["b"].idx(i, 0, n), rhs, np.inf)

    def _add_withdrawal_limits(self):
        for i in range(self.N_i):
            # Wierdly enough, setting horizons causes a effects on HiGHS and MOSEK
            # for n in range(self.N_n):
            for n in range(self.horizons[i]):
                rowDic = {self.vm["w"].idx(i, 1, n): -1,
                          self.vm["x"].idx(i, n): -1,
                          self.vm["b"].idx(i, 1, n): 1}
                self.A.addNewRow(rowDic, 0, np.inf)
                for j in [0, 2, 3]:
                    rowDic = {self.vm["w"].idx(i, j, n): -1,
                              self.vm["b"].idx(i, j, n): 1}
                    self.A.addNewRow(rowDic, 0, np.inf)

    def _add_objective_constraints(self, objective, options):
        if objective == "maxSpending":
            if "fixedSpending" in options:
                spending = u.get_monetary_option(options, "fixedSpending", 1)
                self.B.setRange(self.vm["g"].idx(0), spending, spending)
            if "bequest" in options:
                bequest = u.get_monetary_option(options, "bequest", 1) * self.gamma_n[-1]
            else:
                bequest = 1

            # Bequest constraint now refers only to savings accounts
            # User specifies desired bequest from accounts (fixed assets are separate)
            # Total bequest = accounts - debts + fixed_assets
            # So: accounts >= desired_bequest_from_accounts + debts
            # (fixed_assets are added separately in the total bequest calculation)
            total_bequest_value = bequest + self.remaining_debt_balance

            row = self.A.newRow()
            for i in range(self.N_i):
                row.addElem(self.vm["b"].idx(i, 0, self.N_n), 1)
                row.addElem(self.vm["b"].idx(i, 1, self.N_n), 1 - self.nu)
                row.addElem(self.vm["b"].idx(i, 2, self.N_n), 1)
                row.addElem(self.vm["b"].idx(i, 3, self.N_n), 1 - self.nu)   # HSA: heirs pay ordinary income tax
            self.A.addRow(row, total_bequest_value, np.inf)
        elif objective == "maxBequest":
            spending = u.get_monetary_option(options, "netSpending", 1)
            self.B.setRange(self.vm["g"].idx(0), spending, spending)
        elif objective == "maxHybrid":
            spending_floor = u.get_monetary_option(options, "spendingFloor", 0)
            if spending_floor > 0:
                self.B.setRange(self.vm["g"].idx(0), spending_floor, np.inf)

    def _add_initial_balances(self):
        # Back project balances to the beginning of the year.
        yearSpent = 1 - self.yearFracLeft

        for i in range(self.N_i):
            for j in range(self.N_j):
                backTau = 1 + yearSpent * np.sum(self.tau_kn[:, 0] * self.alpha_ijkn[i, j, :, 0])
                rhs = self.beta_ij[i, j] / backTau
                self.B.setRange(self.vm["b"].idx(i, j, 0), rhs, rhs)

    def _add_surplus_deposit_linking(self, options):
        for i in range(self.N_i):
            fac1 = u.krond(i, 0) * (1 - self.eta) + u.krond(i, 1) * self.eta
            for n in range(self.n_d):
                rowDic = {self.vm["d"].idx(i, n): 1, self.vm["s"].idx(n): -fac1}
                self.A.addNewRow(rowDic, 0, 0)
            fac2 = u.krond(self.i_s, i)
            for n in range(self.n_d, self.N_n):
                rowDic = {self.vm["d"].idx(i, n): 1, self.vm["s"].idx(n): -fac2}
                self.A.addNewRow(rowDic, 0, 0)

        # Prevent surplus on two last year as they have little tax and/or growth consequence.
        disallow = options.get("noLateSurplus", False)
        if disallow:
            self.B.setRange(self.vm["s"].idx(self.N_n - 2), 0, 0)
            self.B.setRange(self.vm["s"].idx(self.N_n - 1), 0, 0)

    def _add_account_balance_carryover(self):
        tau_ijn = np.zeros((self.N_i, self.N_j, self.N_n))
        for i in range(self.N_i):
            for j in range(self.N_j):
                for n in range(self.N_n):
                    tau_ijn[i, j, n] = np.sum(self.alpha_ijkn[i, j, :, n] * self.tau_kn[:, n], axis=0)

        # Weights are normalized on k: sum_k[alpha*(1 + tau)] = 1 + sum_k[alpha*tau]
        Tau1_ijn = 1 + tau_ijn
        Tauh_ijn = 1 + tau_ijn / 2

        for i in range(self.N_i):
            for j in range(self.N_j):
                for n in range(self.N_n):
                    if self.N_i == 2 and self.n_d < self.N_n and i == self.i_d and n == self.n_d - 1:
                        fac1 = 0
                    else:
                        fac1 = 1

                    rhs = fac1 * self.kappa_ijn[i, j, n] * Tauh_ijn[i, j, n]

                    row = self.A.newRow()
                    row.addElem(self.vm["b"].idx(i, j, n + 1), 1)
                    row.addElem(self.vm["b"].idx(i, j, n), -fac1 * Tau1_ijn[i, j, n])
                    row.addElem(self.vm["w"].idx(i, j, n), fac1 * Tau1_ijn[i, j, n])
                    row.addElem(self.vm["d"].idx(i, n), -fac1 * u.krond(j, 0) * Tau1_ijn[i, 0, n])
                    row.addElem(
                        self.vm["x"].idx(i, n),
                        -fac1 * (self.xnet * u.krond(j, 2) - u.krond(j, 1)) * Tau1_ijn[i, j, n],
                    )

                    if self.N_i == 2 and self.n_d < self.N_n and i == self.i_s and n == self.n_d - 1:
                        fac2 = self.phi_j[j]
                        rhs += fac2 * self.kappa_ijn[self.i_d, j, n] * Tauh_ijn[self.i_d, j, n]
                        row.addElem(self.vm["b"].idx(self.i_d, j, n),
                                    -fac2 * Tau1_ijn[self.i_d, j, n])
                        row.addElem(self.vm["w"].idx(self.i_d, j, n),
                                    fac2 * Tau1_ijn[self.i_d, j, n])
                        row.addElem(self.vm["d"].idx(self.i_d, n),
                                    -fac2 * u.krond(j, 0) * Tau1_ijn[self.i_d, 0, n])
                        row.addElem(
                            self.vm["x"].idx(self.i_d, n),
                            -fac2 * (self.xnet * u.krond(j, 2) - u.krond(j, 1)) * Tau1_ijn[self.i_d, j, n],
                        )
                    self.A.addRow(row, rhs, rhs)

    def _add_net_cash_flow(self, options=None):
        tau_0prev = np.roll(self.tau_kn[0, :], 1)
        tau_0prev[tau_0prev < 0] = 0
        for n in range(self.N_n):
            rhs = -self.M_n[n] - self.ACA_n[n]
            if not getattr(self, "_niit_lp", False):
                rhs -= self.J_n[n]
            # Add fixed assets proceeds (positive cash flow)
            rhs += (self.fixed_assets_tax_free_n[n]
                    + self.fixed_assets_ordinary_income_n[n]
                    + self.fixed_assets_capital_gains_n[n])
            # Subtract debt payments (negative cash flow)
            rhs -= self.debt_payments_n[n]
            row = self.A.newRow({self.vm["g"].idx(n): 1})
            row.addElem(self.vm["s"].idx(n), 1)
            row.addElem(self.vm["m"].idx(n), 1)
            if "maca" in self.vm:
                row.addElem(self.vm["maca"].idx(n), 1)
            for i in range(self.N_i):
                if "ssb" in self.vm:
                    # SS own-benefit is an LP variable; spousal/survivor is a parameter offset.
                    ss_income = self._ssa_spousal_offset[i, n]
                    row.addElem(self.vm["ssb"].idx(i, n), -1)
                else:
                    ss_income = self.zetaBar_in[i, n]
                rhs += (
                    self.omega_in[i, n]
                    + self.other_inc_in[i, n]
                    + self.netinv_in[i, n]
                    + ss_income
                    + self.piBar_in[i, n]
                    + self.Lambda_in[i, n]
                )
                row.addElem(self.vm["w"].idx(i, 0, n), -1)
                penalty = 0.1 if n < self.n595[i] else 0
                row.addElem(self.vm["w"].idx(i, 1, n), -1 + penalty)
                # maturation constraints govern; no 10% penalty
                row.addElem(self.vm["w"].idx(i, 2, n), -1)
                # HSA: qualified medical withdrawals are tax-free (simplified model)
                row.addElem(self.vm["w"].idx(i, 3, n), -1)

            for t in range(self.N_t):
                row.addElem(self.vm["f"].idx(t, n), self.theta_tn[t, n])

            # LTCG tax from bracket variables q[1,n] and q[2,n] directly.
            row.addElem(self.vm["q"].idx(1, n), 0.15)
            row.addElem(self.vm["q"].idx(2, n), 0.20)

            # NIIT: when optimize mode, use LP variable Jn; otherwise already in rhs.
            if getattr(self, "_niit_lp", False):
                row.addElem(self.vm["Jn"].idx(n), 1)

            self.A.addRow(row, rhs, rhs)

    def _add_income_profile(self, objective):
        # For maxHybrid the profile is a floor only: spending can freely exceed the profile shape.
        # Slack (lambdha) is ignored — the lower bound uses the bare profile value.
        # For all other objectives both bounds apply with the user-specified slack tolerance.
        if objective == "maxHybrid":
            # Lower bound (floor): g[n] ≥ g[0] × ratio_n  — bare profile, no slack.
            # Upper bound (cap): g[n] ≤ g[0] × (1+λ) × ratio_n  — only when slack > 0.
            # When slack=0 spending roams freely above the floor; slack provides an optional cap.
            spHi = 1 + self.lambdha
            for n in range(1, self.N_n):
                rowDic = {self.vm["g"].idx(0): self.xiBar_n[n],
                          self.vm["g"].idx(n): -self.xiBar_n[0]}
                self.A.addNewRow(rowDic, -np.inf, 0)
                if self.lambdha > 0:
                    rowDic = {self.vm["g"].idx(0): spHi * self.xiBar_n[n],
                              self.vm["g"].idx(n): -self.xiBar_n[0]}
                    self.A.addNewRow(rowDic, 0, np.inf)
        else:
            spLo = 1 - self.lambdha
            spHi = 1 + self.lambdha
            for n in range(1, self.N_n):
                rowDic = {self.vm["g"].idx(0): spLo * self.xiBar_n[n],
                          self.vm["g"].idx(n): -self.xiBar_n[0]}
                self.A.addNewRow(rowDic, -np.inf, 0)
                rowDic = {self.vm["g"].idx(0): spHi * self.xiBar_n[n],
                          self.vm["g"].idx(n): -self.xiBar_n[0]}
                self.A.addNewRow(rowDic, 0, np.inf)

    def _add_taxable_income(self, options=None):
        ss_lp = options is not None and options.get("withSSTaxability", "loop") == "optimize"
        for n in range(self.N_n):
            # Add fixed assets ordinary income
            rhs = self.fixed_assets_ordinary_income_n[n]
            row = self.A.newRow()
            row.addElem(self.vm["e"].idx(n), 1)
            for i in range(self.N_i):
                if ss_lp:
                    # Taxable SS is an LP variable (tss_n); omit the Psi_n*zetaBar parameter.
                    rhs += (self.omega_in[i, n] + self.other_inc_in[i, n]
                            + self.netinv_in[i, n] + self.piBar_in[i, n])
                else:
                    rhs += (self.omega_in[i, n] + self.other_inc_in[i, n]
                            + self.netinv_in[i, n]
                            + self.Psi_n[n] * self.zetaBar_in[i, n] + self.piBar_in[i, n])
                # HSA contributions are pre-tax deductions (reduce ordinary taxable income).
                rhs -= self.kappa_ijn[i, 3, n]
                row.addElem(self.vm["w"].idx(i, 1, n), -1)
                row.addElem(self.vm["x"].idx(i, n), -1)
                # Only positive returns are taxable (interest/dividends); losses don't reduce income.
                fak = np.sum(np.maximum(0, self.tau_kn[1:self.N_k, n]) * self.alpha_ijkn[i, 0, 1:self.N_k, n], axis=0)
                rhs += 0.5 * fak * self.kappa_ijn[i, 0, n]
                row.addElem(self.vm["b"].idx(i, 0, n), -fak)
                row.addElem(self.vm["w"].idx(i, 0, n), fak)
                row.addElem(self.vm["d"].idx(i, n), -fak)
            for t in range(self.N_t):
                row.addElem(self.vm["f"].idx(t, n), 1)
            if ss_lp:
                # t^σ_n = taxable SS LP variable replaces Psi_n*zetaBar_n in the constraint:
                # e_n - t^σ_n + sum_t(f_tn) = non_SS_ordinary_income
                row.addElem(self.vm["tss"].idx(n), -1)
            self.A.addRow(row, rhs, rhs)

    def _configure_exclusion_binary_variables(self, options):
        if not options.get("amoConstraints", True):
            return

        bigM = u.get_numeric_option(options, "bigMamo", BIGM_AMO, min_value=0)

        # A surplus cannot be created from a taxable or tax-exempt withdrawal.
        if options.get("amoSurplus", True):
            for n in range(self.N_n):
                # Make z_0 and z_1 exclusive binary variables.
                dic0 = {self.vm["zx"].idx(n, 0): bigM*self.gamma_n[n],
                        self.vm["w"].idx(0, 0, n): -1,
                        self.vm["w"].idx(0, 2, n): -1}
                if self.N_i == 2:
                    dic1 = {self.vm["w"].idx(1, 0, n): -1,
                            self.vm["w"].idx(1, 2, n): -1}
                    dic0.update(dic1)

                self.A.addNewRow(dic0, 0, np.inf)

                self.A.addNewRow(
                    {self.vm["zx"].idx(n, 1): bigM*self.gamma_n[n],
                     self.vm["s"].idx(n): -1},
                    0, np.inf)

                # As both can be zero, bound as z_0 + z_1 <= 1
                self.A.addNewRow(
                    {self.vm["zx"].idx(n, 0): +1,
                     self.vm["zx"].idx(n, 1): +1},
                    0, 1
                )

        if "maxRothConversion" in options:
            rhsopt = options.get("maxRothConversion")
            if rhsopt != "file":
                rhsopt = u.get_numeric_option(options, "maxRothConversion", 0)
                if rhsopt < -1:
                    return

        # Turning off this constraint for maxRothConversions = 0 makes solution infeasible.
        # A Roth conversion cannot be done in the same year as a Roth withdrawal.
        if options.get("amoRoth", True):
            n595_max = int(np.max(self.n595))   # last year any individual still needs the ladder
            for n in range(self.N_n):
                if n < n595_max:
                    continue   # relax AMO: allow simultaneous conversion + mature Roth withdrawal
                # Make z_2 and z_3 at-most-one binary variables.
                dic0 = {self.vm["zx"].idx(n, 2): bigM*self.gamma_n[n],
                        self.vm["x"].idx(0, n): -1}
                if self.N_i == 2:
                    dic1 = {self.vm["x"].idx(1, n): -1}
                    dic0.update(dic1)

                self.A.addNewRow(dic0, 0, np.inf)

                dic0 = {self.vm["zx"].idx(n, 3): bigM*self.gamma_n[n],
                        self.vm["w"].idx(0, 2, n): -1}
                if self.N_i == 2:
                    dic1 = {self.vm["w"].idx(1, 2, n): -1}
                    dic0.update(dic1)

                self.A.addNewRow(dic0, 0, np.inf)

                self.A.addNewRow(
                    {self.vm["zx"].idx(n, 2): +1,
                     self.vm["zx"].idx(n, 3): +1},
                    0, 1
                )

    def _configure_ss_taxability_lp(self, options):
        """
        Configure SS taxability using the MIP approach (alternative to the SC loop).

        Introduces per-year LP variables p^lo_n, p^hi_n, q_n, t^σ_n and binary variables
        z^σ_{0n}, z^σ_{1n} to compute taxable Social Security exactly within the LP,
        eliminating the need to update Psi_n in the self-consistent loop for SS.

        Filing status is per-year: for a couple, status switches from MFJ to Single
        in year n_d when the first spouse dies. The 50% tier uses q_n ≤ min(Δ𝒫_n, ζ̄_n, p^lo_n),
        matching the IRS formula exactly.

        Provisional income Π_n = (non-SS ordinary income) + Q_n + 0.5·ζ̄_n is built using
        the same coefficient structure as the Medicare MAGI constraint, adjusted for the
        current year n and using 0.5·ζ̄_n instead of the full ζ̄_n.
        """
        if options.get("withSSTaxability", "loop") != "optimize":
            return

        bigM = u.get_numeric_option(options, "bigMss", BIGM_AMO, min_value=0)
        tau_0 = self.tau_kn[0, :]

        for n in range(self.N_n):
            zetaBar_n = np.sum(self.zetaBar_in[:, n])

            # No SS income this year: fix variables to 0 and skip all 8 constraints.
            # tss_n MUST be fixed (it appears in taxable income with -1 coefficient).
            # z0_n, z1_n should be fixed to remove them from MIP branching.
            if zetaBar_n == 0:
                self.B.setRange(self.vm["tss"].idx(n), 0, 0)
                self.B.setRange(self.vm["zs"].idx(n, 0), 0, 0)
                self.B.setRange(self.vm["zs"].idx(n, 1), 0, 0)
                continue

            # Per-year filing status: for couple, switch to Single at n_d.
            status_n = 0 if (self.N_i == 2 and n >= self.n_d) else self.N_i - 1
            ss_lo_n = tx.ssTaxabilityLo[status_n]
            ss_hi_n = tx.ssTaxabilityHi[status_n]
            delta_p_n = ss_hi_n - ss_lo_n

            # === Build Π_n LP coefficients ===
            # Π_n = e_n - t^σ_n + LP_income_terms + params, where the LP income terms
            # mirror the Medicare MAGI constraint (same year n, not n-2) but with 0.5·ζ̄_n.
            # The t^σ_n correction cancels the taxable-SS portion embedded in e_n (with ss_lp,
            # e_n = B_n + t^σ_n), leaving Π_n = B_n + Q_n + 0.5·ζ̄_n independent of t^σ_n.
            tau_prev = tau_0[max(0, n - 1)]

            rhs_pi = (self.fixed_assets_ordinary_income_n[n]
                      + self.fixed_assets_capital_gains_n[n]
                      + 0.5 * zetaBar_n)   # 0.5·SS for provisional income (not full SS)

            pi_row = {}
            pi_row[self.vm["e"].idx(n)] = -1   # subtract e_n
            pi_row[self.vm["tss"].idx(n)] = +1   # add back t^σ_n to cancel its share in e_n

            for i in range(self.N_i):
                # Combined dividend + interest yield for taxable account (equity + bonds/notes/cash).
                afac = (self.mu * self.alpha_ijkn[i, 0, 0, n]
                        + np.sum(self.alpha_ijkn[i, 0, 1:, n] * np.maximum(0, self.tau_kn[1:, n])))
                # Capital gains: price appreciation only (total equity return − dividend rate).
                bfac = self.alpha_ijkn[i, 0, 0, n] * max(0, tau_prev - self.mu)

                w1_idx = self.vm["w"].idx(i, 1, n)
                x_idx = self.vm["x"].idx(i, n)
                b_idx = self.vm["b"].idx(i, 0, n)
                d_idx = self.vm["d"].idx(i, n)
                w0_idx = self.vm["w"].idx(i, 0, n)

                pi_row[w1_idx] = pi_row.get(w1_idx, 0) - 1    # IRA withdrawals (income)
                pi_row[x_idx] = pi_row.get(x_idx, 0) - 1      # Roth conversions (income)
                pi_row[b_idx] = pi_row.get(b_idx, 0) - afac   # beginning balance × yield
                pi_row[d_idx] = pi_row.get(d_idx, 0) - afac   # contributions × yield
                pi_row[w0_idx] = pi_row.get(w0_idx, 0) + (afac - bfac)  # withdrawals (net)

                rhs_pi += (self.omega_in[i, n]
                           + self.other_inc_in[i, n]
                           + self.netinv_in[i, n]
                           + self.piBar_in[i, n]
                           + 0.5 * self.kappa_ijn[i, 0, n] * afac)   # half-period contribution yield
                rhs_pi -= self.kappa_ijn[i, 3, n]   # HSA contributions reduce provisional income

            # Variable index shorthands.
            plo_idx = self.vm["plo"].idx(n)
            phi_idx = self.vm["phi"].idx(n)
            pmin_idx = self.vm["pmin"].idx(n)
            tss_idx = self.vm["tss"].idx(n)
            z0_idx = self.vm["zs"].idx(n, 0)
            z1_idx = self.vm["zs"].idx(n, 1)
            bigMBar = bigM * self.gamma_n[n]

            # === p^lo_n = max(0, Π_n − 𝒫^lo) ===
            # Lower bound ≥ 0 from default variable bounds; explicit inequality enforces the max.
            # Row: p^lo_n + pi_row_coeffs ≥ rhs_pi − ss_lo_n
            row_plo = dict(pi_row)
            row_plo[plo_idx] = row_plo.get(plo_idx, 0) + 1
            self.A.addNewRow(row_plo, rhs_pi - ss_lo_n, np.inf)

            # === p^hi_n = max(0, Π_n − 𝒫^hi) ===
            row_phi = dict(pi_row)
            row_phi[phi_idx] = row_phi.get(phi_idx, 0) + 1
            self.A.addNewRow(row_phi, rhs_pi - ss_hi_n, np.inf)

            # === p^{σ,min}_n = min(Δ𝒫_n, ζ̄_n, p^lo_n) via binary z^σ_{0n} ===
            # Upper bounds: p^{σ,min}_n ≤ min(Δ𝒫_n, ζ̄_n) (setRange) and ≤ p^lo_n (constraint).
            # When ζ̄_n < Δ𝒫_n, the effective upper bound on pmin is ζ̄_n; using Δ𝒫_n in the big-M
            # lower bound of constraint (3b) would force pmin ≥ Δ𝒫_n > ζ̄_n, causing infeasibility.
            p_ub = min(delta_p_n, zetaBar_n)
            self.A.addNewRow({pmin_idx: 1, plo_idx: -1}, -np.inf, 0)              # pmin ≤ p^lo
            # p^{σ,min}_n ≥ min(Δ𝒫_n, ζ̄_n) − M·(1 − z0)  →  pmin − M·z0 ≥ p_ub − M
            self.A.addNewRow({pmin_idx: 1, z0_idx: -bigMBar}, p_ub - bigMBar, np.inf)
            # p^{σ,min}_n ≥ p^lo_n − M·z0  →  pmin − p^lo + M·z0 ≥ 0
            self.A.addNewRow({pmin_idx: 1, plo_idx: -1, z0_idx: bigMBar}, 0, np.inf)
            self.B.setRange(pmin_idx, 0, p_ub)                                    # pmin ≤ min(Δ𝒫_n, ζ̄_n)

            # === t^σ_n = min(0.85·ζ̄_n, 0.5·p^{σ,min}_n + 0.85·p^hi_n) via binary z^σ_{1n} ===
            # Upper bound t^σ_n ≤ 0.5·p^{σ,min}_n + 0.85·p^hi_n.
            self.A.addNewRow({tss_idx: 1, pmin_idx: -0.5, phi_idx: -0.85}, -np.inf, 0)
            # t^σ_n ≥ 0.85·ζ̄_n − M·(1 − z1)  →  t^σ_n − M·z1 ≥ 0.85·ζ̄_n − M
            self.A.addNewRow({tss_idx: 1, z1_idx: -bigMBar}, 0.85 * zetaBar_n - bigMBar, np.inf)
            # t^σ_n ≥ 0.5·p^{σ,min}_n + 0.85·p^hi_n − M·z1  →  tss − 0.5·pmin − 0.85·phi + M·z1 ≥ 0
            self.A.addNewRow({tss_idx: 1, pmin_idx: -0.5, phi_idx: -0.85, z1_idx: bigMBar}, 0, np.inf)
            self.B.setRange(tss_idx, 0, 0.85 * zetaBar_n)                        # t^σ ≤ 0.85·ζ̄

    def _configure_ss_age_variables(self):
        """
        Add SS claiming-age optimization constraints (withSSAges='optimize' mode).

        Two constraint groups per individual i:
          a) AMO (exactly-one claiming month): sum_k zssa[i,k] = 1
          b) Benefit definition: ssb[i,n] = sum_k B_own[i,k,n] * zssa[i,k]  for each n

        For already-claimed individuals (current age >= stored claiming age), all zssa[i,k]
        are fixed to 0/1 via bounds so the optimizer leaves them at the recorded age.
        For individuals with zero PIA, zssa[i,0] is fixed to 1 (irrelevant, B_own=0).

        Notes
        -----
        ssb[i,n] = own SS benefit only; spousal/survivor offsets are added as parameters
        (_ssa_spousal_offset[i,n]) in the cash-flow constraint and updated each SC iteration.
        """
        if not self._ssa_lp:
            return

        vm = self.vm
        N_K = self._ssa_N_K
        B_own = self._ssa_B_own
        thisyear = date.today().year

        for i in range(self.N_i):
            pia_i = int(self.ssecAmounts[i]) if hasattr(self, "ssecAmounts") else 0

            # Determine if this individual has already claimed or is not selected for optimization.
            current_age = thisyear - self.yobs[i] - (self.mobs[i] - 1) / 12
            already_claimed = current_age >= float(self.ssecAges[i])
            not_selected = i not in getattr(self, "_ssa_optimize_set", set(range(self.N_i)))

            if pia_i == 0 or already_claimed or not_selected:
                # Fix to the known/current claiming age (or age 62 if no SS).
                if pia_i == 0:
                    k_fixed = 0   # Arbitrary; B_own is all-zero anyway.
                else:
                    k_fixed = int(round((float(self.ssecAges[i]) - 62.0) * 12))
                    k_fixed = max(0, min(N_K - 1, k_fixed))
                for k in range(N_K):
                    lb = 1 if k == k_fixed else 0
                    self.B.setRange(vm["zssa"].idx(i, k), lb, lb)
            else:
                # Free individual: fix ineligible claiming ages to 0.
                bornOnFirstDays = (self.tobs[i] <= 2)
                eligible = 62.0 if bornOnFirstDays else 62.0 + 1.0 / 12
                for k in range(N_K):
                    if self._ssa_ages_k[k] < eligible:
                        self.B.setRange(vm["zssa"].idx(i, k), 0, 0)

            # a) AMO: exactly one claiming month per individual.
            amo_row = {vm["zssa"].idx(i, k): 1 for k in range(N_K)}
            self.A.addNewRow(amo_row, 1, 1)

            # b) Benefit definition: ssb[i,n] = sum_k B_own[i,k,n] * zssa[i,k]
            for n in range(self.N_n):
                ssb_idx = vm["ssb"].idx(i, n)
                row = {ssb_idx: 1}
                for k in range(N_K):
                    b_val = float(B_own[i, k, n])
                    if b_val != 0.0:
                        row[vm["zssa"].idx(i, k)] = row.get(vm["zssa"].idx(i, k), 0) - b_val
                self.A.addNewRow(row, 0, 0)   # equality: ssb[i,n] - sum_k B_own * zssa = 0

    def _configure_ltcg_constraints(self):
        """
        Configure LTCG tax using LP bracket-allocation variables.

        When self._ltcg_lp is False (default, 'loop' mode):
          Pure LP, no binaries. Uses SC-loop G_n parameter for bracket room.

        When self._ltcg_lp is True ('optimize' mode):
          MILP big-M formulation with G_n as a continuous LP variable and zl binaries
          to select which LTCG bracket applies. Globally optimal within gap.

        Introduces per-year variables q_{pn} (p=0,1,2) partitioning total LTCG across the
        0%, 15%, and 20% capital-gains brackets. Because the LTCG cost function is convex
        (0 < 0.15 < 0.20), the LP naturally minimises tax without binary variables.

        T15_n and T20_n are thresholds on TOTAL taxable income (ordinary + LTCG, after the
        standard deduction). Capital gains are "stacked" on top of ordinary taxable income G_n.
        The bracket constraints are:

        LP mode constraints per year n:
          (1) q[0,n] ≤ max(0, T15_n − G_n)               (cap on 0% allocation)
          (2) q[0,n] + q[1,n] ≤ max(0, T20_n − G_n)      (cap on 0%+15% allocation)
          (3) q[0,n] + q[1,n] + q[2,n] ≥ Q_n             (partition lower bound)

        MILP mode replaces (1) and (2) with big-M binary constraints using gn and zl.

        U_n = 0.15*q[1,n] + 0.20*q[2,n] is computed as a derived quantity after solving.
        """
        tau_0 = self.tau_kn[0, :]
        # Use the same roll as _aggregateResults so the partition constraint is consistent
        # with the Q_n computation used for reporting.
        tau_0prev = np.roll(tau_0, 1)

        for n in range(self.N_n):
            # Per-year filing status: couple switches to Single at n_d.
            status_n = 0 if (self.N_i == 2 and n >= self.n_d) else self.N_i - 1

            # Inflation-adjusted bracket thresholds.
            T15_n = self.gamma_n[n] * tx.capGainRates[status_n][0]
            T20_n = self.gamma_n[n] * tx.capGainRates[status_n][1]

            # Previous year's equity return for capital-gains rate on sold shares.
            # max(0,...) matches capital_gains_rate computation in _aggregateResults.
            cap_rate = max(0.0, tau_0prev[n] - self.mu)

            q0_idx = self.vm["q"].idx(0, n)   # p=0: 0% bracket
            q1_idx = self.vm["q"].idx(1, n)   # p=1: 15% bracket
            q2_idx = self.vm["q"].idx(2, n)   # p=2: 20% bracket

            if self._ltcg_lp:
                # =========================================================
                # MILP mode: G_n is a continuous LP variable (gn), bracket
                # room is encoded via big-M binary constraints with zl.
                # =========================================================
                gn_idx = self.vm["gn"].idx(n)
                zl15_idx = self.vm["zl"].idx(0, n)   # regime binary: G_n < T15
                zl20_idx = self.vm["zl"].idx(1, n)   # regime binary: G_n < T20

                # Big-M: scale with gamma_n[n] (nominal dollars grow with inflation), following
                # the pattern used elsewhere (e.g., bigMBar = bigM * gamma_n[n] in SS taxability).
                # Default: 3*T20_n (already gamma-scaled, safe upper bound on G_n).
                # Using T20_n alone is too tight — G_n can exceed T15+T20 in Roth conversion years.
                base_Mltcg = getattr(self, "_bigMltcg", None)
                if base_Mltcg is None or base_Mltcg <= 0:
                    M_ltcg = 3.0 * T20_n   # gamma_n[n] already embedded in T20_n
                else:
                    M_ltcg = base_Mltcg * self.gamma_n[n]

                # G_n equality: gn = sum_t f_tn  (ordinary taxable income)
                row_gn = {gn_idx: 1}
                for t in range(self.N_t):
                    row_gn[self.vm["f"].idx(t, n)] = row_gn.get(self.vm["f"].idx(t, n), 0) - 1
                self.A.addNewRow(row_gn, 0, 0)

                # Big-M link for zl15: G_n + M*zl15 in [T15, T15+M]
                # Equivalent to: if zl15=0 then G_n = T15 (exactly), if zl15=1 then G_n <= T15+M
                # More precisely: T15 <= G_n + M*zl15 <= T15+M
                # When zl15=0: T15 <= G_n <= T15 → G_n=T15 (at threshold)
                # When zl15=1: T15 <= G_n+M <= T15+M → G_n >= T15-M (always true), G_n <= T15
                # Actually we want: zl15=1 iff G_n >= T15
                # Use: G_n - M*(1-zl15) <= T15  and  G_n >= T15 - M*zl15
                # Simplified: G_n + M*zl15 >= T15  (if zl15=0 → G_n >= T15)
                #             G_n + M*zl15 <= T15+M (always feasible)
                # Better: addNewRow({gn_idx:1, zl15_idx:M_ltcg}, T15_n, T15_n+M_ltcg)
                self.A.addNewRow({gn_idx: 1, zl15_idx: M_ltcg}, T15_n, T15_n + M_ltcg)
                self.A.addNewRow({gn_idx: 1, zl20_idx: M_ltcg}, T20_n, T20_n + M_ltcg)

                # q[0] room15 upper bound: q0 + G_n + M*zl15 <= T15 + M
                # → q0 <= T15 - G_n + M*(1-zl15) (unlimited when zl15=1, i.e. G_n>=T15)
                self.A.addNewRow({q0_idx: 1, gn_idx: 1, zl15_idx: M_ltcg}, -np.inf, T15_n + M_ltcg)
                # q[0] forced zero when G_n >= T15 (zl15=1): q0 - M*zl15 <= 0
                self.A.addNewRow({q0_idx: 1, zl15_idx: -M_ltcg}, -np.inf, 0)

                # q[0]+q[1] room20 upper bound: q0+q1 + G_n + M*zl20 <= T20 + M
                self.A.addNewRow({q0_idx: 1, q1_idx: 1, gn_idx: 1, zl20_idx: M_ltcg}, -np.inf, T20_n + M_ltcg)
                # q[0]+q[1] forced zero when G_n >= T20 (zl20=1): q0+q1 - M*zl20 <= 0
                self.A.addNewRow({q0_idx: 1, q1_idx: 1, zl20_idx: -M_ltcg}, -np.inf, 0)

                # Monotonicity: zl15 <= zl20 (if G_n <= T15 then G_n <= T20, so room for 0% implies room for 15%)
                # zl15 - zl20 <= 0  →  addNewRow({zl15:1, zl20:-1}, -inf, 0)
                self.A.addNewRow({zl15_idx: 1, zl20_idx: -1}, -np.inf, 0)

            else:
                # =========================================================
                # LP (loop) mode: use SC-loop G_n parameter for bracket room.
                # =========================================================
                # LTCG is stacked on top of ordinary taxable income G_n (from previous SC iteration).
                # G_n is initialised to 0 for the first iteration (zero ordinary income assumption).
                # room15_n / room20_n = T15/T20 threshold minus ordinary income already filling the bracket.
                room15_n = max(0.0, T15_n - self.G_n[n])
                room20_n = max(0.0, T20_n - self.G_n[n])
                # (1) q[0,n] ≤ room15_n (enforced via variable upper bound)
                self.B.setRange(q0_idx, 0, room15_n)
                # (2) q[0,n] + q[1,n] ≤ room20_n
                self.A.addNewRow({q0_idx: 1, q1_idx: 1}, -np.inf, room20_n)
                # q[1] upper bound = remaining 15% bracket width after stacking ordinary income.
                self.B.setRange(q1_idx, 0, max(0.0, room20_n - room15_n))
                # q[2] is unbounded above (the 20% bracket has no cap).

            # === Partition lower-bound constraint (both modes): q[0]+q[1]+q[2] ≥ Q_n ===
            # Q_portfolio_n = sum_i alpha_i00n * [mu*(b_i0n + d_in - w_i0n) + cap_rate*w_i0n
            #                                     + 0.5*mu*kappa_i0n]
            # Rearranged: sum_i alpha_i00n * [mu*b_i0n + (cap_rate-mu)*w_i0n + mu*d_in]
            # The kappa half-period correction goes to the RHS.
            rhs_q = self.fixed_assets_capital_gains_n[n]
            row_q = {q0_idx: 1, q1_idx: 1, q2_idx: 1}
            for i in range(self.N_i):
                alpha = self.alpha_ijkn[i, 0, 0, n]
                if alpha == 0:
                    continue
                b_idx = self.vm["b"].idx(i, 0, n)
                w_idx = self.vm["w"].idx(i, 0, n)
                d_idx = self.vm["d"].idx(i, n)
                row_q[b_idx] = row_q.get(b_idx, 0) - alpha * self.mu
                row_q[w_idx] = row_q.get(w_idx, 0) - alpha * (cap_rate - self.mu)
                row_q[d_idx] = row_q.get(d_idx, 0) - alpha * self.mu
                rhs_q += alpha * 0.5 * self.mu * self.kappa_ijn[i, 0, n]

            # (3) q[0]+q[1]+q[2] − Q_portfolio_LP_vars ≥ Q_fixed + kappa_correction
            self.A.addNewRow(row_q, rhs_q, np.inf)

    def _add_magi_lp(self, options):
        """
        Add MAGI equality constraints when withNIIT='optimize'.

        Links vm["magi"].idx(n) to:
          MAGI_n = G_n + e_n + Q_n + zetaBar_n - tss_n
        where:
          - G_n = sum_t f_tn  (ordinary taxable income, via gn LP var or direct f_tn sum)
          - e_n = standard deduction headroom
          - Q_n = portfolio LTCG (sum of q vars minus portfolio LP terms)
          - zetaBar_n = total SS benefits (parameter)
          - tss_n = taxable SS (LP var when withSSTaxability=optimize, else Psi_n*zetaBar_n param)

        Rewritten as equality constraint with all LP vars on LHS:
          magi_n - gn_n - e_n - Q_LP_expr = zetaBar_n - tss_or_param
        """
        if not self._niit_lp:
            return

        tau_0 = self.tau_kn[0, :]
        tau_0prev = np.roll(tau_0, 1)

        zetaBar_n = np.sum(self.zetaBar_in, axis=0)

        for n in range(self.N_n):
            magi_idx = self.vm["magi"].idx(n)
            e_idx = self.vm["e"].idx(n)
            cap_rate = max(0.0, tau_0prev[n] - self.mu)

            # Build MAGI equality row:
            # magi_n = G_n + e_n + Q_n + (1-Psi_n)*zetaBar_n
            # where Q_n = sum_i alpha_i00n*[mu*b + (cr-mu)*w + mu*d] + rhs_q (fixed assets)
            # Rewrite: magi_n - e_n - Q_lp_terms = zetaBar_n - tss_n + fixed_asset_cg
            row = {magi_idx: 1, e_idx: -1}

            # G_n contribution: either via gn LP var or directly from f_tn vars
            if "gn" in self.vm:
                row[self.vm["gn"].idx(n)] = row.get(self.vm["gn"].idx(n), 0) - 1
            else:
                for t in range(self.N_t):
                    f_idx = self.vm["f"].idx(t, n)
                    row[f_idx] = row.get(f_idx, 0) - 1

            # Q_n LP expression: same as partition constraint in _configure_ltcg_constraints
            rhs_magi = self.fixed_assets_capital_gains_n[n]
            q0_idx = self.vm["q"].idx(0, n)
            q1_idx = self.vm["q"].idx(1, n)
            q2_idx = self.vm["q"].idx(2, n)
            row[q0_idx] = row.get(q0_idx, 0) - 1
            row[q1_idx] = row.get(q1_idx, 0) - 1
            row[q2_idx] = row.get(q2_idx, 0) - 1
            for i in range(self.N_i):
                alpha = self.alpha_ijkn[i, 0, 0, n]
                if alpha == 0:
                    continue
                b_idx = self.vm["b"].idx(i, 0, n)
                w_idx = self.vm["w"].idx(i, 0, n)
                d_idx = self.vm["d"].idx(i, n)
                row[b_idx] = row.get(b_idx, 0) + alpha * self.mu
                row[w_idx] = row.get(w_idx, 0) + alpha * (cap_rate - self.mu)
                row[d_idx] = row.get(d_idx, 0) + alpha * self.mu
                rhs_magi -= alpha * 0.5 * self.mu * self.kappa_ijn[i, 0, n]

            # SS: zetaBar_n (parameter) minus taxable SS
            rhs_magi += zetaBar_n[n]
            if "tss" in self.vm:
                # tss is an LP var: subtract it from LHS
                tss_idx = self.vm["tss"].idx(n)
                row[tss_idx] = row.get(tss_idx, 0) + 1  # +1 because moved to LHS: magi - tss = rhs
            else:
                # Use Psi_n parameter: taxable SS = Psi_n * zetaBar_n
                rhs_magi -= self.Psi_n[n] * zetaBar_n[n]

            self.A.addNewRow(row, rhs_magi, rhs_magi)

    def _configure_NIIT_binary_variables(self, options):
        """
        Add NIIT big-M binary constraints when withNIIT='optimize'.

        For each year n, links J_n (NIIT tax) to MAGI_n via a threshold binary zj:
          J_n = max(0, 0.038 * (MAGI_n - T_niit))

        Modeled with binary zj_n:
          zj=1 iff MAGI_n > T_niit (NIIT threshold exceeded)

        Big-M constraints:
          (1) J_n >= 0.038*(MAGI_n - T_niit) - M*(1-zj)  [NIIT rate when above threshold]
          (2) J_n <= M*zj                                  [J=0 when below threshold]
          (3) MAGI_n <= T_niit + M*zj                     [MAGI bounded by threshold when zj=0]
        """
        if not self._niit_lp:
            return

        for n in range(self.N_n):
            # Per-year filing status: couple switches to Single at n_d.
            status_n = 0 if (self.N_i == 2 and n >= self.n_d) else self.N_i - 1
            T_niit = 200000.0 if status_n == 0 else 250000.0  # NOT inflation-adjusted

            # Big-M: scale with gamma_n[n] following the convention used elsewhere in the code.
            # Default: 3*T20_n (already gamma-scaled via T20_n = gamma_n[n]*capGainRates).
            T20_n = self.gamma_n[n] * tx.capGainRates[status_n][1]
            base_Mniit = getattr(self, "_bigMniit", None)
            if base_Mniit is None or base_Mniit <= 0:
                M_niit = 3.0 * T20_n   # gamma_n[n] already embedded in T20_n
            else:
                M_niit = base_Mniit * self.gamma_n[n]

            Jn_idx = self.vm["Jn"].idx(n)
            magi_idx = self.vm["magi"].idx(n)
            zj_idx = self.vm["zj"].idx(n)

            # Bounds on Jn and magi
            self.B.setRange(Jn_idx, 0, M_niit)
            self.B.setRange(magi_idx, 0, M_niit)

            # (1) J_n >= 0.038*(MAGI_n - T_niit) - M*(1-zj)
            #   → J_n - 0.038*MAGI_n + M*zj >= 0.038*(-T_niit) + M*(-1)
            #   Wait, rearranging: J_n - 0.038*MAGI_n - M*zj >= -0.038*T_niit - M
            #   But addNewRow lb <= row <= ub, so:
            #   lb = -0.038*T_niit - M_niit, row = {Jn:1, magi:-0.038, zj:-M_niit}
            self.A.addNewRow(
                {Jn_idx: 1, magi_idx: -0.038, zj_idx: -M_niit},
                -0.038 * T_niit - M_niit, np.inf
            )

            # (2) J_n <= M*zj
            #   → J_n - M*zj <= 0
            self.A.addNewRow({Jn_idx: 1, zj_idx: -M_niit}, -np.inf, 0)

            # (3) MAGI_n <= T_niit + M*zj
            #   → MAGI_n - M*zj <= T_niit
            self.A.addNewRow({magi_idx: 1, zj_idx: -M_niit}, -np.inf, T_niit)

    def _configure_Medicare_binary_variables(self, options):
        if options.get("withMedicare", "loop") != "optimize":
            return

        bigM = u.get_numeric_option(options, "bigMamo", BIGM_AMO, min_value=0)
        Nmed = self.N_n - self.nm
        # Select exactly one IRMAA bracket per year (SOS1 behavior).
        for nn in range(Nmed):
            row = self.A.newRow()
            for q in range(self.N_irmaa):
                row.addElem(self.vm["zm"].idx(nn, q), 1)
            self.A.addRow(row, 1, 1)

        # MAGI decomposition into bracket portions: sum_q h_{q} = MAGI.
        for nn in range(Nmed):
            n = self.nm + nn
            row = self.A.newRow()
            for q in range(self.N_irmaa):
                row.addElem(self.vm["h"].idx(nn, q), 1)

            if n < 2:
                # MAGI for the first two plan years is known (prevMAGI from user-supplied data).
                self.A.addRow(row, self.prevMAGI[n], self.prevMAGI[n])
                # Pre-fix the bracket to match the known MAGI in all solver modes, including
                # Benders.  The correct bracket is deterministic; pre-fixing (Lb == Ub) causes
                # _benders_solve to exclude these zm columns from master_cols automatically.
                # Without pre-fixing, the LP relaxation pushes h-values toward low-premium
                # brackets (maximizer behaviour) so argmax(h) picks the wrong bracket for these
                # years, making the subproblem LP infeasible on the first Benders iteration.
                magi = self.prevMAGI[n]
                qsel = 0
                for q in range(1, self.N_irmaa):
                    if magi > self.Lbar_nq[nn, q - 1]:
                        qsel = q

                for q in range(self.N_irmaa):
                    idx = self.vm["zm"].idx(nn, q)
                    val = 1 if q == qsel else 0
                    self.B.setRange(idx, val, val)
                continue

            n2 = n - 2
            rhs = (self.fixed_assets_ordinary_income_n[n2]
                   + self.fixed_assets_capital_gains_n[n2])

            row.addElem(self.vm["e"].idx(n2), -1)
            for i in range(self.N_i):
                row.addElem(self.vm["w"].idx(i, 1, n2), -1)
                row.addElem(self.vm["x"].idx(i, n2), -1)

                # Dividends and interest gains for year n2. Only positive returns are taxable.
                afac = (self.mu * self.alpha_ijkn[i, 0, 0, n2]
                        + np.sum(self.alpha_ijkn[i, 0, 1:, n2] * np.maximum(0, self.tau_kn[1:, n2])))

                row.addElem(self.vm["b"].idx(i, 0, n2), -afac)
                row.addElem(self.vm["d"].idx(i, n2), -afac)

                # Capital gains on stocks sold from taxable account accrued in year n2 - 1.
                # Capital gains = price appreciation only (total return - dividend rate)
                #  to avoid double taxation of dividends.
                tau_prev = self.tau_kn[0, max(0, n2 - 1)]
                bfac = self.alpha_ijkn[i, 0, 0, n2] * max(0, tau_prev - self.mu)
                row.addElem(self.vm["w"].idx(i, 0, n2), afac - bfac)

                # MAGI includes total Social Security (taxable + non-taxable) for IRMAA.
                sumoni = (self.omega_in[i, n2]
                          + self.other_inc_in[i, n2]
                          + self.netinv_in[i, n2]
                          + self.zetaBar_in[i, n2]
                          + self.piBar_in[i, n2]
                          + 0.5 * self.kappa_ijn[i, 0, n2] * afac)
                rhs += sumoni
                rhs -= self.kappa_ijn[i, 3, n2]   # HSA contributions reduce MAGI

            self.A.addRow(row, rhs, rhs)

        # Bracket bounds: L_{q-1} z_q <= mg_q <= L_q z_q.
        for nn in range(Nmed):
            for q in range(self.N_irmaa):
                mg_idx = self.vm["h"].idx(nn, q)
                zm_idx = self.vm["zm"].idx(nn, q)

                lower = 0 if q == 0 else self.Lbar_nq[nn, q - 1]
                if lower > 0:
                    self.A.addNewRow({mg_idx: 1, zm_idx: -lower}, 0, np.inf)

                if q < self.N_irmaa - 1:
                    upper = self.Lbar_nq[nn, q]
                    self.A.addNewRow({mg_idx: 1, zm_idx: -upper}, -np.inf, 0)
                else:
                    # Upper bound for last bracket so h_qn = 0 when z_q = 0.
                    upper = bigM * self.gamma_n[self.nm + nn]
                    self.A.addNewRow({mg_idx: 1, zm_idx: -upper}, -np.inf, 0)

    def _add_Medicare_costs(self, options):
        if options.get("withMedicare", "loop") != "optimize":
            # In loop mode, Medicare costs are computed outside the solver (M_n).
            # Ensure the in-model Medicare variable (m_n) stays at zero.
            for n in range(self.N_n):
                self.B.setRange(self.vm["m"].idx(n), 0, 0)
            return

        for n in range(self.nm):
            self.B.setRange(self.vm["m"].idx(n), 0, 0)

        Nmed = self.N_n - self.nm
        for nn in range(Nmed):
            n = self.nm + nn
            row = self.A.newRow()
            row.addElem(self.vm["m"].idx(n), 1)
            for q in range(self.N_irmaa):
                row.addElem(self.vm["zm"].idx(nn, q), -self.Cbar_nq[nn, q])
            self.A.addRow(row, 0, 0)

    def _configure_ACA_binary_variables(self, options):
        """
        Build ACA MIP constraints (withACA="optimize" mode only).

        Three constraint groups:
          a) SOS1: exactly one ACA bracket selected per year.
          b) MAGI decomposition: sum_q haca[nn, q] = MAGI_n (current year, no 2-year lag).
          c) Bracket bounds (Big-M): MAGI portion in bracket q is within its FPL thresholds.

        Note: ACA uses current-year MAGI (no 2-year lag like Medicare IRMAA).
        Note: MAGI below 138% FPL qualifies for Medicaid — the LP places it in the lowest
              bracket at the base contribution rate (2.1%) rather than returning full SLCSP.
        """
        if not self._aca_lp:
            return

        bigM = u.get_numeric_option(options, "bigMaca", BIGM_AMO, min_value=0)
        tau_0 = self.tau_kn[0, :]

        # a) SOS1: exactly one bracket selected per year.
        for nn in range(self.n_aca):
            row = self.A.newRow()
            for r in range(tx.N_ACA_R):
                row.addElem(self.vm["za"].idx(nn, r), 1)
            self.A.addRow(row, 1, 1)

        # b) MAGI decomposition: sum_r haca[nn, r] = current-year MAGI.
        for nn in range(self.n_aca):
            n = nn  # ACA uses current year (no lag)
            tau_prev = tau_0[max(0, n - 1)]

            rhs_magi = (self.fixed_assets_ordinary_income_n[n]
                        + self.fixed_assets_capital_gains_n[n])

            row_magi = {}
            row_magi[self.vm["e"].idx(n)] = -1

            for i in range(self.N_i):
                afac = (self.mu * self.alpha_ijkn[i, 0, 0, n]
                        + np.sum(self.alpha_ijkn[i, 0, 1:, n] * np.maximum(0, self.tau_kn[1:, n])))
                bfac = self.alpha_ijkn[i, 0, 0, n] * max(0, tau_prev - self.mu)

                w1_idx = self.vm["w"].idx(i, 1, n)
                x_idx = self.vm["x"].idx(i, n)
                b_idx = self.vm["b"].idx(i, 0, n)
                d_idx = self.vm["d"].idx(i, n)
                w0_idx = self.vm["w"].idx(i, 0, n)

                row_magi[w1_idx] = row_magi.get(w1_idx, 0) - 1
                row_magi[x_idx] = row_magi.get(x_idx, 0) - 1
                row_magi[b_idx] = row_magi.get(b_idx, 0) - afac
                row_magi[d_idx] = row_magi.get(d_idx, 0) - afac
                row_magi[w0_idx] = row_magi.get(w0_idx, 0) + (afac - bfac)

                rhs_magi += (self.omega_in[i, n]
                             + self.other_inc_in[i, n]
                             + self.netinv_in[i, n]
                             + self.zetaBar_in[i, n]    # full SS (not 0.5×SS; ACA uses MAGI)
                             + self.piBar_in[i, n]
                             + 0.5 * self.kappa_ijn[i, 0, n] * afac)
                rhs_magi -= self.kappa_ijn[i, 3, n]     # HSA contributions reduce MAGI

            for r in range(tx.N_ACA_R):
                haca_idx = self.vm["haca"].idx(nn, r)
                row_magi[haca_idx] = row_magi.get(haca_idx, 0) + 1

            self.A.addNewRow(row_magi, rhs_magi, rhs_magi)

        # c) Bracket bounds: Lbar[nn, r-1]*za[r] <= haca[r] <= Lbar[nn, r]*za[r].
        for nn in range(self.n_aca):
            for r in range(tx.N_ACA_R):
                haca_idx = self.vm["haca"].idx(nn, r)
                za_idx = self.vm["za"].idx(nn, r)

                lower = 0 if r == 0 else self.Lbar_aca_nr[nn, r - 1]
                if lower > 0:
                    self.A.addNewRow({haca_idx: 1, za_idx: -lower}, 0, np.inf)

                if r < tx.N_ACA_R - 1:
                    upper = self.Lbar_aca_nr[nn, r]
                    self.A.addNewRow({haca_idx: 1, za_idx: -upper}, -np.inf, 0)
                else:
                    # Last bracket (above 400% FPL): use BigM as upper bound so haca = 0 when za = 0.
                    upper = bigM * self.gamma_n[nn]
                    self.A.addNewRow({haca_idx: 1, za_idx: -upper}, -np.inf, 0)

    def _add_ACA_costs(self, options):
        """
        Add ACA cost constraints for the LP/MIP formulation (optimize mode only).

        In optimize mode: maca_n = sum_{r=0}^{5} cap_pct_r * haca[nn,r] + slcsp_aca_n[nn]*za[nn,6].
        For brackets 0-5: proportional cost. For bracket 6 (above 400% FPL): fixed cost = SLCSP.
        In loop mode: maca variable does not exist; ACA_n (SC loop) goes in the cash-flow RHS.
        """
        if not self._aca_lp:
            return  # Loop mode: no maca variable; ACA_n from SC loop is in the cash-flow RHS.

        # Pin post-ACA years to zero (individual is on Medicare; no ACA cost).
        for n in range(self.n_aca, self.N_n):
            self.B.setRange(self.vm["maca"].idx(n), 0, 0)

        # Cost constraint: maca_n = sum_{r=0}^{5} cap_pct_r * haca[nn,r] + slcsp*za[nn,6].
        # Bracket 6 (>400% FPL): 2026 rules impose full SLCSP (no PTC), not proportional to MAGI.
        for nn in range(self.n_aca):
            row = self.A.newRow({self.vm["maca"].idx(nn): 1})
            for r in range(tx.N_ACA_R - 1):  # r=0..5 only; bracket 6 uses fixed SLCSP
                row.addElem(self.vm["haca"].idx(nn, r), -self.cap_pct_aca_r[r])
            row.addElem(self.vm["za"].idx(nn, tx.N_ACA_R - 1), -self.slcsp_aca_n[nn])
            self.A.addRow(row, 0, 0)
            self.B.setRange(self.vm["maca"].idx(nn), 0, self.slcsp_aca_n[nn])

    def _build_objective_vector(self, objective, options):
        c_arr = np.zeros(self.nvars)

        # Time preference discount: ρ > 0 increases the relative value of near-term spending,
        # counteracting the optimizer's tendency to back-load spending. Has no effect on maxBequest.
        rho = u.get_numeric_option(options, "timePreference", 0.0, min_value=0) / 100.0
        discount_n = np.array([(1.0 / (1.0 + rho)) ** n for n in range(self.N_n)])

        if objective == "maxSpending":
            for n in range(self.N_n):
                c_arr[self.vm["g"].idx(n)] = -discount_n[n] / self.gamma_n[n]
        elif objective == "maxBequest":
            for i in range(self.N_i):
                c_arr[self.vm["b"].idx(i, 0, self.N_n)] = -1
                c_arr[self.vm["b"].idx(i, 1, self.N_n)] = -(1 - self.nu)
                c_arr[self.vm["b"].idx(i, 2, self.N_n)] = -1
                c_arr[self.vm["b"].idx(i, 3, self.N_n)] = -(1 - self.nu)   # HSA: heirs pay ordinary income tax
        elif objective == "maxHybrid":
            # Both terms expressed in present-value dollars for balanced weighting:
            # - Spending: PV-weighted by discount_n/γₙ per year (same as maxSpending)
            # - Bequest: nominal final balances divided by γ_N (terminal inflation factor → PV)
            # This makes h=0.5 a genuine balance point: lifetime PV spending ≈ PV bequest.
            h = float(options.get("spendingWeight", 0.5))
            bequest_pv_scale = self.gamma_n[-1]
            for n in range(self.N_n):
                c_arr[self.vm["g"].idx(n)] -= h * discount_n[n] / self.gamma_n[n]
            for i in range(self.N_i):
                c_arr[self.vm["b"].idx(i, 0, self.N_n)] -= (1 - h) / bequest_pv_scale
                c_arr[self.vm["b"].idx(i, 1, self.N_n)] -= (1 - h) * (1 - self.nu) / bequest_pv_scale
                c_arr[self.vm["b"].idx(i, 2, self.N_n)] -= (1 - h) / bequest_pv_scale
                c_arr[self.vm["b"].idx(i, 3, self.N_n)] -= (1 - h) * (1 - self.nu) / bequest_pv_scale
        else:
            raise RuntimeError("Internal error in objective function.")

        # Turn on epsilon by default to reduce churn and frontload Roth conversions.
        default_epsilon = EPSILON
        epsilon = u.get_numeric_option(options, "epsilon", default_epsilon, min_value=0)
        if epsilon > 0:
            # Penalize Roth conversions to reduce churn.
            for i in range(self.N_i):
                for n in range(self.N_n):
                    c_arr[self.vm["x"].idx(i, n)] += epsilon * (1 + n)

            if self.N_i == 2:
                # Favor withdrawals from spouse 0 by penalizing spouse 1 withdrawals.
                for j in range(self.N_j):
                    for n in range(self.N_n):
                        c_arr[self.vm["w"].idx(1, j, n)] += epsilon

        c = abc.Objective(self.nvars)
        for idx in np.flatnonzero(c_arr):
            c.setElem(idx, c_arr[idx])
        self.c = c

    @_timer
    def runHistoricalRange(self, objective, options, ystart, yend, *, verbose=False, figure=False,
                           progcall=None, reverse=False, roll=0, augmented=False, log_x=False):
        return run_historical_range(
            self, objective, options, ystart, yend, verbose=verbose, figure=figure, progcall=progcall,
            reverse=reverse, roll=roll, augmented=augmented, log_x=log_x)

    @_timer
    def runMC(self, objective, options, N, verbose=False, figure=False, progcall=None, log_x=False):
        return run_mc(
            self, objective, options, N, verbose=verbose, figure=figure, progcall=progcall, log_x=log_x)

    @_timer
    def runStochasticSpending(self, options, scenario_method, *,
                              ystart=None, yend=None, N=None, progcall=None,
                              reverse=False, roll=0,
                              with_longevity=False, sexes=None, seed=None):
        return run_stochastic_spending(
            self, options, scenario_method, ystart=ystart, yend=yend, N=N,
            progcall=progcall, reverse=reverse, roll=roll,
            with_longevity=with_longevity, sexes=sexes, seed=seed)

    @_timer
    def runSpendingFrontier(self, scenario_method, *,
                            ystart=None, yend=None, N=None, progcall=None,
                            reverse=False, roll=0,
                            with_longevity=False, sexes=None, seed=None, **kwargs):
        """
        Run the spending efficient frontier over historical or Monte Carlo scenarios.

        Accepts the same solver options as solve() via keyword arguments.  The
        objective is always maxSpending; passing ``netSpending`` raises an error.

        Parameters
        ----------
        scenario_method : str
            "historical" — sweep ``ystart``..``yend``.
            "mc"         — ``N`` Monte Carlo draws.
        ystart, yend : int, optional
            Start/end years for historical mode.
        N : int, optional
            Number of simulations for Monte Carlo mode.
        progcall : Progress, optional
            Progress callback.
        with_longevity : bool, optional
            Draw random lifespans from SSA tables for each scenario.
        sexes : list of str, optional
            Sex of each individual ('M' or 'F').  Required when with_longevity=True.
        seed : int or None, optional
            Random seed for reproducible longevity draws.
        **kwargs
            Solver options forwarded to solve() (e.g. bequest, maxRothConversion,
            withMedicare, solver, withSSTaxability, previousMAGIs).

        Returns
        -------
        dict — see runStochasticSpending for keys.
        """
        if kwargs.get("objective", "maxSpending") != "maxSpending":
            raise ValueError(
                "runSpendingFrontier only supports 'maxSpending'; "
                f"got objective='{kwargs['objective']}'."
            )
        if "netSpending" in kwargs:
            self.mylog.print("Warning: 'netSpending' is ignored by runSpendingFrontier (maxSpending objective).")
            kwargs.pop("netSpending")
        return run_stochastic_spending(
            self, kwargs, scenario_method, ystart=ystart, yend=yend, N=N,
            progcall=progcall, reverse=reverse, roll=roll,
            with_longevity=with_longevity, sexes=sexes, seed=seed)

    def resolve(self):
        """
        Solve a plan using saved options.
        """
        self.solve(self.objective, self.solverOptions)

        return None

    @_checkConfiguration
    @_timer
    def solve(self, objective, options=None):
        """
        This function builds the necessary constaints and
        runs the optimizer.

        - objective can be 'maxSpending' or 'maxBequest'.

        - options is a dictionary which can include:
            - maxRothConversion: Only allow conversion smaller than amount specified.
            - netSpending: Desired spending amount when optimizing with maxBequest.
            - fixedSpending: Pin first-year spending for maxSpending; slack varies g[n] dynamically.
            - bequest: Value of bequest in today's $ when optimizing with maxSpending.
            - units: Units to use for amounts (1, k, or M).

        All units are in $k, unless specified otherwise.

        Refer to companion document for implementation details.
        """

        if self.rateMethod is None:
            raise RuntimeError("Rate method must be selected before solving.")

        # Assume unsuccessful until problem solved.
        self.caseStatus = "unsuccessful"
        self.convergenceType = "undefined"

        # Check objective and required options.
        knownObjectives = ["maxBequest", "maxSpending", "maxHybrid"]
        knownSolvers = ["default", "HiGHS", "MOSEK"]

        knownOptions = [
            "absTol",
            "amoConstraints",
            "amoRoth",
            "amoSurplus",
            "bequest",
            "bigMamo",    # Big-M value for AMO constraints (default: 5e7)
            "epsilon",
            "gap",
            "maxIter",
            "maxRothConversion",
            "minTaxableBalance",
            "netSpending",
            "noLateSurplus",
            "noRothConversions",
            "oppCostX",
            "previousMAGIs",
            "relTol",
            "solver",
            "fixedSpending",   # Pin first-year spending for maxSpending; optimizer uses slack dynamically
            "spendingFloor",   # Minimum annual spending for maxHybrid (today's $k)
            "spendingSlack",
            "spendingWeight",  # Blend weight h ∈ [0,1] for maxHybrid (1=spending, 0=bequest)
            "timePreference",  # Subjective time discount rate (%/year) to front-load spending
            "startRothConversions",
            "swapRothConverters",
            "maxTime",
            "units",
            "verbose",
            "withACA",        # ACA handling: "loop" (default) or "optimize"
            "bigMaca",        # Big-M for ACA bracket upper bounds (default: BIGM_AMO)
            "withLTCG",       # LTCG handling: "loop" (default) or "optimize"
            "bigMltcg",       # Big-M for LTCG bracket constraints (default: T20_n per year)
            "withNIIT",       # NIIT handling: "loop" (default) or "optimize"
            "bigMniit",       # Big-M for NIIT threshold constraints (default: 3*T20_n per year)
            "bendersMaxIter",      # Maximum Benders iterations (default: 50)
            "withDecomposition",  # MIP decomposition: "none" (default), "sequential", or "benders"
            "withMedicare",
            "withSCLoop",
            "withSSTaxability",
            "withSSAges",          # SS claiming age: "fixed" (default) or "optimize"
        ]
        options = {} if options is None else options

        # We might modify options if required.
        myoptions = dict(options)

        for opt in list(myoptions.keys()):
            if opt not in knownOptions:
                # raise ValueError(f"Option '{opt}' is not one of {knownOptions}.")
                self.mylog.print(f"Ignoring unknown solver option '{opt}'.")
                myoptions.pop(opt)

        if objective not in knownObjectives:
            raise ValueError(f"Objective '{objective}' is not one of {knownObjectives}.")

        if objective == "maxBequest" and "netSpending" not in myoptions:
            raise RuntimeError(f"Objective '{objective}' needs netSpending option.")

        if objective == "maxBequest" and "bequest" in myoptions:
            self.mylog.print("Ignoring bequest option provided.")
            myoptions.pop("bequest")

        if objective == "maxSpending" and "netSpending" in myoptions:
            self.mylog.print("Ignoring netSpending option provided.")
            myoptions.pop("netSpending")

        if objective == "maxSpending" and "bequest" not in myoptions:
            self.mylog.vprint("Using bequest of $1.")

        if objective == "maxHybrid" and "netSpending" in myoptions:
            self.mylog.print("Ignoring netSpending option (not used in maxHybrid; use spendingFloor).")
            myoptions.pop("netSpending")

        if objective == "maxHybrid" and "bequest" in myoptions:
            self.mylog.print("Ignoring bequest option (not used in maxHybrid; use spendingFloor).")
            myoptions.pop("bequest")

        oppCostX = myoptions.get("oppCostX", 0.)
        self.xnet = 1 - oppCostX / 100.

        if "swapRothConverters" in myoptions and "noRothConversions" in myoptions:
            self.mylog.print("Ignoring 'noRothConversions' as 'swapRothConverters' option present.")
            myoptions.pop("noRothConversions")

        # Go easy on MILP - auto gap somehow.
        if "gap" not in myoptions and myoptions.get("withMedicare", "loop") == "optimize":
            fac = 1
            maxRoth = myoptions.get("maxRothConversion", 100)
            if maxRoth != "file" and maxRoth <= 15:
                fac = 10
            # Loosen default MIP gap when Medicare is optimized. Even more if rothX == 0
            gap = fac * MILP_GAP
            myoptions["gap"] = gap
            self.mylog.vprint(f"Using restricted gap of {gap:.1e}.")

        self.prevMAGI = np.zeros(2)
        if "previousMAGIs" in myoptions:
            self.prevMAGI = np.array(u.get_monetary_list_option(myoptions, "previousMAGIs", 2))

        lambdha = myoptions.get("spendingSlack", 0)
        if not (0 <= lambdha <= 50):
            raise ValueError(f"Slack value {lambdha} out of range.")
        self.lambdha = lambdha / 100

        # Reset MAGI to zero.
        self.MAGI_n = np.zeros(self.N_n)
        self.J_n = np.zeros(self.N_n)
        self.M_n = np.zeros(self.N_n)
        self.ACA_n = np.zeros(self.N_n)
        self.maca_n = np.zeros(self.N_n)
        self._aca_lp = False            # Will be set to True in _buildOffsetMap when withACA="optimize"
        self._ltcg_lp = False           # Will be set to True in _buildOffsetMap when withLTCG="optimize"
        self._niit_lp = False           # Will be set to True in _buildOffsetMap when withNIIT="optimize"
        self._ssa_lp = False            # Will be set to True in _buildOffsetMap when withSSAges="optimize"
        self._adjustedParameters = False   # Force fresh parameter setup for each solve()
        self._highs_warm_start = None   # MIP warm-start hint; reset each solve(), updated each SC iter

        self._adjustParameters(self.gamma_n, self.MAGI_n)
        self._buildOffsetMap(myoptions)

        # Process debts and fixed assets
        self.processDebtsAndFixedAssets()

        solver = myoptions.get("solver", self.defaultSolver)
        if solver == "default":
            solver = "MOSEK" if _mosek_available() else "HiGHS"
        if solver not in knownSolvers:
            raise ValueError(f"Unknown solver '{solver}'.")

        if solver == "HiGHS":
            solverMethod = self._milpSolve
        elif solver == "MOSEK":
            solverMethod = self._mosekSolve
        else:
            raise RuntimeError("Internal error in defining solverMethod.")

        self.mylog.vprint(f"Using '{solver}' solver for optimizing {objective}.")
        myoptions_txt = textwrap.fill(f"{myoptions}", initial_indent="\t", subsequent_indent="\t", width=100)
        self.mylog.vprint(f"Solver options:\n{myoptions_txt}.")
        self._scSolve(objective, myoptions, solverMethod)

        self.objective = objective
        self.solverOptions = myoptions

        return None

    def _scSolve(self, objective, options, solverMethod):
        """
        Self-consistent loop, regardless of solver.
        """
        includeMedicare = options.get("withMedicare", "loop") == "loop"
        withSCLoop = options.get("withSCLoop", True)
        ss_val = options.get("withSSTaxability", "loop")
        fixed_psi = float(ss_val) if isinstance(ss_val, (int, float)) else None

        # Convergence uses a relative tolerance tied to MILP gap,
        # with an absolute floor to avoid zero/near-zero objectives.
        gap = u.get_numeric_option(options, "gap", GAP, min_value=0)
        abs_tol = u.get_numeric_option(options, "absTol", ABS_TOL, min_value=0)
        rel_tol = options.get("relTol")
        if rel_tol is None:
            # Keep rel_tol aligned with solver gap to avoid SC loop chasing noise.
            rel_tol = max(REL_TOL, gap / 300)
        # rel_tol = u.get_numeric_option({"relTol": rel_tol}, "relTol", REL_TOL, min_value=0)
        self.mylog.print(f"Using relTol={rel_tol:.1e}, absTol={abs_tol:.1e}, and gap={gap:.1e}.")

        max_iterations = int(u.get_numeric_option(options, "maxIter", MAX_ITERATIONS, min_value=1))
        # self.mylog.print(f"Using maxIter={max_iterations}.")

        if objective == "maxSpending":
            objFac = -1 / self.xi_n[0]
        else:
            objFac = -1 / self.gamma_n[-1]

        it = 0
        old_x = np.zeros(self.nvars)
        old_objfns = [np.inf]
        scaled_obj_history = []  # Track scaled objective values for oscillation detection
        sol_history = []  # Track solutions aligned with scaled_obj_history
        obj_history = []  # Track raw objective values aligned with scaled_obj_history
        # Decomposition dispatch: for sequential mode, replace the monolithic MIP
        # with a hierarchical relax-and-fix solver (HiGHS only).
        decomp_mode = options.get("withDecomposition", "none")
        # Use __func__ comparison to identify solver regardless of bound method identity.
        is_milp = getattr(solverMethod, "__func__", None) is Plan._milpSolve
        is_mosek = getattr(solverMethod, "__func__", None) is Plan._mosekSolve
        is_decomposable = is_milp or is_mosek
        self._decomp_use_mosek = is_mosek   # consumed by _relax_and_fix_solve / _benders_solve
        if decomp_mode == "sequential" and is_decomposable:
            actualSolverMethod = self._relax_and_fix_solve
        elif decomp_mode == "benders" and is_decomposable:
            actualSolverMethod = self._benders_solve
        else:
            if decomp_mode not in ("none", "sequential", "benders"):
                self.mylog.print(f"Unknown withDecomposition mode '{decomp_mode}'; using 'none'.")
            actualSolverMethod = solverMethod

        self._computeNLstuff(None, includeMedicare, fixedPsi=fixed_psi)
        while True:
            objfn, xx, solverSuccess, solverMsg, solgap = actualSolverMethod(objective, options)

            if not solverSuccess or objfn is None:
                self.mylog.print("Solver failed:", solverMsg, solverSuccess)
                break

            if not withSCLoop and it >= 1:
                # When Medicare is in loop mode, M_n was zero in the constraint for this
                # single solve. Update M_n (and J_n) from solution for reporting.
                if includeMedicare:
                    self._computeNLstuff(xx, includeMedicare, fixedPsi=fixed_psi)
                    self.mylog.print(
                        "Warning: Self-consistent loop is off; Medicare premiums are "
                        "computed for display but were not in the budget constraint."
                    )
                break

            # When withSCLoop=False, only update G_n (needed for LTCG bracket accuracy)
            # by passing includeMedicare=False; this preserves the no-Medicare-loop behavior.
            self._computeNLstuff(xx, includeMedicare if withSCLoop else False, fixedPsi=fixed_psi)

            delta = xx - old_x
            # Only consider account balances in dX.
            absSolDiff = np.sum(np.abs(delta[:self.nbals]), axis=0)/self.nbals
            absObjDiff = abs(objFac*(objfn + old_objfns[-1]))
            scaled_obj = objfn * objFac
            scaled_obj_history.append(scaled_obj)
            sol_history.append(xx)
            obj_history.append(objfn)
            self.mylog.vprint(f"Iter: {it:02}; f: {u.d(scaled_obj, f=0)}; gap: {solgap:.1e};"
                              f" |dX|: {absSolDiff:.0f}; |df|: {u.d(absObjDiff, f=0)}")

            # Solution difference is calculated and reported but not used for convergence
            # since it scales with problem size and can prevent convergence for large cases.
            prev_scaled_obj = scaled_obj
            if np.isfinite(old_objfns[-1]):
                prev_scaled_obj = (-old_objfns[-1]) * objFac
            scale = max(1.0, abs(scaled_obj), abs(prev_scaled_obj))
            tol = max(abs_tol, rel_tol * scale)
            # With Medicare in loop mode, the first solve uses M_n=0; require at least
            # one re-solve so the accepted solution had Medicare in the budget.
            if absObjDiff <= tol and (not includeMedicare or it >= 1):
                # Check if convergence was monotonic or oscillatory
                # old_objfns stores -objfn values, so we need to scale them to match displayed values
                # For monotonic convergence, the scaled objective (objfn * objFac) should be non-increasing
                # Include current iteration's scaled objfn value
                scaled_objfns = [(-val) * objFac for val in old_objfns[1:]] + [scaled_obj]
                # Check if scaled objective function is non-increasing (monotonic convergence)
                is_monotonic = all(scaled_objfns[i] <= scaled_objfns[i-1] + tol
                                   for i in range(1, len(scaled_objfns)))
                if is_monotonic:
                    self.convergenceType = "monotonic"
                else:
                    self.convergenceType = "oscillatory"
                self.mylog.print(f"Converged on full solution with {self.convergenceType} behavior.")
                break

            # Check for oscillation (need at least 4 iterations to detect a 2-cycle)
            if it >= 3:
                cycle_len = u.detect_oscillation(scaled_obj_history, tol)
                if cycle_len is not None:
                    # Find the best (maximum) objective in the cycle
                    cycle_values = scaled_obj_history[-cycle_len:]
                    best_idx = np.argmax(cycle_values)
                    best_obj = cycle_values[best_idx]
                    self.convergenceType = f"oscillatory (cycle length {cycle_len})"
                    self.mylog.print(f"Oscillation detected: {cycle_len}-cycle pattern identified.")
                    self.mylog.print(f"Best objective in cycle: {u.d(best_obj, f=2)}")

                    # Select the solution corresponding to the best objective in the detected cycle.
                    cycle_solutions = sol_history[-cycle_len:]
                    cycle_objfns = obj_history[-cycle_len:]
                    xx = cycle_solutions[best_idx]
                    objfn = cycle_objfns[best_idx]
                    self.mylog.print("Accepting best solution from cycle and terminating.")
                    break

                # Stagnation check: chaotic oscillation that never forms a clean cycle.
                # When Medicare is in loop mode, skip iter 0 (M_n=0 gives inflated objective).
                start = 1 if includeMedicare else 0
                valid = scaled_obj_history[start:]
                if len(valid) > STAGNATION_WINDOW:
                    # Exit if the best objective has not improved in the last STAGNATION_WINDOW iters.
                    best_before_window = max(valid[:-STAGNATION_WINDOW])
                    recent_best = max(valid[-STAGNATION_WINDOW:])
                    if recent_best <= best_before_window:
                        best_idx = start + int(np.argmax(valid))
                        xx = sol_history[best_idx]
                        objfn = obj_history[best_idx]
                        self.convergenceType = "oscillatory (stagnation)"
                        self.mylog.print(
                            f"Stagnation detected: no improvement in {STAGNATION_WINDOW} iterations. "
                            "Accepting best solution."
                        )
                        break

            if it >= max_iterations:
                self.convergenceType = "max iteration"
                self.mylog.print("Warning: Exiting loop on maximum iterations.")
                # Use the best solution seen across all valid iterations rather than the last.
                start = 1 if includeMedicare else 0
                valid = scaled_obj_history[start:] if len(scaled_obj_history) > start else scaled_obj_history
                if valid:
                    best_idx = start + int(np.argmax(valid))
                    xx = sol_history[best_idx]
                    objfn = obj_history[best_idx]
                break

            it += 1
            old_objfns.append(-objfn)
            old_x = xx

        if solverSuccess:
            self.mylog.print(f"Self-consistent loop returned after {it+1} iterations.")
            if solverMsg:
                self.mylog.print(solverMsg)
            self.mylog.print(f"Objective: {u.d(objfn * objFac)}")
            # self.mylog.vprint('Upper bound:', u.d(-solution.mip_dual_bound))
            self._aggregateResults(xx)
            self._timestamp = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
            self.caseStatus = "solved"
        else:
            self.mylog.print("Warning: Optimization failed:", solverMsg, solverSuccess)
            self.caseStatus = "unsuccessful"

        return None

    def _run_highs(self, c, Lb, Ub, lbvec, ubvec, a_start, a_index, a_value,
                   integrality, options, warm_x=None):
        """
        Run one HiGHS MIP (or LP when integrality is all-zero) solve directly via highspy.

        Parameters mirror the arrays produced by self.A.to_csr() / self.B.arrays():
          c          — objective coefficients (nvars,)
          Lb, Ub     — variable lower/upper bounds (nvars,)
          lbvec, ubvec — constraint lower/upper bounds (ncons,)
          a_start    — CSR row-starts, length ncons
          a_index    — CSR column indices
          a_value    — CSR non-zero values
          integrality — 0=continuous, 1=integer, per variable (nvars,)
          warm_x     — optional prior solution vector for MIP warm-starting

        Returns (objfn, xx, success, msg, gap) matching the _milpSolve contract.
        """
        import highspy

        time_limit = u.get_numeric_option(options, "maxTime", TIME_LIMIT, min_value=0)
        mygap = u.get_numeric_option(options, "gap", GAP, min_value=0)
        verbose = options.get("verbose", False)

        h = highspy.Highs()
        h.setOptionValue("output_flag", bool(verbose))
        h.setOptionValue("mip_rel_gap", float(mygap))
        h.setOptionValue("time_limit", float(time_limit))
        h.setOptionValue("mip_max_nodes", 1_000_000)
        h.setOptionValue("presolve", "on")

        inf = highspy.kHighsInf
        col_lb = np.where(np.isneginf(Lb), -inf, Lb).astype(np.float64)
        col_ub = np.where(np.isposinf(Ub),  inf, Ub).astype(np.float64)
        row_lb = np.where(np.isneginf(lbvec), -inf, lbvec).astype(np.float64)
        row_ub = np.where(np.isposinf(ubvec),  inf, ubvec).astype(np.float64)

        h.passModel(
            len(c), len(lbvec), len(a_value),
            int(highspy.MatrixFormat.kRowwise),  # 2 — NOT 1 (kColwise)
            int(highspy.ObjSense.kMinimize),     # 1
            0.0,                                 # offset
            c.astype(np.float64),
            col_lb, col_ub,
            row_lb, row_ub,
            a_start.astype(np.int32),
            a_index.astype(np.int32),
            a_value.astype(np.float64),
            integrality.astype(np.int32),
        )

        if warm_x is not None:
            all_idx = np.arange(len(c), dtype=np.int32)
            h.setSolution(len(c), all_idx, warm_x.astype(np.float64))

        h.run()

        ms = h.getModelStatus()
        _, pstatus = h.getInfoValue("primal_solution_status")
        success = (
            ms in (highspy.HighsModelStatus.kOptimal, highspy.HighsModelStatus.kObjectiveBound)
            or pstatus == highspy.kSolutionStatusFeasible
        )

        if success:
            sol = h.getSolution()
            xx = np.array(sol.col_value, dtype=np.float64)
            obj_val = float(h.getObjectiveValue())
            _, gap = h.getInfoValue("mip_gap")
        else:
            xx = np.zeros(len(c))
            obj_val = None
            gap = -1.0

        return obj_val, xx, success, h.modelStatusToString(ms), float(gap)

    def _run_highs_lp_with_duals(self, A, B, c_obj, options, col_overrides=None):
        """
        Solve LP (no integrality) via HiGHS and return primal + row dual variables.
        Used by Benders decomposition for optimality cut generation.

        A, B, c_obj are abcapi objects (ConstraintMatrix, Bounds, Objective).
        col_overrides: optional dict {col_idx: (lb, ub)} to pin specific columns.

        Returns (obj, x, row_dual, success) where row_dual[i] is the dual variable
        for row i (positive = lower bound active, negative = upper bound active).
        Returns (None, zeros, zeros, False) on failure.
        """
        import highspy

        time_limit = u.get_numeric_option(options, "maxTime", TIME_LIMIT, min_value=0)
        verbose = options.get("verbose", False)

        a_start, a_index, a_value = A.to_csr()
        Lb, Ub = B.arrays()
        if col_overrides:
            for col, (lb, ub) in col_overrides.items():
                Lb[col] = lb
                Ub[col] = ub
        lbvec = np.array(A.lb)
        ubvec = np.array(A.ub)
        c = c_obj.arrays()

        h = highspy.Highs()
        h.setOptionValue("output_flag", bool(verbose))
        h.setOptionValue("time_limit", float(time_limit))

        inf = highspy.kHighsInf
        h.passModel(
            len(c), len(lbvec), len(a_value),
            int(highspy.MatrixFormat.kRowwise),
            int(highspy.ObjSense.kMinimize),
            0.0,
            c.astype(np.float64),
            np.where(np.isneginf(Lb), -inf, Lb).astype(np.float64),
            np.where(np.isposinf(Ub), inf, Ub).astype(np.float64),
            np.where(np.isneginf(lbvec), -inf, lbvec).astype(np.float64),
            np.where(np.isposinf(ubvec), inf, ubvec).astype(np.float64),
            a_start.astype(np.int32),
            a_index.astype(np.int32),
            a_value.astype(np.float64),
            np.zeros(len(c), dtype=np.int32),   # LP: all continuous
        )
        h.run()

        ms = h.getModelStatus()
        if ms == highspy.HighsModelStatus.kOptimal:
            sol = h.getSolution()
            return (
                float(h.getObjectiveValue()),
                np.array(sol.col_value, dtype=np.float64),
                np.array(sol.row_dual, dtype=np.float64),
                True,
            )
        return None, np.zeros(len(c)), np.zeros(len(lbvec)), False

    def _build_mosek_task(self, A, B, c_obj, col_overrides=None, int_vars=None, verbose=False):
        """
        Build and populate a MOSEK task from abcapi objects.
        Configures the objective, variable/constraint bounds, and constraint matrix.
        Caller is responsible for setting solver parameters before calling task.optimize().
        Returns (task, ncons, nvars).
        """
        import mosek

        bdic = {
            "fx": mosek.boundkey.fx,
            "fr": mosek.boundkey.fr,
            "lo": mosek.boundkey.lo,
            "ra": mosek.boundkey.ra,
            "up": mosek.boundkey.up,
        }

        Aind, Aval, clb, cub = A.lists()
        ckeys = A.keys()
        vlb, vub = B.arrays()
        vkeys = list(B.keys())     # copy so overrides don't mutate B
        cind, cval = c_obj.lists()
        ncons = A.ncons
        nvars = A.nvars

        if col_overrides:
            for col, (lb, ub) in col_overrides.items():
                vlb[col] = lb
                vub[col] = ub
                vkeys[col] = abc._bound_key(lb, ub)

        task = mosek.Task()
        task.set_Stream(mosek.streamtype.err, lambda t: self.mylog.vprint(t.strip()))
        if verbose:
            task.set_Stream(mosek.streamtype.msg, lambda t: self.mylog.vprint(t.strip()))
        task.appendcons(ncons)
        task.appendvars(nvars)

        for ii in range(len(cind)):
            task.putcj(cind[ii], cval[ii])
        for ii in range(nvars):
            task.putvarbound(ii, bdic[vkeys[ii]], float(vlb[ii]), float(vub[ii]))
        if int_vars:
            for ii in int_vars:
                task.putvartype(int(ii), mosek.variabletype.type_int)
        for i in range(ncons):
            task.putarow(i, Aind[i], Aval[i])
            task.putconbound(i, bdic[ckeys[i]], float(clb[i]), float(cub[i]))
        task.putobjsense(mosek.objsense.minimize)

        return task, ncons, nvars

    def _run_mosek_lp_with_duals(self, A, B, c_obj, options, col_overrides=None):
        """
        Solve LP via MOSEK using abcapi objects; return primal + row dual variables.
        Same return signature as _run_highs_lp_with_duals: (obj, x, row_dual, success).

        A, B, c_obj are abcapi objects (ConstraintMatrix, Bounds, Objective).
        col_overrides: optional dict {col_idx: (lb, ub)} to pin specific columns.
        """
        import mosek

        time_limit = u.get_numeric_option(options, "maxTime", TIME_LIMIT, min_value=0)
        task, ncons, nvars = self._build_mosek_task(A, B, c_obj, col_overrides=col_overrides)
        task.putdouparam(mosek.dparam.optimizer_max_time, float(time_limit))

        try:
            task.optimize()
        except mosek.Error:
            return None, np.zeros(nvars), np.zeros(ncons), False

        solsta = task.getsolsta(mosek.soltype.bas)
        if solsta == mosek.solsta.optimal:
            return (
                float(task.getprimalobj(mosek.soltype.bas)),
                np.array(task.getxx(mosek.soltype.bas)),
                np.array(task.gety(mosek.soltype.bas)),
                True,
            )
        return None, np.zeros(nvars), np.zeros(ncons), False

    def _run_mosek_mip(self, A, B, c_obj, options, lp_relax=False, col_overrides=None):
        """
        Solve MIP (or LP when lp_relax=True) via MOSEK using abcapi objects.
        Same return signature as _run_highs: (obj, x, success, msg, gap).

        A, B, c_obj are abcapi objects (ConstraintMatrix, Bounds, Objective).
        lp_relax: if True, treat all variables as continuous (LP solve).
        col_overrides: optional dict {col_idx: (lb, ub)} to pin specific columns.
        """
        import mosek

        time_limit = u.get_numeric_option(options, "maxTime", TIME_LIMIT, min_value=0)
        mygap = u.get_numeric_option(options, "gap", GAP, min_value=0)
        verbose = options.get("verbose", False)
        int_vars = [] if lp_relax else B.integralityList()
        task, ncons, nvars = self._build_mosek_task(A, B, c_obj, col_overrides=col_overrides,
                                                    int_vars=int_vars, verbose=verbose)
        task.putdouparam(mosek.dparam.mio_max_time, float(time_limit))
        task.putdouparam(mosek.dparam.mio_tol_rel_gap, float(mygap))

        try:
            task.optimize()
        except mosek.Error as e:
            return None, np.zeros(nvars), False, f"MOSEK: {e.msg}", -1.0

        if int_vars:
            sol = mosek.soltype.itg
            solsta = task.getsolsta(sol)
            success = solsta in (mosek.solsta.integer_optimal, mosek.solsta.prim_feas)
            gap = task.getdouinf(mosek.dinfitem.mio_obj_rel_gap) if success else -1.0
        else:
            sol = mosek.soltype.bas
            solsta = task.getsolsta(sol)
            success = (solsta == mosek.solsta.optimal)
            gap = 0.0

        if success:
            return (
                float(task.getprimalobj(sol)),
                np.array(task.getxx(sol)),
                True,
                f"MOSEK: {solsta}",
                float(gap),
            )
        return None, np.zeros(nvars), False, f"MOSEK: {solsta}", -1.0

    def _run_lp_with_duals(self, A, B, c_obj, options, col_overrides=None):
        """Dispatcher: LP solve with dual extraction for Benders (HiGHS or MOSEK)."""
        if getattr(self, "_decomp_use_mosek", False):
            return self._run_mosek_lp_with_duals(A, B, c_obj, options, col_overrides)
        return self._run_highs_lp_with_duals(A, B, c_obj, options, col_overrides)

    def _run_mip(self, A, B, c_obj, options, lp_relax=False, col_overrides=None, update_warm=True):
        """
        Dispatcher: MIP (or LP when lp_relax=True) solve for decomposition methods.
        A, B, c_obj are abcapi objects (ConstraintMatrix, Bounds, Objective).
        For HiGHS, uses and optionally updates self._highs_warm_start.
        For MOSEK, delegates to _run_mosek_mip (no warm-start management needed).
        """
        if getattr(self, "_decomp_use_mosek", False):
            return self._run_mosek_mip(A, B, c_obj, options, lp_relax=lp_relax, col_overrides=col_overrides)
        # HiGHS path: extract CSR arrays from abcapi objects.
        a_start, a_index, a_value = A.to_csr()
        Lb, Ub = B.arrays()
        if col_overrides:
            for col, (lb, ub) in col_overrides.items():
                Lb[col] = lb
                Ub[col] = ub
        lbvec = np.array(A.lb)
        ubvec = np.array(A.ub)
        integrality = np.zeros(A.nvars, dtype=np.int32) if lp_relax else B.integralityArray()
        c = c_obj.arrays()
        warm = self._highs_warm_start if update_warm else None
        result = self._run_highs(c, Lb, Ub, lbvec, ubvec, a_start, a_index, a_value, integrality, options,
                                 warm_x=warm)
        if result[2] and update_warm:
            self._highs_warm_start = result[1].copy()
        return result

    def _milpSolve(self, objective, options):
        """
        Solve using HiGHS directly via highspy, with MIP warm-start between SC iterations.
        The solution from each successful iteration is stored in self._highs_warm_start and
        passed as a hint to the next iteration, reducing branch-and-bound nodes when bracket
        assignments are stable across iterations.
        """
        self._buildConstraints(objective, options)
        a_start, a_index, a_value = self.A.to_csr()
        Lb, Ub = self.B.arrays()
        lbvec = np.array(self.A.lb)
        ubvec = np.array(self.A.ub)
        integrality = self.B.integralityArray()
        c = self.c.arrays()

        result = self._run_highs(c, Lb, Ub, lbvec, ubvec, a_start, a_index, a_value,
                                 integrality, options, warm_x=self._highs_warm_start)
        if result[2]:  # success — store for next SC iteration
            self._highs_warm_start = result[1].copy()
        return result

    def _relax_and_fix_solve(self, objective, options):
        """
        Relax-and-fix MIP heuristic (withDecomposition='sequential').

        1. LP relaxation of the full problem.
        2. Round ALL bracket-selector binaries (zm, za, zs, zl, zj) at once.
           For zm: use MAGI_n from previous SC iteration (solver-independent).
           For za: use argmax of companion haca-block.
           For other families (zs, zl, zj) round the fractional LP value directly.
        3. Single MIP solve with all bracket binaries fixed; zx (Roth exclusion) free.
        4. Fall back to monolithic MIP if LP relaxation fails or the fixed-bracket MIP fails.

        Note: this is a heuristic — result is not proven globally optimal.
        """
        self._buildConstraints(objective, options)

        bracket_names = [name for name in ("zs", "zj", "zm", "za", "zl") if name in self.vm]
        if not bracket_names:
            return self._run_mip(self.A, self.B, self.c, options)

        # LP relaxation.
        lp_result = self._run_mip(self.A, self.B, self.c, options, lp_relax=True, update_warm=False)
        if not lp_result[2]:
            self.mylog.vprint("Decomp: LP relaxation failed; falling back to monolithic.")
            return self._run_mip(self.A, self.B, self.c, options)

        lp_x = lp_result[1]
        self.mylog.vprint(f"Decomp: LP relaxation obj={-lp_result[0]:.0f}.")

        # Round all bracket-selector binaries at once.
        # For zm: use MAGI_n[n-2] (2-year Medicare lag) — solver-independent.
        # For za: use argmax of companion haca-block.
        # For other families (zs, zl, zj): round the fractional LP value directly.
        Lb_all, Ub_all = self.B.arrays()
        col_overrides = {}

        for name in bracket_names:
            blk = self.vm[name]
            if len(blk.shape) == 2:
                Nrows, Nq = blk.shape
                if name == "zm":
                    # Use MAGI from previous SC iteration for bracket selection.
                    # This is solver-independent (avoids LP degeneracy issues with MOSEK).
                    nmstart = self.N_n - Nrows
                    for nn in range(Nrows):
                        n = nmstart + nn
                        magi_src = n - 2
                        mymagi = self.MAGI_n[magi_src] if magi_src >= 0 else 0.0
                        status = (0 if self.N_i == 1
                                  or not (n < self.horizons[0] and n < self.horizons[1])
                                  else 1)
                        best_q = 0
                        for q in range(Nq - 1, -1, -1):
                            if mymagi > self.gamma_n[n] * tx.irmaaBrackets[status][q]:
                                best_q = q
                                break
                        for q in range(Nq):
                            col = blk.idx(nn, q)
                            if Lb_all[col] >= Ub_all[col] - 1e-9:
                                continue
                            v = 1.0 if q == best_q else 0.0
                            col_overrides[col] = (v, v)
                elif name == "za" and "haca" in self.vm and self.vm["haca"].shape == blk.shape:
                    h_blk = self.vm["haca"]
                    for nn in range(Nrows):
                        vals = np.array([lp_x[h_blk.idx(nn, q)] for q in range(Nq)])
                        best_q = int(np.argmax(vals))
                        for q in range(Nq):
                            col = blk.idx(nn, q)
                            if Lb_all[col] >= Ub_all[col] - 1e-9:
                                continue
                            v = 1.0 if q == best_q else 0.0
                            col_overrides[col] = (v, v)
                else:
                    for nn in range(Nrows):
                        vals = np.array([lp_x[blk.idx(nn, q)] for q in range(Nq)])
                        best_q = int(np.argmax(vals))
                        for q in range(Nq):
                            col = blk.idx(nn, q)
                            if Lb_all[col] >= Ub_all[col] - 1e-9:
                                continue
                            v = 1.0 if q == best_q else 0.0
                            col_overrides[col] = (v, v)
            else:
                for col in range(blk.start, blk.end):
                    if Lb_all[col] >= Ub_all[col] - 1e-9:
                        continue
                    v = float(round(lp_x[col]))
                    col_overrides[col] = (v, v)

        # Single MIP solve: bracket binaries fixed, zx free.
        if not getattr(self, "_decomp_use_mosek", False):
            self._highs_warm_start = lp_x   # seed from LP solution
        result = self._run_mip(self.A, self.B, self.c, options, col_overrides=col_overrides)
        if result[2]:
            return result

        self.mylog.vprint("Decomp: fixed-bracket MIP failed; falling back to monolithic.")
        return self._run_mip(self.A, self.B, self.c, options)

    def _benders_solve(self, objective, options):
        """
        Benders decomposition (withDecomposition='benders').

        Master problem: bracket-selector binaries (zm, za, zs, zl, zj).
        Subproblem: continuous variables + Roth exclusion (zx); solved as MIP for UB,
        LP for Benders cut generation.

        For each z* (master assignment):
          - SP LP (zx and continuous relaxed): generates optimality cut via LP duals.
          - SP MIP (zx free, continuous free): provides the true upper bound.

        z* initialization: zm uses MAGI_n[n-2] (2-year Medicare lag, solver-independent);
        za uses argmax(haca) from LP relaxation; zs/zl/zj round LP directly.

        The algorithm terminates when the gap closes, when the master z* stalls,
        when the SP LP is infeasible for the current z*, or when max_iter is reached.
        In all cases the best SP MIP solution found is returned. If the SP LP ever
        becomes infeasible (master assigned a bracket the SP cannot achieve), the
        algorithm stops immediately and returns the last good SP MIP result — this
        avoids cascading no-good cuts that do not converge.
        """
        self._buildConstraints(objective, options)
        nvars = self.A.nvars

        # Master variables: bracket-selector binaries only.
        # zx (Roth exclusion) stays in the SP so it can be optimized freely.
        # Exclude columns already hard-fixed (Lb == Ub), e.g. zm for years with known prevMAGI.
        Lb_all, Ub_all = self.B.arrays()
        master_cols = []
        for name in ("zs", "zj", "zm", "za", "zl"):
            if name in self.vm:
                blk = self.vm[name]
                for col in range(blk.start, blk.end):
                    if Lb_all[col] < Ub_all[col] - 1e-9:
                        master_cols.append(col)

        if not master_cols:
            return self._run_mip(self.A, self.B, self.c, options)

        n_master = len(master_cols)
        master_col_to_pos = {col: pos for pos, col in enumerate(master_cols)}
        master_col_set = set(master_cols)

        # Column-to-row transpose for Benders cut coefficient computation.
        col_rows = [[] for _ in range(nvars)]
        for i, (inds, vals) in enumerate(zip(self.A.Aind, self.A.Aval)):
            for j, v in zip(inds, vals):
                col_rows[j].append((i, float(v)))

        # Master-only rows: rows whose non-zeros lie entirely in master (binary) columns.
        # These are the AMO constraints (sum_q zm[n,q] = 1, etc.) and zl monotonicity.
        master_only_rows = [
            i for i, inds in enumerate(self.A.Aind)
            if inds and all(j in master_col_set for j in inds)
        ]

        # Build master problem: variables = [z_0, ..., z_{n_master-1}, eta].
        mp_nvars = n_master + 1
        eta_pos = n_master
        BIG_ETA = 1e12

        mp_B = abc.Bounds(mp_nvars, 0)
        for pos in range(n_master):
            mp_B.setBinary(pos)
        mp_B.setRange(eta_pos, -BIG_ETA, BIG_ETA)

        mp_c_obj = abc.Objective(mp_nvars)
        mp_c_obj.setElem(eta_pos, 1.0)   # minimize eta

        mp_A_static_rows = []
        for i in master_only_rows:
            rowDic = {master_col_to_pos[j]: v for j, v in zip(self.A.Aind[i], self.A.Aval[i])}
            mp_A_static_rows.append((rowDic, self.A.lb[i], self.A.ub[i]))

        def _build_master_A(cuts):
            mp_A = abc.ConstraintMatrix(mp_nvars)
            for rowDic, lb, ub in mp_A_static_rows:
                mp_A.addNewRow(rowDic, lb, ub)
            for alpha, beta in cuts:
                cut_dic = {eta_pos: 1.0}
                for pos in range(n_master):
                    b = float(beta[pos])
                    if b != 0.0:
                        cut_dic[pos] = -b
                mp_A.addNewRow(cut_dic, float(alpha), np.inf)
            return mp_A

        # Benders parameters.
        max_iter = int(options.get("bendersMaxIter", 50))
        mygap = float(options.get("gap", GAP))
        UB = np.inf
        LB = -np.inf
        best_x = None
        benders_cuts = []

        if objective == "maxSpending":
            display_scale = 1.0 / self.xi_n[0]
        else:
            display_scale = 1.0 / self.gamma_n[-1]

        # Initial LP relaxation: LB and starting z*.
        lp_obj, lp_x, _, lp_ok = self._run_lp_with_duals(self.A, self.B, self.c, options)
        if not lp_ok:
            self.mylog.vprint("Benders: LP relaxation failed; falling back to monolithic.")
            return self._run_mip(self.A, self.B, self.c, options)

        LB = lp_obj
        self.mylog.vprint(f"Benders: LP relaxation obj = {-LB * display_scale:.0f}.")

        # Initialize z* from LP solution.
        # For zm: use MAGI_n[n-2] (2-year Medicare lag) — solver-independent.
        # For za: use argmax of companion haca-block from LP relaxation.
        # For other families (zs, zl, zj): round the fractional LP value directly.
        zm_init_pos = set()
        za_init_pos = set()
        z_star = np.zeros(n_master, dtype=np.float64)

        if "zm" in self.vm:
            z_blk = self.vm["zm"]
            Nrows, Nq = z_blk.shape
            nmstart = self.N_n - Nrows
            for nn in range(Nrows):
                n = nmstart + nn
                magi_src = n - 2
                mymagi = self.MAGI_n[magi_src] if magi_src >= 0 else 0.0
                status = (0 if self.N_i == 1
                          or not (n < self.horizons[0] and n < self.horizons[1])
                          else 1)
                best_q = 0
                for q in range(Nq - 1, -1, -1):
                    if mymagi > self.gamma_n[n] * tx.irmaaBrackets[status][q]:
                        best_q = q
                        break
                for q in range(Nq):
                    col = z_blk.idx(nn, q)
                    if col in master_col_to_pos:
                        pos = master_col_to_pos[col]
                        z_star[pos] = 1.0 if q == best_q else 0.0
                        zm_init_pos.add(pos)

        if "za" in self.vm and "haca" in self.vm and self.vm["za"].shape == self.vm["haca"].shape:
            z_blk, h_blk = self.vm["za"], self.vm["haca"]
            Nrows, Nq = z_blk.shape
            for nn in range(Nrows):
                h_vals = np.array([lp_x[h_blk.idx(nn, q)] for q in range(Nq)])
                best_q = int(np.argmax(h_vals))
                for q in range(Nq):
                    col = z_blk.idx(nn, q)
                    if col in master_col_to_pos:
                        pos = master_col_to_pos[col]
                        z_star[pos] = 1.0 if q == best_q else 0.0
                        za_init_pos.add(pos)

        handled_pos = zm_init_pos | za_init_pos
        for pos, col in enumerate(master_cols):
            if pos not in handled_pos:
                z_star[pos] = float(round(lp_x[col]))

        # Benders main loop.
        for biter in range(max_iter):
            sp_overrides = {col: (float(z_star[pos]), float(z_star[pos]))
                            for pos, col in enumerate(master_cols)}

            # SP LP: for Benders cut generation.
            sp_lp_obj, _, pi, sp_lp_ok = self._run_lp_with_duals(
                self.A, self.B, self.c, options, col_overrides=sp_overrides)

            if not sp_lp_ok:
                # Master assigned a bracket combination the SP cannot satisfy.
                # Stop and return the best MIP solution found so far.
                self.mylog.vprint(f"Benders iter {biter + 1}: SP LP infeasible; terminating.")
                break

            # Benders optimality cut: eta >= alpha + beta^T z (tight at current z*).
            beta = np.array([
                -sum(pi[r] * v for r, v in col_rows[col])
                for col in master_cols
            ])
            alpha = sp_lp_obj - float(beta @ z_star)
            benders_cuts.append((alpha, beta))

            # SP MIP: fix bracket binaries, optimize zx and continuous → true UB.
            sp_mip_res = self._run_mip(self.A, self.B, self.c, options, col_overrides=sp_overrides)
            if sp_mip_res[2] and sp_mip_res[0] is not None and sp_mip_res[0] < UB:
                UB = sp_mip_res[0]
                best_x = sp_mip_res[1].copy()

            if UB < np.inf:
                gap_val = (UB - LB) / max(abs(UB), 1.0)
                self.mylog.vprint(
                    f"Benders iter {biter + 1}: "
                    f"LB={-UB * display_scale:.0f}, UB={-LB * display_scale:.0f}, "
                    f"gap={gap_val:.4f}.")
                if gap_val <= mygap:
                    self.mylog.vprint(f"Benders: converged after {biter + 1} iterations.")
                    break

            # Solve master MIP → new LB and z*.
            mp_A = _build_master_A(benders_cuts)
            mp_res = self._run_mip(mp_A, mp_B, mp_c_obj, options, update_warm=False)
            if not mp_res[2] or mp_res[0] is None:
                self.mylog.vprint(f"Benders iter {biter + 1}: master MIP failed; terminating.")
                break
            LB = max(LB, mp_res[0])

            if UB < np.inf:
                gap_val = (UB - LB) / max(abs(UB), 1.0)
                if gap_val <= mygap:
                    self.mylog.vprint(f"Benders: converged after {biter + 1} iterations.")
                    break

            z_star_new = np.round(mp_res[1][:n_master]).astype(np.float64)
            if np.array_equal(z_star_new, z_star):
                self.mylog.vprint(f"Benders iter {biter + 1}: z* unchanged; terminating.")
                break
            z_star = z_star_new

        if best_x is not None:
            final_gap = (UB - LB) / max(abs(UB), 1.0) if UB < np.inf and LB > -np.inf else -1.0
            return UB, best_x, True, f"Benders ({len(benders_cuts)} cuts)", float(final_gap)

        self.mylog.vprint("Benders: no feasible solution found; falling back to monolithic.")
        return self._run_mip(self.A, self.B, self.c, options)

    def _mosekSolve(self, objective, options):
        """
        Solve problem using MOSEK solver.
        """
        import mosek

        self._buildConstraints(objective, options)
        time_limit = u.get_numeric_option(options, "maxTime", TIME_LIMIT, min_value=0)
        mygap = u.get_numeric_option(options, "gap", GAP, min_value=0)
        verbose = options.get("verbose", False)
        int_vars = self.B.integralityList()

        task, ncons, nvars = self._build_mosek_task(self.A, self.B, self.c,
                                                    int_vars=int_vars, verbose=verbose)
        task.putdouparam(mosek.dparam.mio_max_time, time_limit)        # Default -1
        # task.putdouparam(mosek.dparam.mio_rel_gap_const, 1e-6)       # Default 1e-10
        task.putdouparam(mosek.dparam.mio_tol_rel_gap, mygap)          # Default 1e-4
        # task.putdouparam(mosek.dparam.mio_tol_abs_relax_int, 2e-5)   # Default 1e-5
        # task.putdouparam(mosek.iparam.mio_heuristic_level, 3)        # Default -1

        try:
            trmcode = task.optimize()
        except mosek.Error as e:
            return 0.0, np.zeros(nvars), False, f"MOSEK: {e.msg}", -1

        # Problem MUST contain binary variables to make these calls.
        solsta = task.getsolsta(mosek.soltype.itg)
        solverSuccess = (solsta == mosek.solsta.integer_optimal)
        rel_gap = task.getdouinf(mosek.dinfitem.mio_obj_rel_gap) if solverSuccess else -1

        if solsta == mosek.solsta.integer_optimal:
            solverMsg = "MOSEK: Optimal integer solution found"
        elif solsta == mosek.solsta.prim_feas:
            solverMsg = "MOSEK: Feasible integer solution (not proven optimal)"
        elif solsta == mosek.solsta.unknown:
            symname, desc = mosek.Env.getcodedesc(trmcode)
            solverMsg = f"MOSEK: {symname} - {desc}"
        else:
            solverMsg = f"MOSEK: Solution status {solsta}"

        xx = np.array(task.getxx(mosek.soltype.itg))
        solution = task.getprimalobj(mosek.soltype.itg)
        task.solutionsummary(mosek.streamtype.msg)
        # task.writedata(self._name+'.ptf')

        return solution, xx, solverSuccess, solverMsg, rel_gap

    def _update_Psi_n(self):
        """
        Recompute SS taxability fractions using the IRS provisional income (PI) formula.

        Delegates to tax2026.compute_social_security_taxability() for the pure tax
        computation. A 30% damping blend is applied here to damp potential oscillation
        near threshold boundaries and ensure SC-loop convergence.

        If setSocialSecurity() was called with a tax_fraction override, this method
        returns immediately without modifying Psi_n (the fixed value set on reset is used).
        """
        # Honor explicit override: skip dynamic computation.
        if getattr(self, 'ssecTaxFraction', None) is not None:
            return

        ss_n = np.sum(self.zetaBar_in, axis=0)
        new_Psi_n = tx.compute_social_security_taxability(
            self.N_i, self.MAGI_n, ss_n, ssec_tax_fraction=None, n_d=self.n_d
        )

        # 30% damping blend: damp oscillation near threshold boundaries.
        blended = _PSI_DAMP * new_Psi_n + (1.0 - _PSI_DAMP) * self.Psi_n
        if np.max(np.abs(blended - self.Psi_n)) > 1e-3:
            self.Psi_n = blended

    def _computeNLstuff(self, x, includeMedicare, fixedPsi=None):
        """
        Compute MAGI, Medicare costs, ACA costs, long-term capital gain tax rate, and
        net investment income tax (NIIT).
        """
        if x is None:
            # Reset all nonlinear quantities to their starting values for a fresh solve.
            self.Psi_n = np.ones(self.N_n) * (fixedPsi if fixedPsi is not None else 0.85)
            self.MAGI_n = np.zeros(self.N_n)
            self.G_n = np.zeros(self.N_n)
            self.J_n = np.zeros(self.N_n)
            self.M_n = np.zeros(self.N_n)
            self.ACA_n = np.zeros(self.N_n)
            return

        self._aggregateResults(x, short=True)
        # Psi_n is derived directly from the tss_n LP variable in _aggregateResults
        # when withSSTaxability=="optimize"; skip the SC-loop update in that case.
        # Also skip when fixedPsi is set (numeric withSSTaxability).
        if "tss" not in self.vm and fixedPsi is None:
            self._update_Psi_n()

        # SS claiming-age LP: update zetaBar_in from optimal zssa solution each SC iteration.
        # _ssa_spousal_offset is updated here so next iteration's cash-flow constraint is correct.
        if getattr(self, "_ssa_lp", False) and "zssa" in self.vm:
            zssa_vals = self.vm["zssa"].extract(x)
            new_ages = np.array(self.ssecAges, dtype=float)
            for i in range(self.N_i):
                k_opt = int(np.argmax(zssa_vals[i, :]))
                new_ages[i] = float(self._ssa_ages_k[k_opt])
            new_zeta_in, _ = socsec.compute_social_security_benefits(
                self.ssecAmounts, new_ages, self.yobs, self.mobs, self.tobs, self.horizons,
                self.N_i, self.N_n,
                trim_pct=getattr(self, "ssecTrimPct", 0) or 0,
                trim_year=getattr(self, "ssecTrimYear", None),
                thisyear=date.today().year,
            )
            new_zetaBar_in = new_zeta_in * self.gamma_n[:-1]
            for i in range(self.N_i):
                k_opt = int(np.argmax(zssa_vals[i, :]))
                self._ssa_spousal_offset[i, :] = new_zetaBar_in[i, :] - self._ssa_B_own[i, k_opt, :]
            self.zetaBar_in = new_zetaBar_in

        if not getattr(self, "_niit_lp", False):
            self.J_n = tx.computeNIIT(self.N_i, self.MAGI_n, self.I_n, self.Q_n, self.n_d, self.N_n)
        # LTCG tax is in the LP via q bracket variables; U_n is set by _aggregateResults.
        # Compute Medicare through self-consistent loop.
        if includeMedicare:
            include_part_d = getattr(self, "_include_medicare_part_d", True)
            part_d_base = getattr(self, "_medicare_part_d_base_annual_per_person", 0.0)
            self.M_n = tx.mediCosts(
                self.yobs, self.horizons, self.MAGI_n, self.prevMAGI, self.gamma_n[:-1], self.N_n,
                include_part_d=include_part_d,
                part_d_base_annual_per_person=part_d_base,
            )
        # Compute ACA costs through self-consistent loop (uses current-year MAGI, no 2-year lag).
        # In optimize mode (withACA="optimize"), ACA_n stays zero; maca_n carries the cost.
        if self.slcsp_annual > 0 and not self._aca_lp:
            self.ACA_n = tx.acaCosts(self.yobs, self.horizons, self.MAGI_n, self.gamma_n[:-1],
                                     self.slcsp_annual, self.N_n)

        return None

    def _aggregateResults(self, x, short=False):
        """
        Utility function to aggregate results from solver.
        Process all results from solution vector.
        """
        # Define shortcuts.
        Ni = self.N_i
        Nj = self.N_j
        Nk = self.N_k
        Nn = self.N_n
        n_d = self.n_d
        vm = self.vm

        x = u.roundCents(x)

        # Allocate, slice in, and reshape variables.
        self.b_ijn = vm["b"].extract(x)
        self.b_ijkn = np.zeros((Ni, Nj, Nk, Nn + 1))
        for k in range(Nk):
            self.b_ijkn[:, :, k, :] = self.b_ijn[:, :, :] * self.alpha_ijkn[:, :, k, :]

        self.d_in = vm["d"].extract(x)
        self.e_n = vm["e"].extract(x)
        self.f_tn = vm["f"].extract(x)
        self.g_n = vm["g"].extract(x)
        if "h" in vm:
            self.h_qn = vm["h"].extract(x)
        if "haca" in vm:
            self.haca_qn = vm["haca"].extract(x)
        self.m_n = vm["m"].extract(x)
        self.maca_n = vm["maca"].extract(x) if "maca" in vm else np.zeros(self.N_n)
        self.s_n = vm["s"].extract(x)
        self.w_ijn = vm["w"].extract(x)
        self.x_in = vm["x"].extract(x)

        # Extract SS taxability LP variables and update Psi_n from the LP solution.
        if "tss" in vm:
            self.tss_n = vm["tss"].extract(x)
            ss_n = np.sum(self.zetaBar_in, axis=0)
            mask = ss_n > 0
            self.Psi_n = np.zeros(Nn)
            self.Psi_n[mask] = np.minimum(self.tss_n[mask] / ss_n[mask], 0.85)

        # Extract optimal SS claiming ages from zssa binaries (full aggregation only).
        if "zssa" in vm and not short:
            zssa_vals = vm["zssa"].extract(x)
            for i in range(Ni):
                k_opt = int(np.argmax(zssa_vals[i, :]))
                self.ssecAges[i] = float(self._ssa_ages_k[k_opt])

        self.G_n = np.sum(self.f_tn, axis=0)

        tau_0 = np.array(self.tau_kn[0, :])
        # Last year's rates.
        tau_0prev = np.roll(tau_0, 1)
        # Capital gains = price appreciation only (total return - dividend rate)
        # to avoid double taxation of dividends. No tax harvesting here.
        capital_gains_rate = np.maximum(0, tau_0prev - self.mu)
        self.Q_n = np.sum(
            (
                self.mu
                * (self.b_ijn[:, 0, :Nn] - self.w_ijn[:, 0, :] + self.d_in[:, :] + 0.5 * self.kappa_ijn[:, 0, :Nn])
                + capital_gains_rate * self.w_ijn[:, 0, :]
            )
            * self.alpha_ijkn[:, 0, 0, :Nn],
            axis=0,
        )
        # Add fixed assets capital gains.
        self.Q_n += self.fixed_assets_capital_gains_n
        # Extract LTCG bracket variables and compute derived quantities.
        self.q_pn = vm["q"].extract(x)      # shape (N_p=3, N_n); p=0/1/2 → 0%/15%/20% brackets
        self.U_n = 0.15 * self.q_pn[1, :] + 0.20 * self.q_pn[2, :]
        # When Q_n < T15 the LP may set q[0,n] > Q_n (free 0% bucket), clip for clean reporting.
        total_ltcg = np.maximum(self.Q_n, 0)
        excess = np.maximum(0, self.q_pn[0, :] - total_ltcg)
        self.q_pn[0, :] = np.maximum(0, self.q_pn[0, :] - excess)

        # Extract NIIT LP variable when in optimize mode.
        if "Jn" in vm:
            self.J_n = vm["Jn"].extract(x)

        # Also add back non-taxable part of SS.
        if "magi" in vm:
            self.MAGI_n = vm["magi"].extract(x)
        else:
            self.MAGI_n = (self.G_n + self.e_n + self.Q_n
                           + np.sum((1 - self.Psi_n) * self.zetaBar_in, axis=0))

        # Only positive returns count as interest/dividend income (matches _add_taxable_income).
        I_in = ((self.b_ijn[:, 0, :-1] + self.d_in - self.w_ijn[:, 0, :])
                * np.sum(self.alpha_ijkn[:, 0, 1:, :Nn] * np.maximum(0, self.tau_kn[1:, :]), axis=1))
        # Sum over individuals to share losses across spouses; clamp to non-negative.
        # Also add net investment income from rent/trust (netinv_in) for NIIT purposes.
        self.I_n = np.maximum(0, np.sum(I_in, axis=0)) + np.sum(self.netinv_in, axis=0)

        # Stop after building minimum required for self-consistent loop.
        if short:
            return

        self.T_tn = self.f_tn * self.theta_tn
        self.T_n = np.sum(self.T_tn, axis=0)
        self.P_n = np.zeros(Nn)
        # Add early withdrawal penalty if any.
        for i in range(Ni):
            self.P_n[0:self.n595[i]] += 0.1 * self.w_ijn[i, 1, 0:self.n595[i]]

        self.T_n += self.P_n
        # Compute partial distribution at the passing of first spouse.
        if Ni == 2 and n_d < Nn:
            nx = n_d - 1
            i_d = self.i_d
            part_j = np.zeros(Nj)
            for j in range(Nj):
                ksumj = np.sum(self.alpha_ijkn[i_d, j, :, nx] * self.tau_kn[:, nx], axis=0)
                Tauh = 1 + 0.5 * ksumj
                Tau1 = 1 + ksumj
                part_j[j] = Tauh * self.kappa_ijn[i_d, j, nx] + Tau1 * (
                    self.b_ijn[i_d, j, nx]
                    - self.w_ijn[i_d, j, nx]
                    + self.d_in[i_d, nx] * u.krond(j, 0)
                    + self.x_in[i_d, nx] * (u.krond(j, 2) - u.krond(j, 1))
                )

            self.partialEstate_j = part_j
            partialBequest_j = part_j * (1 - self.phi_j)
            partialBequest_j[1] *= 1 - self.nu   # tax-deferred: heirs pay ordinary income tax
            partialBequest_j[3] *= 1 - self.nu   # HSA: non-spouse heirs include full balance in ordinary income
            self.partialBequest = np.sum(partialBequest_j) / self.gamma_n[n_d]
        else:
            self.partialBequest = 0

        self.rmd_in = self.rho_in * self.b_ijn[:, 1, :-1]
        self.dist_in = self.w_ijn[:, 1, :] - self.rmd_in
        self.dist_in[self.dist_in < 0] = 0

        # Make derivative variables.
        # Putting it all together in a dictionary.
        """
        sourcetypes = [
            'wages',
            'ssec',
            'pension',
            '+dist',
            'RMD',
            'RothX',
            'wdrwl taxable',
            'wdrwl tax-free',
        ]
        """
        sources = {}
        sources["wages"] = self.omega_in
        sources["other inc"] = self.other_inc_in
        sources["net inv"] = self.netinv_in
        sources["ssec"] = self.zetaBar_in
        sources["pension"] = self.piBar_in
        sources["txbl acc wdrwl"] = self.w_ijn[:, 0, :]
        sources["RMD"] = self.rmd_in
        sources["+dist"] = self.dist_in
        sources["RothX"] = self.x_in
        sources["tax-free wdrwl"] = self.w_ijn[:, 2, :]
        sources["HSA wdrwl"] = self.w_ijn[:, 3, :]
        sources["BTI"] = self.Lambda_in
        # Debts and fixed assets (debts are negative as expenses)
        # Show as household totals, not split between individuals
        # Reshape to (1, N_n) to indicate household-level source
        sources["FA ord inc"] = self.fixed_assets_ordinary_income_n.reshape(1, -1)
        sources["FA cap gains"] = self.fixed_assets_capital_gains_n.reshape(1, -1)
        sources["FA tax-free"] = self.fixed_assets_tax_free_n.reshape(1, -1)
        sources["debt pmts"] = -self.debt_payments_n.reshape(1, -1)

        savings = {}
        savings["taxable"] = self.b_ijn[:, 0, :]
        savings["tax-deferred"] = self.b_ijn[:, 1, :]
        savings["tax-free"] = self.b_ijn[:, 2, :]
        savings["hsa"] = self.b_ijn[:, 3, :]

        self.sources_in = sources
        self.savings_in = savings

        estate_j = np.sum(self.b_ijn[:, :, self.N_n], axis=0)
        estate_j[1] *= 1 - self.nu   # tax-deferred: heirs pay ordinary income tax
        estate_j[3] *= 1 - self.nu   # HSA: non-spouse heirs include full balance in ordinary income
        # Subtract remaining debt balance from estate
        total_estate = np.sum(estate_j) - self.remaining_debt_balance
        self.bequest = max(0.0, total_estate) / self.gamma_n[-1]

        self.basis = self.g_n[0] / self.xi_n[0]

        return None

    @property
    def aca_costs_n(self):
        """ACA net premium costs per year: LP result (optimize mode) or SC-loop result (loop mode)."""
        return self.maca_n if self._aca_lp else self.ACA_n

    @_checkCaseStatus
    def estate(self):
        """
        Reports final account balances.
        """
        _estate = np.sum(self.b_ijn[:, :, :, self.N_n], axis=(0, 2))
        _estate[1] *= 1 - self.nu
        self.mylog.vprint(f"Estate value of {u.d(sum(_estate))} at the end of year {self.year_n[-1]}.")

        return None

    @_checkCaseStatus
    def summary(self, N=None):
        """
        Print summary in logs.
        """
        self.mylog.print("SUMMARY ================================================================")
        dic = self.summaryDic(N)
        for key, value in dic.items():
            self.mylog.print(f"{key}: {value}")
        self.mylog.print("------------------------------------------------------------------------")

        return None

    def summaryList(self, N=None):
        """Return summary as a list."""
        return export.build_summary_list(self, N)

    def summaryDf(self, N=None):
        """Return summary as a dataframe."""
        return pd.DataFrame(export.build_summary_dic(self, N), index=[self._name])

    def summaryString(self, N=None):
        """Return summary as a string."""
        return export.build_summary_string(self, N)

    def summaryDic(self, N=None):
        """Return dictionary containing summary of values."""
        return export.build_summary_dic(self, N)

    def showRatesCorrelations(self, tag="", shareRange=False, figure=False):
        """
        Plot correlations between various rates.

        A tag string can be set to add information to the title of the plot.
        """
        if self.rateMethod in [None, "user", "historical average", "conservative", "trailing-30", "optimistic"]:
            self.mylog.print(f"Warning: Cannot plot correlations for {self.rateMethod} rate method.")
            return None

        # Check if rates are constant (all values are the same for each rate type)
        # This can happen with fixed rates
        if self.tau_kn is not None:
            # Check if all rates are constant (no variation)
            rates_are_constant = True
            for k in range(self.N_k):
                # Check if all values in this rate series are (approximately) the same
                rate_std = np.std(self.tau_kn[k])
                # Use a small threshold to account for floating point precision
                if rate_std > 1e-10:  # If standard deviation is non-zero, rates vary
                    rates_are_constant = False
                    break

            if rates_are_constant:
                self.mylog.print("Warning: Cannot plot correlations for constant rates (no variation in rate values).")
                return None

        # For stochastic models, build a large representative sample so that
        # the histograms and scatter plots reflect the method's distribution
        # rather than the properties of the single N_n-year realization.
        # Deterministic models (historical, constant) already have the correct
        # data in tau_kn.
        _N_REPR = 2000
        rateModel = getattr(self, "rateModel", None)
        if rateModel is not None and not rateModel.deterministic:
            repr_series = rateModel.representative_sample(_N_REPR)  # (M, 4) decimal
            display_tau_kn = repr_series.transpose()                # (4, M)
        else:
            display_tau_kn = self.tau_kn

        fig = self._plotter.plot_rates_correlations(self._name, display_tau_kn, self.rateMethod,
                                                    self.rateFrm, self.rateTo, tag, shareRange)

        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    def showRatesDistributions(self, frm=rates.FROM, to=rates.TO, figure=False):
        """
        Plot histograms of the rates distributions.
        """
        fig = self._plotter.plot_rates_distributions(frm, to, rates.SP500, rates.BondsBaa,
                                                     rates.TNotes, rates.Inflation, rates.FROM)
        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    def showRates(self, tag="", figure=False):
        """
        Plot rate values used over the time horizon.

        A tag string can be set to add information to the title of the plot.
        """
        if self.rateMethod is None:
            self.mylog.print("Warning: Rate method must be selected before plotting.")
            return None

        fig = self._plotter.plot_rates(self._name, self.tau_kn, self.year_n,
                                       self.N_k, self.rateMethod, self.rateFrm, self.rateTo, tag)

        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    def showProfile(self, tag="", figure=False):
        """
        Plot spending profile over time.

        A tag string can be set to add information to the title of the plot.
        """
        if self.xi_n is None:
            self.mylog.print("Warning: Profile must be selected before plotting.")
            return None
        title = self._name + "\nSpending Profile"
        if tag:
            title += " - " + tag
        fig = self._plotter.plot_profile(self.year_n, self.xi_n, title, self.inames)

        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showNetSpending(self, tag="", value=None, figure=False):
        """
        Plot net available spending and target over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        value = self._checkValueType(value)
        title = self._name + "\nNet Available Spending"
        if tag:
            title += " - " + tag
        fig = self._plotter.plot_net_spending(self.year_n, self.g_n, self.xi_n, self.xiBar_n,
                                              self.gamma_n, value, title, self.inames)
        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showSavingsRetentionRate(self, tag="", figure=False, log_scale=False):
        """
        Plot savings retention rate (%) over time: fraction of balance NOT drawn each year.

        Retention = 1 − net_draw/b, where net_draw = w_ijn (spending withdrawals)
        minus d_in (taxable deposits) and kappa_ijn[:,1:,:] (contributions to
        tax-deferred/Roth/HSA). Roth conversions excluded — internal transfers.
        Values >100% indicate accumulation years; ~0% in the final year when
        bequest=0 (avoids the scale-distorting 100% withdrawal bar of the inverse).
        When log_scale=True, y = log(retention/100); reference line at 0; bars sum to
        log(b_final/b_initial); sustainability line simplifies to ≈ i_n − r_n.
        """
        net_w = self.w_ijn.copy()
        net_w[:, 0, :] -= self.d_in                            # subtract deposits to taxable
        net_w[:, 1:, :] -= self.kappa_ijn[:, 1:, :self.N_n]   # subtract contributions to deferred/Roth/HSA
        net_draw_n = np.sum(net_w, axis=(0, 1))
        b_n = np.sum(self.b_ijn[:, :, :-1], axis=(0, 1))
        with np.errstate(invalid="ignore", divide="ignore"):
            rate = np.where(b_n > 0, (1 - net_draw_n / b_n) * 100, 0.0)
        # Real break-even: retention rate that preserves real portfolio value
        num_n = np.einsum('ijn,ijkn,kn->n',
                          self.b_ijn[:, :, :self.N_n],
                          self.alpha_ijkn[:, :, :, :self.N_n],
                          self.tau_kn[:, :self.N_n])
        with np.errstate(invalid="ignore", divide="ignore"):
            r_n = np.where(b_n > 0, num_n / b_n, 0.0)
        inflation_n = self.gamma_n[1:self.N_n + 1] / self.gamma_n[:self.N_n]
        sustainability_n = inflation_n / (1.0 + r_n) * 100.0
        if log_scale:
            with np.errstate(invalid="ignore", divide="ignore"):
                rate = np.where(rate > 0, np.log(rate / 100.0), np.nan)
                sustainability_n = np.log(sustainability_n / 100.0)
        title = self._name + "\nSavings Retention Rate"
        if log_scale:
            title += " (log scale)"
        if self.bequest > 0:
            title += f"\n(Note: bequest of {u.d(self.bequest, f=0)} included in balance)"
        if tag:
            title += " - " + tag
        fig = self._plotter.plot_savings_retention_rate(self.year_n, rate, title,
                                                        sustainability_n=sustainability_n,
                                                        log_scale=log_scale)
        if figure:
            return fig
        self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showAssetComposition(self, tag="", value=None, figure=False):
        """
        Plot the composition of each savings account in thousands of dollars
        during the simulation time. This function will generate four
        graphs, one for taxable accounts, one for tax-deferred accounts,
        one for tax-free accounts, and one for HSA accounts.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        value = self._checkValueType(value)
        figures = self._plotter.plot_asset_composition(self.year_n, self.inames, self.b_ijkn,
                                                       self.gamma_n, value, self._name, tag)
        if all(f is None for f in figures):
            return None
        if figure:
            return figures

        for fig in figures:
            if fig is not None:
                self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showGrossIncome(self, tag="", value=None, figure=False):
        """
        Plot income tax and taxable income over time horizon.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        value = self._checkValueType(value)
        tax_brackets = tx.taxBrackets(self.N_i, self.n_d, self.N_n, self.yOBBBA)
        title = self._name + "\nTaxable Ordinary Income vs. Tax Brackets"
        if tag:
            title += " - " + tag
        fig = self._plotter.plot_gross_income(
            self.year_n, self.G_n, self.gamma_n, value, title, tax_brackets
        )
        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    def showAllocations(self, tag="", figure=False):
        """
        Plot desired allocation of savings accounts in percentage
        over simulation time and interpolated by the selected method
        through the interpolateAR() method.

        A tag string can be set to add information to the title of the plot.
        """
        title = self._name + "\nAsset Allocation"
        if tag:
            title += " - " + tag
        figures = self._plotter.plot_allocations(self.year_n, self.inames, self.alpha_ijkn,
                                                 self.ARCoord, title)
        if figure:
            return figures

        for fig in figures:
            self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showAccounts(self, tag="", value=None, figure=False):
        """
        Plot values of savings accounts over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        value = self._checkValueType(value)
        title = self._name + "\nSavings Balance"
        if tag:
            title += " - " + tag
        fig = self._plotter.plot_accounts(self.year_n, self.savings_in, self.gamma_n,
                                          value, title, self.inames)
        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showHSA(self, tag="", value=None, figure=False):
        """
        Plot HSA activity (balance, contributions, withdrawals) over time.

        Returns None if all HSA balances, contributions, and withdrawals are zero.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        hsa_bal = self.b_ijn[:, 3, :]
        hsa_ctrb = self.kappa_ijn[:, 3, :self.N_n]
        hsa_wdrwl = self.w_ijn[:, 3, :]
        if (np.abs(hsa_bal).sum() + np.abs(hsa_ctrb).sum() + np.abs(hsa_wdrwl).sum()) < 1.0:
            return None
        value = self._checkValueType(value)
        title = self._name + "\nHSA Activity"
        if tag:
            title += " - " + tag
        hsa_data = {
            "balance": hsa_bal,
            "contributions": hsa_ctrb,
            "withdrawals": hsa_wdrwl,
        }
        fig = self._plotter.plot_hsa(self.year_n, hsa_data, self.gamma_n, value, title, self.inames)
        if figure:
            return fig
        self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showSources(self, tag="", value=None, figure=False):
        """
        Plot income over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        value = self._checkValueType(value)
        title = self._name + "\nRaw Income Sources"
        if tag:
            title += " - " + tag
        fig = self._plotter.plot_sources(self.year_n, self.sources_in, self.gamma_n,
                                         value, title, self.inames)
        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    @_checkCaseStatus
    def showTaxes(self, tag="", value=None, figure=False):
        """
        Plot income tax paid over time.

        A tag string can be set to add information to the title of the plot.

        The value parameter can be set to *nominal* or *today*, overriding
        the default behavior of setDefaultPlots().
        """
        value = self._checkValueType(value)
        title = self._name + "\nFederal Income Tax"
        if tag:
            title += " - " + tag
        # All taxes: ordinary income, dividends, and NIIT.
        allTaxes = self.T_n + self.U_n + self.J_n
        aca_n = self.aca_costs_n if self.slcsp_annual > 0 else None
        fig = self._plotter.plot_taxes(
            self.year_n, allTaxes, self.m_n + self.M_n, self.gamma_n,
            value, title, self.inames, A_n=aca_n
        )
        if figure:
            return fig

        self._plotter.jupyter_renderer(fig)
        return None

    def saveWorkbook(self, overwrite=False, *, basename=None, saveToFile=True, with_config="no"):
        """
        Save instance in an Excel spreadsheet.
        See export.plan_to_excel for sheet structure and with_config options.
        """
        return export.plan_to_excel(
            self, overwrite=overwrite, basename=basename,
            saveToFile=saveToFile, with_config=with_config
        )

    def saveWorkbookCSV(self, basename):
        """
        Save plan data in CSV format. See saveWorkbook() for related structure.
        """
        return export.plan_to_csv(self, basename, self.mylog)

    def saveConfig(self, basename=None):
        """
        Save parameters in a configuration file.
        """
        if basename is None:
            basename = self._name if self._name.lower().startswith("case_") else "case_" + self._name

        config.saveConfig(self, basename, self.mylog)

        return None
