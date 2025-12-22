import click
from loguru import logger
import owlplanner as owl
from pathlib import Path
from openpyxl import load_workbook


def insert_text_file_as_first_sheet(
    xlsx_path: Path,
    text_path: Path,
    sheet_name: str = "Config (.toml)",
):
    """
    Insert the contents of a text file as the first worksheet in an Excel file.
    Each line goes into its own row (column A).
    """
    wb = load_workbook(xlsx_path)

    # Create new sheet at position 0
    ws = wb.create_sheet(title=sheet_name, index=0)

    with text_path.open("r", encoding="utf-8") as f:
        for row_idx, line in enumerate(f, start=1):
            ws.cell(row=row_idx, column=1, value=line.rstrip())

    wb.save(xlsx_path)


def validate_toml(ctx, param, value: Path):
    if value is None:
        return None

    # If no suffix, append .toml
    if value.suffix == "":
        value = value.with_suffix(".toml")

    # Enforce .toml extension
    if value.suffix.lower() != ".toml":
        raise click.BadParameter("File must have a .toml extension")

    # Check existence AFTER normalization
    if not value.exists():
        raise click.BadParameter(f"File '{value}' does not exist")

    if not value.is_file():
        raise click.BadParameter(f"'{value}' is not a file")

    return value


@click.command(name="run")
@click.argument(
    "filename",
    type=click.Path(exists=False, dir_okay=False, path_type=Path),
    callback=validate_toml,
)
def cmd_run(filename: Path):
    """Run the solver for an input OWL plan file.

    FILENAME is the OWL plan file to run. If no extension is provided,
    .toml will be appended. The file must exist.

    An output Excel file with results will be created in the current directory.
    The output filename is derived from the input filename by appending
    '_results.xlsx' to the stem of the input filename.

    The input TOML file will be inserted as the first worksheet in the output Excel file
    for reference.

    """
    logger.debug(f"Executing the run command with file: {filename}")

    plan = owl.readConfig(str(filename), logstreams="loguru", readContributions=False)
    plan.solve(plan.objective, plan.solverOptions)
    click.echo(f"Case status: {plan.caseStatus}")
    if plan.caseStatus == "solved":
        output_filename = filename.with_name(filename.stem + "_results.xlsx")
        plan.saveWorkbook(basename=output_filename, overwrite=True)
        # Insert TOML file as first worksheet
        insert_text_file_as_first_sheet(
            xlsx_path=output_filename,
            text_path=filename,
        )
        click.echo(f"Results saved to: {output_filename}")
