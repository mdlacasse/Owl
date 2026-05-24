"""
Excel and summary export utilities for Plan.

This module provides plan_to_excel, plan_to_csv, and summary construction
functions. Formatting helpers are also exported for use by saveContributions.

Copyright (C) 2025-2026 The Owl Authors

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import numbers
import numpy as np
import pandas as pd
from io import StringIO
from os.path import isfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from . import config
from . import utils as u
from . import tax2026 as tx
from .rate_models.constants import RATE_DISPLAY_NAMES_SHORT
from .utils import worksheet_age_on_dec_31_or_blank


def _person_index_for_worksheet(sheet_name, inames):
    """Return individual index if sheet is ``{iname}'s …``; else None (household sheet)."""
    for i, iname in enumerate(inames):
        if sheet_name.startswith(f"{iname}'s "):
            return i
    return None


def _last_alive_calendar_year(plan, i):
    """Last calendar year included in the plan horizon for individual ``i``."""
    return int(plan.year_n[0]) + int(plan.horizons[i]) - 1


def _worksheet_age_int_cell(y, plan, i, last_alive_year):
    """Integer age for one row, or None when year is missing or individual is deceased."""
    try:
        yi = int(y)
    except (TypeError, ValueError):
        return None
    v = worksheet_age_on_dec_31_or_blank(
        yi, int(plan.yobs[i]), int(plan.mobs[i]), int(plan.tobs[i]), last_alive_year
    )
    return None if v is None else int(v)


def _insert_age_cols_into_df(df, plan, sheet_name):
    """Insert age column(s) immediately after 'year' in df. Returns modified copy.

    Uses None (not pd.NA) for blank cells so openpyxl can serialize them.
    """
    years = pd.to_numeric(df["year"], errors="coerce")
    idx = df.columns.get_loc("year") + 1
    dfc = df.copy()
    pi = _person_index_for_worksheet(sheet_name, plan.inames)
    persons = [pi] if pi is not None else range(plan.N_i)
    for offset, i in enumerate(persons):
        col = f"age ({plan.inames[i]})"
        last_y = _last_alive_calendar_year(plan, i)
        vals = [_worksheet_age_int_cell(y, plan, i, last_y) for y in years]
        dfc.insert(idx + offset, col, pd.Series(vals, dtype=object, index=dfc.index))
    return dfc


def _format_age_cols_in_ws(ws):
    """Apply integer format to any column whose header starts with 'age ('."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for col_idx, header in enumerate(headers, start=1):
        if isinstance(header, str) and header.startswith("age ("):
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter][1:]:
                cell.number_format = "0"


_FORMAT_STRINGS = {
    "currency": "$#,##0_);[Red]($#,##0)",
    "percent2": "#.00%",
    "percent1": "#.0%",
    "percent0": "#0%",
    "pct_value": "0.00",
}

# Synopsis keys for paired horizon totals: gamma-adjusted vs nominal sums.
SUMMARY_LABEL_TODAY = " (today's $)"
SUMMARY_LABEL_NOMINAL = " (nominal)"

SUMMARY_SECTION_OVERVIEW = "--- Overview ---"
SUMMARY_SECTION_SPENDING = "--- Spending & income (horizon totals) ---"
SUMMARY_SECTION_TAXES = "--- Taxes & premiums (horizon totals) ---"
SUMMARY_SECTION_PARTIAL_BEQUEST = "--- Partial bequest ---"
SUMMARY_SECTION_FINAL_BEQUEST = "--- Final bequest ---"
SUMMARY_SECTION_PLAN = "--- Plan & solver ---"


def _summary_section(dic, title):
    """Insert a visual section divider row (empty value)."""
    dic[title] = ""


def _parse_usd_string(s):
    """Parse ``u.d()`` output or a float; return float or None if not currency."""
    if s is None:
        return None
    if isinstance(s, (int, np.integer)):
        return float(s)
    if isinstance(s, float):
        if np.isnan(s):
            return None
        return float(s)
    if not isinstance(s, str):
        return None
    t = s.strip()
    if not t.startswith("$"):
        return None
    try:
        return float(t[1:].replace(",", ""))
    except ValueError:
        return None


def _summary_currency_pair(dic, label, val_today, val_nominal, prefix=""):
    """Append paired synopsis entries: today's dollars vs nominal dollar totals."""
    dic[f"{prefix}{label}{SUMMARY_LABEL_TODAY}"] = u.d(val_today)
    dic[f"{prefix}{label}{SUMMARY_LABEL_NOMINAL}"] = u.d(val_nominal)


def _save_workbook(wb, basename, overwrite, mylog):
    """Save workbook to file with overwrite prompt."""
    if Path(basename).suffixes == []:
        fname = "workbook" + "_" + basename + ".xlsx"
    else:
        fname = basename

    if not overwrite and isfile(fname):
        mylog.print(f'File "{fname}" already exists.')
        key = input("Overwrite? [Ny] ")
        if key != "y":
            mylog.print("Skipping save and returning.")
            return None

    for _ in range(3):
        try:
            mylog.vprint(f'Saving plan as "{fname}".')
            wb.save(fname)
            break
        except PermissionError:
            mylog.print(f'Failed to save "{fname}": Permission denied.')
            key = input("Close file and try again? [Yn] ")
            if key == "n":
                break
        except Exception as e:
            raise Exception(f"Unanticipated exception {e}.") from e

    return None


def _format_spreadsheet(ws, ftype):
    """Beautify spreadsheet worksheet with appropriate number formatting."""
    if ftype == "summary":
        for col in ws.columns:
            column = col[0].column_letter
            width = max(len(str(col[0].value)) + 20, 40)
            ws.column_dimensions[column].width = width
        return None

    fstring = _FORMAT_STRINGS.get(ftype)
    if fstring is None:
        raise RuntimeError(f"Unknown format: {ftype}.")

    for cell in ws[1] + ws["A"]:
        cell.style = "Pandas"
    for col in ws.columns:
        column = col[0].column_letter
        width = max(len(str(col[0].value)) + 4, 10)
        ws.column_dimensions[column].width = width
        if column == "A":
            for cell in col:
                cell.number_format = "0"
        else:
            for cell in col:
                cell.number_format = fstring

    return None


def _summary_sheet_cell_is_currency_number(v):
    """True if Summary sheet cell should receive Excel currency formatting."""
    if v in (None, "") or isinstance(v, bool):
        return False
    if isinstance(v, numbers.Real):
        try:
            x = float(v)
        except (TypeError, ValueError):
            return False
        return not np.isnan(x)
    return False


def _format_summary_sheet(ws):
    """Wide metric column, fixed currency columns, section merges, number formats."""
    currency_fmt = "$#,##0_);[Red]($#,##0)"
    header_font = Font(bold=True)
    section_font = Font(bold=True, color="FF1565C0")  # --- section divider rows only
    for cell in ws[1]:
        cell.style = "Pandas"
        cell.font = header_font
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row or 2):
        a = row[0]
        av = a.value
        if isinstance(av, str) and av.startswith("---"):
            for cell in row[:3]:
                cell.font = section_font
            ws.merge_cells(start_row=a.row, start_column=1, end_row=a.row, end_column=3)
            a.font = section_font
            a.alignment = Alignment(horizontal="left", vertical="center")
            continue
        b, c = row[1], row[2]
        if _summary_sheet_cell_is_currency_number(b.value):
            b.number_format = currency_fmt
        if _summary_sheet_cell_is_currency_number(c.value):
            c.number_format = currency_fmt


def _format_col_sheet(ws, col_formats, default_fmt=None, lowercase=True):
    """Format a DataFrame sheet: style header, set column widths, apply per-column number formats.

    col_formats : dict mapping column-name string → Excel format string
    default_fmt : format string applied to columns not in col_formats (None = skip them)
    lowercase   : if True, column names are lowercased before lookup (default True)
    """
    from openpyxl.utils import get_column_letter

    for cell in ws[1]:
        cell.style = "Pandas"

    header_row = ws[1]
    col_map = {}
    for idx, cell in enumerate(header_row, start=1):
        col_letter = get_column_letter(idx)
        col_name = (str(cell.value).lower() if lowercase else str(cell.value)) if cell.value else ""
        col_map[col_letter] = col_name
        width = max(len(str(cell.value)) + 4, 10)
        ws.column_dimensions[col_letter].width = width

    for col_letter, col_name in col_map.items():
        fstring = col_formats.get(col_name, default_fmt)
        if fstring is None:
            continue
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == col_letter:
                    cell.number_format = fstring


def _format_debts_sheet(ws):
    """Format Debts sheet with appropriate column formatting."""
    _format_col_sheet(ws, col_formats={
        "year": "0", "term": "0",
        "rate": "#,##0.00",
        "amount": "$#,##0_);[Red]($#,##0)",
    })


def _format_fixed_assets_sheet(ws):
    """Format Fixed Assets sheet with appropriate column formatting."""
    _format_col_sheet(ws, col_formats={
        "yod": "0",
        "rate": "#,##0.00", "commission": "#,##0.00",
        "basis": "$#,##0_);[Red]($#,##0)", "value": "$#,##0_);[Red]($#,##0)",
    })


def _format_federal_income_tax_sheet(ws):
    """Format Federal Income Tax sheet: currency for $ columns, percent for SS % taxed."""
    _format_col_sheet(ws, col_formats={
        "year": "0",
        "SS % taxed": "#.0%",
    }, default_fmt="$#,##0_);[Red]($#,##0)", lowercase=False)


def fixedIncomeStreams(plan, N=None):
    """
    Return fixed income streams per plan year, in nominal dollars.

    Returns a dict with keys "ss", "pension", "wages", "other", "fa_income", "total",
    each a 1D array of shape (N,). Divide by plan.gamma_n[:N] for real (today's) dollars.
    Sum for plan total; mean for annual average (used as the RES floor).
    """
    if N is None:
        N = plan.N_n
    ss = np.sum(plan.zetaBar_in[:, :N], axis=0)
    pension = np.sum(plan.piBar_in[:, :N], axis=0)
    spia = np.sum(plan.spiaBar_in[:, :N], axis=0)
    wages = np.sum(plan.omega_in[:, :N], axis=0)
    other = np.sum(plan.other_inc_in[:, :N], axis=0)
    fa_income = plan.fixed_assets_ordinary_income_n[:N].copy()
    return {
        "ss": ss,
        "pension": pension,
        "spia": spia,
        "wages": wages,
        "other": other,
        "fa_income": fa_income,
        "total": ss + pension + spia + wages + other + fa_income,
    }


def build_summary_dic(plan, N=None):
    """Return dictionary containing summary of plan values.

    Section dividers (keys starting with ``---`` and empty value) group the synopsis.
    Paired horizon totals use two adjacent keys: ``"... (today's $)"`` (scaled by
    ``1/gamma_n``) and ``"... (nominal)"`` (sums of nominal year-by-year flows).
    """
    if N is None:
        N = plan.N_n
    if not (0 < N <= plan.N_n):
        raise ValueError(f"Value N={N} is out of range")

    now = plan.year_n[0]
    dic = {}
    dic["Case name"] = plan._name
    _summary_section(dic, SUMMARY_SECTION_OVERVIEW)
    dic["Net yearly spending basis"] = u.d(plan.g_n[0] / plan.xi_n[0])
    dic["Effective tax rate (plan average)"] = u.pc(plan._actual_effective_tax_rate(), f=1)
    dic[f"Net spending for year {now}"] = u.d(plan.g_n[0])
    dic[f"Net spending remaining in year {now}"] = u.d(plan.g_n[0] * plan.yearFracLeft)

    _summary_section(dic, SUMMARY_SECTION_SPENDING)
    totSpending = np.sum(plan.g_n[:N], axis=0)
    totSpendingNow = np.sum(plan.g_n[:N] / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Total net spending", totSpendingNow, totSpending)

    streams = fixedIncomeStreams(plan, N)
    inv_gamma = 1.0 / plan.gamma_n[:N]
    totFixed = float(np.sum(streams["total"]))
    if totFixed > 0:
        tot_fixed_now = float(np.sum(streams["total"] * inv_gamma))
        _summary_currency_pair(dic, "Total fixed income", tot_fixed_now, totFixed)
        labels = [
            ("ss",        "Social Security"),
            ("pension",   "Pension"),
            ("spia",      "SPIA income"),
            ("wages",     "Wages"),
            ("other",     "Other income"),
            ("fa_income", "Fixed assets income"),
        ]
        for key, label in labels:
            tot = float(np.sum(streams[key]))
            if tot > 0:
                _summary_currency_pair(
                    dic, label, float(np.sum(streams[key] * inv_gamma)), tot, prefix="»  ",
                )

    totRoth = np.sum(plan.x_in[:, :N], axis=(0, 1))
    totRothNow = np.sum(np.sum(plan.x_in[:, :N], axis=0) / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Total Roth conversions", totRothNow, totRoth)

    _summary_section(dic, SUMMARY_SECTION_TAXES)
    taxPaid = np.sum(plan.T_n[:N], axis=0)
    taxPaidNow = np.sum(plan.T_n[:N] / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Total tax paid on ordinary income", taxPaidNow, taxPaid)
    for t in range(plan.N_t):
        taxPaid = np.sum(plan.T_tn[t, :N], axis=0)
        taxPaidNow = np.sum(plan.T_tn[t, :N] / plan.gamma_n[:N], axis=0)
        tname = tx.taxBracketNames[t] if t < len(tx.taxBracketNames) else f"Bracket {t}"
        _summary_currency_pair(dic, f"Tax bracket {tname}", taxPaidNow, taxPaid, prefix="»  ")

    penaltyPaid = np.sum(plan.P_n[:N], axis=0)
    penaltyPaidNow = np.sum(plan.P_n[:N] / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Early withdrawal penalty", penaltyPaidNow, penaltyPaid, prefix="»  ")

    taxPaid = np.sum(plan.U_n[:N], axis=0)
    taxPaidNow = np.sum(plan.U_n[:N] / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Total tax paid on gains and dividends", taxPaidNow, taxPaid)

    taxPaid = np.sum(plan.J_n[:N], axis=0)
    taxPaidNow = np.sum(plan.J_n[:N] / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Total net investment income tax paid", taxPaidNow, taxPaid)

    taxPaid = np.sum(plan.m_n[:N] + plan.M_n[:N], axis=0)
    taxPaidNow = np.sum((plan.m_n[:N] + plan.M_n[:N]) / plan.gamma_n[:N], axis=0)
    _summary_currency_pair(dic, "Total Medicare premiums paid", taxPaidNow, taxPaid)
    hsa_med = np.sum(plan.hsa_medicare_n[:N])
    hsa_med_now = np.sum(plan.hsa_medicare_n[:N] / plan.gamma_n[:N])
    _summary_currency_pair(dic, "»  Covered by HSA", hsa_med_now, hsa_med)

    aca_total = np.sum(plan.aca_costs_n[:N], axis=0)
    if aca_total > 0:
        aca_totalNow = np.sum(plan.aca_costs_n[:N] / plan.gamma_n[:N], axis=0)
        _summary_currency_pair(dic, "Total ACA premiums paid", aca_totalNow, aca_total)

    totDebtPayments = np.sum(plan.debt_payments_n[:N], axis=0)
    if totDebtPayments > 0:
        totDebtPaymentsNow = np.sum(plan.debt_payments_n[:N] / plan.gamma_n[:N], axis=0)
        _summary_currency_pair(dic, "Total debt payments", totDebtPaymentsNow, totDebtPayments)

    if plan.N_i == 2 and plan.n_d < plan.N_n and N == plan.N_n:
        _summary_section(dic, SUMMARY_SECTION_PARTIAL_BEQUEST)
        p_j = plan.partialEstate_j * (1 - plan.phi_j)
        p_j[1] *= 1 - plan.nu   # tax-deferred: heirs pay ordinary income tax
        p_j[3] *= 1 - plan.nu   # HSA: non-spouse heirs include full balance in ordinary income
        nx = plan.n_d - 1
        ynx = plan.year_n[nx]
        ynxNow = 1.0 / plan.gamma_n[nx + 1]
        totOthers = np.sum(p_j)
        q_j = plan.partialEstate_j * plan.phi_j
        totSpousal = np.sum(q_j)
        iname_s = plan.inames[plan.i_s]
        iname_d = plan.inames[plan.i_d]
        dic["Year of partial bequest"] = f"{ynx}"
        _summary_currency_pair(dic, f"Sum of spousal transfer to {iname_s}", ynxNow * totSpousal, totSpousal)
        _summary_currency_pair(
            dic, f"Spousal transfer to {iname_s} - taxable", ynxNow * q_j[0], q_j[0], prefix="»  ",
        )
        _summary_currency_pair(
            dic, f"Spousal transfer to {iname_s} - tax-def", ynxNow * q_j[1], q_j[1], prefix="»  ",
        )
        _summary_currency_pair(
            dic, f"Spousal transfer to {iname_s} - tax-free", ynxNow * q_j[2], q_j[2], prefix="»  ",
        )
        _summary_currency_pair(
            dic, f"Spousal transfer to {iname_s} - HSA", ynxNow * q_j[3], q_j[3], prefix="»  ",
        )
        _summary_currency_pair(
            dic, f"Sum of post-tax non-spousal bequest from {iname_d}", ynxNow * totOthers, totOthers,
        )
        _summary_currency_pair(
            dic,
            f"Post-tax non-spousal bequest from {iname_d} - taxable",
            ynxNow * p_j[0],
            p_j[0],
            prefix="»  ",
        )
        _summary_currency_pair(
            dic,
            f"Post-tax non-spousal bequest from {iname_d} - tax-def",
            ynxNow * p_j[1],
            p_j[1],
            prefix="»  ",
        )
        _summary_currency_pair(
            dic,
            f"Post-tax non-spousal bequest from {iname_d} - tax-free",
            ynxNow * p_j[2],
            p_j[2],
            prefix="»  ",
        )
        _summary_currency_pair(
            dic,
            f"Post-tax non-spousal bequest from {iname_d} - HSA",
            ynxNow * p_j[3],
            p_j[3],
            prefix="»  ",
        )

    if N == plan.N_n:
        _summary_section(dic, SUMMARY_SECTION_FINAL_BEQUEST)
        estate = np.sum(plan.b_ijn[:, :, plan.N_n], axis=0)
        heirsTaxLiability = (estate[1] + estate[3]) * plan.nu   # tax-deferred and HSA
        estate[1] *= 1 - plan.nu   # tax-deferred: heirs pay ordinary income tax
        estate[3] *= 1 - plan.nu   # HSA: non-spouse heirs include full balance in ordinary income
        endyear = plan.year_n[-1]
        lyNow = 1.0 / plan.gamma_n[-1]
        debts = plan.remaining_debt_balance
        savingsEstate = np.sum(estate)
        totEstate = savingsEstate - debts + plan.fixed_assets_bequest_value

        dic["Year of final bequest"] = f"{endyear}"
        _summary_currency_pair(dic, "Total after-tax value of final bequest", lyNow * totEstate, totEstate)
        _summary_currency_pair(
            dic, "After-tax value of savings assets", lyNow * savingsEstate, savingsEstate, prefix="» "
        )
        _summary_currency_pair(
            dic,
            "Fixed assets liquidated at end of plan",
            lyNow * plan.fixed_assets_bequest_value,
            plan.fixed_assets_bequest_value,
            prefix="» ",
        )
        _summary_currency_pair(
            dic, "With heirs assuming tax liability of", lyNow * heirsTaxLiability, heirsTaxLiability, prefix="» ",
        )
        _summary_currency_pair(dic, "After paying remaining debts of", lyNow * debts, debts, prefix="» ")
        _summary_currency_pair(
            dic, "Post-tax final bequest account value - taxable", lyNow * estate[0], estate[0], prefix="»  ",
        )
        _summary_currency_pair(
            dic, "Post-tax final bequest account value - tax-def", lyNow * estate[1], estate[1], prefix="»  ",
        )
        _summary_currency_pair(
            dic, "Post-tax final bequest account value - tax-free", lyNow * estate[2], estate[2], prefix="»  ",
        )
        _summary_currency_pair(
            dic, "Post-tax final bequest account value - HSA", lyNow * estate[3], estate[3], prefix="»  ",
        )

    _summary_section(dic, SUMMARY_SECTION_PLAN)
    dic["Case starting date"] = str(plan.startDate)
    dic["Cumulative inflation factor at end of final year"] = f"{plan.gamma_n[N]:.2f}"
    for i in range(plan.N_i):
        dic[f"{plan.inames[i]:>14}'s life horizon"] = f"{now} -> {now + plan.horizons[i] - 1}"
        dic[f"{plan.inames[i]:>14}'s years planned"] = f"{plan.horizons[i]}"
        if hasattr(plan, "ssecAges") and plan.ssecAmounts[i] > 0:
            age_y = int(plan.ssecAges[i])
            age_m = round((plan.ssecAges[i] - age_y) * 12)
            dic[f"{plan.inames[i]:>14}'s SS claiming age"] = f"{age_y}y {age_m:02d}m"

    dic["Number of decision variables"] = str(plan.A.nvars)
    dic["Number of constraints"] = str(plan.A.ncons)
    dic["Convergence"] = plan.convergenceType
    dic["Case executed on"] = str(plan._timestamp)

    return dic


def build_summary_sheet_df(plan, N=None):
    """Synopsis as three columns (metric, today's dollars, nominal) for Excel."""
    dic = build_summary_dic(plan, N)
    lt = SUMMARY_LABEL_TODAY
    ln = SUMMARY_LABEL_NOMINAL
    rows = []
    items = list(dic.items())
    i = 0
    while i < len(items):
        k, v = items[i]
        if isinstance(k, str) and k.startswith("---") and v == "":
            rows.append({"Metric": k, "Today's $": "", "Nominal $": ""})
            i += 1
            continue
        if k.endswith(lt) and i + 1 < len(items):
            k2, v2 = items[i + 1]
            base = k[: -len(lt)]
            if k2 == base + ln:
                t_num = _parse_usd_string(v)
                n_num = _parse_usd_string(v2)
                rows.append({
                    "Metric": base,
                    "Today's $": t_num if t_num is not None else v,
                    "Nominal $": n_num if n_num is not None else v2,
                })
                i += 2
                continue
        today_cell = _parse_usd_string(v)
        rows.append({
            "Metric": k,
            "Today's $": today_cell if today_cell is not None else v,
            "Nominal $": "",
        })
        i += 1
    return pd.DataFrame(rows)


def build_summary_list(plan, N=None):
    """Return summary as list of key: value strings."""
    dic = build_summary_dic(plan, N)
    return [f"{key}: {value}" for key, value in dic.items()]


def build_summary_string(plan, N=None):
    """Return multi-column synopsis text (aligned metric / today's $ / nominal)."""
    dic = build_summary_dic(plan, N)
    lt = SUMMARY_LABEL_TODAY
    ln = SUMMARY_LABEL_NOMINAL
    w_m, w_v = 58, 22
    sep = "  "

    def fmt_pair_row(base, today_s, nom_s):
        return f"{base:<{w_m}}{sep}{today_s:>{w_v}}{sep}{nom_s:>{w_v}}"

    lines = ["Synopsis", ""]
    lines.append(fmt_pair_row("Metric", "Today's $", "Nominal $"))
    lines.append("-" * (w_m + len(sep) + w_v + len(sep) + w_v))
    items = list(dic.items())
    i = 0
    while i < len(items):
        k, v = items[i]
        if isinstance(k, str) and k.startswith("---") and v == "":
            lines.append("")
            lines.append(k)
            lines.append("")
            i += 1
            continue
        if k.endswith(lt) and i + 1 < len(items):
            k2, v2 = items[i + 1]
            base = k[: -len(lt)]
            if k2 == base + ln:
                lines.append(fmt_pair_row(base, v, v2))
                i += 2
                continue
        val = f"{v}" if v != "" else ""
        lines.append(fmt_pair_row(k, val, ""))
        i += 1
    return "\n".join(lines) + "\n"


def plan_to_excel(plan, overwrite=False, *, basename=None, saveToFile=True, with_config="no"):
    """
    Build Excel workbook from plan. Optionally save to file.

    Returns wb if saveToFile is False, else None.
    """
    if with_config not in {"no", "first", "last"}:
        raise ValueError(f"Invalid with_config option '{with_config}'.")

    wb = Workbook()

    def add_config_sheet(position):
        if with_config == "no" or position != with_config:
            return

        config_buffer = StringIO()
        config.saveConfig(plan, config_buffer, plan.mylog)
        config_buffer.seek(0)

        ws_config = wb.create_sheet(title="Config (.toml)", index=0 if position == "first" else None)
        for row_idx, line in enumerate(config_buffer.getvalue().splitlines(), start=1):
            ws_config.cell(row=row_idx, column=1, value=line)

    real = getattr(plan, "worksheetRealDollars", False)
    inv_gamma = (1.0 / plan.gamma_n[:plan.N_n]) if real else None

    def fillsheet(sheet, dic, datatype, op=lambda x: x, scale=None, sheet_name=None):
        rawData = {}
        rawData["year"] = plan.year_n
        if datatype == "currency":
            for key in dic:
                val = op(dic[key])
                if scale is not None:
                    val = val * scale
                rawData[key] = u.roundCents(val)
        else:
            for key in dic:
                rawData[key] = op(dic[key])
        df = pd.DataFrame(rawData)
        if plan.worksheetShowAges and sheet_name is not None and "year" in df.columns:
            df = _insert_age_cols_into_df(df, plan, sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        _format_spreadsheet(sheet, datatype)
        if plan.worksheetShowAges and sheet_name is not None:
            _format_age_cols_in_ws(sheet)

    ws = wb.active  # Save reference before add_config_sheet("first") may displace it
    add_config_sheet("first")

    ws.title = "Income"
    incomeDic = {
        "net spending": plan.g_n,
        "taxable ord. income": plan.G_n,
        "taxable gains/divs": plan.Q_n,
        "Tax bills + Med.": plan.T_n + plan.U_n + plan.m_n + plan.M_n + plan.J_n + plan.aca_costs_n,
    }
    fillsheet(ws, incomeDic, "currency", scale=inv_gamma, sheet_name="Income")

    cashFlowDic = {
        "net spending": plan.g_n,
        "all wages": np.sum(plan.omega_in, axis=0),
        "all other inc": np.sum(plan.other_inc_in, axis=0),
        "all net inv": np.sum(plan.netinv_in, axis=0),
        "all pensions": np.sum(plan.piBar_in, axis=0),
        "all soc sec": np.sum(plan.zetaBar_in, axis=0),
        "all BTI's": np.sum(plan.Lambda_in, axis=0),
        "FA ord inc": plan.fixed_assets_ordinary_income_n,
        "FA cap gains": plan.fixed_assets_capital_gains_n,
        "FA tax-free": plan.fixed_assets_tax_free_n,
        "debt pmts": -plan.debt_payments_n,
        "all wdrwls": np.sum(plan.w_ijn, axis=(0, 1)),
        "all deposits": -np.sum(plan.d_in, axis=0),
        "ord taxes": -plan.T_n - plan.J_n,
        "div taxes": -plan.U_n,
        "Medicare": -plan.m_n - plan.M_n,
        "ACA premiums": -plan.aca_costs_n,
    }
    ws = wb.create_sheet("Cash Flow")
    fillsheet(ws, cashFlowDic, "currency", scale=inv_gamma, sheet_name="Cash Flow")

    srcDic = {
        "wages": plan.sources_in["wages"],
        "other inc": plan.sources_in["other inc"],
        "net inv": plan.sources_in["net inv"],
        "social sec": plan.sources_in["ssec"],
        "pension": plan.sources_in["pension"],
        "txbl acc wdrwl": plan.sources_in["txbl acc wdrwl"],
        "RMDs": plan.sources_in["RMD"],
        "+distributions": plan.sources_in["+dist"],
        "Roth conv": plan.sources_in["RothX"],
        "tax-free wdrwl": plan.sources_in["tax-free wdrwl"],
        "HSA wdrwl": plan.sources_in["HSA wdrwl"],
        "big-ticket items": plan.sources_in["BTI"],
    }
    for i in range(plan.N_i):
        sname = plan.inames[i] + "'s Sources"
        ws = wb.create_sheet(sname)
        fillsheet(ws, srcDic, "currency", op=lambda x, i=i: x[i], scale=inv_gamma, sheet_name=sname)

    householdSrcDic = {
        "FA ord inc": plan.sources_in["FA ord inc"],
        "FA cap gains": plan.sources_in["FA cap gains"],
        "FA tax-free": plan.sources_in["FA tax-free"],
        "debt pmts": plan.sources_in["debt pmts"],
    }
    ws = wb.create_sheet("Household Sources")
    fillsheet(ws, householdSrcDic, "currency", op=lambda x: x[0], scale=inv_gamma, sheet_name="Household Sources")

    accDic = {
        "taxable bal": plan.b_ijn[:, 0, :-1],
        "taxable ctrb": plan.kappa_ijn[:, 0, :plan.N_n],
        "taxable dep": plan.d_in,
        "taxable wdrwl": plan.w_ijn[:, 0, :],
        "tax-deferred bal": plan.b_ijn[:, 1, :-1],
        "tax-deferred ctrb": plan.kappa_ijn[:, 1, :plan.N_n],
        "tax-deferred wdrwl": plan.w_ijn[:, 1, :],
        "(included RMDs)": plan.rmd_in[:, :],
        "Roth conv": plan.x_in,
        "tax-free bal": plan.b_ijn[:, 2, :-1],
        "tax-free ctrb": plan.kappa_ijn[:, 2, :plan.N_n],
        "tax-free wdrwl": plan.w_ijn[:, 2, :],
        "HSA bal": plan.b_ijn[:, 3, :-1],
        "HSA ctrb": plan.kappa_ijn[:, 3, :plan.N_n],
        "HSA wdrwl": plan.w_ijn[:, 3, :],
    }
    for i in range(plan.N_i):
        aname = plan.inames[i] + "'s Accounts"
        ws = wb.create_sheet(aname)
        fillsheet(ws, accDic, "currency", op=lambda x, i=i: x[i], scale=inv_gamma, sheet_name=aname)
        scale_final = (1.0 / plan.gamma_n[plan.N_n]) if real else 1.0
        final_year = plan.year_n[-1] + 1

        lastRow = [
            final_year,
            float(u.roundCents(plan.b_ijn[i][0][-1] * scale_final)),
            0, 0, 0,
            float(u.roundCents(plan.b_ijn[i][1][-1] * scale_final)),
            0, 0, 0, 0,
            float(u.roundCents(plan.b_ijn[i][2][-1] * scale_final)),
            0, 0,
            float(u.roundCents(plan.b_ijn[i][3][-1] * scale_final)),
            0, 0,
        ]
        if plan.worksheetShowAges:
            last_y = _last_alive_calendar_year(plan, i)
            age_cell = _worksheet_age_int_cell(final_year, plan, i, last_y)  # None or int
            lastRow.insert(1, age_cell)
        ws.append(lastRow)
        _format_spreadsheet(ws, "currency")
        if plan.worksheetShowAges:
            _format_age_cols_in_ws(ws)

    hsa_total_n = np.sum(plan.w_ijn[:, 3, :], axis=0)
    hsa_qme_n = np.maximum(hsa_total_n - plan.hsa_medicare_n, 0.0)
    hsaDic = {
        "Medicare": plan.m_n + plan.M_n,
        "QME": plan.other_medical_n,
        "HSA total wdrwl": hsa_total_n,
        "HSA→Medicare": plan.hsa_medicare_n,
        "HSA→QME": hsa_qme_n,
    }
    for i in range(plan.N_i):
        pname = plan.inames[i]
        hsaDic[f"HSA bal {pname}"] = plan.b_ijn[i, 3, :-1]
        hsaDic[f"HSA ctrb {pname}"] = plan.kappa_ijn[i, 3, :plan.N_n]
        hsaDic[f"HSA wdrwl {pname}"] = plan.w_ijn[i, 3, :]
    ws = wb.create_sheet("HSA")
    fillsheet(ws, hsaDic, "currency", scale=inv_gamma, sheet_name="HSA")

    TxDic = {}
    for t in range(plan.N_t):
        tname = tx.taxBracketNames[t] if t < len(tx.taxBracketNames) else f"Bracket {t}"
        TxDic[tname] = plan.T_tn[t, :]
    TxDic["total"] = plan.T_n
    TxDic["NIIT"] = plan.J_n
    TxDic["LTCG tax"] = plan.U_n
    TxDic["10% penalty"] = plan.P_n
    ss_n = np.sum(plan.zetaBar_in, axis=0)
    TxDic["SS % taxed"] = np.where(ss_n > 0, plan.Psi_n, 0)
    ws = wb.create_sheet("Federal Income Tax")
    rawData = {"year": plan.year_n}
    for key in TxDic:
        if key == "SS % taxed":
            rawData[key] = TxDic[key]
        else:
            val = TxDic[key]
            if real:
                val = val * inv_gamma
            rawData[key] = u.roundCents(val)
    df = pd.DataFrame(rawData)
    if plan.worksheetShowAges and "year" in df.columns:
        df = _insert_age_cols_into_df(df, plan, "Federal Income Tax")
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _format_federal_income_tax_sheet(ws)
    if plan.worksheetShowAges:
        _format_age_cols_in_ws(ws)

    jDic = {"taxable": 0, "tax-deferred": 1, "tax-free": 2, "hsa": 3}
    kDic = {"stocks": 0, "C bonds": 1, "T notes": 2, "common": 3}
    year_n = np.append(plan.year_n, [plan.year_n[-1] + 1])
    for i in range(plan.N_i):
        ws = wb.create_sheet(plan.inames[i] + "'s Allocations")
        rawData = {}
        rawData["year"] = year_n
        for jkey in jDic:
            for kkey in kDic:
                rawData[jkey + "/" + kkey] = 100 * plan.alpha_ijkn[i, jDic[jkey], kDic[kkey], :]
        df = pd.DataFrame(rawData)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        _format_spreadsheet(ws, "pct_value")

    ratesDic = {name: 100 * plan.tau_kn[k] for k, name in enumerate(RATE_DISPLAY_NAMES_SHORT)}
    ws = wb.create_sheet("Rates")
    fillsheet(ws, ratesDic, "pct_value")

    ws = wb.create_sheet("Summary")
    df = build_summary_sheet_df(plan, plan.N_n)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _format_summary_sheet(ws)
    add_config_sheet("last")

    if saveToFile:
        if basename is None:
            basename = plan._name
        if real:
            basename = basename + "_real"
        _save_workbook(wb, basename, overwrite, plan.mylog)
        return None

    return wb


def plan_to_csv(plan, basename, mylog):
    """Build plan data and write to CSV file."""
    planData = {}
    planData["year"] = plan.year_n
    planData["net spending"] = plan.g_n
    planData["taxable ord. income"] = plan.G_n
    planData["taxable gains/divs"] = plan.Q_n
    planData["Tax bills + Med."] = plan.T_n + plan.U_n + plan.m_n + plan.M_n + plan.J_n + plan.aca_costs_n
    planData["all wages"] = np.sum(plan.omega_in, axis=0)
    planData["all other inc"] = np.sum(plan.other_inc_in, axis=0)
    planData["all net inv"] = np.sum(plan.netinv_in, axis=0)
    planData["all pensions"] = np.sum(plan.piBar_in, axis=0)
    planData["all soc sec"] = np.sum(plan.zetaBar_in, axis=0)
    planData["all BTI's"] = np.sum(plan.Lambda_in, axis=0)
    planData["FA ord inc"] = plan.fixed_assets_ordinary_income_n
    planData["FA cap gains"] = plan.fixed_assets_capital_gains_n
    planData["FA tax-free"] = plan.fixed_assets_tax_free_n
    planData["debt pmts"] = -plan.debt_payments_n
    planData["all wdrwls"] = np.sum(plan.w_ijn, axis=(0, 1))
    planData["all deposits"] = -np.sum(plan.d_in, axis=0)
    planData["ord taxes"] = -plan.T_n - plan.J_n
    planData["div taxes"] = -plan.U_n
    planData["Medicare"] = -plan.m_n - plan.M_n
    planData["ACA premiums"] = -plan.aca_costs_n
    planData["QME"] = plan.other_medical_n
    planData["HSA total wdrwl"] = np.sum(plan.w_ijn[:, 3, :], axis=0)
    planData["HSA→Medicare"] = plan.hsa_medicare_n
    planData["HSA→QME"] = np.maximum(planData["HSA total wdrwl"] - plan.hsa_medicare_n, 0.0)

    for i in range(plan.N_i):
        planData[plan.inames[i] + " txbl bal"] = plan.b_ijn[i, 0, :-1]
        planData[plan.inames[i] + " txbl dep"] = plan.d_in[i, :]
        planData[plan.inames[i] + " txbl wrdwl"] = plan.w_ijn[i, 0, :]
        planData[plan.inames[i] + " tx-def bal"] = plan.b_ijn[i, 1, :-1]
        planData[plan.inames[i] + " tx-def ctrb"] = plan.kappa_ijn[i, 1, :plan.N_n]
        planData[plan.inames[i] + " tx-def wdrl"] = plan.w_ijn[i, 1, :]
        planData[plan.inames[i] + " (RMD)"] = plan.rmd_in[i, :]
        planData[plan.inames[i] + " Roth conv"] = plan.x_in[i, :]
        planData[plan.inames[i] + " tx-free bal"] = plan.b_ijn[i, 2, :-1]
        planData[plan.inames[i] + " tx-free ctrb"] = plan.kappa_ijn[i, 2, :plan.N_n]
        planData[plan.inames[i] + " tax-free wdrwl"] = plan.w_ijn[i, 2, :]
        planData[plan.inames[i] + " HSA bal"] = plan.b_ijn[i, 3, :-1]
        planData[plan.inames[i] + " HSA ctrb"] = plan.kappa_ijn[i, 3, :plan.N_n]
        planData[plan.inames[i] + " HSA wdrwl"] = plan.w_ijn[i, 3, :]
        planData[plan.inames[i] + " big-ticket items"] = plan.Lambda_in[i, :]

    for k, name in enumerate(RATE_DISPLAY_NAMES_SHORT):
        planData[name] = 100 * plan.tau_kn[k]

    df = pd.DataFrame(planData)

    while True:
        try:
            fname = "worksheet" + "_" + basename + ".csv"
            df.to_csv(fname)
            break
        except PermissionError:
            mylog.print(f'Failed to save "{fname}": Permission denied.')
            key = input("Close file and try again? [Yn] ")
            if key == "n":
                break
        except Exception as e:
            raise Exception(f"Unanticipated exception: {e}.") from e

    return None
