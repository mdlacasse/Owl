"""Generate the modest-portfolio example case (Case_cameron.toml + HFP_cameron.xlsx).

Cameron is Dana scaled to roughly a tenth of the savings, with no home and a lifetime
low-to-middle earner's Social Security record rather than Dana's. The point of the case
is the regime it lands in: a household whose benefit stays below the Social Security
taxability threshold, whose remaining ordinary income sits under the standard deduction,
and whose capital gains sit entirely in the 0% bracket pays no tax in any year. Solved
with the old exclusion binaries this case took 577 seconds; the same case in a state with
no income tax took 0.06.

    uv run python scripts/make_lowwealth_case.py

Copyright (C) 2024-2026 Martin-D. Lacasse and The Owl Authors

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.
"""

import openpyxl

HFP_OUT = "examples/HFP_cameron.xlsx"
TOML_OUT = "examples/Case_cameron.toml"

FIRST_YEAR, LAST_YEAR = 2026, 2052

PERSON_COLUMNS = [
    "year",
    "anticipated wages",
    "other inc",
    "net inv",
    "taxable ctrb",
    "401k ctrb",
    "Roth 401k ctrb",
    "IRA ctrb",
    "Roth IRA ctrb",
    "HSA ctrb",
    "Roth conv",
    "big-ticket items",
]
DEBT_COLUMNS = ["active", "name", "type", "year", "term", "amount", "rate"]
ASSET_COLUMNS = ["active", "name", "type", "year", "basis", "value", "rate", "yod", "commission"]

TOML = r'''case_name = "cameron"
description = "Cameron is a single retiree of modest means: 66, retired, renting in New York, with about $119,000 of savings and a lifetime low-to-middle earner's Social Security record. There is no pension and no wage income. Claiming is deferred to 70, which raises the benefit to about $1,550/month.\n\nThe plan settles on roughly $19,000 a year of spending in today's dollars. Savings carry the first four years and drop to about $52,000 by the time Social Security starts, then hold near that level for the rest of the horizon, with a $20,000 bequest left at the end.\n\nThe case also covers a regime the other examples miss. Every other shipped scenario has enough income to owe federal tax; Cameron does not. The benefit is small enough that provisional income never reaches the threshold where Social Security becomes taxable, the rest of the ordinary income is absorbed by the standard deduction, and every realized gain sits inside the 0% long-term capital gains bracket, so the plan pays nothing in tax in any year of the horizon.\n\nThat matters more than it sounds. Owl distinguishes the three account types only through their tax treatment, so at a zero marginal rate a dollar is worth the same wherever it sits, and the optimizer becomes indifferent between plans that look very different on paper. New York compounds it: a state that taxes income carries its own deduction, and New York adds a retirement income exclusion on top, so at zero income those are free to move as well. Cameron is the regression case for that degeneracy -- the objective must stay stable even though the individual withdrawal path is not uniquely determined.\n"

[basic_info]
status = "single"
names = [ "Cameron",]
date_of_birth = [ "1960-03-20",]
life_expectancy = [ 92,]
start_date = "2026-01-01"
state = "NY"

[savings_assets]
taxable_savings_balances = [ 15.0,]
tax_deferred_savings_balances = [ 94.0,]
tax_free_savings_balances = [ 10.0,]
hsa_savings_balances = [ 0.0,]
beneficiary_fractions = [ 1.0, 1.0, 1.0, 1.0,]
spousal_surplus_deposit_fraction = 0.5

[household_financial_profile]
HFP_file_name = "HFP_cameron.xlsx"

[fixed_income]
pension_monthly_amounts = [ 0,]
pension_ages = [ 65.0,]
pension_indexed = [ false,]
pension_survivor_fraction = [ 0.0,]
social_security_pia_amounts = [ 1_250,]
social_security_ages = [ 70.0,]

[rates_selection]
heirs_rate_on_tax_deferred_estate = 30.0
dividend_rate = 1.7
obbba_expiration_year = 2032
method = "historical_average"
to = 2025
from = 1928
reverse_sequence = false
roll_sequence = 0

[asset_allocation]
interpolation_method = "s-curve"
type = "individual"
interpolation_center = 15.0
interpolation_width = 5.0
generic = [ [ [ 60.0, 40.0, 0.0, 0.0,], [ 60.0, 40.0, 0.0, 0.0,],],]

[optimization_parameters]
spending_profile = "flat"
objective = "maxSpending"

[solver_options]
noRothConversions = "none"
startRothConversions = 2026
bequest = 20
solver = "default"
spendingSlack = 0
withMedicare = "loop"
withACA = "loop"
withLTCG = "loop"
withNIIT = "loop"
withDecomposition = "none"
withSSTaxability = "loop"
withSSAges = "fixed"

[results]
default_plots = "today"
worksheet_show_ages = false
worksheet_hide_zero_columns = false
worksheet_real_dollars = false
'''


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cameron"
    ws.append(PERSON_COLUMNS)
    for year in range(FIRST_YEAR, LAST_YEAR + 1):
        ws.append([year] + [0] * (len(PERSON_COLUMNS) - 1))

    wb.create_sheet("Debts").append(DEBT_COLUMNS)

    assets = wb.create_sheet("Fixed Assets")
    assets.append(ASSET_COLUMNS)
    # No home: Cameron rents, which is what keeps taxable income under the deduction.

    wb.save(HFP_OUT)
    with open(TOML_OUT, "w") as f:
        f.write(TOML)
    print(f"Wrote {HFP_OUT} and {TOML_OUT}")


if __name__ == "__main__":
    main()
