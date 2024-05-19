'''

A Python program for exploring scenarios in retirement planning.
Calculations are done on a yearly basis.

Currently, it supports single filers and married couples
filing jointly. Feel free to modify for other cases.

Copyright --  Martin-D. Lacasse (2023)

Disclaimer: This program comes with no guarantee. Use at your own risk.

'''
######################################################################
# External modules required:
import sys
import datetime
import numpy as np
import math

# Our own modules:
import utils as u
import rates
import tax2024 as tx

######################################################################


def setVerbose(state):
    '''
    Control verbosity of calculations. True or False for now.
    '''
    u.setVerbose(state)


class Plan:
    def __init__(self, YOB, expectancy):
        '''
        Constructor requires two lists: the first list containing the
        years of birth of each spouse and the second, life expectancies.
        To be clear: for singles, a list of one entry, for married
        couples, a list of two entries.
        This information will determine the size of arrays required
        for the calculations.
        '''
        # Name defaults to Untitled.
        self._name = 'Untitled'

        # Check validity of inputs.
        self.count = len(YOB)
        assert self.count == len(expectancy)
        assert 0 < self.count and self.count <= 2

        self.status = ['single', 'married'][self.count - 1]

        self.yob = YOB
        self.expectancy = expectancy

        now = datetime.date.today().year
        # Compute both life horizons through a comprehension.
        # Add one to account for inclusive bounds.
        # We assume that individual lives until the end of the year during
        # which she/he reached the age of life expectancy.
        self.horizons = [YOB[i] + expectancy[i] - now + 1 for i in range(self.count)]
        # Add one more year as we are computing balances for next year and
        # the beginning of last year will contain estate value.
        self.span = max(self.horizons) + 1

        u.vprint(
            'Preparing scenario of',
            self.span - 1,
            'years',
            'for',
            self.count,
            'individual' + ['', 's'][self.count - 1],
        )

        # Variables starting with a 'y' are tracking yearly values.
        # Initialize variables that track values year after year:
        self.yyear = np.array(range(now, now + self.span))

        if self.count == 1:
            age = _age(YOB[0])
            self.y2ages = np.zeros((self.count, self.span), dtype=int)
            self.y2ages[0][:] = range(age, age + self.span)
        elif self.count == 2:
            ages = np.array([_age(YOB[0]), _age(YOB[1])])
            self.y2ages = np.array(
                [
                    range(ages[0], ages[0] + self.span),
                    range(ages[1], ages[1] + self.span),
                ]
            )

        self.y2ages = self.y2ages.transpose()

        now = datetime.date.today().year
        u.vprint('Current ages in', now, ':', self.y2ages[0])

        self.n2balances = {}
        for aType in ['taxable', 'tax-deferred', 'tax-free']:
            self.n2balances[aType] = np.zeros((self.count))

        # By defaults, spouses are full mutual beneficiaries.
        self.beneficiary = np.ones((2))

        # In case Excel file is not read yet.
        if self.count == 1:
            self.names = ['Person 1']
        else:
            self.names = ['Person 1', 'Person 2']

        self.timeLists = None
        self.timeListsFileName = None

        # Default value for distribution split between spouse is auto.
        self.split = 'auto'

        # Net income target during retirement.
        self.target = None
        self.profile = 'flat'

        self.pensionAmount = None
        self.pensionAge = None

        self.ssecAmount = None
        self.ssecAge = None

        # Heir tax rate on taxable portion of estate.
        self.heirsTxRate = 0
        self.setHeirsTaxRate(25)

        self.survivorFraction = 0
        self.setSurvivorFraction(0.6)

        # Variables needed for handling assets ratios.
        self.coordinatedAR = 'none'
        self.y2assetRatios = {}
        self.boundsAR = {}
        for aType in ['taxable', 'tax-deferred', 'tax-free', 'coordinated']:
            self.y2assetRatios[aType] = np.zeros((self.span, self.count, 4))
            self.boundsAR[aType] = np.zeros((2, self.count, 4))

        # Set default values on bounds. k is initial(0) and final(1).
        u.vprint('Using default values for assets allocation ratios.')
        for k in [0, 1]:
            if self.count == 1:
                self._setAR([[0, 25, 50, 25]], [[60, 40, 0, 0]], [[60, 40, 0, 0]], k)
            else:
                self._setAR(
                    [[0, 25, 50, 25], [0, 25, 50, 25]],
                    [[60, 40, 0, 0], [60, 40, 0, 0]],
                    [[60, 40, 0, 0], [60, 40, 0, 0]],
                    k,
                )

        self.interpolateAR()

        self.reset()
        self.setRates('default')

    def reset(self):
        '''
        Reset variables that will change from one run to the other.
        '''
        self.rates = None
        self.rateMethod = None
        self.rateFrm = None
        self.rateTo = None
        self.rateValues = None

        self.y2accounts = None
        self.y2source = None
        self.yincome = None

        # Track if run was successful. Successful until it fails.
        self.success = True

        return

    def setName(self, name):
        '''
        Provide a name to the plan. This name is used
        to distinguish the output.
        '''
        self._name = name

        return

    def setSurvivorFraction(self, fraction):
        '''
        Set fraction of income desired for survivor spouse.
        '''
        assert 0 <= fraction and fraction <= 1.0
        u.vprint('Survivor spouse income fraction set to', fraction)
        self.survivorFraction = fraction

        return

    def setHeirsTaxRate(self, rate):
        '''
        Set the heirs tax rate on the tax-deferred portion of the estate.
        '''
        assert 0 <= rate and rate <= 100
        rate /= 100
        u.vprint(
            'Heirs tax rate on tax-deferred portion of estate set to', pc(rate, f=0)
        )
        self.heirsTxRate = rate

        return

    def setInitialAR(self, *, taxable, taxDeferred, taxFree):
        '''
        Set values of assets allocation ratios in each of the accounts
        for the first horizon year.
        '''
        self._setAR(taxable, taxDeferred, taxFree, 0)

        return

    def setFinalAR(self, *, taxable, taxDeferred, taxFree):
        '''
        Set values of assets allocation ratios in each of the accounts
        for the last horizon year.
        '''
        self._setAR(taxable, taxDeferred, taxFree, 1)

        return

    def _setAR(self, taxable, taxDeferred, taxFree, k):
        '''
        Utility function for setting initial time and final time values
        on assets allocation ratios.
        '''
        # Make sure we have proper entries.
        assert len(taxable) == self.count
        assert len(taxDeferred) == self.count
        assert len(taxFree) == self.count

        for i in range(self.count):
            assert len(taxable[i]) == 4
            assert abs(sum(taxable[i]) - 100) < 0.01

            assert len(taxDeferred[i]) == 4
            assert abs(sum(taxDeferred[i]) - 100) < 0.01

            assert len(taxFree[i]) == 4
            assert abs(sum(taxFree[i]) - 100) < 0.01

        which = ['Initial', 'Final'][k]

        u.vprint(
            which,
            'assets allocation ratios set to: (%)\n',
            taxable,
            '\n',
            taxDeferred,
            '\n',
            taxFree,
        )
        self.boundsAR['taxable'][k] = np.array(taxable) / 100
        self.boundsAR['tax-deferred'][k] = np.array(taxDeferred) / 100
        self.boundsAR['tax-free'][k] = np.array(taxFree) / 100
        self.coordinatedAR = 'none'

        return

    def setCoordinatedAR(self, *, initial, final):
        '''
        Set bounds for portfolios coordinated between assets and spouses.
        Scope of coordination is taken from size of arrays.
        '''
        if len(initial) == self.count:
            scope = 'individual'
            assert len(final) == self.count

            for i in range(self.count):
                assert len(initial[i]) == 4
                assert len(final[i]) == 4
                assert abs(sum(initial[i]) - 100) < 0.01
                assert abs(sum(final[i]) - 100) < 0.01

            self.boundsAR['coordinated'][0] = np.array(initial) / 100
            self.boundsAR['coordinated'][1] = np.array(final) / 100
        elif len(initial) == 4:
            scope = 'both'
            assert len(final) == 4
            assert abs(sum(initial) - 100) < 0.01
            assert abs(sum(final) - 100) < 0.01

            for i in range(self.count):
                self.boundsAR['coordinated'][0][i] = np.array(initial) / 100
                self.boundsAR['coordinated'][1][i] = np.array(final) / 100
        else:
            u.xprint('Lists provided have wrong length:', len(initial))

        self.coordinatedAR = scope
        u.vprint(
            'Coordinating',
            scope,
            'assets allocation ratios (%):\n',
            'initial:',
            initial,
            '\n   final:',
            final,
        )

        return

    def interpolateAR(self, method='linear', center=15, width=5):
        '''
        Interpolate assets allocation ratios from initial value (today) to
        final value (at the end of horizon).

        Two interpolation methods are supported: linear and s-curve.
        Linear is a straight line between now and the end of the simulation.
        Hyperbolic tangent give a smooth "S" curve centered at point "c"
        with a width "w". Center point defaults to 15 years and width to
        5 years. This means that the transition from initial to final
        will start occuring in 10 years (15-5) and will end in 20 years (15+5).
        '''
        if method not in ['linear', 's-curve', 'none']:
            u.xprint('Method', method, 'not supported.')

        if self.coordinatedAR == 'both':
            accType = 'coordinated'
            # Use longest-lived spouse for both time scales.
            maxhorizons = np.ones(self.count, dtype=int) * max(self.horizons)
            if method == 'linear':
                self._linInterp(accType, 1, maxhorizons)
            elif method == 's-curve':
                self._tanhInterp(accType, 1, maxhorizons, center, width)
        elif self.coordinatedAR == 'individual':
            accType = 'coordinated'
            if method == 'linear':
                self._linInterp(accType, self.count, self.horizons)
            elif method == 's-curve':
                self._tanhInterp(accType, self.count, self.horizons, center, width)
        elif self.coordinatedAR == 'none':
            for accType in ['taxable', 'tax-deferred', 'tax-free']:
                if method == 'linear':
                    self._linInterp(accType, self.count, self.horizons)
                elif method == 's-curve':
                    self._tanhInterp(accType, self.count, self.horizons, center, width)
        else:
            u.xprint('Unknown coordination:', self.coordinatedAR)

        u.vprint('Interpolated assets allocation ratios using', method, 'method.')

        return

    def _linInterp(self, accType, count, numPoints):
        '''
        Utility function to interpolate multiple cases using
        a linear interpolation. Range goes one more year than
        horizon as year passed death includes estate.
        '''
        for who in range(count):
            for j in range(4):
                dat = np.linspace(
                    self.boundsAR[accType][0][who][j],
                    self.boundsAR[accType][1][who][j],
                    numPoints[who] + 1,
                )
                for n in range(numPoints[who] + 1):
                    self.y2assetRatios[accType][n][who][j] = dat[n]
                for n in range(numPoints[who] + 1, self.span):
                    self.y2assetRatios[accType][n][who][j] = 0

        return

    def _tanhInterp(self, accType, count, numPoints, c, w):
        '''
        Utility function to interpolate multiple cases using hyperbolic
        tangent interpolation. "c" is the center where the inflection point
        is, and "w" is the width of the transition.
        '''
        for who in range(count):
            for j in range(4):
                t = np.linspace(0, numPoints[who], numPoints[who] + 1)
                a = self.boundsAR[accType][0][who][j]
                b = self.boundsAR[accType][1][who][j]
                dat = a + 0.5 * (b - a) * (1 + np.tanh((t - c) / w))
                for n in range(numPoints[who] + 1):
                    self.y2assetRatios[accType][n][who][j] = dat[n]
                for n in range(numPoints[who] + 1, self.span):
                    self.y2assetRatios[accType][n][who][j] = 0

        return

    def _balanceAR(self, n):
        '''
        Coordinate assets allocation ratios amongst different accounts.
        '''
        if self.coordinatedAR == 'none':
            return
        elif self.coordinatedAR == 'individual':
            for who in range(self.count):
                if n > self.horizons[who]:
                    continue
                c = self.y2assetRatios['coordinated'][n][who]
                X = self.y2accounts['taxable'][n][who]
                Y = self.y2accounts['tax-deferred'][n][who]
                Z = self.y2accounts['tax-free'][n][who]

                (
                    x,
                    y,
                    z,
                ) = _balance(c, X, Y, Z)
                T = X + Y + Z + 0.01
                u.vprint(
                    'Global assets allocation:',
                    pc((X * x[0] + Y * y[0] + Z * z[0]) / T),
                    pc((X * x[1] + Y * y[1] + Z * z[1]) / T),
                    pc((X * x[2] + Y * y[2] + Z * z[2]) / T),
                    pc((X * x[3] + Y * y[3] + Z * z[3]) / T),
                )

                self.y2assetRatios['taxable'][n][who] = x
                self.y2assetRatios['tax-deferred'][n][who] = y
                self.y2assetRatios['tax-free'][n][who] = z
        elif self.coordinatedAR == 'both':
            c = self.y2assetRatios['coordinated'][n][0]
            X = sum(self.y2accounts['taxable'][n])
            Y = sum(self.y2accounts['tax-deferred'][n])
            Z = sum(self.y2accounts['tax-free'][n])

            x, y, z = _balance(c, X, Y, Z)
            T = X + Y + Z + 0.01
            # print('both AR:', x, y, z)
            u.vprint(
                'Global assets allocation:',
                pc((X * x[0] + Y * y[0] + Z * z[0]) / T),
                pc((X * x[1] + Y * y[1] + Z * z[1]) / T),
                pc((X * x[2] + Y * y[2] + Z * z[2]) / T),
                pc((X * x[3] + Y * y[3] + Z * z[3]) / T),
            )

            for who in range(self.count):
                self.y2assetRatios['taxable'][n][who] = x
                self.y2assetRatios['tax-deferred'][n][who] = y
                self.y2assetRatios['tax-free'][n][who] = z
        else:
            u.xprint('Unknown coordination keyword:', self.coordinatedAR)

        return

    def setRates(self, method, frm=rates.FROM, to=rates.TO, values=None):
        '''
        Generate rates for return and inflation based on the method and
        years selected. Note that last bound is included.

        The following methods are available:
        default, fixed, realistic, conservative, average, and historical.

        For *fixed*, values must be provided.
        For *average*, and *historical*, a range of years can be provided.

        '''
        dr = rates.rates()
        dr.setMethod(method, frm, to, values)
        self.rateMethod = method
        self.rateFrm = frm
        self.rateTo = to
        self.rateValues = values
        # Remember that all coded calculations are forward looking.
        # No reference to years before today can be done.
        self.rates = dr.genSeries(frm, to, self.span)
        # u.vprint('Generated rate series of', len(self.rates))

        return

    def setAssetBalances(
        self, *, taxable, taxDeferred, taxFree, beneficiary, units=None
    ):
        '''
        Four entries must be provided. The first three are lists
        containing the balance of all assets in each category for
        each spouse. The last one is the fraction of assets left to
        the other spouse as a beneficiary. For single individuals,
        these lists will contain only one entry and the beneficiary
        value is not relevant.

        Units of 'k', or 'M'
        '''
        assert len(taxable) == self.count
        assert len(taxDeferred) == self.count
        assert len(taxFree) == self.count
        assert len(beneficiary) == self.count

        fac = _getUnits(units)

        _rescale(taxable, fac)
        _rescale(taxDeferred, fac)
        _rescale(taxFree, fac)

        self.n2balances['taxable'][:] = taxable
        self.n2balances['tax-deferred'][:] = taxDeferred
        self.n2balances['tax-free'][:] = taxFree
        self.beneficiary = beneficiary

        u.vprint('Taxable balances:', taxable)
        u.vprint('Tax-deferred balances:', taxDeferred)
        u.vprint('Tax-free balances:', taxFree)
        u.vprint('Beneficiary:', beneficiary)

        return

    def _initializeAccounts(self):
        for aType in ['taxable', 'tax-deferred', 'tax-free']:
            self.y2accounts[aType][0][:] = self.n2balances[aType]

        return

    # Asset ratios are lists with 3 values: stock, bonds, and fixed assets.
    def setAssetRatiosI(self, *, taxableR, taxDeferredR, taxFreeR):
        '''
        Provide asset ratios between stock, bonds, and fixed assets
        for each asset category and each spouse. For a married couple
        with a 60/40 tax-deferred portfolio, 80/20 tax-free, and
        taxable account with 50% bonds and 50% in fixed assets would
        be as follows:

        plan = retirement.Plan(...)

        plan.setAssetRatiosI(
            taxableR = [[0, 50, 50], [0, 50, 50]],
            taxDeferredR = [[60, 40, 0], [60, 40, 0]],
            taxFreeR = [[80, 20, 0], [80, 20, 0]]
            )

        For single individuals, the list contains only one entry
        with three values. Each triplet values must add up to 100%.
        '''
        self._testAssetRatios(self, taxableR, taxDeferredR, taxFreeR)

        # Convert from percent to decimal.
        self.taxableR[0] = np.array(taxableR) / 100
        self.taxDeferredR[0] = np.array(taxDeferredR) / 100
        self.taxFreeR[0] = np.array(taxFreeR) / 100

        return

    def _testAssetRatios(self, taxableR, taxDeferredR, taxFreeR):
        '''
        Determine if entries are correct.
        '''
        assert len(taxableR) == self.count
        assert len(taxDeferredR) == self.count
        assert len(taxFreeR) == self.count

        for i in range(self.count):
            assert (
                len(taxableR[i]) == 3
                and len(taxDeferredR[i]) == 3
                and len(taxFreeR[i]) == 3
            )
            assert (
                abs(sum(taxableR[i]) - 100) < 0.0000001
                and abs(sum(taxDeferredR[i]) - 100) < 0.0000001
                and abs(sum(taxFreeR[i]) - 100) < 0.0000001
            )

        return

    def readContributions(self, filename):
        '''
        Provide the name of the file containing the financial events over the
        anticipated life span determined by the assumed longevity.
        File can be an excel, or odt file with one tab named after
        each spouse and must have the following column headers:

                'year',
                'anticipated wages',
                'ctrb taxable',
                'ctrb 401k',
                'ctrb Roth 401k',
                'ctrb IRA',
                'ctrb Roth IRA',
                'Roth X',
                'big ticket items'

        in any order. A template is provided as an example.
        '''
        self.names, self.timeLists = _readTimeLists(filename, self.count)

        _checkTimeLists(self.names, self.timeLists, self.horizons)
        self.timeListsFileName = filename

        return

    def clearRothX(self):
        '''
        Remove all Roth conversions that might be in the plan.
        '''
        for i in range(self.count):
            for n in range(self.horizons[i] + 1):
                self.timeLists[i]['Roth X'][n] = 0

        return

    def setSpousalSplit(self, split):
        '''
        Specify ration of withdrawals between spousal accounts.
        Default value is 'auto'.
        '''
        if split != 'auto':
            split = float(split)
            assert 0 <= split and split <= 1

        self.split = split
        u.vprint('Using spousal split of', split)

        return

    def _getSplit(self, oldsplit, amount, n, surviving):
        '''
        Calculate auto split based on spousal asset ratio.
        '''
        if surviving == 1:
            return oldsplit
        elif isinstance(self.split, float) or isinstance(self.split, int):
            return self.split
        elif self.split == 'auto':
            s1 = np.sum(self.y2accounts['taxable'][n][:])
            s2 = np.sum(self.y2accounts['tax-deferred'][n][:])
            amount = abs(amount)
            if amount <= s1:
                autoSplit = (self.y2accounts['taxable'][n][0]) / (s1 + 1)
            elif amount <= s2:
                autoSplit = (
                    self.y2accounts['taxable'][n][0]
                    + self.y2accounts['tax-deferred'][n][0]
                ) / (s1 + s2 + 1)
            else:
                s3 = np.sum(self.y2accounts['tax-free'][n][:])
                autoSplit = (
                    self.y2accounts['taxable'][n][0]
                    + self.y2accounts['tax-deferred'][n][0]
                    + self.y2accounts['tax-free'][n][0]
                ) / (s1 + s2 + s3 + 1)

            return autoSplit
        else:
            u.xprint('Unknown split keyword:', self.split)

    def setDesiredIncome(self, income, profile, units=None):
        '''
        Set the net target income in retirement following the specified
        profile. Profile can be 'flat' or 'smile'.
        '''
        income *= _getUnits(units)
        self.target = income
        if profile == 'smile' or profile == 'flat':
            self.profile = profile
        else:
            u.xprint('Unknown profile keyword: ', profile)

        u.vprint('Using desired net income of', d(income), 'with a', profile, 'profile')

        return

    def setPension(self, amounts, ages, units=None):
        '''
        Set amounts of fixed income pensions if any, and age at which
        it will start.
        '''
        assert len(amounts) == self.count
        assert len(ages) == self.count

        fac = _getUnits(units)
        _rescale(amounts, fac)

        self.pensionAmount = amounts
        self.pensionAge = ages
        u.vprint('Setting pension of', amounts, 'at age(s)', ages)

        return

    def _computePension(self, n, i):
        '''
        Compute pension (non-indexed) for individual i in plan year n.
        '''
        refAge = self.pensionAge[i]

        if self.y2ages[n][i] >= refAge:
            return self.pensionAmount[i]

        return 0.0

    def setSocialSecurity(self, amounts, ages, units=None):
        '''
        Set yearly amounts of SS income if any, and age at which
        it will start.
        '''
        assert len(amounts) == self.count
        assert len(ages) == self.count

        fac = _getUnits(units)
        _rescale(amounts, fac)

        self.ssecAmount = amounts
        self.ssecAge = ages
        u.vprint('Setting SSA of', amounts, 'at age(s)', ages)

        return

    def _computeSS(self, n, who):
        '''
        Compute social security benefits (indexed) given age,
        inflation, and predicted amount.
        '''
        if self.y2ages[n][who] < self.ssecAge[who]:
            return 0

        now = datetime.date.today().year
        refIndex = self.yob[who] + self.ssecAge[who] - now

        # Inflation is computed from benefit start year.
        return tx.inflationAdjusted(self.ssecAmount[who], n, self.rates, refIndex)

    def _transferWealth(self, year, late):
        '''
        Transfer fraction of assets from one spouse to the other.
        Spouses can inherit IRA and Roth and treat them as their own.
        '''
        xfer = self.beneficiary
        other = (late + 1) % 2
        u.vprint(
            'Transfering',
            xfer[late],
            'fraction of %s\'s' % self.names[late],
            'wealth to',
            self.names[other],
        )
        for key in self.y2accounts.keys():
            self.y2accounts[key][year][other] += (
                self.beneficiary[late] * self.y2accounts[key][year][late]
            )
            self.y2accounts[key][year][late] = 0

        return

    def totals(self):
        '''
        Returns total amounts either paid or received in today's $.
        '''
        totTaxes = 0
        totIrmaa = 0
        totNetIncome = 0
        totRothX = np.zeros((self.count))
        totSSA = np.zeros((self.count))
        totPension = np.zeros((self.count))
        totWages = np.zeros((self.count))
        for n in range(self.span):
            fac = 1.0 / tx.inflationAdjusted(1.0, n, self.rates)
            totTaxes += fac * self.yincome['taxes'][n]
            totIrmaa += fac * self.yincome['irmaa'][n]
            totNetIncome += fac * self.yincome['net'][n]
            for who in range(self.count):
                totSSA[who] += fac * self.y2source['ssec'][n][who]
                totPension[who] += fac * self.y2source['pension'][n][who]
                totWages[who] += fac * self.y2source['wages'][n][who]
                totRothX[who] += fac * self.y2source['RothX'][n][who]

        print(
            '\nTotals for %r (in today\'s $):\n' % self._name,
            '\tNet income:',
            d(totNetIncome),
            '\n',
            '\tIncome taxes:',
            d(totTaxes),
            '\n',
            '\tIRMAA:',
            d(totIrmaa),
        )
        for who in range(self.count):
            print(self.names[who])
            print(
                '\tWages:',
                d(totWages[who]),
                '\n',
                '\tSocial security:',
                d(totSSA[who]),
                '\n',
                '\tPension:',
                d(totPension[who]),
                '\n',
                '\tRoth conversions:',
                d(totRothX[who]),
            )

        return

    def run(self, verbose=None):
        '''
        Run a simulation given the underlying assumptions for
        the next horizon years determined by life expectancies.
        '''
        if verbose is not None:
            prevState = u.setVerbose(verbose)

        # Keep data in [year][who] for now.
        # We'll transpose later if needed when plotting.
        self.y2accounts = {}
        for aType in ['taxable', 'tax-deferred', 'tax-free']:
            self.y2accounts[aType] = np.zeros((self.span, self.count))

        self._initializeAccounts()

        # Shorter names for class variables:
        ya2taxable = self.y2accounts['taxable']
        ya2taxDef = self.y2accounts['tax-deferred']
        ya2taxFree = self.y2accounts['tax-free']
        filingStatus = self.status

        # All sources of income.
        self.y2source = {
            'rmd': np.zeros((self.span, self.count)),
            'ssec': np.zeros((self.span, self.count)),
            'pension': np.zeros((self.span, self.count)),
            'div': np.zeros((self.span, self.count)),
            'wages': np.zeros((self.span, self.count)),
            'taxable': np.zeros((self.span, self.count)),
            'dist': np.zeros((self.span, self.count)),
            'tax-free': np.zeros((self.span, self.count)),
            'RothX': np.zeros((self.span, self.count)),
            'bti': np.zeros((self.span, self.count)),
        }

        # Use shorter names:
        ys2wages = self.y2source['wages']
        ys2rmd = self.y2source['rmd']
        ys2pension = self.y2source['pension']
        ys2ssec = self.y2source['ssec']
        ys2div = self.y2source['div']
        ys2dist = self.y2source['dist']
        ys2RothX = self.y2source['RothX']
        ys2txfree = self.y2source['tax-free']
        ys2txbl = self.y2source['taxable']
        ys2bti = self.y2source['bti']

        # Beware of names: tax-free income includes
        # distributions from taxable, tax-free account,
        # and part of SS income.
        self.yincome = {
            'RothX': np.zeros((self.span)),
            'gross': np.zeros((self.span)),
            'taxes': np.zeros((self.span)),
            'irmaa': np.zeros((self.span)),
            'net': np.zeros((self.span)),
            'target': np.zeros((self.span)),
            'taxable': np.zeros((self.span)),
            'tax-free': np.zeros((self.span)),
        }

        # Shorter names:
        yRothX = self.yincome['RothX']
        yincomeTax = self.yincome['taxes']
        yirmaa = self.yincome['irmaa']
        ytargetIncome = self.yincome['target']
        ynetIncome = self.yincome['net']
        ygrossIncome = self.yincome['gross']
        ytaxfreeIncome = self.yincome['tax-free']
        ytaxableIncome = self.yincome['taxable']

        # Assumptions/methods:
        # - Simulation starts at the beginning of the year;
        # - Roth conversions are made at the beginning of the year;
        # - Contributions are made in the middle of the year;
        # - Withdrawals are made in the middle of the year;
        # - Taxes are paid in their current year (as estmated tax);
        # - All balances are calculated for beginning of next year.

        rawTarget = self.target

        # For each year ahead:
        u.vprint(
            'Computing plan %r' % self._name,
            'for next',
            self.span - 2,
            'years for',
            ' and '.join(str(x) for x in self.names),
        )

        # Keep track of surviving spouses.
        surviving = self.count
        # Use 1 as default for deposit and withdrawal ratios.
        wdrlRatio = 1
        depRatio = 1
        # y2accounts have one more value for allowing values[n+1].
        for n in range(self.span - 1):
            u.vprint(
                '-------',
                self.yyear[n],
                ' -----------------------------------------------',
            )

            # Balance portfolio with desired assets allocations
            # considering account balances.
            self._balanceAR(n)

            # Annual tracker for taxable distribution related to big items.
            btiEvent = 0
            for i in range(self.count):
                # Is this nth year more than i's life horizon?
                if n >= self.horizons[i]:
                    u.vprint('Skipping', self.names[i], 'in', self.yyear[n])
                    continue

                # Perform requested Roth conversions early in the year.
                # Keep Roth conversions separately as they are not true income
                # but are taxable events.
                # We will add them separately to taxable income we call gross.
                reqRoth = self.timeLists[i]['Roth X'][n]
                assert reqRoth >= 0
                tmp = min(reqRoth, ya2taxDef[n][i])
                if tmp != reqRoth:
                    u.vprint(
                        'WARNING:',
                        'Insufficient funds for',
                        d(reqRoth),
                        'Roth conversion for',
                        self.names[i],
                        'in',
                        self.yyear[n],
                    )
                if tmp > 0:
                    u.vprint(
                        self.names[i],
                        'requested Roth conversion:',
                        d(reqRoth),
                        ' Performed:',
                        d(tmp),
                    )
                    ya2taxDef[n][i] -= tmp
                    ya2taxFree[n][i] += tmp
                    ys2RothX[n][i] = tmp
                    yRothX[n] += tmp

                # Add anticipated wages for the year.
                tmp = self.timeLists[i]['anticipated wages'][n]
                if tmp > 0:
                    u.vprint(self.names[i], 'reported wages of', d(tmp))
                    ys2wages[n][i] += tmp
                    ytaxableIncome[n] += tmp

                # Add contributions and growth to taxable account.
                # Year-end growth assumes contributions are in midyear.
                # Use += to avoid overwriting spousal inheritance.
                # Else, arrays were initialized to zero.
                ctrb = self.timeLists[i]['ctrb taxable'][n]
                growth = (ya2taxable[n][i] + 0.5 * ctrb) * _pfReturn(
                    self.y2assetRatios['taxable'], self.rates, n, i
                )
                ys2div[n][i] = min(0, growth)
                ya2taxable[n + 1][i] += ya2taxable[n][i] + ctrb + growth
                ytaxableIncome[n] += min(0, growth)
                if ctrb > 0 or growth > 0:
                    u.vprint(
                        self.names[i],
                        'Taxable account growth:',
                        d(ya2taxable[n][i]),
                        '->',
                        d(ya2taxable[n + 1][i]),
                    )

                # Same for tax-deferred, including RMDs on year-end balance.
                ctrb = (
                    self.timeLists[i]['ctrb 401k'][n] + self.timeLists[i]['ctrb IRA'][n]
                )

                growth = (ya2taxDef[n][i] + 0.5 * ctrb) * _pfReturn(
                    self.y2assetRatios['tax-deferred'], self.rates, n, i
                )

                ya2taxDef[n + 1][i] += ya2taxDef[n][i] + ctrb + growth

                if ctrb > 0 or growth > 0:
                    u.vprint(
                        self.names[i],
                        'Tax-deferred account growth:',
                        d(ya2taxDef[n][i]),
                        '->',
                        d(ya2taxDef[n + 1][i]),
                    )

                rmd = ya2taxDef[n + 1][i] * tx.rmdFraction(self.yyear[n], self.yob[i])

                if rmd > 0:
                    u.vprint(self.names[i], 'RMD:', d(rmd))

                ya2taxDef[n + 1][i] -= rmd
                ys2rmd[n][i] = rmd
                ytaxableIncome[n] += rmd

                # And contributions to tax-free accounts:
                ctrb = (
                    self.timeLists[i]['ctrb Roth 401k'][n]
                    + self.timeLists[i]['ctrb Roth IRA'][n]
                )

                growth = (ya2taxFree[n][i] + 0.5 * ctrb) * _pfReturn(
                    self.y2assetRatios['tax-free'], self.rates, n, i
                )

                ya2taxFree[n + 1][i] += ya2taxFree[n][i] + ctrb + growth

                if ctrb > 0 or growth > 0:
                    u.vprint(
                        self.names[i],
                        'Tax-free account growth:',
                        d(ya2taxFree[n][i]),
                        '->',
                        d(ya2taxFree[n + 1][i]),
                    )

                # Compute fixed income for this year:
                ys2pension[n][i] = self._computePension(n, i)
                ytaxableIncome[n] += ys2pension[n][i]
                ys2ssec[n][i] = self._computeSS(n, i)
                # Assume our revenues are such that 85% of SS is taxable.
                # Fix if needs arises.
                ytaxfreeIncome[n] += 0.15 * ys2ssec[n][i]
                ytaxableIncome[n] += 0.85 * ys2ssec[n][i]

                # Big-ticket items can be positive or negative.
                # They do not contribute to income,
                # but withdrawals can be taxable.
                # Take it from the account of bearer: use a split of (i+1)%2.
                bti = self.timeLists[i]['big ticket items'][n]
                if bti != 0:
                    u.vprint(self.names[i], 'requested big-ticket item of', d(bti))
                    amounts, total = _smartBanking(
                        bti,
                        ya2taxable,
                        ya2taxDef,
                        ya2taxFree,
                        n + 1,
                        (i + 1) % 2,
                        self.names,
                        self.y2ages,
                        True,
                    )
                    if total != abs(bti):
                        u.vprint(
                            'WARNING: Insufficient funds for BTI for',
                            self.names[i],
                            'in',
                            self.yyear[n],
                        )
                        u.vprint('\tRequested:', d(bti), 'Performed:', d(total))
                        self.success = False

                    # A BTI is not an income unless we created a taxable event
                    # that we track separately (as we do for Roth conversions).
                    # Total is stored in first value of amounts, followed
                    # by values per individual.
                    btiEvent += amounts['tax-def'][0]
                    ys2dist[n][:] += amounts['tax-def'][1:]
                    ys2txfree[n][:] += amounts['tax-free'][1:]
                    ys2txbl[n][:] += amounts['taxable'][1:]
                    ys2bti[n][i] = math.copysign(total, bti)

            # Compute couple's income needs following profile based on
            # oldest spouse's timeline.
            adjustedTarget = rawTarget * _spendingAdjustment(
                np.max(self.y2ages[n][:]), self.profile
            )
            ytargetIncome[n] = tx.inflationAdjusted(adjustedTarget, n, self.rates)

            gross = ytaxableIncome[n] + yRothX[n] + btiEvent
            estimatedTax = tx.incomeTax(
                gross, self.yob, filingStatus, self.yyear[n], self.rates
            )
            netInc = ytaxfreeIncome[n] + ytaxableIncome[n] - estimatedTax
            gap = netInc - ytargetIncome[n]
            u.vprint(
                'Net income target:',
                d(ytargetIncome[n]),
                ' Unadj. net:',
                d(netInc),
                ' Delta:',
                d(gap),
            )
            u.vprint(
                'Taxable:',
                d(ytaxableIncome[n]),
                ' Gross:',
                d(gross),
                ' Est. Taxes:',
                d(estimatedTax),
            )

            if gap >= 0:
                if surviving == 2:
                    # Deposit surplus following this year's income ratio.
                    depRatio = (
                        ys2wages[n][0] + ys2ssec[n][0] + ys2pension[n][0] + ys2rmd[n][0]
                    ) / (
                        sum(
                            ys2wages[n][:]
                            + ys2pension[n][:]
                            + ys2ssec[n][:]
                            + ys2rmd[n][:]
                        )
                        + 1
                    )

                u.vprint(
                    'Depositing',
                    d(gap),
                    'in taxable accounts with ratio',
                    '{:.2f}'.format(depRatio),
                )
                _smartBanking(
                    gap,
                    ya2taxable,
                    ya2taxDef,
                    ya2taxFree,
                    n + 1,
                    depRatio,
                    self.names,
                    self.y2ages,
                    True,
                )
                yincomeTax[n] = estimatedTax
                ygrossIncome[n] = gross
                # Medicare IRMAA looks back 2 years.
                irmaaIncome = ygrossIncome[max(0, n - 2)]
                for i in range(self.count):
                    if self.y2ages[n][i] >= 65 and n <= self.horizons[i]:
                        yirmaa[n] += tx.irmaa(
                            irmaaIncome, filingStatus, self.yyear[n], self.rates
                        )

                ynetIncome[n] = netInc
                u.vprint(
                    'Adj. Income:\n Gross taxable:',
                    d(ygrossIncome[n]),
                    'Tax bill:',
                    d(yincomeTax[n]),
                    'IRMAA:',
                    d(yirmaa[n]),
                    '\n Net:',
                    d(ynetIncome[n]),
                    'Tax free:',
                    d(ytaxfreeIncome[n]),
                )
            else:
                # Solve amount to withdraw self-consistently.
                # Try at most thirty two times.
                # Typically 10 or less iterations are ok.
                # Goal is to reconcile withdrawal in after-tax money,
                # given an existing taxable income.
                withdrawal = gap
                # Adjust withdrawal ratio every year,
                # monitoring surviving spouses.
                wdrlRatio = self._getSplit(wdrlRatio, withdrawal, n + 1, surviving)

                for k in range(32):
                    amounts, total = _smartBanking(
                        withdrawal,
                        ya2taxable,
                        ya2taxDef,
                        ya2taxFree,
                        n + 1,
                        wdrlRatio,
                        self.names,
                        self.y2ages,
                        False,
                    )

                    # Zeroth component of amounts countains total.
                    txfree = amounts['taxable'][0] + amounts['tax-free'][0]
                    txbl = amounts['tax-def'][0]
                    totaxblIncome = yRothX[n] + ytaxableIncome[n] + btiEvent + txbl
                    estimatedTax = tx.incomeTax(
                        totaxblIncome, self.yob, filingStatus, self.yyear[n], self.rates
                    )

                    netInc = (
                        txfree
                        + txbl
                        + ytaxfreeIncome[n]
                        + ytaxableIncome[n]
                        - estimatedTax
                    )

                    # No point to loop if we are running out of money.
                    if abs(total - abs(withdrawal)) > 1:
                        print('WARNING: Running out of money in year', self.yyear[n])
                        self.success = False
                        break

                    delta = netInc - ytargetIncome[n]
                    if delta >= -1:
                        u.vprint(
                            'Solved with', k, 'iteration(s) and delta of', d(delta, 2)
                        )
                        break

                    withdrawal += delta
                else:
                    u.xprint('Could not converge on withdrawal.')

                # Now commit withdrawal and pay taxes.
                amounts, total = _smartBanking(
                    withdrawal,
                    ya2taxable,
                    ya2taxDef,
                    ya2taxFree,
                    n + 1,
                    wdrlRatio,
                    self.names,
                    self.y2ages,
                    True,
                )

                u.vprint(
                    'Performed withdrawal of',
                    d(total),
                    'using split of',
                    '{:.2f}'.format(wdrlRatio),
                )

                txfree = amounts['taxable'][0] + amounts['tax-free'][0]
                txbl = amounts['tax-def'][0]
                ytaxableIncome[n] += txbl
                ys2dist[n][:] += amounts['tax-def'][1:]
                ys2txfree[n][:] += amounts['tax-free'][1:]
                ys2txbl[n][:] += amounts['taxable'][1:]
                ytaxfreeIncome[n] += txfree
                yincomeTax[n] = estimatedTax
                ynetIncome[n] = ytaxfreeIncome[n] + ytaxableIncome[n] - yincomeTax[n]
                ygrossIncome[n] = ytaxableIncome[n] + yRothX[n] + btiEvent
                # IRMAA looks back 2 years for income.
                irmaaIncome = ygrossIncome[max(0, n - 2)]
                yirmaa[n] = tx.irmaa(
                    irmaaIncome, filingStatus, self.yyear[n], self.rates
                )
                u.vprint('\t...of which', d(txbl), 'is taxable.')
                u.vprint(
                    'Adj. Income:\n Gross taxable:',
                    d(ygrossIncome[n]),
                    'Tax bill:',
                    d(yincomeTax[n]),
                    'IRMAA:',
                    d(yirmaa[n]),
                    '\n Net:',
                    d(ynetIncome[n]),
                    'Tax free:',
                    d(ytaxfreeIncome[n]),
                )

            # Now check if anyone passed? Then transfer wealth at year-end.
            for j in range(self.count):
                if n == self.horizons[j] - 1:
                    u.vprint(self.names[j], 'has passed.')
                    surviving -= 1
                    if surviving == 0:
                        if self.count == 2:
                            u.vprint('Both spouses have passed.')
                        if verbose is not None:
                            u.setVerbose(prevState)
                        return

                    self._transferWealth(n + 1, j)

                    # Split becomes binary at death of one spouse.
                    wdrlRatio = j
                    depRatio = j
                    filingStatus = 'single'
                    # Reduce target income.
                    u.vprint(
                        'Reducing net income to',
                        pc(self.survivorFraction, f=0),
                        'of original target',
                    )
                    rawTarget *= self.survivorFraction

            if not self.success:
                u.vprint('==================================================')
                u.vprint('Aborting scenario early due to account exhaustion.')
                u.vprint('==================================================')
                break

        if verbose is not None:
            u.setVerbose(prevState)

        return

    def showAssetDistribution(self, tag=''):
        '''
        Plot the distribution of each savings account in thousands of dollars
        during the simulation time. This function will generate three
        graphs, one for taxable accounts, one the tax-deferred accounts,
        and one for tax-free accounts.

        A tag string can be set to add information to the title of the plot.
        '''

        y2stack = {}
        assetDic = {'stocks': 0, 'C bonds': 1, 'T notes': 2, 'common': 3}
        for acType in ['taxable', 'tax-deferred', 'tax-free']:
            stackNames = []
            for key in assetDic:
                name = key + ' / ' + acType
                stackNames.append(name)
                y2stack[name] = np.zeros((self.count, self.span))
                for i in range(self.count):
                    y2stack[name][i][:] = (
                        self.y2accounts[acType].transpose()[i][:]
                        * self.y2assetRatios[acType].transpose(1, 2, 0)[i][
                            assetDic[key]
                        ][:]
                    )
                y2stack[name] = y2stack[name].transpose()

            title = self._name + '\nAssets Distribution - ' + acType
            if tag != '':
                title += ' - ' + tag

            self._stackPlot(title, range(self.count), y2stack, stackNames, 'upper left')

        return

    def showAllocations(self, tag=''):
        '''
        Plot desired allocation of savings accounts in percentage
        over simulation time and interpolated by the selected method
        through the interpolateAR() method.

        A tag string can be set to add information to the title of the plot.
        '''
        count = self.count
        if self.coordinatedAR == 'both':
            acList = ['coordinated']
            count = 1
        elif self.coordinatedAR == 'individual':
            acList = ['coordinated']
        else:
            acList = ['taxable', 'tax-deferred', 'tax-free']

        assetDic = {'stocks': 0, 'C bonds': 1, 'T notes': 2, 'common': 3}
        for i in range(count):
            y2stack = {}
            for acType in acList:
                stackNames = []
                for key in assetDic:
                    aname = key + ' / ' + acType
                    stackNames.append(aname)
                    y2stack[aname] = np.zeros((count, self.span))
                    y2stack[aname][i][:] = self.y2assetRatios[acType].transpose(
                        1, 2, 0
                    )[i][assetDic[key]][:]
                    y2stack[aname] = y2stack[aname].transpose()

                    title = self._name + '\nAssets Allocations (%) - ' + acType
                    if self.coordinatedAR == 'both':
                        title += ' both'
                    else:
                        title += ' ' + self.names[i]

                if tag != '':
                    title += ' - ' + tag

                self._stackPlot(
                    title, [i], y2stack, stackNames, 'upper left', 'percent'
                )

        return

    def showAccounts(self, tag=''):
        '''
        Plot values of savings accounts over time.

        A tag string can be set to add information to the title of the plot.
        '''
        title = self._name + '\nSavings Balance'
        if tag != '':
            title += ' - ' + tag
        types = ['taxable', 'tax-deferred', 'tax-free']

        self._stackPlot(title, range(self.count), self.y2accounts, types, 'upper left')

        return

    def showSources(self, tag=''):
        '''
        Plot income over time.

        A tag string can be set to add information to the title of the plot.
        '''
        title = self._name + '\nRaw Income Sources'
        if tag != '':
            title += ' - ' + tag

        types = [
            'wages',
            'ssec',
            'pension',
            'dist',
            'rmd',
            'RothX',
            'div',
            'taxable',
            'tax-free',
        ]

        self._stackPlot(title, range(self.count), self.y2source, types, 'upper left')

        return

    def _stackPlot(self, title, irange, accounts, types, location, dtype='dollars'):
        '''
        Core function for stacked plots.
        '''
        import matplotlib.pyplot as plt
        import matplotlib.ticker as tk

        accountValues = {}
        for aType in types:
            for i in irange:
                tmp = accounts[aType].transpose()[i]
                if sum(tmp) > 0.01:
                    accountValues[aType + ' ' + self.names[i]] = tmp

        if len(accountValues) == 0:
            print('Nothing to plot for', title)
            return

        fig, ax = plt.subplots(figsize=(6, 4))
        plt.grid(visible='both')

        ax.stackplot(
            self.yyear, accountValues.values(), labels=accountValues.keys(), alpha=0.8
        )
        ax.legend(loc=location, reverse=True, fontsize=8, ncol=2, framealpha=0.7)
        # ax.legend(loc=location)
        ax.set_title(title)
        ax.set_xlabel('year')
        if dtype == 'dollars':
            ax.set_ylabel('k$')
            ax.get_yaxis().set_major_formatter(
                tk.FuncFormatter(lambda x, p: format(int(x / 1000), ','))
            )
        elif dtype == 'percent':
            ax.set_ylabel('%')
            ax.get_yaxis().set_major_formatter(
                tk.FuncFormatter(lambda x, p: format(int(100 * x), ','))
            )
        else:
            u.xprint('Unknown dtype:', dtype)

        # plt.show()
        return fig, ax

    def showNetIncome(self, tag=''):
        '''
        Plot net income and target over time.

        A tag string can be set to add information to the title of the plot.
        '''
        title = self._name + '\nNet Income vs. Target'
        if tag != '':
            title += ' - ' + tag

        style = {'net': '-', 'target': ':'}
        self._lineIncomePlot(style, title)

        return

    def _lineIncomePlot(self, style, title):
        '''
        Core line plotter function.
        '''
        import matplotlib.pyplot as plt
        import matplotlib.ticker as tk

        fig, ax = plt.subplots(figsize=(6, 4))
        plt.grid(visible='both')

        for aType in style:
            ax.plot(self.yyear, self.yincome[aType], label=aType, ls=style[aType])

        ax.legend(loc='upper left', reverse=True, fontsize=8, framealpha=0.7)
        # ax.legend(loc='upper left')
        ax.set_title(title)
        ax.set_xlabel('year')
        ax.set_ylabel('k$')
        ax.get_yaxis().set_major_formatter(
            tk.FuncFormatter(lambda x, p: format(int(x / 1000), ','))
        )

        # plt.show()
        return fig, ax

    def showGrossIncome(self, tag=''):
        '''
        Plot income tax and taxable income over time horizon.

        A tag string can be set to add information to the title of the plot.
        '''
        import matplotlib.pyplot as plt

        title = self._name + '\nGross Income vs. Tax Brackets'
        if tag != '':
            title += ' - ' + tag

        style = {'gross': '-'}

        fig, ax = self._lineIncomePlot(style, title)

        data = tx.taxBrackets(self.status, self.horizons, self.rates)

        '''
        myyears = np.array([2022, 2025, 2026, 2052])
        tax2428 = np.array([178000, 220000, 205000, 400000])
        tax3233 = np.array([340000, 405000, 310000, 600000])
        tax35 = np.array([432000, 500000, 580000, 1000000])
        '''

        for key in data:
            ax.plot(self.yyear, data[key], label=key, ls=':')

        plt.grid(visible='both')
        ax.legend(loc='upper left', reverse=True, fontsize=8, framealpha=0.7)
        # ax.legend(loc='upper left')

        return

    def showTaxes(self, tag=''):
        '''
        Plot income tax paid over time.

        A tag string can be set to add information to the title of the plot.
        '''
        title = self._name + '\nIncome Tax and IRMAA'
        if tag != '':
            title += ' - ' + tag

        data = {'irmaa': '-', 'taxes': '-'}

        fig, ax = self._lineIncomePlot(data, title)

        # plt.show()
        # return fig, ax
        return

    def showRates(self, tag=''):
        '''
        Plot rate values used over the time horizon.

        A tag string can be set to add information to the title of the plot.
        '''
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(6, 4))
        plt.grid(visible='both')
        title = self._name + '\nReturn & Inflation Rates (' + str(self.rateMethod)
        if self.rateMethod in ['historical', 'stochastic', 'average']:
            title += ' ' + str(self.rateFrm) + '-' + str(self.rateTo)
        elif self.rateMethod == 'fixed':
            title += str(self.rateMethod)
        title += ')'

        if tag != '':
            title += ' - ' + tag

        rateName = [
            'S&P500 including dividends',
            'Baa Corporate bonds',
            '10-y Treasury notes',
            'Inflation',
        ]
        ltype = ['-', '-.', ':', '--']
        for i in range(4):
            data = 100 * self.rates.transpose()[i]
            label = rateName[i] + ' <' + '{:.2f}'.format(np.mean(data)) + '>'
            ax.plot(self.yyear, data, label=label, ls=ltype[i % 4])

        ax.legend(loc='upper left', reverse=False, fontsize=8, framealpha=0.7)
        # ax.legend(loc='upper left')
        ax.set_title(title)
        ax.set_xlabel('year')
        ax.set_ylabel('%')

        # plt.show()
        # return fig, ax
        return

    def saveInstance(self, basename, overwrite=False):
        '''
        Save instance in an Excel spreadsheet.
        The first worksheet will contain income in the following
        fields in columns:
            - target income
            - net income
            - taxable income
            - gross income
            - tax bill (federal only)
            - IRMAA bill
        for all the years for the time span of the plan.

        The second worksheet contains the rates
        used for the plan as follows:
            - S&P 500
            - Corporate Baa bonds
            - Treasury notes (10y)
            - Inflation.

        The subsequent worksheets has the sources for each
        spouse as follows:
            - taxable account withdrawals
            - RMDs
            - distributions
            - Roth conversions
            - tax-free withdrawals
            - big-ticket items.

        The subsequent worksheets contains the balances
        of the savings accounts for each spouse:
            - taxable savings account
            - tax-deferred account
            - tax-free account.

        The last worksheets will contain the allocation
        ratios for each spouse:
            - S&P 500
            - Corporate Baa bonds
            - Treasury notes (10y)
            - inflation-indexed assets.

        '''
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        ws = wb.active
        ws.title = 'Income'

        rawData = {}
        rawData['year'] = self.yyear[:-1]
        incDic = {
            'target income': 'target',
            'net income': 'net',
            'taxable income': 'taxable',
            'gross income': 'gross',
            'tax bill': 'taxes',
            'IRMAA bill': 'irmaa',
        }
        for key in incDic:
            rawData[key] = self.yincome[incDic[key]][:-1]

        # We need to work by row.
        df = pd.DataFrame(rawData)
        for rows in dataframe_to_rows(df, index=False, header=True):
            ws.append(rows)

        _formatSpreadsheet(ws, 'currency')

        # Save rates on a different sheet.
        ws = wb.create_sheet('Rates')
        rawData = {}
        rawData['year'] = self.yyear[:-1]
        ratesDic = {'S&P 500': 0, 'Corporate Baa': 1, 'T Bonds': 2, 'inflation': 3}

        for key in ratesDic:
            rawData[key] = self.rates.transpose()[ratesDic[key]][:-1]

        # We need to work by row.
        df = pd.DataFrame(rawData)
        for rows in dataframe_to_rows(df, index=False, header=True):
            ws.append(rows)

        _formatSpreadsheet(ws, 'percent2')

        # Save sources.
        srcDic = {
            'txbl acc. wdrwl': 'taxable',
            'RMD': 'rmd',
            'distribution': 'dist',
            'Roth conversion': 'RothX',
            'tax-free wdrwl': 'tax-free',
            'big-ticket items': 'bti',
        }

        for i in range(self.count):
            sname = self.names[i] + '\'s Sources'
            ws = wb.create_sheet(sname)
            rawData = {}
            rawData['year'] = self.yyear[:-1]
            for key in srcDic:
                rawData[self.names[i] + ' ' + key] = self.y2source[
                    srcDic[key]
                ].transpose()[i][:-1]

            df = pd.DataFrame(rawData)
            for rows in dataframe_to_rows(df, index=False, header=True):
                ws.append(rows)

            _formatSpreadsheet(ws, 'currency')

        # Save account balances.
        for i in range(self.count):
            sname = self.names[i] + '\'s Accounts'
            ws = wb.create_sheet(sname)
            rawData = {}
            rawData['year'] = self.yyear[:-1]
            for acType in ['taxable', 'tax-deferred', 'tax-free']:
                rawData[self.names[i] + ' ' + acType] = self.y2accounts[
                    acType
                ].transpose()[i][:-1]
            df = pd.DataFrame(rawData)
            for rows in dataframe_to_rows(df, index=False, header=True):
                ws.append(rows)

            _formatSpreadsheet(ws, 'currency')

        # Save assets allocation ratios.
        astDic = {'S&P 500': 0, 'Corporate Baa': 1, 'T Bonds': 2, 'Common assets': 3}
        for i in range(self.count):
            sname = self.names[i] + '\'s AR'
            ws = wb.create_sheet(sname)
            rawData = {}
            rawData['year'] = self.yyear[:-1]
            for acType in ['taxable', 'tax-deferred', 'tax-free']:
                for ast in astDic:
                    rawData[
                        ast + ' / ' + self.names[i] + ' ' + acType
                    ] = self.y2assetRatios[acType].transpose(1, 2, 0)[i][astDic[ast]][
                        :-1
                    ]

            df = pd.DataFrame(rawData)
            for rows in dataframe_to_rows(df, index=False, header=True):
                ws.append(rows)

            _formatSpreadsheet(ws, 'percent0')

        _saveWorkbook(wb, basename, overwrite)

        return

    def saveInstanceCSV(self, basename):
        import pandas as pd

        planData = {}

        # Start with single entries.
        planData['year'] = self.year
        planData['target income'] = self.yincome['target']
        planData['net income'] = self.yincome['net']
        planData['tax bill'] = self.yincome['taxes']

        for i in range(self.count):
            planData[self.names[i] + ' txbl acc. wrdwl'] = self.y2source[
                'taxable'
            ].transpose()[i]
            planData[self.names[i] + ' RMD'] = self.y2source['rmd'].transpose()[i]
            planData[self.names[i] + ' distribution'] = self.y2source[
                'dist'
            ].transpose()[i]
            planData[self.names[i] + ' Roth conversion'] = self.y2source[
                'RothX'
            ].transpose()[i]
            planData[self.names[i] + ' tax-free wdrwl'] = self.y2source[
                'tax-free'
            ].transpose()[i]
            planData[self.names[i] + ' big-ticket items'] = self.y2source[
                'bti'
            ].transpose()[i]

        df = pd.DataFrame(planData)

        while True:
            try:
                fname = 'plan' + '_' + basename + '.csv'
                df.to_csv(fname)
                # Requires xlwt which is obsolete
                # df.to_excel(fname)
                break
            except PermissionError:
                print('Failed to save', fname, '. Permission denied.')
                key = input('Try again? [Yn] ')
                if key == 'n':
                    break
            except Exception:
                u.xprint('Unanticipated exception', Exception)

        return

    def showAndSave(self, filename=None):
        '''
        Final statement for scripts desiring to save events in excel file.
        '''
        import matplotlib.pyplot as plt
        import os.path as path

        plt.show(block=False)
        plt.pause(0.001)
        while True:
            key = input("[Enter] 'r' to repeat, 's' to save, or 'x' to exit: ")
            if key == 'r':
                plt.close("all")
                break
            elif key == 's':
                if filename is None:
                    filename = path.basename(sys.argv[0][:-3])
                self.saveInstance(filename)
                break
            elif key == 'x':
                sys.exit(0)

        return

    def show(self, pause=2, block=False):
        '''
        Use to show graphs between runs.
        '''
        import matplotlib.pyplot as plt

        plt.show(block=block)
        if isInJupyter() is False:
            plt.pause(pause)

        plt.close('all')

        return

    def _estate(self, taxRate):
        '''
        Back-end method to estate() method.
        '''

        total = sum(self.y2accounts['taxable'][-1][:])
        total += sum(self.y2accounts['tax-free'][-1][:])
        total += (1 - taxRate) * sum(self.y2accounts['tax-deferred'][-1][:])

        div = tx.inflationAdjusted(1.0, self.span - 1, self.rates)
        value = total / div

        return value, (div - 1)

    def estate(self, taxRate=None):
        '''
        Return estimated post-tax value of total of assets at
        the end of the run in today's $. The heirs tax rate provided
        is used to determine an approximate value and provide
        the tax-deferred portion of the portfolio an after-tax
        value. The inflation rates used during the simulation
        are re-used to bring the net value in today's $.
        Cumulative inflation factor is returned as well as the
        estate value.
        '''
        if taxRate is None:
            taxRate = self.heirsTxRate
        else:
            taxRate /= 100

        val, percent = self._estate(taxRate)

        now = datetime.date.today().year
        print(
            '%r' % self._name,
            self.yyear[-2],
            'Estate: (%d $) %s (nominal %s),' % (now, d(val), d(val * (1 + percent))),
            '\n\t\tcum. infl.: %s, heirs tax rate: %s' % (pc(percent), pc(taxRate)),
        )

        return

    def _saveConfig(self, fileName):
        '''
        Save plan parameters in a configuration file.
        '''
        import configparser

        config = configparser.ConfigParser()

        config['Who'] = {
            'Count': str(self.count),
            'Names': ','.join(str(k) for k in self.names),
        }

        # Parameters getting one value for each spouse.
        config['YOB'] = {}
        config['Life expectancy'] = {}
        config['Beneficiary'] = {}
        config['Pension amounts'] = {}
        config['Pension ages'] = {}
        config['Social security amounts'] = {}
        config['Social security ages'] = {}
        config['Asset balances'] = {}
        config['Initial allocation ratios'] = {}
        config['Final allocation ratios'] = {}

        if self.coordinatedAR == 'none':
            listAR = ['taxable', 'tax-deferred', 'tax-free']
        else:
            listAR = ['coordinated']

        for i in range(self.count):
            config['YOB'][self.names[i]] = str(self.yob[i])
            config['Life expectancy'][self.names[i]] = str(self.expectancy[i])
            config['Beneficiary'][self.names[i]] = str(self.beneficiary[i])
            config['Pension amounts'][self.names[i]] = str(self.pensionAmount[i])
            config['Pension ages'][self.names[i]] = str(self.pensionAge[i])
            config['Social security amounts'][self.names[i]] = str(self.ssecAmount[i])
            config['Social security ages'][self.names[i]] = str(self.ssecAge[i])
            for aType in ['taxable', 'tax-deferred', 'tax-free']:
                config['Asset balances'][aType + ' ' + self.names[i]] = str(
                    self.n2balances[aType][i]
                )
            for aType in listAR:
                config['Initial allocation ratios'][
                    aType + ' ' + self.names[i]
                ] = ', '.join(str(100 * k) for k in self.boundsAR[aType][0][i])
                config['Final allocation ratios'][
                    aType + ' ' + self.names[i]
                ] = ', '.join(str(100 * k) for k in self.boundsAR[aType][1][i])

        # Joint parameters.
        config['Parameters'] = {
            'Name': self._name,
            'Target': str(self.target),
            'Profile': str(self.profile),
            'Heirs rate on tax-deferred estate': str(100 * self.heirsTxRate),
            'Spousal split': str(self.split),
            'Survivor fraction': str(self.survivorFraction),
            'Time lists file name': str(self.timeListsFileName),
            'Coordinated allocations': str(self.coordinatedAR),
        }

        config['Rates'] = {
            'Method': str(self.rateMethod),
            'From': str(self.rateFrm),
            'To': str(self.rateTo),
        }
        if self.rateMethod == 'fixed':
            config['Rates']['values'] = ', '.join(str(k) for k in self.rateValues)

        with open(fileName + '.cfg', 'w') as configfile:
            config.write(configfile)

        return

    def _runOnce(
        self, stype, frm=rates.FROM, to=rates.TO, rates=None, myplots=[], tag=''
    ):
        '''
        Run one instance of a simulation.
        '''
        self.reset()
        self.setRates(stype, frm, to, rates)

        self.run()

        plotDic = {
            'rates': self.showRates,
            'net income': self.showNetIncome,
            'sources': self.showSources,
            'taxes': self.showTaxes,
            'gross income': self.showGrossIncome,
            'accounts': self.showAccounts,
            'distribution': self.showAssetDistribution,
        }

        for pl in myplots:
            plotDic[pl](tag)

        return self

    def runHistorical(self, frm, to, myplots=[], tag=''):
        '''
        Run historical simulation from each year in the rates provided.
        '''
        to = frm + self.span - 1
        N = rates.TO - self.span - frm + 2
        now = datetime.date.today().year
        estateResults = np.zeros(N)
        successCount = 0
        for i in range(N):
            print('--------------------------------------------')
            print('Running case # %d (%r)' % (i, frm + i))
            self._runOnce('historical', frm + i, to + i, myplots=myplots, tag=tag)
            print('Plan success:', self.success)
            if self.success:
                successCount += 1

            # Use heirs tax rate provided on taxable part of estate.
            val, percent = self._estate(self.heirsTxRate)
            print(
                '%r' % self._name,
                self.yyear[-2],
                'Estate: (%d $) %s (nominal %s),'
                % (now, d(val), d(val * (1 + percent))),
                '\n\t\tcum. infl.: %s, heirs tax rate: %s'
                % (pc(percent), pc(self.heirsTxRate)),
            )
            estateResults[i] = val
            if len(myplots) > 0:
                # Number of seconds to wait.
                # For embedding in jupyter set to a low value.
                # plan.show(0.000001)
                self.show(2)
                # plan.showAndSave()

        print('============================================')
        print(
            'Success rate: %d out of %d (%s)' % (successCount, N, pc(successCount / N))
        )
        print('Median estate value (%d $): %s' % (now, d(np.median(estateResults))))

        showHistogram(estateResults, binDiv=4)

        return

    def runMonteCarlo(self, N, frm=rates.FROM, to=rates.TO, myplots=[]):
        '''
        Run N simulations using a stochastic sinulation.
        '''
        prevState = u.setVerbose(False)
        estateResults = np.zeros(N)
        successCount = 0
        for i in range(N):
            print('--------------------------------------------')
            print('Running case #', i)
            self._runOnce('stochastic', 1940, 2023, myplots=myplots)
            print('Plan success:', self.success)
            if self.success:
                successCount += 1

            # Rely on self.heirsTxRate for rate.
            estate, factor = self._estate(self.heirsTxRate)
            print(
                self.yyear[-2],
                'Estate: (today\'s $)',
                d(estate),
                '\n\tcum. infl.:',
                pc(factor),
                ', heirs tax rate:',
                pc(self.heirsTxRate),
            )
            estateResults[i] = estate

        print('============================================')
        print(
            'Success rate:', successCount, 'out of', N, '(' + pc(successCount / N) + ')'
        )
        print('Median estate value (today\'s $): ', d(np.median(estateResults)))

        showHistogram(estateResults, binDiv=5)
        u.setVerbose(prevState)

        return


######################################################################
def d(value, f=0):
    '''
    Return a string formatting number in $ currency.
    '''
    mystr = '${:,.' + str(f) + 'f}'

    return mystr.format(value)


def pc(value, f=1, mul=100):
    '''
    Return a string formatting number in percent.
    '''
    mystr = '{:.' + str(f) + 'f}%'

    return mystr.format(mul * value)


def _rescale(vals, fac):
    '''
    Rescale elements of a list by factor fac.
    '''
    for i in range(len(vals)):
        vals[i] *= fac

    return


def _getUnits(units):
    '''
    Return proper factor for units.
    '''
    if units is None:
        fac = 1
    elif units in {'k', 'K'}:
        fac = 1000
    elif units in {'m', 'M'}:
        fac = 1000000
    else:
        u.xprint('Unknown units', units)

    return fac


def _formatSpreadsheet(ws, ftype):
    '''
    Utility function to beautify spreadsheet.
    '''
    if ftype == 'currency':
        fstring = u'$#,##0_);[Red]($#,##0)'
    elif ftype == 'percent2':
        fstring = u'#.00%'
    elif ftype == 'percent0':
        fstring = u'#0%'
    else:
        u.xprint('Unknown format:', ftype)

    for cell in ws[1] + ws['A']:
        cell.style = 'Pandas'
    for col in ws.columns:
        column = col[0].column_letter
        # col[0].style = 'Title'
        width = len(str(col[0].value)) + 4
        ws.column_dimensions[column].width = width
        if column != 'A':
            for cell in col:
                cell.number_format = fstring

    return


def savePlan(plan, fileName):
    '''
    Save plan configuration parameters to a file.
    '''
    u.vprint('Saving plan config as', fileName + '.cfg')

    plan._saveConfig(fileName)

    return


def readPlan(fileName):
    '''
    Read plan configuration parameters from a file.
    '''
    import configparser

    u.vprint('Reading plan config from', fileName + '.cfg')

    config = configparser.ConfigParser()
    config.read(fileName + '.cfg')

    count = int(config['Who']['count'])
    names = config['Who']['names'].split(',')
    coordinatedAR = config['Parameters']['coordinated allocations']
    name = config['Parameters']['name']

    # Parameters getting one value for each spouse.
    yob = []
    expectancy = []
    beneficiary = []
    pensionAmounts = []
    pensionAges = []
    ssecAmounts = []
    ssecAges = []

    n2balances = {}
    initialAR = {}
    finalAR = {}

    if coordinatedAR == 'none':
        listAR = ['taxable', 'tax-deferred', 'tax-free']
    else:
        listAR = ['coordinated']

    for aType in ['taxable', 'tax-deferred', 'tax-free']:
        n2balances[aType] = []

    for aType in listAR:
        initialAR[aType] = []
        finalAR[aType] = []

    for i in range(count):
        yob.append(int(config['YOB'][names[i]]))
        expectancy.append(int(config['Life expectancy'][names[i]]))
        beneficiary.append(float(config['Beneficiary'][names[i]]))
        pensionAmounts.append(float(config['Pension amounts'][names[i]]))
        pensionAges.append(int(config['Pension ages'][names[i]]))
        ssecAmounts.append(int(config['Social security amounts'][names[i]]))
        ssecAges.append(int(config['Social security ages'][names[i]]))
        for aType in ['taxable', 'tax-deferred', 'tax-free']:
            n2balances[aType].append(
                float(config['Asset balances'][aType + ' ' + names[i]])
            )
        for aType in listAR:
            initialAR[aType].append(
                config['Initial allocation ratios'][aType + ' ' + names[i]].split(',')
            )
            finalAR[aType].append(
                config['Final allocation ratios'][aType + ' ' + names[i]].split(',')
            )

    # Convert those strings to float.
    for aType in listAR:
        for k in range(len(initialAR[aType])):
            initialAR[aType][k] = [float(j) for j in initialAR[aType][k]]
            finalAR[aType][k] = [float(j) for j in finalAR[aType][k]]

    plan = Plan(yob, expectancy)
    plan.setName(name)
    plan.setPension(pensionAmounts, pensionAges)
    plan.setSocialSecurity(ssecAmounts, ssecAges)
    plan.setAssetBalances(
        taxable=n2balances['taxable'],
        taxDeferred=n2balances['tax-deferred'],
        taxFree=n2balances['tax-free'],
        beneficiary=beneficiary,
    )

    if coordinatedAR == 'none':
        plan.setInitialAR(
            taxable=initialAR['taxable'],
            taxDeferred=initialAR['tax-deferred'],
            taxFree=initialAR['tax-free'],
        )
        plan.setFinalAR(
            taxable=finalAR['taxable'],
            taxDeferred=finalAR['tax-deferred'],
            taxFree=finalAR['tax-free'],
        )
    elif coordinatedAR == 'individual':
        plan.setCoordinatedAR(
            initial=initialAR['coordinated'], final=finalAR['coordinated']
        )
    elif coordinatedAR == 'both':
        plan.setCoordinatedAR(
            initial=initialAR['coordinated'][0], final=finalAR['coordinated'][0]
        )
    else:
        u.xprint('Unknown coordination type:', coordinatedAR)

    plan.interpolateAR()

    plan.setDesiredIncome(
        float(config['Parameters']['Target']), config['Parameters']['Profile']
    )
    plan.setHeirsTaxRate(
        float(config['Parameters']['Heirs rate on tax-deferred estate'])
    )
    plan.setSpousalSplit(config['Parameters']['Spousal split'])
    plan.setSurvivorFraction(float(config['Parameters']['Survivor fraction']))

    timeListsFileName = config['Parameters']['Time lists file name']
    plan.readContributions(timeListsFileName)

    method = config['Rates']['Method']
    frm = int(config['Rates']['From'])
    to = int(config['Rates']['To'])
    values = None
    if method == 'fixed':
        values = config['Rates']['values'].split(',')
        values = np.array([float(j) for j in values])

    plan.setRates(method, frm, to, values)

    return plan


def _saveWorkbook(wb, basename, overwrite=False):
    '''
    Utility function to save XL workbook.
    '''
    import os.path as path

    fname = 'instance' + '_' + basename + '.xlsx'

    if overwrite is False and path.isfile(fname):
        print('File ', fname, ' already exists.')
        key = input('Overwrite? [Ny] ')
        if key != 'y':
            print('Skipping save and returning.')
            return

    while True:
        try:
            u.vprint('Saving plan as', fname)
            wb.save(fname)
            break
        except PermissionError:
            print('Failed to save', fname, '. Permission denied.')
            key = input('Try again? [Yn] ')
            if key == 'n':
                break
        except Exception:
            u.xprint('Unanticipated exception', Exception)

    return


def _pfReturn(assetRatios, rates, year, who):
    '''
    Return annual rate of return depending on portfolio asset ratio.

    For the first year, keep fraction of days remaining in the year.
    '''
    # Only consider the time remaining this year.
    if year == 0:
        fac = 1 - datetime.datetime.today().timetuple().tm_yday / 365
    else:
        fac = 1

    return (
        assetRatios[year][who][0] * fac * rates[year][0]
        + assetRatios[year][who][1] * fac * rates[year][1]
        + assetRatios[year][who][2] * fac * rates[year][2]
        + assetRatios[year][who][3] * fac * rates[year][3]
    )


def _age(yob, refYear=0):
    '''
    Return age of individual in year provided. If no year
    is provided, current year will be used.
    '''
    if refYear == 0:
        refYear = datetime.date.today().year

    assert refYear >= yob

    return refYear - yob


def _spendingAdjustment(age, profile='flat'):
    '''
    Return spending profile for age provided.
    Profile can be 'flat' or 'smile'.
    '''
    assert age <= 100
    # Return 1 for age before 65 of flat profile.
    if age <= 65 or profile == 'flat':
        return 1.0

    if profile == 'smile':
        # From the gogo years to the no-go years.
        table = [
            1.000,
            1.010,
            1.015,
            1.010,
            1.000,
            0.993,
            0.978,
            0.960,
            0.940,
            0.918,
            0.895,
            0.871,
            0.848,
            0.825,
            0.804,
            0.785,
            0.769,
            0.757,
            0.748,
            0.744,
            0.745,
            0.752,
            0.766,
            0.787,
            0.815,
            0.852,
            0.899,
            0.955,
            1.021,
            1.059,
            1.100,
            1.121,
            1.141,
            1.151,
            1.161,
            1.171,
        ]
        return table[age - 65]

    u.xprint('In spendingAdjustment: Unknown profile', profile)


def _readTimeLists(filename, n):
    '''
    Read listed parameters from an excel spreadsheet through pandas.
    Use one sheet for each individual with the following columns.
    Supports xls, xlsx, xlsm, xlsb, odf, ods, and odt file extensions.
    '''
    import pandas as pd

    # Expected headers in each excel sheet, one per individual.
    timeHorizonItems = [
        'year',
        'anticipated wages',
        'ctrb taxable',
        'ctrb 401k',
        'ctrb Roth 401k',
        'ctrb IRA',
        'ctrb Roth IRA',
        'Roth X',
        'big ticket items',
    ]

    timeLists = []
    names = []
    now = datetime.date.today().year
    # Read all worksheets in memory but only process first n.
    dfDict = pd.read_excel(filename, sheet_name=None)
    i = 0
    for name in dfDict.keys():
        u.vprint('Reading time horizon for', name, '...')
        names.append(name)
        # Only consider lines after this year.
        dfDict[name] = dfDict[name][dfDict[name]['year'] >= now]
        # Replace empty (NaN) cells with 0 value.
        dfDict[name].fillna(0, inplace=True)

        timeLists.append({})
        # Transfer values from dataframe to lists
        for item in timeHorizonItems:
            timeLists[i][item] = dfDict[name][item].tolist()

        i += 1
        if i >= n:
            break

    u.vprint('Successfully read time horizons from file', filename)

    return names, timeLists


def _checkTimeLists(names, timeLists, horizons):
    '''
    Make sure that time horizons contain all years up to life expectancy.
    '''
    if len(names) == 2:
        # Verify that both sheets start on the same year.
        if timeLists[0]['year'][0] != timeLists[1]['year'][0]:
            u.xprint('Time horizons not starting on same year.')

    # Verify that year range covers life expectancy for each individual
    now = datetime.date.today().year
    for i in range(len(names)):
        yend = now + horizons[i]
        if timeLists[i]['year'][-1] < yend:
            u.xprint(
                'Time horizon for',
                names[i],
                'is too short.\n\tIt should end in',
                yend,
                'but ends in',
                timeLists[i]['year'][-1],
            )

    timeHorizonItems = [
        'year',
        'anticipated wages',
        'ctrb taxable',
        'ctrb 401k',
        'ctrb Roth 401k',
        'ctrb IRA',
        'ctrb Roth IRA',
        'Roth X',
    ]

    # Verify that all numbers except bti are positive.
    for i in range(len(names)):
        for n in range(horizons[i]):
            for item in timeHorizonItems:
                assert timeLists[i][item][i] >= 0

    return


def _balance(c, X, Y, Z):
    '''
    Core function to coordinate assets allocation ratios amongst
    different accounts.
    X is taxable, Y is tax-deferred, and Z tax-free are balances.
    x, y, and z are corresponding allocation ratios.

    For c[], this is the overall desired allocation:
    0 is stock, 1 C bonds, 2 T notes, and 3 common assets are ratios.
    '''
    x = np.zeros((4))
    y = np.zeros((4))
    z = np.zeros((4))
    # Total balance.
    T = X + Y + Z

    if T < 0.01:
        return x, y, z

    # Maximize stocks in tax-free account, followed by tax-deferred account.
    z[0] = c[0] * T / (Z + 0.01)
    z[0] = min(z[0], 1.0)
    y[0] = (c[0] * T - z[0] * Z) / (Y + 0.01)
    y[0] = min(y[0], 1.0)
    x[0] = (c[0] * T - y[0] * Y - z[0] * Z) / (X + 0.01)

    # Maximize bonds in tax-free account, followed by tax-deferred account.
    z[1] = c[1] * T / (Z + 0.01)
    z[1] = min(z[1], 1.0 - z[0])
    y[1] = (c[1] * T - z[1] * Z) / (Y + 0.01)
    y[1] = min(y[1], 1.0 - y[0])
    x[1] = (c[1] * T - y[1] * Y - z[1] * Z) / (X + 0.01)
    x[1] = min(x[1], 1.0 - x[0])

    # Maximize fixed assets in taxable account.
    x[3] = c[3] * T / (X + 0.01)
    x[3] = min(x[3], 1.0 - x[0] - x[1])
    y[3] = (c[3] * T - x[3] * X) / (Y + 0.01)
    y[3] = min(y[3], 1.0 - y[0] - y[1])
    z[3] = (c[3] * T - x[3] * X - y[3] * Y) / (Z + 0.01)
    z[3] = min(z[3], 1.0 - z[0] - z[1])

    # Treasury notes get the rest.
    if X > 0.01:
        x[2] = 1.0 - sum(x)
    if Y > 0.01:
        y[2] = 1.0 - sum(y)
    if Z > 0.01:
        z[2] = 1.0 - sum(z)

    return x, y, z


def _smartBanking(
    amount, taxable, taxdef, taxfree, year, wdrlRatio, names, ages, commit=True
):
    '''
    Deposit/withdraw amount from given accounts. Return dictionary
    of lists itemizing amount taken from taxable, tax-deferred,
    and tax-free accounts, itemized by total and then spousal accounts.
    Total amount from all accounts is second value returned.
    If commit is False, amounts are calculated without changing
    the account values. Withdrawal ratio x controls relative amount
    taken from respective spousal accounts, with reference to first
    spouse (other is 1-x).
    '''
    assert 0 <= wdrlRatio and wdrlRatio <= 1.0
    amounts = {'taxable': [], 'tax-def': [], 'tax-free': []}
    totAmount = 0
    for i in range(len(names)):
        subAmount = (wdrlRatio - 2 * i * wdrlRatio + i) * amount
        itemized = _smartBankingSub(
            subAmount, taxable, taxdef, taxfree, year, i, names, ages[year], commit
        )

        amounts['taxable'].append(itemized[0])
        amounts['tax-def'].append(itemized[1])
        amounts['tax-free'].append(itemized[2])
        totAmount += itemized[3]

    # Store per-account total in first list entry.
    for numlist in amounts.values():
        numlist.insert(0, sum(numlist[:]))

    return amounts, totAmount


def _smartBankingSub(
    amount, taxable, taxdef, taxfree, year, i, names, ages, commit=True
):
    '''
    Deposit/withdraw amount from given accounts.
    Return amounts withdrawn from relative accounts:
    taxable, tax-deferred, and tax-free accounts and total withdrawn.
    If commit is False, amounts are calculated without changing account values.
    '''
    if amount == 0:
        return [0, 0, 0, 0]

    # Is this a deposit? Put it in taxable account.
    if amount > 0:
        if commit:
            taxable[year][i] += amount
        return [0, 0, 0, amount]

    # This is a withdrawal. Change sign and start from taxable account.
    withdrawal = abs(amount)
    if withdrawal <= taxable[year][i]:
        if commit:
            taxable[year][i] -= withdrawal
        return [withdrawal, 0, 0, withdrawal]

    portion1 = taxable[year][i]
    if commit:
        taxable[year][i] = 0
    remain = withdrawal - portion1

    # Then go through other accounts. First tax-deferred.
    if remain <= taxdef[year][i]:
        if commit:
            _checkPenalty('tax-deferred', remain, names, ages, i)
            taxdef[year][i] -= remain
        return [portion1, remain, 0, withdrawal]

    portion2 = taxdef[year][i]
    if commit:
        _checkPenalty('tax-deferred', portion2, names, ages, i)
        taxdef[year][i] = 0
    remain -= portion2

    # Then tax-free account. Only portion2 is taxable.
    if remain <= taxfree[year][i]:
        if commit:
            _checkPenalty('tax-free', remain, names, ages, i)
            taxfree[year][i] -= remain
        return [portion1, portion2, remain, withdrawal]

    portion3 = taxfree[year][i]
    if commit:
        _checkPenalty('tax-free', portion3, names, ages, i)
        taxfree[year][i] = 0
    remain -= portion3

    if commit:
        u.vprint('WARNING: Withdrawal of', d(amount), 'in year', year, 'for', names[i])
        u.vprint('         short of', d(remain), 'as all accounts were exhausted!')

    return [portion1, portion2, portion3, withdrawal - remain]


def _checkPenalty(account, amount, names, ages, i):
    '''
    Withdrawal from savings accounts before the age of 59 1/2
    will trigger a 10% penalty. Check for the age.
    '''
    if ages[i] < 60:
        u.vprint(
            'WARNING: Withdrawal of',
            d(amount),
            'from',
            names[i],
            account,
            'account will trigger a 10% penalty!',
        )

    return


def showHistogram(data, tag='', binDiv=10):
    '''
    Plot estate results of a Monte Carlo simulation.
    '''
    import matplotlib.pyplot as plt
    import matplotlib.ticker as tk

    nbins = int(len(data) / binDiv)
    fig, ax = plt.subplots(tight_layout=True)

    title = 'Estate Value Distribution'
    if tag != '':
        title += ' - ' + tag

    ax.set_title(title)
    label = 'median: ' + d(np.median(data))
    ax.hist(data, bins=nbins, label=label)
    ax.set_ylabel('N')
    ax.legend(loc='upper right', reverse=False, fontsize=8, framealpha=0.7)
    ax.set_xlabel('today\'s k$')
    ax.get_xaxis().set_major_formatter(
        tk.FuncFormatter(lambda x, p: format(int(x / 1000), ','))
    )

    plt.show()

    return


def showRateDistributions(frm=rates.FROM, to=rates.TO):
    '''
    Plot histograms of the rates distributions.
    '''
    import matplotlib.pyplot as plt

    title = 'Rates from ' + str(frm) + ' to ' + str(to)
    # Bring year values to indices.
    frm -= rates.FROM
    to -= rates.FROM

    nbins = int((to - frm) / 4)
    fig, ax = plt.subplots(1, 4, sharey=True, sharex=True, tight_layout=True)

    dat0 = np.array(rates.SP500[frm:to])
    dat1 = np.array(rates.BondsBaa[frm:to])
    dat2 = np.array(rates.TNotes[frm:to])
    dat3 = np.array(rates.Inflation[frm:to])

    fig.suptitle(title)
    ax[0].set_title('S&P500')
    label = '<>: ' + pc(np.mean(dat0), 2, 1)
    ax[0].hist(dat0, bins=nbins, label=label)
    ax[0].legend(loc='upper left', fontsize=8, framealpha=0.7)

    ax[1].set_title('BondsBaa')
    label = '<>: ' + pc(np.mean(dat1), 2, 1)
    ax[1].hist(dat1, bins=nbins, label=label)
    ax[1].legend(loc='upper left', fontsize=8, framealpha=0.7)

    ax[2].set_title('TNotes')
    label = '<>: ' + pc(np.mean(dat2), 2, 1)
    ax[2].hist(dat1, bins=nbins, label=label)
    ax[2].legend(loc='upper left', fontsize=8, framealpha=0.7)

    ax[3].set_title('Inflation')
    label = '<>: ' + pc(np.mean(dat3), 2, 1)
    ax[3].hist(dat3, bins=nbins, label=label)
    ax[3].legend(loc='upper left', fontsize=8, framealpha=0.7)

    plt.show()

    return fig, ax


def _printDot(trials):
    '''
    Print a dot for every 100 trials.
    '''
    if trials % 100 == 0:
        print('.', end='', flush=True)
        if trials % 1000 == 0:
            print()


def _amountAnnealRoth(p2, baseValue, txrate, minConv, startConv, only):
    '''
    Determine best Roth conversions through mixed sweeping approach,
    starting from large conversions being reduced by half over the
    simulation. Starting conversion amount is startConv.
    Minimum conversion considered is minConv.
    For married couples, the *only* parameter is used to restrict
    conversions to the individual named.
    '''
    import random

    random.seed()
    # Make conversions proportional to first-year balances for spouses.
    ratio = [1, 0]

    if p2.count == 2:
        ratio[0] = p2.y2accounts['tax-deferred'][0][0] / sum(
            p2.y2accounts['tax-deferred'][0][:]
        )
        ratio[1] = 1 - ratio[0]

    print(
        'Starting Roth optimizer.',
        '\n\tThis calculation typically takes less than a few minutes.',
    )
    print('\tEach dot represents 100 different scenarios tested:')

    # Allow all to make conversions.
    if only is None:
        whos = range(p2.count)
    else:
        who = p2.names.index(only)
        whos = range(who, who + 1)

    myConv = startConv
    maxValue = baseValue
    bestX = np.zeros((p2.span, p2.count), dtype=int)
    trials = 0
    blank = 0
    i = 0
    who = whos[i]
    while myConv >= minConv:
        nMax = -1
        rothXmax = 0
        loopMax = maxValue
        # Don't allow conversions during last year alive.
        for n in range(p2.horizons[who]):
            rothX = tx.inflationAdjusted(myConv * ratio[who], n, p2.rates)

            if rothX > p2.y2accounts['tax-deferred'][n][who]:
                break

            p2.timeLists[who]['Roth X'][n] += rothX
            p2.run()
            newValue, mul2 = p2._estate(txrate)
            p2.timeLists[who]['Roth X'][n] -= rothX

            if newValue > loopMax:
                loopMax = newValue
                nMax = n
                rothXmax = rothX

            trials += 1
            _printDot(trials)

        if nMax >= 0:
            maxValue = loopMax
            p2.timeLists[who]['Roth X'][nMax] += rothXmax
            bestX[nMax][who] += rothXmax
        else:
            blank += 1

        # Alternating between individuals if needed.
        i = (i + 1) % len(whos)
        who = whos[i]

        # If nothing happened during last round:
        # reduce conversion amount and look for overshoots.
        if blank == len(whos):
            myConv /= 2
            blank = 0

            # Backtracking - remove overshoots.
            success = True
            while success:
                success = False
                for j in whos:
                    for n in reversed(range(p2.horizons[j])):
                        rothX = tx.inflationAdjusted(myConv * ratio[j], n, p2.rates)

                        if rothX < p2.timeLists[j]['Roth X'][n]:
                            p2.timeLists[j]['Roth X'][n] -= rothX

                            p2.run()
                            newValue, mul2 = p2._estate(txrate)

                            if newValue >= maxValue:
                                maxValue = newValue
                                success = True
                            else:
                                p2.timeLists[j]['Roth X'][n] += rothX

                            trials += 1
                            _printDot(trials)

    print('\nReturning after', trials, 'trials.')

    return bestX


def optimizeRoth(p, txrate, minConv=10, startConv=64000, only=None):
    '''
    Determines optimal Roth conversions.
    Goal is to maximize estate given a tax-deferred heirs tax rate.
    A heirs tax rate for the tax-deferred account must be provided.
    Then, startConv is the largest conversion amount which is
    considered at the beginning. When no more wealth-positive
    conversion can be made, the the amount is reduced by half
    and conversions are adjusted by the new half-amount until no
    more wealth-positive conversions can be made.

    Due to this division, it is better to choose a number for
    startConv that is related to a power of 2. Choosing 64,000
    or 128,000 for example.

    For married couples, the *only* parameter is used to restrict
    conversions to the individual named, as not all plans allow for
    Roth conversions.
    '''
    p2 = clone(p)
    p2.setName(p._name + ' (Roth optimized)')

    txrate /= 100

    prevState = u.setVerbose(False)

    # Start by zeroing all RothX in cloned plan.
    p2.clearRothX()
    p2.run()
    baseValue, mul = p2._estate(txrate)

    bestX = _amountAnnealRoth(p2, baseValue, txrate, minConv, startConv, only)
    p2.run()
    newValue, mul = p2._estate(txrate)
    now = datetime.date.today().year
    print(
        p2.yyear[-2],
        'Estate (in %d $) increased from' % now,
        d(baseValue),
        'to',
        d(newValue),
        '(+%s)' % d(newValue - baseValue),
    )

    u.setVerbose(prevState)

    return p2, bestX


def clone(plan, name=''):
    '''
    Return an almost identical copy of plan.
    Only the name wiil be modified and appended '(copy)',
    unless a new name is provided as an argument.
    '''
    import copy

    newplan = copy.deepcopy(plan)
    if name != '':
        newplan.setName(name)
    else:
        newplan.setName(plan._name + ' (copy)')

    return newplan


def isInJupyter():
    '''
    Boolean function determining if we are in interactive Python.
    '''
    from IPython import get_ipython

    try:
        name = get_ipython().__class__.__name__
        if isinstance(name, type(None)) or name == 'NoneType':
            return False
    except NameError:
        return False

    return True
